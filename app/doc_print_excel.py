# -*- coding: utf-8 -*-
"""全模块 Excel 打印模板适配层（PRINT-TEMPLATE-F03）。

目标（2026-08-25 需求）：所有单据、列表、报表的打印模块都支持
「在线 Excel 格式编辑」（参考简道云 Excel 打印模板）：
- 每种单据/列表/报表都有内置默认 Excel 模板（带 {order.*}/{item.*} 占位符），
  启动时幂等同步进 excel_print_template 表，模板中心可在线编辑；
- 打印时按所选（或默认）模板填充数据生成 .xlsx 下载；
- 无模板时回退各模块原有 HTML/硬编码 Excel 版式，行为不回退。

架构：
- DOC_EXCEL_PRINT_TYPES：单据注册表（表头字段 + 明细列 + 加载器）
- TABLE_EXCEL_PRINT_TYPES：列表/报表注册表（平铺行，{item.*} 逐行填充）
- LABEL_EXCEL_PRINT_TYPES：标签注册表（PRINT-TEMPLATE-F04，每物料一行 +
  {img_barcode:item.*} 条码图片列，打印出口 /label/batch_print_excel）
- generate_builtin_template()：按注册表生成规范的内置 .xlsx 模板
  （标题合并居中 + 表头字段区 + 明细占位符行 + 签名行；明细行下方不做
  合并单元格，避免 openpyxl insert_rows 不移动合并区导致版式错位）
- ensure_builtin_excel_doc_templates()：启动幂等同步（缺行补行、缺文件补文件）
- render_doc_excel_print() / render_table_excel_print()：路由侧统一入口

占位符沿用 print_fill 引擎规则：{order.<attr路径>} / {item.<attr路径>} /
{total_quantity} / {total_amount}（仅模型有 total_amount 的单据）/ {print_date}。
"""
from __future__ import annotations

import io
import os
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# ==================== 单据注册表 ====================
# header: [(标签, order 属性路径)]；columns: [(列名, item 属性路径, 列宽)]
# totals: 'quantity' 输出 {total_quantity}；'amount' 输出 {total_amount}（模型须有 total_amount）

DOC_EXCEL_PRINT_TYPES = {
    'check': {
        'label': '盘点单',
        'header': [('单据编号', 'check_no'), ('日期', 'date'), ('仓库', 'warehouse'),
                   ('状态', 'status'), ('备注', 'remark'), ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'material.unit.name', 6), ('账面库存', 'system_stock', 10),
                    ('实盘库存', 'actual_stock', 10), ('差异', 'difference', 8),
                    ('差异原因', 'reason', 16)],
        'totals': (),
    },
    'transfer': {
        'label': '调拨单',
        'header': [('调拨单号', 'transfer_no'), ('日期', 'date'), ('调出仓库', 'from_warehouse'),
                   ('调入仓库', 'to_warehouse'), ('状态', 'status'), ('备注', 'remark'),
                   ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'unit.name', 6), ('数量', 'quantity', 8),
                    ('单价', 'price', 10), ('金额', 'amount', 10), ('备注', 'remark', 14)],
        'totals': ('quantity',),
    },
    'requisition': {
        'label': '领料申请单',
        'header': [('申请单号', 'req_no'), ('日期', 'date'), ('生产工单', 'production_order'),
                   ('用途', 'purpose'), ('领料人', 'picker'), ('仓库', 'warehouse'),
                   ('状态', 'status'), ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'unit.name', 6), ('申请数量', 'quantity', 10),
                    ('已发数量', 'issued_quantity', 10), ('备注', 'remark', 14)],
        'totals': ('quantity',),
    },
    'purchase_order': {
        'label': '采购订单',
        'header': [('订单号', 'order_no'), ('日期', 'date'), ('供应商', 'supplier.name'),
                   ('联系人', 'supplier.contact'), ('电话', 'supplier.phone'),
                   ('预计到货', 'expected_date'), ('合同编号', 'contract_no'),
                   ('工程名称', 'project_name'), ('备注', 'remark'),
                   ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'material.unit.name', 6), ('数量', 'quantity', 8),
                    ('已入库', 'received_quantity', 8), ('单价', 'price', 10),
                    ('金额', 'amount', 10), ('合同编号', 'contract_no', 12),
                    ('备注', 'remark', 12)],
        'totals': ('quantity', 'amount'),
    },
    'sales_order': {
        'label': '销售订单',
        'header': [('订单号', 'order_no'), ('日期', 'date'), ('客户', 'customer.name'),
                   ('仓库', 'warehouse'), ('交货日期', 'delivery_date'),
                   ('合同编号', 'contract_no'), ('工程名称', 'project_name'),
                   ('备注', 'remark'), ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'material.unit.name', 6), ('数量', 'quantity', 8),
                    ('已发货', 'shipped_quantity', 8), ('单价', 'price', 10),
                    ('金额', 'amount', 10), ('税率%', 'tax_rate', 8),
                    ('合同编号', 'contract_no', 12), ('备注', 'remark', 12)],
        'totals': ('quantity', 'amount'),
    },
    'adjustment': {
        'label': '库存调整单',
        'header': [('调整单号', 'adjustment_no'), ('日期', 'date'), ('调整类型', 'adjustment_type'),
                   ('仓库', 'warehouse'), ('状态', 'status'), ('备注', 'remark'),
                   ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'unit.name', 6), ('库位', 'location', 10),
                    ('调整数量', 'quantity', 10), ('调整原因', 'reason', 18)],
        'totals': ('quantity',),
    },
    'subcontract': {
        'label': '委外加工单',
        'header': [('单号', 'order_no'), ('日期', 'date'), ('加工商', 'supplier.name'),
                   ('联系人', 'contact'), ('电话', 'phone'), ('交期', 'deadline'),
                   ('仓库', 'warehouse'), ('状态', 'status'), ('备注', 'remark')],
        'columns': [('物料编码', 'material.code', 12), ('物料名称', 'material.name', 20),
                    ('规格', 'material.spec', 14), ('单位', 'unit.name', 6),
                    ('数量', 'quantity', 8), ('已回厂', 'returned_quantity', 8),
                    ('损耗', 'loss', 8)],
        'totals': ('quantity', 'amount'),
    },
    'subcontract_issue': {
        'label': '委外发料单',
        'header': [('发料单号', 'issue_no'), ('日期', 'date'),
                   ('委外单号', 'subcontract_order.order_no'), ('加工商', 'supplier.name'),
                   ('仓库', 'warehouse'), ('库位', 'location'), ('状态', 'status'),
                   ('备注', 'remark'), ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('物料名称', 'material.name', 20),
                    ('规格', 'material.spec', 14), ('单位', 'unit.name', 6),
                    ('数量', 'quantity', 8), ('备注', 'remark', 16)],
        'totals': ('quantity',),
    },
    'subcontract_receive': {
        'label': '委外收货单',
        'header': [('收货单号', 'receive_no'), ('日期', 'date'),
                   ('委外单号', 'subcontract_order.order_no'), ('加工商', 'supplier.name'),
                   ('仓库', 'warehouse'), ('状态', 'status'), ('备注', 'remark'),
                   ('操作人', 'operator.username')],
        'columns': [('物料编码', 'material.code', 12), ('物料名称', 'material.name', 20),
                    ('规格', 'material.spec', 14), ('单位', 'unit.name', 6),
                    ('数量', 'quantity', 8), ('报废', 'scrap_quantity', 8),
                    ('单价', 'price', 10), ('金额', 'amount', 10), ('备注', 'remark', 12)],
        'totals': ('quantity',),
    },
    'after_sale_out': {
        'label': '售后出库单',
        'header': [('单号', 'order_no'), ('日期', 'date'), ('客户', 'customer'),
                   ('联系人', 'contact'), ('电话', 'phone'), ('仓库', 'warehouse'),
                   ('售后原因', 'reason'), ('状态', 'status'), ('备注', 'remark')],
        'columns': [('物料编码', 'material.code', 12), ('品牌', 'material.brand', 10),
                    ('物料名称', 'material.name', 20), ('规格', 'material.spec', 14),
                    ('单位', 'material.unit.name', 6), ('数量', 'quantity', 8),
                    ('单价', 'price', 10), ('金额', 'amount', 10),
                    ('合同编号', 'contract_no', 12), ('备注', 'remark', 12)],
        'totals': ('quantity', 'amount'),
    },
}

# 标签注册表（PRINT-TEMPLATE-F04）：物料标签 Excel 模板，每物料一行，
# 条码列用 {img_barcode:item.<路径>} 图片占位符（填充时嵌入 600DPI 条码 PNG）。
# columns: [(列名, 行属性, 列宽)]；barcode_column: (列名, 行属性, 列宽)
LABEL_EXCEL_PRINT_TYPES = {
    'material_label': {
        'label': '物料标签',
        'columns': [('物料编码', 'code', 14), ('物料名称', 'name', 22),
                    ('规格型号', 'spec', 16), ('单位', 'unit_name', 8),
                    ('分类', 'category_name', 12), ('供应商', 'supplier_name', 18),
                    ('单价', 'price', 10), ('当前库存', 'stock', 10)],
        'barcode_column': ('条码', 'barcode', 20),
    },
}

# 列表/报表注册表：行数据为平铺命名空间，{item.<列key>} 逐行填充。
# sheets: [{name, columns:[(列名, 行属性, 列宽)]}]（报表可多工作表）
TABLE_EXCEL_PRINT_TYPES = {
    'stock_query': {
        'target_type': 'list',
        'label': '库存查询列表',
        'sheets': [{
            'name': '库存查询',
            'columns': [('仓库', 'warehouse', 12), ('物料编码', 'code', 12),
                        ('品牌', 'brand', 10), ('物料名称', 'name', 20),
                        ('规格', 'spec', 14), ('单位', 'unit', 6),
                        ('当前库存', 'stock', 10), ('单价', 'price', 10),
                        ('库存金额', 'amount', 12)],
        }],
    },
    'report_stock': {
        'target_type': 'report',
        'label': '库存报表',
        'sheets': [{
            'name': '库存报表',
            'columns': [('仓库', 'warehouse', 12), ('物料编码', 'code', 12),
                        ('物料名称', 'name', 20), ('规格', 'spec', 14),
                        ('单位', 'unit', 6), ('库存数量', 'stock', 10),
                        ('单价', 'price', 10), ('库存金额', 'amount', 12)],
        }],
    },
    'report_inout': {
        'target_type': 'report',
        'label': '出入库统计报表',
        'sheets': [
            {'name': '入库统计',
             'columns': [('单据编号', 'order_no', 14), ('日期', 'date', 12),
                         ('供应商', 'supplier', 16), ('合同编号', 'contract_no', 12),
                         ('工程名称', 'project_name', 14), ('物料编码', 'material_code', 12),
                         ('物料名称', 'material_name', 18), ('数量', 'quantity', 8),
                         ('金额', 'amount', 10)]},
            {'name': '领料统计',
             'columns': [('单据编号', 'order_no', 14), ('日期', 'date', 12),
                         ('领料部门', 'supplier', 16), ('合同编号', 'contract_no', 12),
                         ('工程名称', 'project_name', 14), ('物料编码', 'material_code', 12),
                         ('物料名称', 'material_name', 18), ('数量', 'quantity', 8),
                         ('金额', 'amount', 10)]},
        ],
    },
}

# ==================== 内置模板生成 ====================

_THIN = Side(style='thin', color='000000')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FONT = Font(name='微软雅黑', size=16, bold=True)
_HEADER_FONT = Font(name='微软雅黑', size=11, bold=True)
_BODY_FONT = Font(name='微软雅黑', size=11)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center')


def _write_sheet_layout(ws, title, header, columns, totals):
    """在空白工作表上绘制内置模板版式（标题+表头字段区+明细占位符行+签名行）。

    明细占位符行下方不使用合并单元格（openpyxl insert_rows 不移动合并区）。
    """
    ncols = len(columns)
    for idx, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 标题（合并居中，位于明细行上方，安全）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, title).font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.row_dimensions[1].height = 30

    # 表头字段区：每行 2 对（标签+占位符），不合并，稳妥
    row = 2
    pairs = header or []
    for i in range(0, len(pairs), 2):
        label1, path1 = pairs[i]
        ws.cell(row, 1, f'{label1}：').font = _BODY_FONT
        ws.cell(row, 2, '{order.%s}' % path1).font = _BODY_FONT
        if i + 1 < len(pairs) and ncols >= 4:
            label2, path2 = pairs[i + 1]
            mid = min(4, ncols - 1)
            ws.cell(row, mid, f'{label2}：').font = _BODY_FONT
            ws.cell(row, mid + 1, '{order.%s}' % path2).font = _BODY_FONT
        row += 1

    # 明细表头
    for c, (name, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row, c, name)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    row += 1

    # 明细占位符行（引擎识别的模板行）
    for c, (_, path, _) in enumerate(columns, start=1):
        cell = ws.cell(row, c, '{item.%s}' % path)
        cell.font = _BODY_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    row += 1

    # 合计行（含订单级占位符，同时充当明细块边界行）
    if totals:
        ws.cell(row, 1, '合计').font = _HEADER_FONT
        ws.cell(row, 1).border = _BORDER
        for c, (name, path, _) in enumerate(columns, start=1):
            cell = ws.cell(row, c)
            cell.border = _BORDER
            if 'quantity' in totals and path in ('quantity', 'system_stock'):
                cell.value = '{total_quantity}'
                cell.font = _HEADER_FONT
            elif 'amount' in totals and path == 'amount':
                cell.value = '{total_amount}'
                cell.font = _HEADER_FONT
        row += 1

    # 签名行（订单级占位符，无合并）
    ws.cell(row, 1, '制单：{order.operator.username}').font = _BODY_FONT
    if ncols >= 4:
        ws.cell(row, min(4, ncols), '打印日期：{print_date}').font = _BODY_FONT
    return ws


def _write_label_sheet_layout(ws, spec):
    """绘制内置标签模板版式：标题 + 字段表头 + 明细占位符行（含条码图片列）。"""
    barcode_col = spec.get('barcode_column')
    columns = list(spec['columns']) + ([barcode_col] if barcode_col else [])
    ncols = len(columns)
    for idx, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, spec['label']).font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.row_dimensions[1].height = 30

    for c, (name, _, _) in enumerate(columns, start=1):
        cell = ws.cell(2, c, name)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    for c, (_, path, _) in enumerate(columns, start=1):
        if barcode_col and path == barcode_col[1]:
            value = '{img_barcode:item.%s}' % path
        else:
            value = '{item.%s}' % path
        cell = ws.cell(3, c, value)
        cell.font = _BODY_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[3].height = 48  # 为条码图片预留行高
    return ws


def generate_builtin_template(target_code):
    """按注册表生成内置默认 .xlsx 模板字节流；未注册的 target_code 返回 None。"""
    if target_code in DOC_EXCEL_PRINT_TYPES:
        spec = DOC_EXCEL_PRINT_TYPES[target_code]
        wb = Workbook()
        ws = wb.active
        ws.title = spec['label']
        _write_sheet_layout(ws, spec['label'], spec['header'], spec['columns'],
                            spec.get('totals') or ())
    elif target_code in LABEL_EXCEL_PRINT_TYPES:
        spec = LABEL_EXCEL_PRINT_TYPES[target_code]
        wb = Workbook()
        ws = wb.active
        ws.title = spec['label']
        _write_label_sheet_layout(ws, spec)
    elif target_code in TABLE_EXCEL_PRINT_TYPES:
        spec = TABLE_EXCEL_PRINT_TYPES[target_code]
        wb = Workbook()
        first = True
        for sheet in spec['sheets']:
            ws = wb.active if first else wb.create_sheet()
            first = False
            ws.title = sheet['name']
            _write_sheet_layout(ws, spec['label'], [], sheet['columns'], ())
    else:
        return None
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _builtin_template_abspath(static_folder, target_code):
    """返回内置模板文件绝对路径；不存在则按注册表生成（幂等）。"""
    rel_dir = os.path.join('uploads', 'print_templates')
    abs_dir = os.path.join(static_folder, rel_dir)
    os.makedirs(abs_dir, exist_ok=True)
    path = os.path.join(abs_dir, f'builtin_{target_code}_default.xlsx')
    if not os.path.exists(path):
        content = generate_builtin_template(target_code)
        if content is None:
            return None
        with open(path, 'wb') as f:
            f.write(content.read())
    return path


# ==================== 启动幂等同步 ====================

def _default_static_folder():
    """app 包内的 static 目录（不依赖 Flask current_app，模块导入期可用）。"""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')


def _default_db_path():
    """解析实际 sqlite 数据库文件路径（不依赖 Flask 应用上下文）。

    优先复用 app._resolve_sqlite_db_path（app.py 模块执行到本函数接线点时，
    该函数已定义，from-import 安全）；失败回退 app 包旁 instance/inventory.db。
    非 sqlite / 内存库返回 None（无需同步）。
    """
    try:
        from app import _resolve_sqlite_db_path
        path = _resolve_sqlite_db_path()
        if path:
            return path
    except Exception:  # noqa: BLE001 - app 尚未就绪时走回退
        pass
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'instance', 'inventory.db')


def ensure_builtin_excel_doc_templates(db_path=None, static_folder=None):
    """为所有已注册的单据/列表/报表补齐内置默认 Excel 模板（幂等）。

    仿照 app._ensure_default_print_templates_unconditional：独立 raw sqlite
    连接、独立于迁移开关无条件执行——start_wms_offline.bat 默认
    WMS_NO_DB_TOUCH=1 会跳过 initialize_database 里的 ORM 版同步，若在
    模块导入期用 ORM/current_app 会直接失败，导致线上永远没有内置模板。

    同步规则：
    - 该 target_code 没有任何模板记录 → 生成内置文件 + 建行（is_default=1）
    - 有记录但 excel_template_path 为空，或指向 builtin_ 副本但文件丢失
      → 补文件并回指内置副本
    - 组内无 is_default=1 → 把第一条记录置为默认
    用户自建模板（非 builtin_ 前缀且文件存在）一律不动。
    任何失败仅记日志，不阻断启动（打印路由仍有各模块原有回退版式）。
    """
    import logging
    import sqlite3

    conn = None
    try:
        if static_folder is None:
            static_folder = _default_static_folder()
        if db_path is None:
            db_path = _default_db_path()
        if not db_path or not os.path.exists(db_path):
            return

        registry = [(code, 'document', spec['label'])
                    for code, spec in DOC_EXCEL_PRINT_TYPES.items()]
        registry += [(code, spec['target_type'], spec['label'])
                     for code, spec in TABLE_EXCEL_PRINT_TYPES.items()]
        registry += [(code, 'label', spec['label'])
                     for code, spec in LABEL_EXCEL_PRINT_TYPES.items()]

        conn = sqlite3.connect(db_path, timeout=60)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute('PRAGMA busy_timeout=60000')
        tbl = cur.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='excel_print_template'"
        ).fetchone()
        if tbl is None:
            return

        from print_fill import template_file_abspath

        changed = False
        for code, target_type, label in registry:
            path = _builtin_template_abspath(static_folder, code)
            if not path:
                continue
            url_path = '/static/uploads/print_templates/' + os.path.basename(path)
            rows = cur.execute(
                "SELECT id, excel_template_path, is_default FROM excel_print_template"
                " WHERE target_type=? AND target_code=? ORDER BY id",
                (target_type, code),
            ).fetchall()
            if not rows:
                cur.execute(
                    "INSERT INTO excel_print_template"
                    " (name, target_type, target_code, template_type,"
                    "  excel_template_path, is_default, created_at, updated_at)"
                    " VALUES (?, ?, ?, 'excel', ?, 1,"
                    "         datetime('now'), datetime('now'))",
                    (f'系统默认{label}模板', target_type, code, url_path),
                )
                changed = True
                continue
            has_default = any(r['is_default'] for r in rows)
            for row in rows:
                cur_path = (row['excel_template_path'] or '').strip()
                need_repoint = False
                if not cur_path:
                    need_repoint = True
                elif cur_path.startswith('/static/uploads/print_templates/builtin_'):
                    # 内置副本丢失时回指重新生成的内置模板（用户自建模板不动）
                    existing = template_file_abspath(cur_path, static_folder)
                    if not existing or not os.path.exists(existing):
                        need_repoint = True
                if need_repoint:
                    cur.execute(
                        "UPDATE excel_print_template SET excel_template_path=?,"
                        " updated_at=datetime('now') WHERE id=?",
                        (url_path, row['id']),
                    )
                    changed = True
                if not has_default:
                    cur.execute(
                        "UPDATE excel_print_template SET is_default=1,"
                        " updated_at=datetime('now') WHERE id=?",
                        (row['id'],),
                    )
                    has_default = True
                    changed = True
        if changed:
            conn.commit()
            logging.getLogger(__name__).info(
                '[DB] 全模块内置 Excel 打印模板已同步（%d 个目标）', len(registry))
    except Exception as e:  # noqa: BLE001 - 同步失败不阻断启动
        try:
            logging.getLogger(__name__).error(
                'ensure_builtin_excel_doc_templates 同步失败: %s', e, exc_info=True)
        except Exception:  # noqa: BLE001
            pass
        if conn is not None:
            try:
                conn.rollback()
            except Exception:  # noqa: BLE001
                pass
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass


# ==================== 打印渲染 ====================

def resolve_excel_template(target_type, target_code, template_id=None):
    """返回指定（或默认）Excel 模板记录；无记录返回 None。

    ORM 查询失败（无应用上下文、表缺失等）时返回 None，调用方回退内置模板，
    保证打印路由在任何环境下不 500。
    """
    try:
        from app import ExcelPrintTemplate
        query = ExcelPrintTemplate.query.filter_by(
            target_type=target_type, target_code=target_code)
        if template_id:
            template = query.filter_by(id=template_id).first()
            if template:
                return template
        return query.filter_by(is_default=True).first() or \
            query.order_by(ExcelPrintTemplate.updated_at.desc()).first()
    except Exception:  # noqa: BLE001 - 查询失败回退内置模板
        return None


def _template_path_or_builtin(template, static_folder, target_code):
    """模板记录的文件路径；记录缺失/文件丢失时回退内置模板文件。"""
    from print_fill import template_file_abspath
    if template and template.excel_template_path:
        path = template_file_abspath(template.excel_template_path, static_folder)
        if path and os.path.exists(path):
            return path
    return _builtin_template_abspath(static_folder, target_code)


def render_doc_excel_print(target_code, order, items=None, template_id=None,
                           static_folder=None, date_str=None):
    """按注册表填充单据 Excel 模板，返回 (BytesIO, 文件名)；未注册返回 None。"""
    spec = DOC_EXCEL_PRINT_TYPES.get(target_code)
    if spec is None or order is None:
        return None
    from datetime import datetime
    from print_fill import build_filled_print_excel
    path = _template_path_or_builtin(
        resolve_excel_template('document', target_code, template_id),
        static_folder, target_code)
    if not path:
        return None
    output = build_filled_print_excel(
        path, order,
        items=items if items is not None else getattr(order, 'items', None),
        date_str=date_str or datetime.now().strftime('%Y-%m-%d'),
    )
    order_no = ''
    for attr in ('order_no', 'check_no', 'transfer_no', 'req_no',
                 'adjustment_no', 'issue_no', 'receive_no'):
        value = getattr(order, attr, None)
        if value:
            order_no = value
            break
    filename = f"{spec['label']}_{order_no or getattr(order, 'id', '')}.xlsx"
    return output, filename


def render_table_excel_print(target_code, sheet_rows, template_id=None,
                             static_folder=None, date_str=None, title_extra=None):
    """按注册表填充列表/报表 Excel 模板。

    sheet_rows: {工作表名: [SimpleNamespace 行]}；单表模板也可直接传行列表。
    返回 (BytesIO, 文件名)；未注册返回 None。
    """
    spec = TABLE_EXCEL_PRINT_TYPES.get(target_code)
    if spec is None:
        return None
    from datetime import datetime
    from openpyxl import load_workbook
    from print_fill import _Filler  # 复用引擎内核（同包私有工具）
    path = _template_path_or_builtin(
        resolve_excel_template(spec['target_type'], target_code, template_id),
        static_folder, target_code)
    if not path:
        return None
    if isinstance(sheet_rows, (list, tuple)):
        sheet_rows = {spec['sheets'][0]['name']: sheet_rows}
    order = SimpleNamespace(
        title=(spec['label'] + (f'（{title_extra}）' if title_extra else '')),
        operator=SimpleNamespace(username=''),
    )
    wb = load_workbook(path)
    filler_rows = sheet_rows or {}
    for ws in wb.worksheets:
        _Filler(order, filler_rows.get(ws.title, []),
                date_str or datetime.now().strftime('%Y-%m-%d')).fill(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{spec['label']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


# ==================== 报表通用模板打印（PRINT-TEMPLATE-F04 A4） ====================
#
# 每种报表（REPORT_DEFINITIONS）都可在线设计 Excel 模板：
# - target_code = report_<报表类型>，target_type = 'report'；
# - 内置模板按报表 columns 动态生成（表头标题 + 列名行 + {item.<field>} 行），
#   首次打印时由路由侧自动登记到 excel_print_template 表，之后即可在模板
#   中心在线编辑/上传变体；
# - 无用户模板时回退动态内置模板，行为与通用导出一致（全量行）。

def report_target_code(report_type):
    """报表类型 → 模板 target_code。"""
    return f'report_{report_type}'


def generate_report_builtin_template(title, columns):
    """按报表列定义动态生成内置 .xlsx 模板字节流。

    columns: [{'field': ..., 'title': ...}]（REPORT 构建器返回结构）。
    """
    specs = []
    for col in columns or []:
        field = (col.get('field') or '').strip()
        if not field:
            continue
        col_title = (col.get('title') or field).strip()
        width = min(30, max(10, len(col_title) * 2 + 4))
        specs.append((col_title, field, width))
    if not specs:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = (title or '报表')[:31]
    _write_sheet_layout(ws, title or '报表', [], specs, ())
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _report_builtin_abspath(static_folder, report_type, title, columns):
    """报表动态内置模板文件绝对路径；不存在则按 columns 生成（幂等）。"""
    if not static_folder:
        return None
    abs_dir = os.path.join(static_folder, 'uploads', 'print_templates')
    os.makedirs(abs_dir, exist_ok=True)
    path = os.path.join(abs_dir, f'builtin_report_{report_type}_default.xlsx')
    if not os.path.exists(path):
        content = generate_report_builtin_template(title, columns)
        if content is None:
            return None
        with open(path, 'wb') as f:
            f.write(content.read())
    return path


def render_report_excel_print(report_type, title, columns, rows,
                              template_id=None, static_folder=None,
                              date_str=None):
    """按模板填充报表 Excel（target_code=report_<报表类型>）。

    rows: dict 行列表（键与 columns 的 field 对齐；引擎支持 dict 取值）。
    返回 (BytesIO, 文件名, 模板文件绝对路径)；生成失败返回 None。
    第三个返回值供路由侧把动态内置模板登记进模板中心（幂等）。
    """
    from datetime import datetime
    from openpyxl import load_workbook
    from print_fill import _Filler, template_file_abspath  # 同包私有工具复用
    target_code = report_target_code(report_type)
    template = resolve_excel_template('report', target_code, template_id)
    path = None
    if template and template.excel_template_path:
        candidate = template_file_abspath(template.excel_template_path,
                                          static_folder)
        if candidate and os.path.exists(candidate):
            path = candidate
    if path is None:
        path = _report_builtin_abspath(static_folder, report_type, title,
                                       columns)
    if not path:
        return None
    order = SimpleNamespace(
        title=title or '报表', operator=SimpleNamespace(username=''),
        total_amount=0)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        _Filler(order, list(rows or []),
                date_str or datetime.now().strftime('%Y-%m-%d')).fill(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{title or report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename, path


def render_label_excel_print(target_code, rows, template_id=None,
                             static_folder=None, date_str=None):
    """按注册表填充标签 Excel 模板（每物料一行，含条码图片）。

    rows: dict 或 SimpleNamespace 行（字段与 LABEL 注册表 columns 对齐：
    code/name/spec/unit_name/category_name/supplier_name/stock/price/barcode）。
    返回 (BytesIO, 文件名)；未注册返回 None。
    """
    spec = LABEL_EXCEL_PRINT_TYPES.get(target_code)
    if spec is None:
        return None
    from datetime import datetime
    from openpyxl import load_workbook
    from print_fill import _Filler  # 复用引擎内核（同包私有工具）
    path = _template_path_or_builtin(
        resolve_excel_template('label', target_code, template_id),
        static_folder, target_code)
    if not path:
        return None
    normalized = []
    for row in rows or []:
        if isinstance(row, dict):
            normalized.append(SimpleNamespace(**row))
        else:
            normalized.append(row)
    order = SimpleNamespace(
        title=spec['label'], operator=SimpleNamespace(username=''),
        total_amount=0)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        _Filler(order, normalized,
                date_str or datetime.now().strftime('%Y-%m-%d')).fill(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{spec['label']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename
