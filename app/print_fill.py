# -*- coding: utf-8 -*-
"""Excel 打印模板填充引擎（PRINT-TEMPLATE-F01）。

打印模板仅允许 Excel：用户上传带 `{占位符}` 的 .xlsx 模板，打印时选择模板后，
本模块把单据数据填充进模板并生成可下载的 .xlsx。

支持的占位符：
- 单据级：{order.order_no}、{order.date}、{order.supplier.name}、
  {order.supplier.contact}、{order.supplier.phone}、{order.supplier.address}、
  {order.customer}、{order.purpose}、{order.remark}、{order.picker}、
  {order.operator.username}、{total_quantity}、{total_amount}、{print_date}
- 明细级：{item.material.code}、{item.material.name}、{item.material.spec}、
  {item.material.brand}、{item.material.unit.name}、{item.quantity}、
  {item.price}、{item.amount}、{item.contract_no}、{item.project_name}、
  {item.remark}
- 图片级（PRINT-TEMPLATE-F04，标签/单据条码）：
  {img_barcode:item.barcode}、{img_qrcode:item.code}、{img_barcode:order.order_no}
  ——单元格内容恰好是图片占位符时，填充为 600DPI 条码/二维码 PNG 图片；
  数据为空或图片生成失败时回退为数据文本，绝不输出占位符原文

模板内第一处含 `{item.` 的行视为"明细模板行"，其下方到首个含订单级占位符
（如 {total_*} / {order.} / {print_date}）之前的行为"示例数据行"。填充时按实际
明细条数在明细块上复制/删除行，并从明细模板行复制样式与会合计性保证版式一致。
"""
from __future__ import annotations

import io
import os
import re
import zipfile
from copy import copy
from datetime import datetime

from openpyxl import load_workbook

_PLACEHOLDER_RE = re.compile(r'\{([^{}]+)\}')
# 明细扩展时判断"明细块结束边界"用到的订单级占位符前缀
_ORDER_LEVEL_HINTS = ('{total_', '{order.', '{print_date}', '{today}')

# 合法占位符：order.* / item.* 为通配（按属性路径解析），其余为显式白名单
_EXPLICIT_TOKENS = ('total_quantity', 'total_amount', 'print_date', 'today')

# 图片占位符（PRINT-TEMPLATE-F04）：{img_barcode:item.barcode} /
# {img_qrcode:item.code} / {img_barcode:order.order_no}，填充时嵌入 PNG 图片
_IMG_PLACEHOLDER_RE = re.compile(
    r'^img_(barcode|qrcode):(item|order)\.[A-Za-z0-9_]+(\.[A-Za-z0-9_]+)*$')
_IMG_ITEM_HINTS = ('{img_barcode:item.', '{img_qrcode:item.')

# 上传时防超大/解压炸弹/畸形文件的安全阈值
_MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024  # 50MB 解压后上限
_MAX_TEMPLATE_ROWS = 2000
_MAX_TEMPLATE_COLS = 100


def _is_supported_placeholder(token):
    if token.startswith('order.') or token.startswith('item.') \
            or token in _EXPLICIT_TOKENS:
        return True
    return bool(_IMG_PLACEHOLDER_RE.match(token))


def _has_item_placeholder(text):
    """文本是否含明细级占位符（含图片型），用于定位明细模板行。"""
    if not isinstance(text, str):
        return False
    if '{item.' in text:
        return True
    return any(hint in text for hint in _IMG_ITEM_HINTS)


def _match_image_cell(text):
    """单元格内容恰好是一个图片占位符时返回 (kind, scope, attr_path)，否则 None。"""
    if not isinstance(text, str):
        return None
    full = _PLACEHOLDER_RE.fullmatch(text.strip())
    if not full:
        return None
    match = _IMG_PLACEHOLDER_RE.match(full.group(1))
    if not match:
        return None
    path = full.group(1).split(':', 1)[1].split('.', 1)[1]
    return match.group(1), match.group(2), path


def _render_placeholder_image(kind, data):
    """把图片占位符的数据渲染为 PIL Image；数据为空/生成失败返回 None。"""
    text = str(data or '').strip()
    if not text:
        return None
    try:
        buf = io.BytesIO()
        if kind == 'barcode':
            import barcode
            from barcode.writer import ImageWriter
            code128 = barcode.get_barcode_class('code128')
            code128(text, writer=ImageWriter()).write(buf, {
                'module_width': 0.33, 'module_height': 15.0, 'font_size': 10,
                'text_distance': 5, 'quiet_zone': 2, 'dpi': 600})
        else:
            import qrcode
            qrcode.make(text).save(buf, format='PNG')
        buf.seek(0)
        from PIL import Image as PILImage
        return PILImage.open(buf)
    except Exception:  # noqa: BLE001 图片生成失败一律回退文本
        return None


def _embed_placeholder_image(ws, row, col, kind, data, fallback_text):
    """在指定单元格嵌入条码/二维码图片；失败时回退写入数据文本。"""
    cell = ws.cell(row, col)
    image = _render_placeholder_image(kind, data)
    if image is None:
        cell.value = fallback_text
        return
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    xl_image = XLImage(image)
    if kind == 'barcode':
        xl_image.width, xl_image.height = 180, 48
    else:
        xl_image.width = xl_image.height = 72
    cell.value = None
    ws.add_image(xl_image, f'{get_column_letter(col)}{row}')
    if ws.row_dimensions[row].height is None:
        ws.row_dimensions[row].height = 40 if kind == 'barcode' else 58


def validate_template_file(raw: bytes) -> str:
    """校验上传的打印模板原始字节，返回错误信息；合法则返回空串。

    覆盖三类问题：
    - 损坏/非 xlsx（openpyxl 打开失败或解压非法）→ 明确报错而非运行时失败
    - 超大文件 / zip 解压炸弹（解压后体积超限）→ 拒绝
    - 模板里含引擎不支持的 {占位符} → 拒绝，避免打印时静默输出原文
    """
    if not raw:
        return '打印模板文件为空'
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
        total = sum(i.file_size for i in zf.infolist())
        zf.close()
    except zipfile.BadZipFile:
        return '不是有效的 Excel 文件（仅支持 .xlsx）'
    if total > _MAX_UNCOMPRESSED_BYTES:
        return '打印模板文件解压后过大，请压缩后重新上传'
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=False)
    except zipfile.BadZipFile:
        return '不是有效的 Excel 文件（仅支持 .xlsx）'
    except Exception:  # noqa: BLE001 开放异常：解析层面的任何失败都按坏文件处理
        return 'Excel 文件损坏或无法解析，请重新上传有效的 .xlsx 模板'
    try:
        for ws in workbook.worksheets:
            if (ws.max_row or 0) > _MAX_TEMPLATE_ROWS or (ws.max_column or 0) > _MAX_TEMPLATE_COLS:
                return f'工作表「{ws.title}」规模过大（行>%d 或列>%d），请精简模板' % (
                    _MAX_TEMPLATE_ROWS, _MAX_TEMPLATE_COLS)
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for token in _PLACEHOLDER_RE.findall(cell.value):
                            if not _is_supported_placeholder(token):
                                return '模板包含不支持的占位符 {%s}，请使用 order.* / item.* / %s' % (
                                    token, ' / '.join(_EXPLICIT_TOKENS))
    finally:
        workbook.close()
    return ''


def template_file_abspath(excel_template_path, static_folder):
    """把模板记录的 excel_template_path（静态 URL）解析为绝对文件路径。"""
    if not excel_template_path:
        return None
    p = excel_template_path.replace('\\', '/')
    if p.startswith('/static/'):
        rel = p[len('/static/'):]
    elif p.startswith('uploads/'):
        rel = p
    else:
        # 历史/兜底：仅保留文件名放入 uploads/print_templates
        rel = os.path.join('uploads', 'print_templates', os.path.basename(p))
    return os.path.join(static_folder, rel)


def _num(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _fmt(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return '是' if value else '否'
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # 保留数值，供 Excel 计算与对齐
        return value
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)


def _resolve_path(obj, parts):
    for part in parts:
        if obj is None:
            return None
        if isinstance(obj, dict):
            # 报表/列表平铺 dict 行（PRINT-TEMPLATE-F04 报表模板打印）
            attr = obj.get(part)
        else:
            attr = getattr(obj, part, None)
        if callable(attr):
            attr = attr()
        obj = attr
    return obj


class _Filler:
    """对单个工作表执行占位符填充与明细行扩展。"""

    def __init__(self, order, items, date_str):
        self.order = order
        self.items = items or []
        self.date_str = date_str or datetime.now().strftime('%Y-%m-%d')

    # ---------- 解析 ----------
    def order_value(self, token):
        """解析订单级占位符；明细/图片占位符返回 _KEEP 哨兵。"""
        if token.startswith('item.'):
            return _KEEP
        if _IMG_PLACEHOLDER_RE.match(token):
            return _KEEP  # 图片占位符由 _embed 阶段单独处理
        if token == 'total_quantity':
            return _fmt(sum(_num(i.quantity) for i in self.items))
        if token == 'total_amount':
            return self.order.total_amount
        if token in ('print_date', 'today'):
            return self.date_str
        if token.startswith('order.'):
            return _fmt(_resolve_path(self.order, token.split('.')[1:]))
        return _fmt(_resolve_path(self.order, [token]))

    def item_value(self, token, item):
        if token.startswith('item.'):
            return _fmt(_resolve_path(item, token.split('.')[1:]))
        if _IMG_PLACEHOLDER_RE.match(token):
            return _KEEP  # 图片占位符由 _embed 阶段单独处理
        return self.order_value(token)

    @staticmethod
    def _replace_cell(text, resolver):
        """替换一个单元格中的占位符。

        整个单元格恰好是一个占位符时返回原始解析值，保留数值类型；
        否则在文本中逐个替换，未命中的 (item 或缺失) 保留原占位符文本。
        """
        if not isinstance(text, str):
            return text
        stripped = text.strip()
        full = _PLACEHOLDER_RE.fullmatch(stripped)
        if full:
            value = resolver(full.group(1))
            if value is _KEEP or value is None:
                return text
            return value

        def _repl(m):
            value = resolver(m.group(1))
            if value is _KEEP or value is None:
                return m.group(0)
            return str(value)

        return _PLACEHOLDER_RE.sub(_repl, text)

    # ---------- 填充 ----------
    def fill(self, ws):
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # 1) 定位"明细模板行"
        template_row = None
        col_cells = {}
        for r in range(1, max_row + 1):
            has_item = any(
                _has_item_placeholder(ws.cell(r, c).value)
                for c in range(1, max_col + 1)
            )
            if has_item:
                template_row = r
                break

        # 无明细占位符：纯订单级模板
        if template_row is None:
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, str):
                        cell.value = self._replace_cell(cell.value, self.order_value)
            self._embed_order_images(ws)
            return

        for c in range(1, max_col + 1):
            v = ws.cell(template_row, c).value
            if _has_item_placeholder(v):
                col_cells[c] = v

        # 边界须在订单级填充（step 2）之前基于原始模板计算：
        # 因为 step 2 会先把边界行的 {total_*}/{order.} 占位符替换成数值，
        # 若之后才探测边界，原始占位符已消失，无法识别明细块的结束行。
        boundary = self._find_boundary(ws, template_row, max_row, max_col)

        # 2) 订单级填充（跳过明细模板行；含 item 占位符的单元格仍可填充其中的订单级占位符）
        for r in range(1, max_row + 1):
            if r == template_row:
                continue
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                if isinstance(cell.value, str):
                    cell.value = self._replace_cell(cell.value, self.order_value)

        # 3) 明细行扩展
        sample_count = max(0, boundary - template_row - 1 if boundary else 0)
        item_count = len(self.items)

        if item_count == 0:
            for c, orig in col_cells.items():
                ws.cell(template_row, c).value = self._replace_cell(orig, lambda t: '')
            self._embed_order_images(ws)
            return

        capacity = sample_count + 1
        if item_count < capacity:
            # 删除多余的示例行
            ws.delete_rows(template_row + item_count, capacity - item_count)
        elif item_count > capacity:
            # 在示例块之后（边界行之前）插入缺失行
            ws.insert_rows(template_row + sample_count + 1, item_count - capacity)

        self._write_items(ws, template_row, col_cells, item_count)
        self._embed_order_images(ws)

    def _embed_order_images(self, ws):
        """嵌入工作表中残留的订单级图片占位符（行扩展完成后调用）。"""
        for row in ws.iter_rows():
            for cell in row:
                parsed = _match_image_cell(cell.value)
                if not parsed or parsed[1] != 'order':
                    continue
                kind, _, path = parsed
                data = _resolve_path(self.order, path.split('.'))
                _embed_placeholder_image(
                    ws, cell.row, cell.column, kind, data, _fmt(data))

    def _find_boundary(self, ws, template_row, max_row, max_col):
        """返回明细块结束边界的第一个订单级占位符行号（不含）；找不到返回 None。"""
        for r in range(template_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                lowered = v
                if any(hint in lowered for hint in _ORDER_LEVEL_HINTS):
                    return r
                # 含非 item 占位符也算边界（更低优先，避免误吞纯文本示例行）
                if _contains_order_placeholder(v):
                    return r
        return None

    def _write_items(self, ws, template_row, col_cells, item_count):
        template_height = ws.row_dimensions[template_row].height
        for index in range(item_count):
            item = self.items[index]
            row_idx = template_row + index
            if template_height:
                ws.row_dimensions[row_idx].height = template_height
            for c, orig in col_cells.items():
                cell = ws.cell(row_idx, c)
                parsed = _match_image_cell(orig)
                if parsed and parsed[1] == 'item':
                    # 明细级图片占位符：按本行明细数据嵌入条码/二维码
                    kind, _, path = parsed
                    data = _resolve_path(item, path.split('.'))
                    _embed_placeholder_image(
                        ws, row_idx, c, kind, data, _fmt(data))
                else:
                    cell.value = self._replace_cell(
                        orig, lambda t, item=item: self.item_value(t, item)
                    )
                source = ws.cell(template_row, c)
                cell._style = copy(source._style)


def _contains_order_placeholder(text):
    """文本是否含非 item 的占位符（用于边界探测的兜底）。"""
    if not isinstance(text, str):
        return False
    for token in _PLACEHOLDER_RE.findall(text):
        if not token.startswith('item.'):
            return True
    return False


_KEEP = object()


def build_filled_print_excel(template_path, order, items=None, date_str=None):
    """按订单数据填充 Excel 模板，返回可下载的 BytesIO。

    template_path：模板 .xlsx 的绝对路径；
    order：InOrder 或 OutOrder（含 items / supplier / operator 等关系）；
    items：默认取 order.items，也可显式传入。
    """
    workbook = load_workbook(template_path)
    filler = _Filler(order, items if items is not None else order.items, date_str)
    for ws in workbook.worksheets:
        filler.fill(ws)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


# ==================== 在线网格编辑（PRINT-TEMPLATE-F02） ====================
#
# 浏览器"在线修改打印模板"：把 .xlsx 模板序列化为 JSON 网格供前端编辑，
# 前端改完单元格值后回写。合并单元格只在左上角锚点处可编辑，样式/合并保持
# 原样，不在此处增删——避免前端破坏版式。

_SCALAR_TYPES = (str, int, float, bool)


def _cell_to_json(value):
    """把单元格值转为可 JSON 序列化的标量；其他类型（公式对象等）转字符串。"""
    if isinstance(value, _SCALAR_TYPES) and value is not None:
        if isinstance(value, bool):
            return value
        if isinstance(value, float):
            from math import isfinite
            return value if isfinite(value) else None
        return value
    if value is None:
        return None
    return str(value)


def _sheet_merge_anchors(ws):
    """返回该工作表所有合并区域的锚点（左上角单元格）集合 {(min_row, min_col)}。"""
    anchors = set()
    for rng in ws.merged_cells.ranges:
        anchors.add((rng.min_row, rng.min_col))
    return anchors


# ==================== 网格样式序列化（PRINT-TEMPLATE-F05-A1） ====================

_GRID_COLOR_RE = re.compile(r'^[0-9A-Fa-f]{6}$')
_GRID_H_ALIGN = ('left', 'center', 'right')
_GRID_V_ALIGN = ('top', 'center', 'bottom')
_GRID_BORDER_EDGES = ('top', 'right', 'bottom', 'left')


def _color_to_hex(color):
    """openpyxl Color → '#RRGGBB'；theme/indexed/非法值返回 None。"""
    if color is None or getattr(color, 'type', None) != 'rgb':
        return None
    rgb = color.rgb
    if not isinstance(rgb, str):
        return None
    if len(rgb) == 8:
        rgb = rgb[2:]
    if not _GRID_COLOR_RE.match(rgb):
        return None
    return '#' + rgb.upper()


def _cell_style_to_json(cell):
    """提取单元格样式为网格 JSON style 对象；无显式样式返回 None。

    输出键（全可选）：bold/italic/underline、font_name、font_size、
    font_color/bg_color（#RRGGBB）、h_align/v_align、wrap、
    border={top/right/bottom/left: 边样式}。
    """
    if not getattr(cell, 'has_style', False):
        return None
    style = {}
    font = cell.font
    if font is not None:
        if font.bold:
            style['bold'] = True
        if font.italic:
            style['italic'] = True
        if font.underline:
            style['underline'] = True
        if font.name:
            style['font_name'] = str(font.name)
        if font.size:
            try:
                style['font_size'] = float(font.size)
            except (TypeError, ValueError):
                pass
        font_color = _color_to_hex(font.color)
        if font_color:
            style['font_color'] = font_color
    fill = cell.fill
    if fill is not None and getattr(fill, 'patternType', None) == 'solid':
        bg_color = _color_to_hex(fill.fgColor)
        if bg_color:
            style['bg_color'] = bg_color
    alignment = cell.alignment
    if alignment is not None:
        if alignment.horizontal in _GRID_H_ALIGN:
            style['h_align'] = alignment.horizontal
        if alignment.vertical in _GRID_V_ALIGN:
            style['v_align'] = alignment.vertical
        if alignment.wrap_text:
            style['wrap'] = True
    border = cell.border
    if border is not None:
        edges = {}
        for edge in _GRID_BORDER_EDGES:
            side = getattr(border, edge, None)
            side_style = getattr(side, 'style', None) if side is not None else None
            if side_style:
                edges[edge] = side_style
        if edges:
            style['border'] = edges
    return style or None


def _sheet_dimensions_to_json(ws):
    """提取显式设置的列宽/行高：{列号: 宽} / {行号: 高}（仅 customWidth/显式行高）。"""
    from openpyxl.utils import column_index_from_string
    col_widths = {}
    for letter, dim in (ws.column_dimensions or {}).items():
        width = getattr(dim, 'width', None)
        if width is None or not getattr(dim, 'customWidth', False):
            continue
        try:
            col_widths[str(column_index_from_string(letter))] = float(width)
        except (TypeError, ValueError):
            continue
    row_heights = {}
    for row, dim in (ws.row_dimensions or {}).items():
        height = getattr(dim, 'height', None)
        if height is None:
            continue
        try:
            row_heights[str(int(row))] = float(height)
        except (TypeError, ValueError):
            continue
    return col_widths, row_heights


def serialize_print_template_grid(template_path):
    """把打印模板 .xlsx 序列化为浏览器网格 JSON。

    返回：{sheets: [{name, max_row, max_col, merges, cells, col_widths, row_heights}]}
    - merges: [{row, col, rowspan, colspan}]（仅给前端展示只读提示，编辑锚点=左上角）
    - cells: [{row, col, value, merged, style?}]，含非空单元格与带显式样式的
      单元格（如合计行的空值边框格）；合并区仅在锚点给出。
    - col_widths/row_heights: {列号/行号: 数值}，仅显式设置过的尺寸。
    style 结构见 `_cell_style_to_json`（PRINT-TEMPLATE-F05-A1 新增，向后兼容：
    旧前端不读 style/col_widths/row_heights 字段即忽略）。
    """
    workbook = load_workbook(template_path, data_only=False)
    try:
        sheets = []
        for ws in workbook.worksheets:
            anchors = _sheet_merge_anchors(ws)
            merges = []
            for rng in ws.merged_cells.ranges:
                merges.append({
                    'row': rng.min_row,
                    'col': rng.min_col,
                    'rowspan': rng.max_row - rng.min_row + 1,
                    'colspan': rng.max_col - rng.min_col + 1,
                })
            cells = []
            for r in range(1, (ws.max_row or 1) + 1):
                for c in range(1, (ws.max_column or 1) + 1):
                    if (r, c) in anchors:
                        continue  # 非锚点合并格，值在锚点处读取
                    cell = ws.cell(r, c)
                    value = _cell_to_json(cell.value)
                    style = _cell_style_to_json(cell)
                    if (value is None or value == '') and not style:
                        continue
                    item = {
                        'row': r, 'col': c, 'value': value,
                        'merged': False,
                    }
                    if style:
                        item['style'] = style
                    cells.append(item)
            # 合并锚点单元格（值为本身就是该区域的值）
            for (r, c) in sorted(anchors):
                anchor_cell = ws.cell(r, c)
                value = _cell_to_json(anchor_cell.value)
                item = {'row': r, 'col': c,
                        'value': value if value is not None else '',
                        'merged': True}
                style = _cell_style_to_json(anchor_cell)
                if style:
                    item['style'] = style
                cells.append(item)
            cells.sort(key=lambda x: (x['row'], x['col']))
            col_widths, row_heights = _sheet_dimensions_to_json(ws)
            sheets.append({
                'name': ws.title,
                'max_row': ws.max_row or 1,
                'max_col': ws.max_column or 1,
                'merges': merges,
                'cells': cells,
                'col_widths': col_widths,
                'row_heights': row_heights,
            })
        return {'sheets': sheets}
    finally:
        workbook.close()


def _validate_grid_value(value):
    """校验网格待回写值，返回错误信息；合法返回空串。"""
    if value is None:
        return ''
    if isinstance(value, bool):
        return ''
    if isinstance(value, (int, float)):
        return ''
    if isinstance(value, str):
        if len(value) > 32767:
            return '单元格文本过长，单格最多 32767 字符'
        for token in _PLACEHOLDER_RE.findall(value):
            if not _is_supported_placeholder(token):
                return '模板包含不支持的占位符 {%s}，请使用 order.* / item.* / %s' % (
                    token, ' / '.join(_EXPLICIT_TOKENS))
        return ''
    return '不支持的单元格值类型'


# ==================== 网格样式回写（PRINT-TEMPLATE-F05-A2） ====================

_GRID_STYLE_KEYS = frozenset({
    'bold', 'italic', 'underline', 'font_name', 'font_size',
    'font_color', 'bg_color', 'h_align', 'v_align', 'wrap', 'border',
})
_GRID_STYLE_COLOR_RE = re.compile(r'^#[0-9A-Fa-f]{6}$')
_GRID_BORDER_VALUES = frozenset({
    'thin', 'medium', 'thick', 'dashed', 'double', 'hair', 'dotted',
})


def _validate_grid_style(style):
    """校验单个单元格样式对象，返回错误信息；合法返回空串。

    语义：键存在且有值=设置；键存在且值为 None=清除；键不存在=保持不变。
    """
    if not isinstance(style, dict):
        return '样式必须是对象'
    unknown = set(style) - _GRID_STYLE_KEYS
    if unknown:
        return '不支持的样式键：%s' % '、'.join(sorted(unknown))
    for key in ('bold', 'italic', 'underline', 'wrap'):
        if key in style and style[key] is not None \
                and not isinstance(style[key], bool):
            return '%s 必须是布尔值' % key
    if 'font_name' in style and style['font_name'] is not None:
        name = style['font_name']
        if not isinstance(name, str) or not name.strip() or len(name) > 30:
            return 'font_name 必须是 1-30 字符的字体名'
    if 'font_size' in style and style['font_size'] is not None:
        size = style['font_size']
        if isinstance(size, bool) or not isinstance(size, (int, float)) \
                or not (6 <= size <= 72):
            return 'font_size 必须在 6-72 之间'
    for key in ('font_color', 'bg_color'):
        if key in style and style[key] is not None:
            color = style[key]
            if not isinstance(color, str) or not _GRID_STYLE_COLOR_RE.match(color):
                return '%s 必须是 #RRGGBB 格式' % key
    if 'h_align' in style and style['h_align'] is not None \
            and style['h_align'] not in _GRID_H_ALIGN:
        return 'h_align 只支持 left/center/right'
    if 'v_align' in style and style['v_align'] is not None \
            and style['v_align'] not in _GRID_V_ALIGN:
        return 'v_align 只支持 top/center/bottom'
    if 'border' in style and style['border'] is not None:
        border = style['border']
        if not isinstance(border, dict):
            return 'border 必须是对象'
        unknown_edges = set(border) - set(_GRID_BORDER_EDGES)
        if unknown_edges:
            return '不支持的边框边：%s' % '、'.join(sorted(unknown_edges))
        for edge, edge_style in border.items():
            if edge_style is None or edge_style == 'none':
                continue
            if edge_style not in _GRID_BORDER_VALUES:
                return '边框样式 %s 不受支持' % edge_style
    return ''


def _hex_to_argb(color):
    """'#RRGGBB' → 'FFRRGGBB'。"""
    return 'FF' + color[1:].upper()


def _apply_grid_style(cell, style):
    """把校验过的样式对象应用到单元格（None 值=清除，缺省=保持）。"""
    from copy import copy as _copy
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    font_keys = ('bold', 'italic', 'underline', 'font_name',
                 'font_size', 'font_color')
    if any(k in style for k in font_keys):
        font = _copy(cell.font)
        if 'bold' in style:
            font.bold = style['bold'] or None
        if 'italic' in style:
            font.italic = style['italic'] or None
        if 'underline' in style:
            font.underline = 'single' if style['underline'] else None
        if 'font_name' in style:
            font.name = style['font_name'] or None
        if 'font_size' in style:
            font.size = style['font_size'] or None
        if 'font_color' in style:
            font.color = (_hex_to_argb(style['font_color'])
                          if style['font_color'] else None)
        cell.font = font

    if 'bg_color' in style:
        if style['bg_color']:
            cell.fill = PatternFill(patternType='solid',
                                    fgColor=_hex_to_argb(style['bg_color']))
        else:
            cell.fill = PatternFill(fill_type=None)

    align_keys = ('h_align', 'v_align', 'wrap')
    if any(k in style for k in align_keys):
        alignment = _copy(cell.alignment)
        if 'h_align' in style:
            alignment.horizontal = style['h_align'] or None
        if 'v_align' in style:
            alignment.vertical = style['v_align'] or None
        if 'wrap' in style:
            alignment.wrap_text = bool(style['wrap']) or None
        cell.alignment = alignment

    if 'border' in style and style['border'] is not None:
        border = _copy(cell.border)
        for edge, edge_style in style['border'].items():
            side = Side(style=None) if edge_style in (None, 'none') \
                else Side(style=edge_style)
            setattr(border, edge, side)
        cell.border = border


def _validate_grid_dimension(value, label, low, high):
    """校验列宽/行高数值，返回错误信息；合法返回空串。"""
    if isinstance(value, bool) or not isinstance(value, (int, float)) \
            or not (low <= value <= high):
        return '%s 必须在 %s-%s 之间' % (label, low, high)
    return ''


def apply_print_template_grid(template_path, sheets_data):
    """把前端编辑后的网格写回 .xlsx 模板。

    sheets_data: [{name, upserts, del_rows, styles?, col_widths?, row_heights?}]
    - upserts 写/覆盖单元格值（'' 或 None 清除单元格）；保留样式与合并。
    - styles（PRINT-TEMPLATE-F05-A2）: [[row, col, style], ...] 回写单元格样式，
      语义「键存在有值=设置 / 键存在为 None=清除 / 键缺失=保持」，白名单校验。
    - col_widths/row_heights（F05-A2）: {列号/行号: 数值} 回写尺寸。
    - del_rows 删除指定行（整行，序号由小到大），用于清理多余示例行。
    全程占位符/样式白名单校验，任何非法输入立即报错且不落盘（原子性）；
    旧 payload（仅 upserts/del_rows）保持兼容。
    """
    sheets_data = sheets_data or []
    workbook = load_workbook(template_path)
    try:
        sheets_map = {ws.title: ws for ws in workbook.worksheets}
        for sheet in sheets_data:
            name = (sheet.get('name') or '').strip()
            ws = sheets_map.get(name)
            if ws is None:
                raise ValueError(f'工作表「{name}」不存在')
            row_map = {}
            for (r, c, value) in (sheet.get('upserts') or []):
                r = int(r)
                c = int(c)
                if r < 1 or c < 1:
                    raise ValueError(f'单元格坐标无效：第{r}行第{c}列')
                if r > _MAX_TEMPLATE_ROWS or c > _MAX_TEMPLATE_COLS:
                    raise ValueError(f'单元格坐标超出上限：第{r}行第{c}列')
                msg = _validate_grid_value(value)
                if msg:
                    raise ValueError(f'第{r}行第{c}列：{msg}')
                row_map.setdefault(r, {})[c] = value
            # 先处理单元格值（行索引删除前基于原始行号）
            for r, cols in sorted(row_map.items()):
                for c, value in cols.items():
                    cell = ws.cell(r, c)
                    if value is None or (isinstance(value, str) and value.strip() == ''):
                        cell.value = None
                    else:
                        cell.value = value
            # 样式回写（先全部校验再逐格应用，任一非法即整体不落盘）
            style_ops = []
            for item in (sheet.get('styles') or []):
                if not (isinstance(item, (list, tuple)) and len(item) == 3):
                    raise ValueError('每个样式项必须是 [行, 列, 样式]')
                r, c, style = int(item[0]), int(item[1]), item[2]
                if r < 1 or c < 1 or r > _MAX_TEMPLATE_ROWS \
                        or c > _MAX_TEMPLATE_COLS:
                    raise ValueError(f'样式坐标超出上限：第{r}行第{c}列')
                msg = _validate_grid_style(style)
                if msg:
                    raise ValueError(f'第{r}行第{c}列样式：{msg}')
                style_ops.append((r, c, style))
            for r, c, style in style_ops:
                _apply_grid_style(ws.cell(r, c), style)
            # 列宽/行高回写
            from openpyxl.utils import get_column_letter
            for col, width in (sheet.get('col_widths') or {}).items():
                c = int(col)
                if c < 1 or c > _MAX_TEMPLATE_COLS:
                    raise ValueError(f'列号超出上限：第{c}列')
                msg = _validate_grid_dimension(width, '列宽', 1, 200)
                if msg:
                    raise ValueError(f'第{c}列：{msg}')
                ws.column_dimensions[get_column_letter(c)].width = float(width)
            for row, height in (sheet.get('row_heights') or {}).items():
                r = int(row)
                if r < 1 or r > _MAX_TEMPLATE_ROWS:
                    raise ValueError(f'行号超出上限：第{r}行')
                msg = _validate_grid_dimension(height, '行高', 1, 500)
                if msg:
                    raise ValueError(f'第{r}行：{msg}')
                ws.row_dimensions[r].height = float(height)
            del_rows = sorted(set(int(x) for x in (sheet.get('del_rows') or [])), reverse=True)
            for r in del_rows:
                if r < 1:
                    continue
                ws.delete_rows(r, 1)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    finally:
        workbook.close()