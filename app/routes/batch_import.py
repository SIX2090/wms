#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 批量导入（batch_import）域路由。
#
# 批量拆分模式：为避免 endpoint 前缀化导致大量 url_for 引用改动，
# 采用「register_batch_import_routes(app)」直接在 app 上注册路由，endpoint 名保持不变
# （如 batch_import_page、import_out_order、import_in_order），与 app.py 内原有
# url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（InOrder、Material、Supplier、api_error 等）在各路由函数内
#   延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

from datetime import datetime

from flask import flash, jsonify, redirect, render_template, request, url_for
from flask_login import login_required

from db import db
from utils import require_role, validate_excel_extension, validate_excel_size, round_to_2_decimals


# no-test:reason=路由注册辅助函数，能力由 batch_import 各路由测试覆盖
def register_batch_import_routes(app):
    @app.route('/import/out_order', methods=['POST'])
    @require_role('warehouse', 'purchase')
    @login_required
    def import_out_order():
        from app import (
            Department,
            Material,
            OutOrder,
            OutOrderItem,
            Unit,
            api_error,
            current_user,
        )
        file = request.files.get('file')
        if not file:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error('请选择要导入的领料单文件')
            flash('请选择要导入的领料单文件', 'danger')
            return redirect(url_for('batch_import_page'))
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_ext_msg)
            flash(_ext_msg, 'danger')
            return redirect(url_for('batch_import_page'))
        # m-03：限制 Excel 上传 ≤ 5MB
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_size_msg)
            flash(_size_msg, 'danger')
            return redirect(url_for('batch_import_page'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            header_row = [str(cell).strip() if cell else '' for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col_map = {}
            for idx, h in enumerate(header_row):
                if not h:
                    continue
                if '单据编号' in h or '领料单号' in h or '出库单号' in h or '订单编号' in h:
                    col_map['order_no'] = idx
                elif h == '日期' or '日期' in h:
                    col_map['date'] = idx
                elif h == '用途' or '用途' in h:
                    col_map['purpose'] = idx
                elif '部门' in h or '领料' in h:
                    col_map['department'] = idx
                elif '物料编码' in h or '编码' in h:
                    col_map['material_code'] = idx
                elif '物料名称' in h or '名称' in h:
                    col_map['material_name'] = idx
                elif h == '规格' or '规格' in h:
                    col_map['spec'] = idx
                elif h == '单位' or '单位' in h:
                    col_map['unit'] = idx
                elif '数量' in h:
                    col_map['quantity'] = idx
                elif '单价' in h or '价格' in h:
                    col_map['price'] = idx
                elif '金额' in h or '总额' in h:
                    col_map['amount'] = idx
                elif '备注' in h:
                    col_map['remark'] = idx
            if 'order_no' not in col_map:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(f'Excel表头缺少"单据编号"列。检测到的表头：{", ".join(header_row)}')
                flash(f'Excel表头缺少"单据编号"列', 'danger')
                return redirect(url_for('batch_import_page'))
            count = 0
            skip = 0
            skip_details = []
            warnings = []
            current_order = None
            current_items = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                def get_val(key):
                    if key not in col_map:
                        return ''
                    idx = col_map[key]
                    if idx >= len(row):
                        return ''
                    return str(row[idx]).strip() if row[idx] is not None else ''
                def get_num(key):
                    if key not in col_map:
                        return 0
                    idx = col_map[key]
                    if idx >= len(row) or row[idx] is None:
                        return 0
                    try:
                        return round_to_2_decimals(row[idx])
                    except (ValueError, TypeError):
                        return 0
                order_no = get_val('order_no')
                if not order_no:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：单据编号为空')
                    continue
                if current_order and order_no != current_order.order_no:
                    for item in current_items:
                        db.session.add(item)
                    count += 1
                    current_order = None
                    current_items = []
                if not current_order:
                    dept_name = get_val('department')
                    department = Department.query.filter_by(name=dept_name).first() if dept_name else None
                    if not department and dept_name:
                        department = Department.query.filter_by(code=dept_name).first()
                    date_str = get_val('date')
                    date_val = None
                    if date_str:
                        try:
                            date_val = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                date_val = datetime.strptime(date_str, '%Y/%m/%d').date()
                            except ValueError:
                                try:
                                    date_val = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()
                                except ValueError:
                                    pass
                    if date_val is None and 'date' in col_map:
                        raw_date = row[col_map['date']] if col_map['date'] < len(row) else None
                        if raw_date is not None:
                            if hasattr(raw_date, 'date'):
                                date_val = raw_date.date()
                            elif hasattr(raw_date, 'year'):
                                date_val = raw_date
                    current_order = OutOrder(
                        order_no=order_no,
                        date=date_val,
                        purpose='领料单' if get_val('purpose') == '生产出库' else get_val('purpose'),
                        business_type='领料单',
                        department_id=department.id if department else None,
                        remark=get_val('remark'),
                        operator_id=current_user.id
                    )
                    if OutOrder.query.filter_by(order_no=order_no).first():
                        skip += 1
                        skip_details.append(f'第{row_idx}行：单据编号{order_no}已存在')
                        current_order = None
                        continue
                    db.session.add(current_order)
                    db.session.flush()
                material_code = get_val('material_code')
                if material_code:
                    material = Material.query.filter_by(code=material_code).first()
                    if not material:
                        material = Material(
                            code=material_code,
                            name=get_val('material_name'),
                            spec=get_val('spec')
                        )
                        db.session.add(material)
                        db.session.flush()
                        warnings.append(f'自动创建物料：{material_code}')
                    unit_name = get_val('unit')
                    unit = Unit.query.filter_by(name=unit_name).first() if unit_name else None
                    if not unit and unit_name:
                        unit = Unit.query.filter_by(code=unit_name).first()
                    if not unit and unit_name:
                        unit = Unit(code=unit_name, name=unit_name)
                        db.session.add(unit)
                        db.session.flush()
                        warnings.append(f'自动创建单位：{unit_name}')
                    if material and unit and not material.unit_id:
                        material.unit_id = unit.id
                    qty = get_num('quantity')
                    prc = get_num('price')
                    amt = get_num('amount')
                    if amt == 0:
                        amt = round_to_2_decimals(qty * prc)
                    item = OutOrderItem(
                        out_order_id=current_order.id,
                        material_id=material.id if material else None,
                        quantity=qty,
                        price=prc,
                        amount=amt,
                        remark=(get_val('remark') or '').strip() or None
                    )
                    current_items.append(item)
            if current_order and current_items:
                for item in current_items:
                    db.session.add(item)
                count += 1
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(f'导入失败：{str(e)}')
                flash(f'导入失败：{str(e)}', 'danger')
                return redirect(url_for('batch_import_page'))
            msg = f'领料单导入成功，共导入 {count} 张单据'
            if skip:
                msg += f'，跳过 {skip} 条'
            if skip_details:
                warnings.append(f'跳过详情：{"; ".join(skip_details[:20])}')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                resp = {'status': 'success', 'msg': msg, 'count': count}
                if warnings:
                    resp['warnings'] = '；'.join(warnings)
                return jsonify(resp)
            flash(msg, 'success')
            for w in warnings:
                flash(w, 'warning')
        except Exception as e:
            db.session.rollback()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(f'领料单导入失败：{str(e)}')
            flash(f'领料单导入失败：{str(e)}', 'danger')
        return redirect(url_for('batch_import_page'))

    @app.route('/import/in_order', methods=['POST'])
    @require_role('warehouse', 'purchase')
    @login_required
    def import_in_order():
        from app import (
            InOrder,
            InOrderItem,
            Material,
            Supplier,
            Unit,
            api_error,
            current_user,
        )
        file = request.files.get('file')
        if not file:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error('请选择要导入的入库单文件')
            flash('请选择要导入的入库单文件', 'danger')
            return redirect(url_for('batch_import_page'))
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_ext_msg)
            flash(_ext_msg, 'danger')
            return redirect(url_for('batch_import_page'))
        # m-03：限制 Excel 上传 ≤ 5MB
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_size_msg)
            flash(_size_msg, 'danger')
            return redirect(url_for('batch_import_page'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            header_row = [str(cell).strip() if cell else '' for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col_map = {}
            for idx, h in enumerate(header_row):
                if not h:
                    continue
                if '单据编号' in h or '入库单号' in h or '订单编号' in h:
                    col_map['order_no'] = idx
                elif h == '日期' or '日期' in h:
                    col_map['date'] = idx
                elif h == '用途' or '用途' in h:
                    col_map['purpose'] = idx
                elif '供应商' in h:
                    col_map['supplier'] = idx
                elif '物料编码' in h or '编码' in h:
                    col_map['material_code'] = idx
                elif '物料名称' in h or '名称' in h:
                    col_map['material_name'] = idx
                elif h == '规格' or '规格' in h:
                    col_map['spec'] = idx
                elif h == '单位' or '单位' in h:
                    col_map['unit'] = idx
                elif '数量' in h:
                    col_map['quantity'] = idx
                elif '单价' in h or '价格' in h:
                    col_map['price'] = idx
                elif '金额' in h or '总额' in h:
                    col_map['amount'] = idx
                elif '备注' in h:
                    col_map['remark'] = idx
            if 'order_no' not in col_map:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(f'Excel表头缺少"单据编号"列。检测到的表头：{", ".join(header_row)}')
                flash(f'Excel表头缺少"单据编号"列', 'danger')
                return redirect(url_for('batch_import_page'))
            count = 0
            skip = 0
            skip_details = []
            warnings = []
            current_order = None
            current_items = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                def get_val(key):
                    if key not in col_map:
                        return ''
                    idx = col_map[key]
                    if idx >= len(row):
                        return ''
                    return str(row[idx]).strip() if row[idx] is not None else ''
                def get_num(key):
                    if key not in col_map:
                        return 0
                    idx = col_map[key]
                    if idx >= len(row) or row[idx] is None:
                        return 0
                    try:
                        return round_to_2_decimals(row[idx])
                    except (ValueError, TypeError):
                        return 0
                order_no = get_val('order_no')
                if not order_no:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：单据编号为空')
                    continue
                if current_order and order_no != current_order.order_no:
                    for item in current_items:
                        db.session.add(item)
                    count += 1
                    current_order = None
                    current_items = []
                if not current_order:
                    supplier_name = get_val('supplier')
                    supplier = Supplier.query.filter_by(name=supplier_name).first() if supplier_name else None
                    if not supplier and supplier_name:
                        supplier = Supplier.query.filter_by(code=supplier_name).first()
                    if not supplier and supplier_name:
                        supplier = Supplier(code=supplier_name, name=supplier_name)
                        db.session.add(supplier)
                        db.session.flush()
                        warnings.append(f'自动创建供应商：{supplier_name}')
                    date_str = get_val('date')
                    date_val = None
                    if date_str:
                        try:
                            date_val = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                date_val = datetime.strptime(date_str, '%Y/%m/%d').date()
                            except ValueError:
                                try:
                                    date_val = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()
                                except ValueError:
                                    pass
                    if date_val is None and 'date' in col_map:
                        raw_date = row[col_map['date']] if col_map['date'] < len(row) else None
                        if raw_date is not None:
                            if hasattr(raw_date, 'date'):
                                date_val = raw_date.date()
                            elif hasattr(raw_date, 'year'):
                                date_val = raw_date
                    current_order = InOrder(
                        order_no=order_no,
                        date=date_val,
                        purpose=get_val('purpose'),
                        supplier_id=supplier.id if supplier else None,
                        remark=get_val('remark'),
                        operator_id=current_user.id
                    )
                    if InOrder.query.filter_by(order_no=order_no).first():
                        skip += 1
                        skip_details.append(f'第{row_idx}行：单据编号{order_no}已存在')
                        current_order = None
                        continue
                    db.session.add(current_order)
                    db.session.flush()
                material_code = get_val('material_code')
                if material_code:
                    material = Material.query.filter_by(code=material_code).first()
                    if not material:
                        material = Material(
                            code=material_code,
                            name=get_val('material_name'),
                            spec=get_val('spec')
                        )
                        db.session.add(material)
                        db.session.flush()
                        warnings.append(f'自动创建物料：{material_code}')
                    unit_name = get_val('unit')
                    unit = Unit.query.filter_by(name=unit_name).first() if unit_name else None
                    if not unit and unit_name:
                        unit = Unit.query.filter_by(code=unit_name).first()
                    if not unit and unit_name:
                        unit = Unit(code=unit_name, name=unit_name)
                        db.session.add(unit)
                        db.session.flush()
                        warnings.append(f'自动创建单位：{unit_name}')
                    if material and unit and not material.unit_id:
                        material.unit_id = unit.id
                    qty = get_num('quantity')
                    prc = get_num('price')
                    amt = get_num('amount')
                    if amt == 0:
                        amt = round_to_2_decimals(qty * prc)
                    item = InOrderItem(
                        in_order_id=current_order.id,
                        material_id=material.id if material else None,
                        quantity=qty,
                        price=prc,
                        amount=amt
                    )
                    current_items.append(item)
            if current_order and current_items:
                for item in current_items:
                    db.session.add(item)
                count += 1
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(f'导入失败：{str(e)}')
                flash(f'导入失败：{str(e)}', 'danger')
                return redirect(url_for('batch_import_page'))
            msg = f'入库单导入成功，共导入 {count} 张单据'
            if skip:
                msg += f'，跳过 {skip} 条'
            if skip_details:
                warnings.append(f'跳过详情：{"; ".join(skip_details[:20])}')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                resp = {'status': 'success', 'msg': msg, 'count': count}
                if warnings:
                    resp['warnings'] = '；'.join(warnings)
                return jsonify(resp)
            flash(msg, 'success')
            for w in warnings:
                flash(w, 'warning')
        except Exception as e:
            db.session.rollback()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(f'入库单导入失败：{str(e)}')
            flash(f'入库单导入失败：{str(e)}', 'danger')
        return redirect(url_for('batch_import_page'))

    @app.route('/batch_import')
    @login_required
    def batch_import_page():
        _module_type = request.args.get('type', '').strip().lower() or None
        return render_template('batch_import.html', module_type=_module_type)

    # P1-类别 B：基础资料 /import 与 /export 便捷入口（统一跳转集中式 /batch_import）
    @app.route('/user/import', methods=['POST'])
    @login_required
    @require_role('admin')
    def user_import_stub():
        return redirect(url_for('batch_import_page', type='user'))

    @app.route('/user/export')
    @login_required
    @require_role('admin')
    def user_export_stub():
        return redirect(url_for('batch_import_page', type='user'))

    @app.route('/label_template/import', methods=['POST'])
    @login_required
    @require_role('admin')
    def label_template_import_stub():
        return redirect(url_for('batch_import_page', type='label_template'))

    @app.route('/label_template/export')
    @login_required
    @require_role('admin')
    def label_template_export_stub():
        return redirect(url_for('batch_import_page', type='label_template'))

    @app.route('/opening_stock/import', methods=['POST'])
    @login_required
    def opening_stock_import_stub():
        return redirect(url_for('batch_import_page', type='opening_stock'))

    @app.route('/opening_stock/export')
    @login_required
    def opening_stock_export_stub():
        return redirect(url_for('batch_import_page', type='opening_stock'))