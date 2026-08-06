#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 物料清单（bom）域路由。
#
# 批量拆分模式：与销售（sales）/售后出库（after_sale_out）域一致，采用
# 「register_bom_routes(app)」直接在 app 上注册路由，endpoint 名保持不变
# （bom_list、bom_detail、print_bom、bom_add_page、save_bom_table、add_bom、
# update_bom、add_bom_item、delete_bom_item、delete_bom、batch_delete_bom、
# calculate_bom_cost、export_bom、import_bom、create_requisition_from_bom 等），
# 与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / flask_login / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（BOM、BOMItem、ProductionRequisition、ProductionRequisitionItem、
#   各辅助函数 _bom_query_from_args / _render_bom_form / serialize_bom /
#   calculate_bom_cost_value / _material_from_payload / _clean_int /
#   _material_row_common / _render_generic_document_print / _fmt_date / api_error /
#   generate_order_no / log_operation 等）在各路由函数内延迟导入（请求期才执行），
#   避免 app.py 模块加载期触发循环导入。
# - 日志复用 register_bom_routes(app) 传入的 app.logger（与 app.py 原实现一致）。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

from flask import jsonify, render_template, request, send_file
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 bom_* 各路由测试覆盖
def register_bom_routes(app):
    @app.route('/bom')
    @login_required
    def bom_list():
        from app import Unit, _bom_query_from_args, serialize_bom
        query, filters, sort_by, sort_order = _bom_query_from_args()
        boms = query.all()
        units = Unit.query.all()
        return render_template('bom.html', boms=[serialize_bom(bom) for bom in boms], units=units, filters=filters, sort_by=sort_by, sort_order=sort_order)

    @app.route('/bom/<int:id>')
    @login_required
    def bom_detail(id):
        from sqlalchemy.orm import joinedload, selectinload
        from app import BOM, BOMItem, Material, _render_bom_form
        bom = BOM.query.options(
            selectinload(BOM.items).joinedload(BOMItem.material).joinedload(Material.unit),
            selectinload(BOM.items).joinedload(BOMItem.unit)
        ).get_or_404(id)
        return _render_bom_form(bom)

    @app.route('/bom/<int:id>/print')
    @login_required
    def print_bom(id):
        from datetime import datetime
        from sqlalchemy.orm import joinedload, selectinload
        from app import (BOM, BOMItem, Material, _fmt_date, _material_row_common,
                         _render_generic_document_print)
        bom = BOM.query.options(
            selectinload(BOM.items).joinedload(BOMItem.material).joinedload(Material.unit),
            selectinload(BOM.items).joinedload(BOMItem.unit)
        ).get_or_404(id)
        rows = [
            _material_row_common(
                item,
                price=item.unit_cost or 0,
                amount=item.total_cost or 0,
                extra={'usage': item.usage or ''}
            )
            for item in bom.items
        ]
        return _render_generic_document_print({
            'title': 'BOM清单',
            'subtitle': 'BILL OF MATERIALS',
            'number_label': 'BOM编号',
            'number': bom.bom_no,
            'date_label': '打印日期',
            'date': datetime.now().strftime('%Y-%m-%d'),
            'status': bom.status,
            'info': [
                ('产品编码', bom.product_code),
                ('产品名称', bom.product_name),
                ('版本', bom.version),
                ('层级', bom.level or 1),
                ('总成本', f'{bom.total_cost or 0:.2f}'),
                ('创建时间', _fmt_date(bom.created_at)),
            ],
            'remark': bom.remark or '',
            'columns': [
                ('code', '物料编码', ''),
                ('name', '物料名称', ''),
                ('spec', '规格', ''),
                ('unit', '单位', 'center'),
                ('quantity', '用量', 'right'),
                ('price', '单位成本', 'right money'),
                ('amount', '成本金额', 'right money'),
                ('usage', '用途', ''),
                ('remark', '备注', ''),
            ],
            'rows': rows,
            'total_amount': bom.total_cost or 0,
            'signatures': ['制单', '审核', '生产', '仓库'],
        })

    @app.route('/bom/add', methods=['GET'])
    @login_required
    def bom_add_page():
        from app import _render_bom_form
        return _render_bom_form()

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/save_table', methods=['POST'])
    @require_role('production')
    @login_required
    def save_bom_table():
        from app import (BOM, BOMItem, _clean_int, _material_from_payload,
                         api_error, calculate_bom_cost_value, generate_order_no,
                         log_operation, parse_float_value, round_to_2_decimals)
        data = request.get_json(silent=True) or {}
        order_id = _clean_int(data.get('order_id'))
        bom_no = (data.get('order_no') or data.get('bom_no') or '').strip() or generate_order_no('BOM')
        header = data.get('header') or {}
        items_data = data.get('items') or []

        product_code = (header.get('product_code') or data.get('product_code') or '').strip()
        product_name = (header.get('product_name') or data.get('product_name') or '').strip()
        version = (header.get('version') or data.get('version') or 'V1.0').strip()
        status = (header.get('status') or data.get('status') or 'active').strip()
        remark = (header.get('remark') or data.get('remark') or '').strip()

        if not product_code or not product_name:
            return api_error('请输入产品编码和产品名称')
        if status not in ('active', 'inactive'):
            status = 'active'
        if not items_data:
            return api_error('请至少填写一条BOM明细')

        try:
            if order_id:
                bom = db.session.get(BOM, order_id)
                if not bom:
                    return api_error('BOM不存在，请刷新后重试')
                duplicate = BOM.query.filter(BOM.bom_no == bom_no, BOM.id != order_id).first()
                if duplicate:
                    return api_error('BOM编号已存在')
            else:
                bom = BOM.query.filter_by(bom_no=bom_no).first()
                if not bom:
                    bom = BOM(bom_no=bom_no)
                    db.session.add(bom)

            bom.bom_no = bom_no
            bom.product_code = product_code
            bom.product_name = product_name
            bom.version = version
            bom.status = status
            bom.remark = remark
            db.session.flush()
            BOMItem.query.filter_by(bom_id=bom.id).delete()

            for item_data in items_data:
                material = _material_from_payload(item_data)
                if not material:
                    return api_error(f'物料不存在：{item_data.get("code") or ""}')
                quantity = round_to_2_decimals(parse_float_value(item_data.get('quantity'), 0))
                if quantity <= 0:
                    return api_error(f'物料 {material.code} 的数量必须大于0')
                unit_id = _clean_int(item_data.get('unit_id')) or material.unit_id
                unit_cost = round_to_2_decimals(parse_float_value(item_data.get('price'), material.price or 0))
                db.session.add(BOMItem(
                    bom_id=bom.id,
                    material_id=material.id,
                    quantity=quantity,
                    unit_id=unit_id,
                    unit_cost=unit_cost,
                    total_cost=round_to_2_decimals(quantity * unit_cost),
                    usage=(item_data.get('usage') or '').strip(),
                    remark=(item_data.get('remark') or '').strip()
                ))

            calculate_bom_cost_value(bom, refresh_unit_cost=False)
            db.session.commit()
            log_operation('保存BOM', f'BOM：{bom.bom_no}', 'bom', bom.id)
            return jsonify({'status': 'success', 'msg': '保存成功', 'id': bom.id, 'order_no': bom.bom_no})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'保存BOM失败: {e}')
            return api_error('保存失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/add', methods=['POST'])
    @require_role('production')
    @login_required
    def add_bom():
        from app import BOM, api_error, generate_order_no, log_operation
        try:
            product_code = (request.form.get('product_code') or '').strip()
            product_name = (request.form.get('product_name') or '').strip()
            version = (request.form.get('version') or '1.0').strip()
            remark = (request.form.get('remark') or '').strip()

            if not product_code or not product_name:
                return api_error('请输入成品编码和成品名称')

            bom_no = generate_order_no('BOM')
            bom = BOM(
                bom_no=bom_no,
                product_code=product_code,
                product_name=product_name,
                version=version,
                remark=remark,
                status='active'
            )
            db.session.add(bom)
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'数据库操作失败: {e}')
                return jsonify({'status': 'error', 'msg': '操作失败，请稍后重试'}), 500
            log_operation('新增BOM', f'BOM：{bom_no}', 'bom', bom.id)
            return jsonify({'status': 'success', 'msg': 'BOM 新增成功'})
        except Exception as e:
            db.session.rollback()
            return api_error('操作失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/update', methods=['POST'])
    @require_role('production')
    @login_required
    def update_bom(id):
        from app import BOM, api_error
        bom = BOM.query.get_or_404(id)
        try:
            bom.product_code = (request.form.get('product_code') or bom.product_code).strip()
            bom.product_name = (request.form.get('product_name') or bom.product_name).strip()
            bom.version = (request.form.get('version') or bom.version).strip()
            bom.status = (request.form.get('status') or bom.status).strip()
            bom.remark = (request.form.get('remark') or bom.remark).strip()
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'数据库操作失败: {e}')
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            return jsonify({'status': 'success'})
        except Exception as e:
            db.session.rollback()
            return api_error('操作失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/item/add', methods=['POST'])
    @require_role('production')
    @login_required
    def add_bom_item(id):
        from app import (BOM, BOMItem, Material, Unit, api_error,
                         calculate_bom_cost_value, parse_float_value, round_to_2_decimals)
        bom = BOM.query.get_or_404(id)
        try:
            material_code = (request.form.get('material_code') or '').strip()
            # 用 parse_float_value 兜底：传入 "abc" 或负数时回落到 0，
            # 触发下方 quantity <= 0 检查返回明确错误，避免 ValueError 500
            quantity = parse_float_value(request.form.get('quantity'), 0)
            unit_id = request.form.get('unit_id')
            usage = (request.form.get('usage') or '').strip()
            remark = (request.form.get('remark') or '').strip()

            material = Material.query.filter_by(code=material_code).first()
            if not material:
                return api_error('物料编码不存在')
            if quantity <= 0:
                return api_error('用量必须大于 0')

            unit = db.session.get(Unit, unit_id) if unit_id else None
            unit_cost = material.price or 0
            item = BOMItem(
                bom_id=id,
                material_id=material.id,
                quantity=quantity,
                unit_id=unit.id if unit else None,
                unit_cost=unit_cost,
                total_cost=round_to_2_decimals(quantity * unit_cost),
                usage=usage,
                remark=remark
            )
            db.session.add(item)
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'数据库操作失败: {e}')
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            calculate_bom_cost_value(bom)
            db.session.commit()
            return jsonify({'status': 'success', 'msg': 'BOM 明细新增成功', 'id': item.id})
        except Exception as e:
            db.session.rollback()
            return api_error('操作失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/add_item', methods=['POST'])
    @require_role('production')
    @login_required
    def add_bom_item_alias(id):
        return add_bom_item(id)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/item/<int:item_id>/delete', methods=['POST'])
    @require_role('production')
    @login_required
    def delete_bom_item(id, item_id):
        from app import BOM, BOMItem, api_error, calculate_bom_cost_value
        item = BOMItem.query.get_or_404(item_id)
        if item.bom_id != id:
            return api_error('BOM')
        db.session.delete(item)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        bom = db.session.get(BOM, id)
        if bom:
            calculate_bom_cost_value(bom)
            db.session.commit()
        return jsonify({'status': 'success'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/delete_item/<int:item_id>', methods=['POST'])
    @require_role('production')
    @login_required
    def delete_bom_item_alias(id, item_id):
        return delete_bom_item(id, item_id)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/delete', methods=['POST'])
    @require_role('production')
    @login_required
    def delete_bom(id):
        from app import BOM, BOMItem
        bom = BOM.query.get_or_404(id)
        BOMItem.query.filter_by(bom_id=id).delete()
        db.session.delete(bom)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/delete/<int:id>', methods=['POST'])
    @require_role('production')
    @login_required
    def delete_bom_alias(id):
        return delete_bom(id)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/batch_delete', methods=['POST'])
    @require_role('production')
    @login_required
    def batch_delete_bom():
        from app import BOM, BOMItem
        ids = (request.get_json(silent=True) or {}).get('ids', [])
        for bid in ids:
            BOMItem.query.filter_by(bom_id=bid).delete()
            BOM.query.filter_by(id=bid).delete()
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:id>/calculate_cost', methods=['POST'])
    @require_role('production')
    @login_required
    def calculate_bom_cost(id):
        from app import BOM, api_error, calculate_bom_cost_value
        bom = BOM.query.get_or_404(id)
        try:
            total_cost = calculate_bom_cost_value(bom)
            db.session.commit()
            return jsonify({'status': 'success', 'msg': f'成本已更新，总成本：{total_cost:.2f}', 'total_cost': total_cost})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'计算BOM成本失败: {e}')
            return api_error('计算失败，请稍后重试')

    @app.route('/bom/export')
    @login_required
    def export_bom():
        import io
        from openpyxl import Workbook
        from app import _bom_query_from_args
        wb = Workbook()
        ws = wb.active
        ws.title = 'BOM'
        ws.append(['BOM编号', '产品编码', '产品名称', '版本', '状态', '层级', '物料数', '总成本', '创建时间', '备注'])
        query, _, _, _ = _bom_query_from_args()
        for bom in query.all():
            status_label = '激活' if bom.status == 'active' else ('停用' if bom.status == 'inactive' else (bom.status or ''))
            ws.append([
                bom.bom_no,
                bom.product_code,
                bom.product_name,
                bom.version,
                status_label,
                bom.level or 1,
                len(bom.items or []),
                bom.total_cost or 0,
                bom.created_at.strftime('%Y-%m-%d %H:%M') if bom.created_at else '',
                bom.remark or '',
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='BOM.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/import', methods=['POST'])
    @require_role('production')
    @login_required
    def import_bom():
        from app import (BOM, BOMItem, Material, Unit, api_error,
                         calculate_bom_cost_value, generate_order_no,
                         round_to_2_decimals, validate_excel_extension, validate_excel_size)
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的 BOM 文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            header_row = [str(cell).strip() if cell else '' for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col_map = {}
            for idx, header in enumerate(header_row):
                if not header:
                    continue
                if '产品编码' in header or '成品编码' in header or 'product_code' in header.lower():
                    col_map['product_code'] = idx
                elif '产品名称' in header or '成品名称' in header or 'product_name' in header.lower():
                    col_map['product_name'] = idx
                elif '版本' in header or 'version' in header.lower():
                    col_map['version'] = idx
                elif '物料编码' in header or '子件编码' in header or '材料编码' in header or 'material_code' in header.lower():
                    col_map['material_code'] = idx
                elif '物料名称' in header or '子件名称' in header or '材料名称' in header or 'material_name' in header.lower():
                    col_map['material_name'] = idx
                elif '规格' in header or 'spec' in header.lower():
                    col_map['spec'] = idx
                elif '单位' in header or 'unit' in header.lower():
                    col_map['unit'] = idx
                elif '数量' in header or '用量' in header or 'quantity' in header.lower():
                    col_map['quantity'] = idx
                elif '用途' in header or 'usage' in header.lower():
                    col_map['usage'] = idx
                elif '备注' in header or 'remark' in header.lower():
                    col_map['remark'] = idx

            # Backward compatibility for the old three-column BOM import format.
            if 'product_code' not in col_map and header_row:
                col_map['product_code'] = 0
            if 'product_name' not in col_map and len(header_row) > 1:
                col_map['product_name'] = 1
            if 'version' not in col_map and len(header_row) > 2:
                col_map['version'] = 2

            if 'product_code' not in col_map or 'product_name' not in col_map:
                msg = f'Excel表头缺少必要列（产品编码、产品名称）。检测到的表头：{", ".join(header_row)}'
                return api_error(msg)

            # no-test:reason=BOM 导入的内部取数辅助函数，由 import_bom 路由测试覆盖
            def get_val(row, key):
                if key not in col_map:
                    return ''
                idx = col_map[key]
                if idx >= len(row) or row[idx] is None:
                    return ''
                return str(row[idx]).strip()

            # no-test:reason=BOM 导入的内部取数辅助函数，由 import_bom 路由测试覆盖
            def get_num(row, key, default=0):
                value = get_val(row, key)
                if value == '':
                    return default
                try:
                    return round_to_2_decimals(value)
                except (ValueError, TypeError):
                    return default

            boms_by_key = {}
            count = 0
            item_count = 0
            skip = 0
            skip_details = []
            warnings = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                product_code = get_val(row, 'product_code')
                product_name = get_val(row, 'product_name')
                version = get_val(row, 'version') or '1.0'
                if not product_code or not product_name:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：产品编码或产品名称为空')
                    continue

                bom_key = (product_code, version)
                bom = boms_by_key.get(bom_key)
                if not bom:
                    bom = BOM.query.filter_by(product_code=product_code, version=version).first()
                    if bom:
                        warnings.append(f'第{row_idx}行：BOM {product_code}/{version} 已存在，追加明细')
                        bom.product_name = product_name
                    else:
                        bom = BOM(
                            bom_no=generate_order_no('BOM'),
                            product_code=product_code,
                            product_name=product_name,
                            version=version,
                            status='active'
                        )
                        db.session.add(bom)
                        db.session.flush()
                        count += 1
                    boms_by_key[bom_key] = bom

                material_code = get_val(row, 'material_code')
                if not material_code:
                    continue

                quantity = get_num(row, 'quantity', 1)
                if quantity <= 0:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：物料 {material_code} 数量必须大于0')
                    continue

                material = Material.query.filter_by(code=material_code).first()
                if not material:
                    material = Material(
                        code=material_code,
                        name=get_val(row, 'material_name') or material_code,
                        spec=get_val(row, 'spec')
                    )
                    db.session.add(material)
                    db.session.flush()
                    warnings.append(f'自动创建物料：{material_code}')
                elif get_val(row, 'material_name') and not material.name:
                    material.name = get_val(row, 'material_name')
                if get_val(row, 'spec') and not material.spec:
                    material.spec = get_val(row, 'spec')

                unit_name = get_val(row, 'unit')
                unit = None
                if unit_name:
                    unit = Unit.query.filter_by(name=unit_name).first() or Unit.query.filter_by(code=unit_name).first()
                    if not unit:
                        unit = Unit(code=unit_name, name=unit_name)
                        db.session.add(unit)
                        db.session.flush()
                        warnings.append(f'自动创建单位：{unit_name}')
                if not unit and material.unit:
                    unit = material.unit
                if not unit:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：物料 {material_code} 缺少单位')
                    continue
                if not material.unit_id:
                    material.unit_id = unit.id

                existing_item = BOMItem.query.filter_by(bom_id=bom.id, material_id=material.id).first()
                unit_cost = material.price or 0
                if existing_item:
                    existing_item.quantity = quantity
                    existing_item.unit_id = unit.id
                    existing_item.unit_cost = unit_cost
                    existing_item.total_cost = round_to_2_decimals(quantity * unit_cost)
                    existing_item.usage = get_val(row, 'usage')
                    existing_item.remark = get_val(row, 'remark')
                else:
                    item = BOMItem(
                        bom_id=bom.id,
                        material_id=material.id,
                        quantity=quantity,
                        unit_id=unit.id,
                        unit_cost=unit_cost,
                        total_cost=round_to_2_decimals(quantity * unit_cost),
                        usage=get_val(row, 'usage'),
                        remark=get_val(row, 'remark')
                    )
                    db.session.add(item)
                    item_count += 1

            for bom in boms_by_key.values():
                calculate_bom_cost_value(bom)

            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'数据库操作失败: {e}')
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            msg = f'BOM 导入成功，共导入 {count} 个BOM，{item_count} 条明细'
            if skip:
                msg += f'，跳过 {skip} 行'
            if skip_details:
                warnings.append(f'跳过详情：{"; ".join(skip_details[:20])}')
            resp = {'status': 'success', 'msg': msg, 'count': count, 'item_count': item_count}
            if warnings:
                resp['warnings'] = '；'.join(warnings)
            return jsonify(resp)
        except Exception as e:
            db.session.rollback()
            return api_error(f'BOM 导入失败：{str(e)}')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/bom/<int:bom_id>/create_requisition', methods=['POST'])
    @require_role('production')
    @login_required
    def create_requisition_from_bom(bom_id):
        from flask_login import current_user
        from app import (BOM, ProductionRequisition, ProductionRequisitionItem,
                         generate_order_no, get_default_warehouse)
        bom = BOM.query.get_or_404(bom_id)
        # 复用 generate_order_no('REQ')，避免原先基于秒级时间戳生成 req_no
        # 在并发或同秒点击时产生重复单号，触发 unique 约束失败
        req_no = generate_order_no('REQ')
        # BUG-2026-08-05-008：BOM 下推工单领料单自动带入默认仓库
        _default_wh = get_default_warehouse()
        requisition = ProductionRequisition(
            req_no=req_no,
            bom_id=bom_id,
            warehouse=_default_wh.name if _default_wh else None,
            status='pending',
            operator_id=current_user.id
        )
        db.session.add(requisition)
        db.session.flush()
        for item in bom.items:
            req_item = ProductionRequisitionItem(
                requisition_id=requisition.id,
                material_id=item.material_id,
                quantity=item.quantity,
                unit_id=item.unit_id
            )
            db.session.add(req_item)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success', 'id': requisition.id})