#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 物料分类（category）域路由。
#
# 批量拆分模式：与供应商域一致，采用「register_<domain>_routes(app)」直接在 app
# 上注册路由，endpoint 名保持不变（如 category_list），与 app.py 内原有 url_for
# 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（MaterialCategory 模型、辅助函数等）在各路由函数内延迟导入
#   （请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import current_app, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 category_* 各路由测试覆盖
def register_category_routes(app):
    @app.route('/category')
    @login_required
    def category_list():
        from app import (
            MaterialCategory,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
            build_category_parent_options,
            build_category_tree_rows,
        )
        search, status_filter, sort_by, sort_order = _get_master_list_filters('code')
        allowed_sorts = {'id', 'code', 'name', 'created_at'}
        query = _apply_simple_search(MaterialCategory.query, MaterialCategory, search, ['code', 'name'])
        query, sort_by = _apply_master_order(query, MaterialCategory, sort_by, sort_order, allowed_sorts, 'code')
        categories = query.all()
        all_categories = MaterialCategory.query.order_by(MaterialCategory.code.asc(), MaterialCategory.name.asc(), MaterialCategory.id.asc()).all()
        display_rows = build_category_tree_rows(categories if search else all_categories)
        parent_options = build_category_parent_options(all_categories)
        return render_template(
            'category.html',
            categories=categories,
            category_rows=display_rows,
            parent_options=parent_options,
            filters={'search': search, 'status': status_filter},
            sort_by=sort_by,
            sort_order=sort_order,
        )

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/category/add', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_category():
        from app import MaterialCategory, api_error
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        parent_id = request.form.get('parent_id', type=int) or None
        if not code:
            return api_error('请输入分类编码')
        if not name:
            return api_error('请输入分类名称')
        if MaterialCategory.query.filter_by(code=code).first():
            return api_error('分类编码已存在')
        if MaterialCategory.query.filter_by(name=name).first():
            return api_error('分类名称已存在')
        if parent_id and not db.session.get(MaterialCategory, parent_id):
            return api_error('上级分类不存在')
        cat = MaterialCategory(code=code, name=name, parent_id=parent_id)
        db.session.add(cat)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'添加分类失败: {e}')
            return jsonify({'status': 'error', 'msg': '添加失败'}), 500
        return jsonify({'status': 'success', 'msg': '分类新增成功', 'id': cat.id, 'code': cat.code, 'name': cat.name})

    @app.route('/category/<int:id>')
    @login_required
    def get_category(id):
        from app import MaterialCategory, api_error
        cat = db.session.get(MaterialCategory, id)
        if not cat:
            return api_error('分类不存在')
        return jsonify({'status': 'success', 'category': {'id': cat.id, 'code': cat.code, 'name': cat.name, 'parent_id': cat.parent_id or ''}})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/category/edit/<int:id>', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def edit_category(id):
        from app import MaterialCategory, api_error
        cat = db.session.get(MaterialCategory, id)
        if not cat:
            return api_error('分类不存在')
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        parent_id = request.form.get('parent_id', type=int) or None
        if not code:
            return api_error('请输入分类编码')
        if not name:
            return api_error('请输入分类名称')
        existing = MaterialCategory.query.filter_by(code=code).first()
        if existing and existing.id != id:
            return api_error('分类编码已存在')
        existing = MaterialCategory.query.filter_by(name=name).first()
        if existing and existing.id != id:
            return api_error('分类名称已存在')
        if parent_id == id:
            return api_error('上级分类不能选择自己')
        if parent_id:
            parent = db.session.get(MaterialCategory, parent_id)
            if not parent:
                return api_error('上级分类不存在')
            current_parent = parent
            while current_parent:
                if current_parent.parent_id == id:
                    return api_error('上级分类不能选择自己的下级分类')
                current_parent = current_parent.parent
        cat.code = code
        cat.name = name
        cat.parent_id = parent_id
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'编辑分类失败: {e}')
            return jsonify({'status': 'error', 'msg': '编辑失败'}), 500
        return jsonify({'status': 'success', 'msg': '分类编辑成功'})

    @app.route('/category/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_category():
        # P2-B: Pydantic BaseModel 存量迁移示范（A8 规则模式）
        from pydantic import BaseModel, Field
        from app import MaterialCategory, Material, api_error

        class DeleteCategoryRequest(BaseModel):
            ids: list[int] = Field(default_factory=list, description='待删除分类 ID 列表')

        payload = request.get_json(silent=True) or {}
        try:
            req = DeleteCategoryRequest.model_validate(payload)
        except Exception as exc:
            return jsonify({'status': 'error', 'msg': f'参数校验失败：{exc}'}), 400
        ids = req.ids
        id_set = set(ids)
        for id in ids:
            cat = db.session.get(MaterialCategory, id)
            if cat:
                if MaterialCategory.query.filter(
                    MaterialCategory.parent_id == cat.id,
                    ~MaterialCategory.id.in_(id_set),
                ).first():
                    return api_error(f'分类 {cat.name} 下还有子分类，请先删除或调整子分类')
                # F-02：分类被物料引用时禁止硬删，避免外键悬空
                mat_n = Material.query.filter_by(category_id=cat.id).count()
                if mat_n > 0:
                    return jsonify({'status': 'error',
                                    'msg': f'分类“{cat.name}”已被 {mat_n} 个物料引用，禁止删除'})
                db.session.delete(cat)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success'})

    @app.route('/category/download_template')
    @login_required
    def download_category_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '物料分类导入模板'
        ws.append(['分类编码', '分类名称', '上级分类编码'])
        ws.append(['100', '原材料', ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='category_template.xlsx', as_attachment=True)

    @app.route('/category/export')
    @login_required
    def export_category():
        from openpyxl import Workbook
        from app import (
            MaterialCategory,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '物料分类数据'
        ws.append(['分类编码', '分类名称', '上级分类编码'])
        search, status_filter, sort_by, sort_order = _get_master_list_filters('code')
        query = _apply_simple_search(MaterialCategory.query, MaterialCategory, search, ['code', 'name'])
        query, _ = _apply_master_order(query, MaterialCategory, sort_by, sort_order, {'id', 'code', 'name', 'created_at'}, 'code')
        for c in query.all():
            ws.append([c.code, c.name, c.parent.code if c.parent else ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='categories.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/category/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_category():
        from app import (
            MaterialCategory,
            api_error,
            validate_excel_extension,
            validate_excel_size,
        )
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的分类文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                parent_code = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                if not code or not name:
                    skip += 1
                    continue
                if MaterialCategory.query.filter_by(code=code).first():
                    skip += 1
                    continue
                if MaterialCategory.query.filter_by(name=name).first():
                    skip += 1
                    continue
                parent_id = None
                if parent_code:
                    parent = MaterialCategory.query.filter_by(code=parent_code).first()
                    if not parent:
                        skip += 1
                        continue
                    parent_id = parent.id
                cat = MaterialCategory(code=code, name=name, parent_id=parent_id)
                db.session.add(cat)
                count += 1
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            msg = f'分类导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条（重复或格式错误）'
            return jsonify({'status': 'success', 'msg': msg})
        except Exception:
            db.session.rollback()
            return api_error('导入失败，请稍后重试')