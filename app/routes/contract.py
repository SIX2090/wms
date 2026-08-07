#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 合同（contract）域路由。
#
# 批量拆分模式：与员工（employee）域一致，采用「register_contract_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 contract_list、api_contracts_search），
# 与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（Contract 模型、_contract_delete_blockers、_format_delete_blockers、
#   api_error、validate_excel_extension、validate_excel_size 等）在各路由函数内
#   延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import current_app, jsonify, render_template, request, send_file
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 contract_* 各路由测试覆盖
def register_contract_routes(app):
    @app.route('/contract')
    @login_required
    def contract_list():
        from app import Contract
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        per_page = max(1, per_page)
        if per_page not in [10, 20, 50, 100, 200]:
            per_page = 20
        search = (request.args.get('search') or '').strip()
        status = (request.args.get('status') or '').strip()
        # BUG-F02-01 修复：合同列表默认按 contract_no 升序（合同没有 code 字段，用单号排序）
        sort_by = (request.args.get('sort') or 'contract_no').strip()
        sort_order = (request.args.get('order') or 'asc').strip()
        query = Contract.query
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(
                Contract.contract_no.ilike(like),
                Contract.project_name.ilike(like),
                Contract.remark.ilike(like),
            ))
        if status in ('active', 'inactive'):
            query = query.filter(Contract.status == status)
        sort_map = {
            'contract_no': Contract.contract_no,
            'project_name': Contract.project_name,
            'status': Contract.status,
            'created_at': Contract.created_at,
        }
        sort_col = sort_map.get(sort_by, Contract.created_at)
        query = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        return render_template('contract.html', contracts=pagination.items, pagination=pagination,
                               filters={'search': search, 'status': status},
                               sort_by=sort_by, sort_order=sort_order, per_page=per_page)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/contract/add', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_contract():
        from app import Contract, api_error
        contract_no = request.form.get('contract_no', '').strip()
        if not contract_no:
            return api_error('请输入合同编号')
        project_name = request.form.get('project_name', '').strip()
        if not project_name:
            return api_error('请输入工程名称')
        if Contract.query.filter_by(contract_no=contract_no).first():
            return api_error('合同编号已存在')
        contract = Contract(
            contract_no=contract_no,
            project_name=project_name,
            status=request.form.get('status', 'active'),
            remark=request.form.get('remark', '').strip() or None
        )
        try:
            db.session.add(contract)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'新增合同失败: {e}')
            return jsonify({'status': 'error', 'msg': '新增失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '新增成功'})

    @app.route('/contract/<int:id>')
    @login_required
    def get_contract(id):
        from app import Contract
        contract = db.session.get(Contract, id)
        if not contract:
            return jsonify({'status': 'error', 'msg': '合同不存在'}), 404
        return jsonify({
            'status': 'success',
            'contract': {
                'id': contract.id,
                'contract_no': contract.contract_no,
                'project_name': contract.project_name,
                'status': contract.status,
                'remark': contract.remark
            }
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/contract/<int:id>/edit', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def edit_contract(id):
        from app import Contract, api_error
        contract = db.session.get(Contract, id)
        if not contract:
            return jsonify({'status': 'error', 'msg': '合同不存在'}), 404
        contract_no = request.form.get('contract_no', '').strip()
        if not contract_no:
            return api_error('请输入合同编号')
        project_name = request.form.get('project_name', '').strip()
        if not project_name:
            return api_error('请输入工程名称')
        existing = Contract.query.filter(Contract.contract_no == contract_no, Contract.id != id).first()
        if existing:
            return api_error('合同编号已存在')
        contract.contract_no = contract_no
        contract.project_name = project_name
        contract.status = request.form.get('status', 'active')
        contract.remark = request.form.get('remark', '').strip() or None
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'编辑合同失败: {e}')
            return jsonify({'status': 'error', 'msg': '编辑失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '编辑成功'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/contract/<int:id>/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_contract(id):
        from app import Contract, _contract_delete_blockers, _format_delete_blockers, api_error
        contract = db.session.get(Contract, id)
        if not contract:
            return jsonify({'status': 'error', 'msg': '合同不存在'}), 404
        blockers = _contract_delete_blockers(contract)
        if blockers:
            return api_error('该合同已有业务数据，不能删除：' + _format_delete_blockers(blockers))
        try:
            db.session.delete(contract)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'删除合同失败: {e}')
            return jsonify({'status': 'error', 'msg': '删除失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '删除成功'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/contract/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def batch_delete_contract_master():
        from app import Contract, _contract_delete_blockers, _format_delete_blockers
        ids = (request.get_json(silent=True) or {}).get('ids', [])
        deleted = 0
        blocked = []
        for item_id in ids:
            contract = db.session.get(Contract, int(item_id)) if str(item_id).isdigit() else None
            if contract:
                blockers = _contract_delete_blockers(contract)
                if blockers:
                    blocked.append(f'{contract.contract_no}（{_format_delete_blockers(blockers)}）')
                    continue
                db.session.delete(contract)
                deleted += 1
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'批量删除合同失败: {e}')
            return jsonify({'status': 'error', 'msg': '删除失败，请稍后重试'}), 500
        if blocked:
            return jsonify({
                'status': 'error' if deleted == 0 else 'success',
                'msg': f'已删除 {deleted} 个合同，以下合同有关联数据未删除：' + '；'.join(blocked)
            })
        return jsonify({'status': 'success', 'msg': f'已删除 {deleted} 个合同'})

    @app.route('/contract/api/list')
    @login_required
    def contract_api_list():
        # 合同列表API - 只返回启用的合同（给下拉用）
        from app import Contract
        contracts = Contract.query.filter_by(status='active').order_by(Contract.contract_no).all()
        return jsonify({
            'contracts': [{
                'id': c.id,
                'contract_no': c.contract_no,
                'project_name': c.project_name
            } for c in contracts]
        })

    @app.route('/api/contracts')
    @login_required
    def api_contracts_search():
        # 合同搜索API - 给单据头合同搜索框用，支持 keyword 模糊匹配 contract_no 或 project_name
        from app import Contract
        keyword = (request.args.get('keyword') or request.args.get('q') or '').strip()
        query = Contract.query
        if keyword:
            like = f'%{keyword}%'
            query = query.filter(db.or_(
                Contract.contract_no.ilike(like),
                Contract.project_name.ilike(like),
            ))
        contracts = query.order_by(Contract.contract_no).limit(50).all()
        # 返回标准信封 {status, data}，与 api.js 统一请求层对齐；
        # 否则 WMS.api.getContracts() 会因缺少 status==='success' 被当作业务失败而 reject。
        return jsonify({
            'status': 'success',
            'data': {
                'contracts': [{
                    'id': c.id,
                    'contract_no': c.contract_no,
                    'project_name': c.project_name,
                    'remark': c.remark or ''
                } for c in contracts]
            }
        })

    @app.route('/contract/download_template')
    @login_required
    def download_contract_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '合同档案导入模板'
        ws.append(['合同编号', '工程名称', '状态', '备注'])
        ws.append(['HD260713', '厚街医院新医疗综合大楼变配电工程', 'active', ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='contract_template.xlsx', as_attachment=True)

    @app.route('/contract/export')
    @login_required
    def export_contract():
        from openpyxl import Workbook
        from app import Contract
        wb = Workbook()
        ws = wb.active
        ws.title = '合同档案数据'
        ws.append(['合同编号', '工程名称', '状态', '备注', '创建时间'])
        search = (request.args.get('search') or '').strip()
        status = (request.args.get('status') or '').strip()
        query = Contract.query
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(
                Contract.contract_no.ilike(like),
                Contract.project_name.ilike(like),
                Contract.remark.ilike(like),
            ))
        if status in ('active', 'inactive'):
            query = query.filter(Contract.status == status)
        for contract in query.order_by(Contract.created_at.desc()).all():
            status_label = '启用' if contract.status == 'active' else ('停用' if contract.status == 'inactive' else (contract.status or ''))
            ws.append([
                contract.contract_no,
                contract.project_name,
                status_label,
                contract.remark or '',
                contract.created_at.strftime('%Y-%m-%d %H:%M') if contract.created_at else ''
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='contracts.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/contract/import', methods=['POST'])
    @require_role('admin')
    @login_required
    def import_contract():
        # F-03：合同/工程批量导入。表头需含：合同编号、工程名称、状态、备注。按合同编号查重，已存在则更新。
        from openpyxl import load_workbook
        from app import Contract, validate_excel_extension, validate_excel_size
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'status': 'error', 'msg': '请选择文件'}), 400
        _ext_ok, _ext_msg = validate_excel_extension(f.filename)
        if not _ext_ok:
            return jsonify({'status': 'error', 'msg': _ext_msg or '仅支持 .xlsx / .xls 文件'}), 400
        # m-03：限制 Excel 上传 ≤ 5MB
        _size_ok, _size_msg = validate_excel_size(f)
        if not _size_ok:
            return jsonify({'status': 'error', 'msg': _size_msg}), 400
        try:
            wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return jsonify({'status': 'error', 'msg': '文件无数据行'}), 400
            header = [str(c).strip() if c else '' for c in rows[0]]
            required = ['合同编号', '工程名称']
            missing = [c for c in required if c not in header]
            if missing:
                return jsonify({'status': 'error', 'msg': f'缺少必要列: {missing}'}), 400
            idx = {col: header.index(col) for col in header if col}
            added, updated, skipped = 0, 0, 0
            for r in rows[1:]:
                if not r or not r[0]:
                    skipped += 1
                    continue
                d = {col: (r[i] if i < len(r) and r[i] is not None else '') for col, i in idx.items()}
                cno = str(d.get('合同编号', '')).strip()
                pname = str(d.get('工程名称', '')).strip()
                if not cno or not pname:
                    skipped += 1
                    continue
                status_raw = str(d.get('状态', '')).strip()
                status = 'active' if status_raw in ('', '启用', 'active') else ('inactive' if status_raw in ('停用', 'inactive') else status_raw)
                remark = str(d.get('备注', '')).strip() or None
                existing = Contract.query.filter_by(contract_no=cno).first()
                if existing:
                    existing.project_name = pname
                    existing.status = status
                    existing.remark = remark
                    updated += 1
                else:
                    db.session.add(Contract(contract_no=cno, project_name=pname, status=status, remark=remark))
                    added += 1
            db.session.commit()
            msg = f'导入完成：新增 {added} 条，更新 {updated} 条'
            if skipped:
                msg += f'，跳过空行 {skipped} 条'
            return jsonify({'status': 'success', 'msg': msg})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'合同导入失败: {e}')
            return jsonify({'status': 'error', 'msg': f'导入失败：{e}'}), 500