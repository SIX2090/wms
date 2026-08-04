#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 客户（customer）域路由。
#
# 批量拆分模式：与员工/供应商/物料分类域一致，采用「register_<domain>_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 customer_list），与 app.py 内
# 原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（Customer 模型、OutOrder/AfterSaleOutOrder/InOrder/SalesOrder
#   模型、辅助函数等）在各路由函数内延迟导入（请求期才执行），避免 app.py 模块
#   加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import current_app, jsonify, render_template, request, send_file
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 customer_* 各路由测试覆盖
def register_customer_routes(app):
    @app.route('/customer')
    @login_required
    def customer_list():
        from app import (
            Customer,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
        )
        search, status_filter, sort_by, sort_order = _get_master_list_filters('code')
        allowed_sorts = {'id', 'code', 'name', 'contact', 'phone', 'created_at'}
        query = _apply_simple_search(Customer.query, Customer, search, ['code', 'name', 'contact', 'phone', 'address'])
        query, sort_by = _apply_master_order(query, Customer, sort_by, sort_order, allowed_sorts, 'code')
        customers = query.all()
        return render_template('customer.html', customers=customers, filters={'search': search, 'status': status_filter}, sort_by=sort_by, sort_order=sort_order)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/customer/add', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_customer():
        from app import Customer, api_error, sanitize_text_input
        # BUG-2026-07-29-002/009: 客户主数据字段走 sanitize_text_input
        code = sanitize_text_input(request.form.get('code'), max_len=50)
        name = sanitize_text_input(request.form.get('name'), max_len=100)
        if not code:
            return api_error('请输入客户编号')
        if not name:
            return api_error('请输入客户名称')
        # BUG-F02-02 修复：客户主数据长度截断防护（与供应商一致）
        if len(code) > 50:
            return api_error(f'客户编号不能超过 50 个字符（当前 {len(code)}）')
        if len(name) > 100:
            return api_error(f'客户名称不能超过 100 个字符（当前 {len(name)}）')
        contact = sanitize_text_input(request.form.get('contact'), max_len=50)
        if len(contact) > 50:
            return api_error(f'联系人不能超过 50 个字符（当前 {len(contact)}）')
        phone = sanitize_text_input(request.form.get('phone'), max_len=20)
        if len(phone) > 20:
            return api_error(f'电话不能超过 20 个字符（当前 {len(phone)}）')
        address = sanitize_text_input(request.form.get('address'), max_len=200)
        if len(address) > 200:
            return api_error(f'地址不能超过 200 个字符（当前 {len(address)}）')
        if Customer.query.filter_by(code=code).first():
            return api_error('客户编号已存在')
        if Customer.query.filter_by(name=name).first():
            return api_error('客户名称已存在')
        customer = Customer(
            code=code,
            name=name,
            contact=contact or None,
            phone=phone or None,
            address=address or None
        )
        db.session.add(customer)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"数据库操作失败: {e}")
            return api_error("操作失败", code=500)
        return jsonify({'status': 'success', 'id': customer.id, 'name': customer.name})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/customer/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_customer():
        from app import (
            AfterSaleOutOrder,
            Customer,
            InOrder,
            OutOrder,
            SalesOrder,
            api_error,
        )
        ids = request.json.get('ids', [])
        for id in ids:
            customer = db.session.get(Customer, id)
            if customer:
                # 客户在出库单/售后出库单中以名称字符串引用（非外键），
                # 直接删除会让历史单据的客户名失去主数据支撑（下拉选不到、报表对不上），
                # 因此需要检查是否有单据使用该客户名
                if OutOrder.query.filter(OutOrder.customer == customer.name).first():
                    return api_error(f'客户“{customer.name}”已有关联出库单，不能删除')
                if AfterSaleOutOrder.query.filter(AfterSaleOutOrder.customer == customer.name).first():
                    return api_error(f'客户“{customer.name}”已有关联售后出库单，不能删除')
                # m-01：采购入库单/销售订单等以 customer_id 外键引用时也要校验，避免硬删后外键悬空
                if hasattr(InOrder, 'customer_id') and \
                        InOrder.query.filter_by(customer_id=customer.id).count() > 0:
                    return jsonify({'status': 'error',
                                    'msg': f'客户“{customer.name}”已被采购入库单(其他入库)引用，不能删除'})
                if hasattr(SalesOrder, 'customer_id') and \
                        SalesOrder.query.filter_by(customer_id=customer.id).count() > 0:
                    return jsonify({'status': 'error',
                                    'msg': f'客户“{customer.name}”已被销售订单引用，不能删除'})
                if hasattr(AfterSaleOutOrder, 'customer_id') and \
                        AfterSaleOutOrder.query.filter_by(customer_id=customer.id).count() > 0:
                    return jsonify({'status': 'error',
                                    'msg': f'客户“{customer.name}”已被售后出库单引用，不能删除'})
                db.session.delete(customer)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success'})

    @app.route('/customer/<int:customer_id>')
    @require_role('warehouse')
    @login_required
    def get_customer(customer_id):
        """M-01：行级编辑 - 返回客户详情 JSON。"""
        from app import Customer
        customer = db.session.get(Customer, customer_id)
        if not customer:
            return jsonify({'status': 'error', 'msg': '客户不存在'}), 404
        return jsonify({
            'status': 'success',
            'customer': {
                'id': customer.id, 'code': customer.code, 'name': customer.name,
                'contact': customer.contact or '', 'phone': customer.phone or '',
                'address': customer.address or ''
            }
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/customer/<int:customer_id>/edit', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def edit_customer(customer_id):
        """M-01：客户行级编辑。"""
        from app import Customer, api_error
        customer = db.session.get(Customer, customer_id)
        if not customer:
            return jsonify({'status': 'error', 'msg': '客户不存在'}), 404
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        if not code:
            return api_error('请输入客户编号')
        if not name:
            return api_error('请输入客户名称')
        # BUG-F02-02 修复：客户编辑入口同样 5 字段长度校验
        if len(code) > 50:
            return jsonify({'status': 'error', 'msg': f'客户编号不能超过 50 个字符（当前 {len(code)}）'}), 400
        if len(name) > 100:
            return jsonify({'status': 'error', 'msg': f'客户名称不能超过 100 个字符（当前 {len(name)}）'}), 400
        contact = (request.form.get('contact') or '').strip()
        if len(contact) > 50:
            return jsonify({'status': 'error', 'msg': f'联系人不能超过 50 个字符（当前 {len(contact)}）'}), 400
        phone = (request.form.get('phone') or '').strip()
        if len(phone) > 20:
            return jsonify({'status': 'error', 'msg': f'电话不能超过 20 个字符（当前 {len(phone)}）'}), 400
        address = (request.form.get('address') or '').strip()
        if len(address) > 200:
            return jsonify({'status': 'error', 'msg': f'地址不能超过 200 个字符（当前 {len(address)}）'}), 400
        dup = Customer.query.filter_by(code=code).first()
        if dup and dup.id != customer_id:
            return api_error('客户编号已存在')
        dup = Customer.query.filter_by(name=name).first()
        if dup and dup.id != customer_id:
            return api_error('客户名称已存在')
        customer.code = code
        customer.name = name
        customer.contact = contact or None
        customer.phone = phone or None
        customer.address = address or None
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'编辑客户失败: {e}')
            return jsonify({'status': 'error', 'msg': '编辑失败'}), 500
        return jsonify({'status': 'success', 'msg': '客户编辑成功'})

    @app.route('/customer/download_template')
    @login_required
    def download_customer_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '客户导入模板'
        ws.append(['客户编号', '客户名称', '联系人', '电话', '地址'])
        ws.append(['CUS-001', '示例客户', '张三', '13800138000', '广州市天河区'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='customer_template.xlsx', as_attachment=True)

    @app.route('/customer/export')
    @login_required
    def export_customer():
        from openpyxl import Workbook
        from app import (
            Customer,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '客户数据'
        ws.append(['客户编号', '客户名称', '联系人', '电话', '地址'])
        search, status_filter, sort_by, sort_order = _get_master_list_filters('code')
        query = _apply_simple_search(Customer.query, Customer, search, ['code', 'name', 'contact', 'phone', 'address'])
        query, _ = _apply_master_order(query, Customer, sort_by, sort_order, {'id', 'code', 'name', 'contact', 'phone', 'created_at'}, 'code')
        for customer in query.all():
            ws.append([customer.code, customer.name, customer.contact or '', customer.phone or '', customer.address or ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='customers.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/customer/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_customer():
        from app import Customer, api_error, validate_excel_extension, validate_excel_size
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的客户文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row and row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not code or not name:
                    skip += 1
                    continue
                if Customer.query.filter_by(code=code).first() or Customer.query.filter_by(name=name).first():
                    skip += 1
                    continue
                customer = Customer(
                    code=code,
                    name=name,
                    contact=str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    phone=str(row[3]).strip() if len(row) > 3 and row[3] else '',
                    address=str(row[4]).strip() if len(row) > 4 and row[4] else ''
                )
                db.session.add(customer)
                count += 1
            db.session.commit()
            msg = f'客户导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条（重复或格式错误）'
            return jsonify({'status': 'success', 'msg': msg, 'count': count})
        except Exception:
            db.session.rollback()
            return api_error('客户导入失败')