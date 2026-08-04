#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 部门（department）域路由。
#
# 批量拆分模式：与员工/供应商/物料分类域一致，采用「register_department_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 department_list），与 app.py 内
# 原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（Department 模型、_department_query_from_args 等辅助函数）在各
#   路由函数内延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import current_app, jsonify, render_template, request, send_file
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 department_* 各路由测试覆盖
def register_department_routes(app):
    @app.route('/department')
    @login_required
    def department_list():
        from app import Department, _department_query_from_args
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        # per_page 必须有下限保护，传入 0 或负数会让 paginate 抛 ValueError 导致接口 500
        per_page = max(1, per_page)
        if per_page not in [10, 20, 50, 100, 200]:
            per_page = 20
        query, filters, sort_by, sort_order = _department_query_from_args()
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        departments = pagination.items
        return render_template('department.html', departments=departments, pagination=pagination, filters=filters, sort_by=sort_by, sort_order=sort_order, per_page=per_page)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/department/add', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_department():
        from app import Department, api_error
        code = request.form.get('code', '').strip()
        if not code:
            return api_error('请输入部门编码')

        name = request.form.get('name', '').strip()
        if not name:
            return api_error('请输入部门名称')

        if Department.query.filter_by(code=code).first():
            return api_error('部门编码已存在')

        if Department.query.filter_by(name=name).first():
            return api_error('部门名称已存在')

        department = Department(
            code=code,
            name=name,
            status=request.form.get('status', 'active'),
            remark=request.form.get('remark', '').strip() or None
        )
        try:
            db.session.add(department)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'新增部门失败: {e}')
            return jsonify({'status': 'error', 'msg': '新增失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '新增成功'})

    @app.route('/department/<int:id>')
    @login_required
    def get_department(id):
        from app import Department
        department = db.session.get(Department, id)
        if not department:
            return jsonify({'status': 'error', 'msg': '部门不存在'}), 404

        return jsonify({
            'status': 'success',
            'department': {
                'id': department.id,
                'code': department.code,
                'name': department.name,
                'status': department.status,
                'remark': department.remark
            }
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/department/<int:id>/edit', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def edit_department(id):
        from app import Department, api_error
        department = db.session.get(Department, id)
        if not department:
            return jsonify({'status': 'error', 'msg': '部门不存在'}), 404

        code = request.form.get('code', '').strip()
        if not code:
            return api_error('请输入部门编码')

        name = request.form.get('name', '').strip()
        if not name:
            return api_error('请输入部门名称')

        # 检查编码是否重复（排除自己）
        existing = Department.query.filter(Department.code == code, Department.id != id).first()
        if existing:
            return api_error('部门编码已存在')

        # 检查名称是否重复（排除自己）
        existing = Department.query.filter(Department.name == name, Department.id != id).first()
        if existing:
            return api_error('部门名称已存在')

        department.code = code
        department.name = name
        department.status = request.form.get('status', 'active')
        department.remark = request.form.get('remark', '').strip() or None

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'编辑部门失败: {e}')
            return jsonify({'status': 'error', 'msg': '编辑失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '编辑成功'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/department/<int:id>/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_department(id):
        from app import Department, _department_delete_blockers, _format_delete_blockers, api_error
        department = db.session.get(Department, id)
        if not department:
            return jsonify({'status': 'error', 'msg': '部门不存在'}), 404

        blockers = _department_delete_blockers(department)
        if blockers:
            return api_error('该部门已有业务数据，不能删除：' + _format_delete_blockers(blockers))

        try:
            db.session.delete(department)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'删除部门失败: {e}')
            return jsonify({'status': 'error', 'msg': '删除失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '删除成功'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/department/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def batch_delete_department_master():
        from app import Department, _department_delete_blockers, _format_delete_blockers
        ids = (request.get_json(silent=True) or {}).get('ids', [])
        deleted = 0
        blocked = []
        for item_id in ids:
            department = db.session.get(Department, int(item_id)) if str(item_id).isdigit() else None
            if department:
                blockers = _department_delete_blockers(department)
                if blockers:
                    blocked.append(f'{department.name}（{_format_delete_blockers(blockers)}）')
                    continue
                db.session.delete(department)
                deleted += 1
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'批量删除部门失败: {e}')
            return jsonify({'status': 'error', 'msg': '删除失败，请稍后重试'}), 500
        if blocked:
            return jsonify({
                'status': 'error' if deleted == 0 else 'success',
                'msg': f'已删除 {deleted} 个部门，以下部门有关联数据未删除：' + '；'.join(blocked)
            })
        return jsonify({'status': 'success', 'msg': f'已删除 {deleted} 个部门'})

    @app.route('/department/api/list')
    @login_required
    def department_api_list():
        """部门列表API - 只返回启用的部门"""
        from app import Department
        departments = Department.query.filter_by(status='active').order_by(Department.code).all()
        return jsonify({
            'departments': [{
                'id': d.id,
                'code': d.code,
                'name': d.name
            } for d in departments]
        })

    @app.route('/department/download_template')
    @login_required
    def download_department_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '部门导入模板'
        ws.append(['部门编码', '部门名称', '状态', '备注'])
        ws.append(['D001', '生产部', 'active', ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='department_template.xlsx', as_attachment=True)

    @app.route('/department/export')
    @login_required
    def export_department():
        from openpyxl import Workbook
        from app import _department_query_from_args
        wb = Workbook()
        ws = wb.active
        ws.title = '部门数据'
        ws.append(['部门编码', '部门名称', '状态', '备注'])
        query, _, _, _ = _department_query_from_args()
        for department in query.all():
            status_label = '启用' if department.status == 'active' else ('停用' if department.status == 'inactive' else (department.status or ''))
            ws.append([department.code, department.name, status_label, department.remark or ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='departments.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/department/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_department():
        from app import Department, api_error, validate_excel_extension, validate_excel_size
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的部门文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row and row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not code or not name:
                    skip += 1
                    continue
                if Department.query.filter_by(code=code).first() or Department.query.filter_by(name=name).first():
                    skip += 1
                    continue
                department = Department(
                    code=code,
                    name=name,
                    status=str(row[2]).strip() if len(row) > 2 and row[2] else 'active',
                    remark=str(row[3]).strip() if len(row) > 3 and row[3] else ''
                )
                db.session.add(department)
                count += 1
            db.session.commit()
            msg = f'部门导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条（重复或格式错误）'
            return jsonify({'status': 'success', 'msg': msg, 'count': count})
        except Exception:
            db.session.rollback()
            return api_error('部门导入失败')