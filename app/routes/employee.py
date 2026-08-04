#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 员工（employee）域路由。
#
# 批量拆分模式：与供应商/物料分类域一致，采用「register_<domain>_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 employee_list），与 app.py 内
# 原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（Employee 模型、Department 模型、辅助函数等）在各路由函数内
#   延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import current_app, jsonify, render_template, request, send_file
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 employee_* 各路由测试覆盖
def register_employee_routes(app):
    @app.route('/employee')
    @login_required
    def employee_list():
        from app import (
            Department,
            Employee,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
        )
        search, status_filter, sort_by, sort_order = _get_master_list_filters('name')
        allowed_sorts = {'id', 'code', 'name', 'position', 'phone', 'created_at'}
        query = _apply_simple_search(Employee.query, Employee, search, ['name', 'position', 'phone', 'code'])
        # M-04：支持按部门筛选
        dept_filter = (request.args.get('department_id') or '').strip()
        if dept_filter:
            try:
                query = query.filter(Employee.department_id == int(dept_filter))
            except ValueError:
                pass
        query, sort_by = _apply_master_order(query, Employee, sort_by, sort_order, allowed_sorts, 'name')
        employees = query.all()
        departments = Department.query.order_by(Department.code.asc()).all()
        return render_template('employee.html', employees=employees, departments=departments,
                               filters={'search': search, 'status': status_filter, 'department_id': dept_filter},
                               sort_by=sort_by, sort_order=sort_order)

    @app.route('/employee/<int:employee_id>')
    @login_required
    def get_employee(employee_id):
        """M-01：行级编辑 - 返回员工详情 JSON。"""
        from app import Employee
        emp = db.session.get(Employee, employee_id)
        if not emp:
            return jsonify({'status': 'error', 'msg': '员工不存在'}), 404
        return jsonify({
            'status': 'success',
            'employee': {
                'id': emp.id,
                'code': emp.code or '',
                'name': emp.name,
                'position': emp.position or '',
                'phone': emp.phone or '',
                'department_id': emp.department_id or '',
            }
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/employee/add', methods=['POST'])
    @require_role('admin')
    @login_required
    def add_employee():
        from app import Department, Employee
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        if not name:
            return jsonify({"status": "error", "msg": "姓名不能为空"}), 400
        # M-04：编码唯一性校验
        if code:
            if Employee.query.filter_by(code=code).first():
                return jsonify({"status": "error", "msg": f"员工编码“{code}”已存在"}), 400
        dept_id = request.form.get('department_id', '').strip()
        dept_id_int = None
        if dept_id:
            try:
                dept_id_int = int(dept_id)
                if not db.session.get(Department, dept_id_int):
                    return jsonify({"status": "error", "msg": "所选部门不存在"}), 400
            except ValueError:
                return jsonify({"status": "error", "msg": "部门参数非法"}), 400
        employee = Employee(
            code=code or None,
            name=name,
            position=request.form.get('position'),
            phone=request.form.get('phone'),
            department_id=dept_id_int,
        )
        db.session.add(employee)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/employee/<int:employee_id>/edit', methods=['POST'])
    @require_role('admin')
    @login_required
    def edit_employee(employee_id):
        """M-01：员工行级编辑。"""
        from app import Department, Employee
        emp = db.session.get(Employee, employee_id)
        if not emp:
            return jsonify({'status': 'error', 'msg': '员工不存在'}), 404
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        if not name:
            return jsonify({'status': 'error', 'msg': '姓名不能为空'}), 400
        if code:
            dup = Employee.query.filter_by(code=code).first()
            if dup and dup.id != emp.id:
                return jsonify({'status': 'error', 'msg': f'员工编码“{code}”已存在'}), 400
        dept_id = request.form.get('department_id', '').strip()
        dept_id_int = None
        if dept_id:
            try:
                dept_id_int = int(dept_id)
                if not db.session.get(Department, dept_id_int):
                    return jsonify({'status': 'error', 'msg': '所选部门不存在'}), 400
            except ValueError:
                return jsonify({'status': 'error', 'msg': '部门参数非法'}), 400
        emp.code = code or None
        emp.name = name
        emp.position = request.form.get('position')
        emp.phone = request.form.get('phone')
        emp.department_id = dept_id_int
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'编辑员工失败: {e}')
            return jsonify({'status': 'error', 'msg': '操作失败'}), 500
        return jsonify({'status': 'success'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/employee/delete', methods=['POST'])
    @require_role('admin')
    @login_required
    def delete_employee():
        from app import Employee, SalesOrder
        ids = request.json.get('ids', [])
        # F-01：员工被业务单据引用时禁止硬删
        # SalesOrder.salesperson_id 为 FK employee.id；其他单据 operator_id 引用 user.id（不在此校验）
        for id in ids:
            emp = db.session.get(Employee, id)
            if emp:
                blockers = []
                if hasattr(SalesOrder, 'salesperson_id'):
                    n = SalesOrder.query.filter_by(salesperson_id=emp.id).count()
                    if n:
                        blockers.append(f'销售订单(salesperson_id) 引用 {n} 次')
                if blockers:
                    return jsonify({'status': 'error',
                                    'msg': f'员工“{emp.name}”已被业务单据引用，禁止删除：\n' + '\n'.join(blockers)}), 409
                db.session.delete(emp)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"数据库操作失败: {e}")
            return jsonify({"status": "error", "msg": "操作失败"}), 500
        return jsonify({'status': 'success'})

    @app.route('/employee/download_template')
    @login_required
    def download_employee_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '员工导入模板'
        ws.append(['员工编码', '姓名', '职位', '电话', '部门编码'])
        ws.append(['E001', '张三', '工程师', '13800138000', 'D001'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='employee_template.xlsx', as_attachment=True)

    @app.route('/employee/export')
    @login_required
    def export_employee():
        from openpyxl import Workbook
        from app import (
            Department,
            Employee,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '员工数据'
        ws.append(['员工编码', '姓名', '职位', '电话', '部门编码', '部门名称'])
        search, status_filter, sort_by, sort_order = _get_master_list_filters('name')
        query = _apply_simple_search(Employee.query, Employee, search, ['name', 'position', 'phone', 'code'])
        dept_filter = (request.args.get('department_id') or '').strip()
        if dept_filter:
            try:
                query = query.filter(Employee.department_id == int(dept_filter))
            except ValueError:
                pass
        query, _ = _apply_master_order(query, Employee, sort_by, sort_order, {'id', 'code', 'name', 'position', 'phone', 'created_at'}, 'name')
        for e in query.all():
            dept = e.department
            ws.append([e.code or '', e.name, e.position or '', e.phone or '',
                       dept.code if dept else '', dept.name if dept else ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='employees.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/employee/import', methods=['POST'])
    @require_role('admin')
    @login_required
    def import_employee():
        """M-04：员工批量导入，按表头名匹配列。支持：员工编码、姓名、职位、电话、部门编码。
        姓名必填；员工编码可空，已存在则更新；部门编码按 Department.code 匹配。"""
        from app import (
            Department,
            Employee,
            api_error,
            validate_excel_extension,
            validate_excel_size,
        )
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的员工文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return jsonify({'status': 'error', 'msg': '文件无数据行'}), 400
            header = [str(c).strip() if c else '' for c in rows[0]]
            if '姓名' not in header:
                return jsonify({'status': 'error', 'msg': '缺少必要列：姓名'}), 400
            idx = {col: header.index(col) for col in header if col}
            added, updated, skipped = 0, 0, 0
            for r in rows[1:]:
                if not r or not r[idx['姓名']]:
                    skipped += 1
                    continue
                d = {col: (r[i] if i < len(r) and r[i] is not None else '') for col, i in idx.items()}
                code = str(d.get('员工编码', '')).strip()
                name = str(d.get('姓名', '')).strip()
                position = str(d.get('职位', '')).strip() or None
                phone = str(d.get('电话', '')).strip() or None
                dept_code = str(d.get('部门编码', '')).strip()
                dept_id_int = None
                if dept_code:
                    dept = Department.query.filter_by(code=dept_code).first()
                    if dept:
                        dept_id_int = dept.id
                if code:
                    existing = Employee.query.filter_by(code=code).first()
                    if existing:
                        existing.name = name
                        existing.position = position
                        existing.phone = phone
                        existing.department_id = dept_id_int
                        updated += 1
                        continue
                db.session.add(Employee(
                    code=code or None,
                    name=name,
                    position=position or '',
                    phone=phone or '',
                    department_id=dept_id_int,
                ))
                added += 1
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f'员工导入失败: {e}')
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            msg = f'导入完成：新增 {added} 条，更新 {updated} 条'
            if skipped:
                msg += f'，跳过空行 {skipped} 条'
            return jsonify({'status': 'success', 'msg': msg})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'员工导入异常: {e}')
            return jsonify({'status': 'error', 'msg': f'导入失败：{e}'}), 500