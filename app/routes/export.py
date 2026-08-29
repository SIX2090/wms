#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 导出（export）域路由。
#
# 批量拆分模式：为避免 endpoint 前缀化导致大量 url_for 引用改动，
# 采用「register_<domain>_routes(app)」直接在 app 上注册路由，endpoint 名保持不变
# （如 export_in_order），与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（模型、_workbook_response、_get_order_list_filters 等）在各路由
#   函数内延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import request, send_file
from flask_login import login_required


# no-test:reason=路由注册辅助函数，能力由 export_* 各路由测试覆盖
def register_export_routes(app):
    @app.route('/export/template/bom')
    @login_required
    def export_bom_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'BOM导入模板'
        ws.append(['产品编码', '产品名称', '版本', '物料编码', '物料名称', '规格', '单位', '数量', '用途', '备注'])
        ws.append(['PROD001', '示例产品', '1.0', 'MAT001', '示例材料A', '规格A', '个', 2, '主料', ''])
        ws.append(['PROD001', '示例产品', '1.0', 'MAT002', '示例材料B', '规格B', '个', 1, '辅料', ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='bom_import_template.xlsx', as_attachment=True)

    @app.route('/export/template/requisition')
    @login_required
    def export_requisition_template():
        from app import _workbook_response
        return _workbook_response(
            'requisition_template.xlsx',
            '工单领料导入模板',
            ['单据编号', '日期', '工单', '用途', 'BOM编号', '物料编码', '物料名称', '规格', '单位', '数量', '备注'],
            [['REQ24010001', '2024-01-01', 'MO24010001', '工单领料', '', 'MAT001', '示例物料', '规格A', '个', 10, '']],
        )

    @app.route('/export/template/subcontract')
    @app.route('/subcontract/download_template')
    @login_required
    def export_subcontract_template():
        from app import _workbook_response
        return _workbook_response(
            'subcontract_template.xlsx',
            '委外加工导入模板',
            ['单据编号', '日期', '加工厂商', '联系人', '电话', '交货期限', '物料编码', '物料名称', '规格', '单位', '数量', '备注'],
            [['SC24010001', '2024-01-01', '示例加工厂', '王五', '13800138000', '2024-01-10', 'MAT001', '示例物料', '规格A', '个', 10, '']],
        )

    @app.route('/export/template/subcontract_issue')
    @app.route('/subcontract_issue/download_template')
    @login_required
    def export_subcontract_issue_template():
        from app import _workbook_response
        return _workbook_response(
            'subcontract_issue_template.xlsx',
            '委外发料导入模板',
            ['发料单号', '日期', '委外加工单号', '加工厂商', '物料编码', '物料名称', '规格', '单位', '数量', '备注'],
            [['SF24010001', '2024-01-01', 'SC24010001', '示例加工厂', 'MAT001', '示例物料', '规格A', '个', 10, '']],
        )

    @app.route('/export/template/subcontract_receive')
    @app.route('/subcontract_receive/download_template')
    @login_required
    def export_subcontract_receive_template():
        from app import _workbook_response
        return _workbook_response(
            'subcontract_receive_template.xlsx',
            '委外入库导入模板',
            ['入库单号', '日期', '委外加工单号', '加工厂商', '物料编码', '物料名称', '规格', '单位', '收货数量', '报废数量', '单价', '备注'],
            [['SR24010001', '2024-01-01', 'SC24010001', '示例加工厂', 'MAT001', '示例物料', '规格A', '个', 10, 0, 5, '']],
        )

    @app.route('/export/template/adjustment')
    @login_required
    def export_adjustment_template():
        from app import _workbook_response
        return _workbook_response(
            'adjustment_template.xlsx',
            '库存调整导入模板',
            ['单据编号', '日期', '调整类型', '物料编码', '物料名称', '规格', '单位', '数量', '库位', '原因', '备注'],
            [['ADJ24010001', '2024-01-01', '盘盈', 'MAT001', '示例物料', '规格A', '个', 10, 'A01', '盘点差异', '']],
        )

    @app.route('/export/template/check')
    @login_required
    def export_check_template():
        from app import _workbook_response
        return _workbook_response(
            'check_template.xlsx',
            '库存盘点导入模板',
            ['单据编号', '日期', '物料编码', '物料名称', '规格', '单位', '系统库存', '实际库存', '差异原因', '备注'],
            [['CK24010001', '2024-01-01', 'MAT001', '示例物料', '规格A', '个', 100, 98, '盘点差异', '']],
        )

    @app.route('/export/template/purchase_request')
    @login_required
    def export_purchase_request_template():
        from app import _workbook_response
        return _workbook_response(
            'purchase_request_template.xlsx',
            '采购申请导入模板',
            ['申请编号', '日期', '申请人', '部门', '紧急程度', '期望到货', '申请原因', '物料编码', '物料名称', '规格', '单位', '数量', '预估单价', '推荐供应商', '备注'],
            [['PR24010001', '2024-01-01', '张三', '采购部', '普通', '2024-01-10', '采购备货', 'MAT001', '示例物料', '规格A', '个', 10, 5, '示例供应商', '']],
        )

    @app.route('/export/template/purchase_order')
    @login_required
    def export_purchase_order_template():
        from app import _workbook_response
        return _workbook_response(
            'purchase_order_template.xlsx',
            '采购单导入模板',
            ['采购单号', '日期', '供应商', '预计到货', '物料编码', '物料名称', '规格', '单位', '数量', '单价', '备注'],
            [['PO24010001', '2024-01-01', '示例供应商', '2024-01-10', 'MAT001', '示例物料', '规格A', '个', 10, 5, '']],
        )

    @app.route('/export/in_order')
    @app.route('/in_order/export')
    @login_required
    def export_in_order():
        from openpyxl import Workbook
        from sqlalchemy.orm import joinedload, selectinload
        from app import (
            InOrder,
            InOrderItem,
            Material,
            _apply_header_or_item_contract_filters,
            _apply_order_partner_text_filter,
            _apply_in_order_search,
            _apply_status_date_filters,
            _get_order_list_filters,
            resolve_request_warehouse,
        )
        from db import db
        wb = Workbook()
        ws = wb.active
        ws.title = '入库单'
        ws.append(['单据编号', '日期', '业务类型', '用途', '供应商', '客户', '仓库', '库位', '物料编码', '物料名称', '规格', '单位', '数量', '单价', '金额', '客供', '合同编号', '工程名称', '状态', '备注'])
        status_filter, search, date_start, date_end, sort_by, sort_order = _get_order_list_filters(('pending', 'completed'))
        type_alias = {'purchase_in': '采购入库', 'product_in': '产品入库', 'other_in': '其他入库'}
        business_type_filter = type_alias.get((request.args.get('type') or '').strip(), (request.args.get('business_type') or '').strip())
        if business_type_filter not in ('采购入库', '产品入库', '其他入库'):
            business_type_filter = ''
        contract_no_filter = (request.args.get('contract_no') or '').strip()
        project_name_filter = (request.args.get('project_name') or '').strip()
        allowed_sorts = {'order_no', 'date', 'supplier_id', 'purpose', 'status', 'created_at', 'total_amount'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        query = db.session.query(InOrder).outerjoin(InOrderItem, InOrderItem.in_order_id == InOrder.id).options(
            joinedload(InOrder.supplier),
            joinedload(InOrder.customer),
            selectinload(InOrder.items).joinedload(InOrderItem.material).joinedload(Material.unit),
        )
        query = _apply_status_date_filters(query, InOrder, status_filter, date_start, date_end)
        query = _apply_in_order_search(query, search)
        supplier_id = request.args.get('supplier_id', type=int) or 0
        # BUG-2026-08-29-003：与列表页一致，supplier_id 为空时按文本模糊匹配往来单位
        supplier_name_filter = (request.args.get('supplier_name') or '').strip()
        if supplier_id:
            query = query.filter(InOrder.supplier_id == supplier_id)
        else:
            query = _apply_order_partner_text_filter(query, InOrder, supplier_name_filter)
        if business_type_filter:
            query = query.filter(InOrder.business_type == business_type_filter)
        query = _apply_header_or_item_contract_filters(
            query, InOrder, InOrderItem, 'in_order_id',
            contract_no_filter=contract_no_filter,
            project_name_filter=project_name_filter,
        )
        warehouse, warehouse_error = resolve_request_warehouse(request.args)
        if warehouse_error and request.args.get('warehouse_id'):
            from app import api_error
            return api_error(warehouse_error, 400)
        if warehouse:
            query = query.filter(InOrder.warehouse == warehouse.name)
        else:
            query = query.filter(db.false())
        sort_col = getattr(InOrder, sort_by, InOrder.created_at)
        query = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc(), InOrder.id.desc()).distinct()
        orders = query.all()
        for order in orders:
            if order.items:
                for item in order.items:
                    ws.append([
                        order.order_no,
                        order.date.strftime('%Y-%m-%d') if order.date else '',
                        order.business_type or '采购入库',
                        order.purpose or '',
                        order.supplier.name if order.supplier else '',
                        order.customer.name if order.customer else '',
                        order.warehouse or '',
                        order.location or '',
                        item.material.code if item.material else '',
                        item.material.name if item.material else '',
                        item.material.spec if item.material else '',
                        item.material.unit.name if item.material and item.material.unit else '',
                        item.quantity or 0,
                        item.price or 0,
                        item.amount or 0,
                        '是' if item.is_customer_supplied else '否',
                        item.contract_no or order.contract_no or '',
                        item.project_name or order.project_name or '',
                        '未审核/待完成' if order.status == 'pending' else ('已完成' if order.status == 'completed' else (order.status or '')),
                        order.remark or ''
                    ])
            else:
                ws.append([
                    order.order_no,
                    order.date.strftime('%Y-%m-%d') if order.date else '',
                    order.purpose or '',
                    order.supplier.name if order.supplier else '',
                    order.warehouse or '',
                    '', '', '', '', 0, 0, 0,
                    '未审核/待完成' if order.status == 'pending' else ('已完成' if order.status == 'completed' else (order.status or '')),
                    order.remark or ''
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='in_orders.xlsx', as_attachment=True)

    @app.route('/export/purchase_request')
    @app.route('/purchase_request/export')
    @login_required
    def export_purchase_request():
        from openpyxl import Workbook
        from sqlalchemy.orm import selectinload
        from app import (
            Material,
            PurchaseRequest,
            PurchaseRequestItem,
            _apply_purchase_request_search,
            _apply_status_date_filters,
            _get_order_list_filters,
            _status_label,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '采购申请单'
        ws.append(['申请编号', '日期', '申请人', '部门', '紧急程度', '物料编码', '物料名称', '规格', '单位', '数量', '预估单价', '预估金额', '状态'])
        status_filter, search, date_start, date_end, sort_by, sort_order = _get_order_list_filters(('pending', 'approved', 'rejected', 'completed'))
        allowed_sorts = {'request_no', 'date', 'applicant', 'department', 'urgency', 'expected_date', 'status', 'created_at', 'total_amount'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        query = PurchaseRequest.query.options(
            selectinload(PurchaseRequest.items).joinedload(PurchaseRequestItem.material).joinedload(Material.unit)
        )
        query = _apply_status_date_filters(query, PurchaseRequest, status_filter, date_start, date_end)
        query = _apply_purchase_request_search(query, search)
        sort_col = getattr(PurchaseRequest, sort_by, PurchaseRequest.created_at)
        requests_list = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc()).all()
        for req in requests_list:
            if req.items:
                for item in req.items:
                    ws.append([
                        req.request_no,
                        req.date.strftime('%Y-%m-%d') if req.date else '',
                        req.applicant or '',
                        req.department or '',
                        req.urgency or '',
                        item.material.code if item.material else '',
                        item.material.name if item.material else '',
                        item.material.spec if item.material else '',
                        item.material.unit.name if item.material and item.material.unit else '',
                        item.quantity or 0,
                        item.estimated_price or 0,
                        item.estimated_amount or 0,
                        _status_label(req.status)
                    ])
            else:
                ws.append([
                    req.request_no,
                    req.date.strftime('%Y-%m-%d') if req.date else '',
                    req.applicant or '',
                    req.department or '',
                    req.urgency or '',
                    '', '', '', '', 0, 0, 0,
                    _status_label(req.status)
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='purchase_requests.xlsx', as_attachment=True)

    @app.route('/export/after_sale_out')
    @app.route('/after_sale_out/export')
    @login_required
    def export_after_sale_out():
        from openpyxl import Workbook
        from sqlalchemy.orm import selectinload
        from app import (
            AfterSaleOutOrder,
            AfterSaleOutOrderItem,
            Material,
            _apply_status_date_filters,
            _get_order_list_filters,
            _status_from_search_keyword,
        )
        from db import db
        wb = Workbook()
        ws = wb.active
        ws.title = '售后出库单'
        ws.append(['单据编号', '日期', '客户', '联系人', '电话', '原因', '物料编码', '物料名称', '规格', '单位', '数量', '单价', '金额', '状态', '备注'])
        status_filter, search, date_start, date_end, sort_by, sort_order = _get_order_list_filters(('pending', 'completed'))
        allowed_sorts = {'order_no', 'date', 'customer', 'reason', 'status', 'created_at', 'total_amount'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        query = AfterSaleOutOrder.query.options(
            selectinload(AfterSaleOutOrder.items).joinedload(AfterSaleOutOrderItem.material).joinedload(Material.unit)
        )
        query = _apply_status_date_filters(query, AfterSaleOutOrder, status_filter, date_start, date_end)
        if search:
            search_like = f'%{search}%'
            status_from_search = _status_from_search_keyword(search, ('pending', 'completed'))
            conditions = [
                AfterSaleOutOrder.order_no.like(search_like),
                AfterSaleOutOrder.customer.like(search_like),
                AfterSaleOutOrder.contact.like(search_like),
                AfterSaleOutOrder.phone.like(search_like),
                AfterSaleOutOrder.reason.like(search_like),
                AfterSaleOutOrder.remark.like(search_like),
                Material.code.like(search_like),
                Material.name.like(search_like),
                Material.spec.like(search_like),
            ]
            if status_from_search:
                conditions.append(AfterSaleOutOrder.status == status_from_search)
            query = query.outerjoin(
                AfterSaleOutOrderItem, AfterSaleOutOrderItem.after_sale_out_order_id == AfterSaleOutOrder.id
            ).outerjoin(Material, AfterSaleOutOrderItem.material_id == Material.id).filter(db.or_(*conditions)).distinct()
        sort_col = getattr(AfterSaleOutOrder, sort_by, AfterSaleOutOrder.created_at)
        orders = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc()).all()
        for order in orders:
            if order.items:
                for item in order.items:
                    ws.append([
                        order.order_no,
                        order.date.strftime('%Y-%m-%d') if order.date else '',
                        order.customer or '',
                        order.contact or '',
                        order.phone or '',
                        order.reason or '',
                        item.material.code if item.material else '',
                        item.material.name if item.material else '',
                        item.material.spec if item.material else '',
                        item.material.unit.name if item.material and item.material.unit else '',
                        item.quantity or 0,
                        item.price or 0,
                        item.amount or 0,
                        '待完成' if order.status == 'pending' else ('已完成' if order.status == 'completed' else (order.status or '')),
                        order.remark or ''
                    ])
            else:
                ws.append([
                    order.order_no,
                    order.date.strftime('%Y-%m-%d') if order.date else '',
                    order.customer or '',
                    order.contact or '',
                    order.phone or '',
                    order.reason or '',
                    '', '', '', '', 0, 0, 0,
                    '待完成' if order.status == 'pending' else ('已完成' if order.status == 'completed' else (order.status or '')),
                    order.remark or ''
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='after_sale_outs.xlsx', as_attachment=True)

    @app.route('/export/template/material')
    @login_required
    def export_material_template():
        from openpyxl import Workbook
        from app import inventory_alert_enabled
        wb = Workbook()
        ws = wb.active
        ws.title = '物料导入模板'
        headers = ['物料编码', '物料名称', '品牌', '规格', '单位', '分类', '供应商', '单价']
        example = ['MAT001', '示例物料', 'ABB', '规格A', '个', '原材料', '示例供应商', 0]
        if inventory_alert_enabled():
            headers.extend(['最低库存', '安全库存'])
            example.extend([0, 0])
        ws.append(headers)
        ws.append(example)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='material_template.xlsx', as_attachment=True)

    @app.route('/export/template/in_order')
    @login_required
    def export_in_order_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '入库单导入模板'
        ws.append(['单据编号', '日期', '业务类型', '用途', '仓库', '库位', '供应商', '客户', '物料编码', '物料名称', '规格', '单位', '数量', '单价', '金额', '客供', '合同编号', '工程名称', '备注'])
        ws.append(['RK20240101001', '2024-01-01', '采购入库', '采购到货', '一号仓库', '', '示例供应商', '', 'MAT001', '示例物料', '规格A', '个', '100', '10.00', '1000.00', '否', 'HT-001', '项目A', '示例备注'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='in_order_template.xlsx', as_attachment=True)

    @app.route('/export/template/out_order')
    @login_required
    def export_out_order_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '领料单导入模板'
        ws.append(['单据编号', '日期', '用途', '领料部门', '领料人', '物料编码', '物料名称', '规格', '单位', '数量', '单价', '金额', '备注'])
        ws.append(['CK20240101001', '2024-01-01', '领料单', '生产车间', '张三', 'MAT001', '示例物料', '规格A', '个', '100', '10.00', '1000.00', '示例备注'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='out_order_template.xlsx', as_attachment=True)
