#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 物料（material）域路由。
#
# 批量拆分模式：与供应商域、分类域一致，采用「register_material_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 material_list），与 app.py 内
# 原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils / sqlalchemy），不导入 app，避免循环导入。
# - app.py 内部定义（Material 模型、_material_low_stock_filter、api_error、
#   sanitize_text_input、validate_excel_extension、validate_excel_size 等）在各路由
#   函数内延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io
import math
from datetime import datetime

from flask import current_app, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from sqlalchemy.orm import joinedload

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 material_* 各路由测试覆盖
def register_material_routes(app):
    @app.route('/material')
    @require_role('warehouse')
    @login_required
    def material_list():
        from app import (
            Material,
            MaterialCategory,
            Supplier,
            Unit,
            _material_low_stock_filter,
            build_category_tree_rows,
            inventory_alert_enabled,
        )
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        # per_page 必须有下限保护，传入 0 或负数会让 paginate 抛 ValueError 导致接口 500
        per_page = max(1, per_page)
        per_page = per_page if per_page in [10, 20, 50, 100, 200] else 20
        search = request.args.get('search', '').strip()
        category_id = request.args.get('category_id', type=int) or 0
        stock_filter = (request.args.get('stock_filter') or '').strip()
        if not inventory_alert_enabled() and stock_filter in {'low', 'normal'}:
            stock_filter = ''
        # BUG-F02-01 修复：默认按 code 升序，与其他基础资料一致
        sort_by = request.args.get('sort', 'code')
        sort_order = request.args.get('order', 'asc')
        allowed_sorts = {'code', 'name', 'brand', 'spec', 'category_id', 'supplier_id', 'stock', 'price', 'created_at'}
        if sort_by not in allowed_sorts:
            sort_by = 'code'
        if sort_order not in ('asc', 'desc'):
            sort_order = 'asc'
        all_categories = MaterialCategory.query.order_by(MaterialCategory.code.asc(), MaterialCategory.name.asc(), MaterialCategory.id.asc()).all()
        category_rows = build_category_tree_rows(all_categories)
        category_descendants = {}
        for row in reversed(category_rows):
            cat = row['category']
            ids = [cat.id]
            for child in all_categories:
                if child.parent_id == cat.id:
                    ids.extend(category_descendants.get(child.id, [child.id]))
            category_descendants[cat.id] = sorted(set(ids))

        query = Material.query.options(joinedload(Material.category), joinedload(Material.unit), joinedload(Material.supplier))
        if search:
            query = query.filter(
                db.or_(
                    Material.code.like(f'%{search}%'),
                    Material.name.like(f'%{search}%'),
                    Material.brand.like(f'%{search}%'),
                    Material.spec.like(f'%{search}%')
                )
            )
        if category_id:
            query = query.filter(Material.category_id.in_(category_descendants.get(category_id, [category_id])))
        if stock_filter == 'low':
            query = query.filter(_material_low_stock_filter())
        sort_column = getattr(Material, sort_by, Material.created_at)
        if sort_order == 'asc':
            query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(sort_column.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        materials = pagination.items
        units = Unit.query.order_by(Unit.code.asc(), Unit.name.asc()).all()
        suppliers = Supplier.query.order_by(Supplier.code.asc(), Supplier.name.asc()).all()
        return render_template('material.html', materials=materials,
                             categories=all_categories, category_rows=category_rows, units=units, suppliers=suppliers,
                             pagination=pagination, sort_by=sort_by, sort_order=sort_order, per_page=per_page,
                             stock_filter=stock_filter, category_id=category_id)

    @app.route('/material/api/list')
    @login_required
    def material_api_list():
        """物料列表API - 支持分页"""
        from app import Material
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 100, type=int)
        # per_page 必须有下限保护，传入 0 或负数会让 paginate 抛 ValueError 导致接口 500
        per_page = max(1, min(per_page, 500))  # 限制最大每页数量

        pagination = Material.query.paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({
            'materials': [{
                'id': m.id,
                'code': m.code,
                'name': m.name,
                'brand': m.brand or '',
                'spec': m.spec
            } for m in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })

    @app.route('/material/api/all')
    @login_required
    def material_api_all():
        """分页返回物料完整数据，避免旧接口静默截断。"""
        from app import Material, serialize_material
        page = max(1, request.args.get('page', 1, type=int) or 1)
        per_page = request.args.get('per_page', 2000, type=int) or 2000
        per_page = min(max(1, per_page), 2000)
        pagination = Material.query.options(joinedload(Material.unit)).order_by(
            Material.code.asc()
        ).paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({
            'status': 'success',
            'materials': [serialize_material(m) for m in pagination.items],
            'total': pagination.total,
            'page': pagination.page,
            'per_page': pagination.per_page,
            'pages': pagination.pages,
            'truncated': pagination.page < pagination.pages,
            'next_page': pagination.next_num if pagination.has_next else None,
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/add', methods=['GET', 'POST'])
    @require_role('warehouse')
    @login_required
    def add_material():
        from app import (
            Material,
            MaterialCategory,
            MAX_REASONABLE_PRICE,
            MAX_REASONABLE_STOCK,
            Supplier,
            Unit,
            add_stock_transaction,
            api_error,
            get_default_warehouse,
            location_management_enabled,
            log_audit,
            material_name_spec_exists,
            parse_bounded_number,
            parse_float_value,
            parse_int_value,
            sanitize_text_input,
            save_upload_image,
            serialize_supplier,
            serialize_unit,
            update_location_inventory,
        )
        from utils import sync_material_primary_image
        if request.method == 'GET':
            # Add new
            categories = MaterialCategory.query.all()
            units = Unit.query.all()
            suppliers = Supplier.query.all()
            return render_template('material_add.html',
                                 categories=categories,
                                 units=[serialize_unit(unit) for unit in units],
                                 suppliers=[serialize_supplier(supplier) for supplier in suppliers])

        # POST
        # BUG-2026-07-29-002/009: name/code/spec/brand/purpose/remark 走 sanitize_text_input
        # 去除 < > 与 NUL 字节，防止存储型 XSS 与 NUL 静默吞掉；截断到列宽上限。
        code = sanitize_text_input(request.form.get('code', ''), max_len=50)
        if not code:
            return api_error('请输入物料编码')

        name = sanitize_text_input(request.form.get('name', ''), max_len=100)
        if not name:
            return api_error('请输入物料名称')

        if Material.query.filter_by(code=code).first():
            return api_error('物料编码已存在')

        spec = sanitize_text_input(request.form.get('spec'), max_len=100)
        brand = sanitize_text_input(request.form.get('brand'), max_len=100)
        # BUG-F02-02 修复：物料主数据长度截断防护
        # 防止 DB 静默截断（DB 列宽：code=50/name=100/brand=100/spec=100/purpose=200/remark=500）
        if len(code) > 50:
            return api_error(f'物料编码不能超过 50 个字符（当前 {len(code)}）')
        if len(name) > 100:
            return api_error(f'物料名称不能超过 100 个字符（当前 {len(name)}）')
        if len(brand) > 100:
            return api_error('品牌不能超过 100 个字符')
        if len(spec) > 100:
            return api_error(f'物料规格不能超过 100 个字符（当前 {len(spec)}）')
        purpose = sanitize_text_input(request.form.get('purpose'), max_len=200)
        if len(purpose) > 200:
            return api_error(f'用途不能超过 200 个字符（当前 {len(purpose)}）')
        remark = sanitize_text_input(request.form.get('remark'), max_len=500)
        if len(remark) > 500:
            return api_error(f'备注不能超过 500 个字符（当前 {len(remark)}）')
        if material_name_spec_exists(name, spec):
            return api_error('物料名称和规格不能同时重复')

        # BUG-2026-07-29-005: 库存/价格上限收紧至 99999999.99（拒绝 12 位以上大数）
        initial_stock = parse_bounded_number(request.form.get('stock'), 0, maximum=MAX_REASONABLE_STOCK)
        if initial_stock is None:
            return api_error(f'初始库存必须是 0 至 {MAX_REASONABLE_STOCK:,.2f} 的有限数字')
        initial_price = parse_bounded_number(request.form.get('price'), 0, maximum=MAX_REASONABLE_PRICE)
        if initial_price is None:
            return api_error(f'参考价格必须是 0 至 {MAX_REASONABLE_PRICE:,.2f} 的有限数字')

        image_file = request.files.get('image')
        image_path = None
        if image_file and image_file.filename:
            image_path, image_error = save_upload_image(image_file, subfolder='material_images')
            if image_error:
                return jsonify({'status': 'error', 'msg': image_error}), 400

        expiry_date = request.form.get('expiry_date')
        try:
            expiry_date_parsed = datetime.strptime(expiry_date, '%Y-%m-%d').date() if expiry_date else None
        except ValueError:
            expiry_date_parsed = None

        material = Material(
            code=code,
            name=name,
            category_id=request.form.get('category_id') or None,
            unit_id=request.form.get('unit_id') or None,
            supplier_id=request.form.get('supplier_id') or None,
            brand=brand or None,
            spec=spec,
            stock=initial_stock,
            purpose=request.form.get('purpose'),
            min_stock=parse_float_value(request.form.get('min_stock'), 0),
            max_stock=parse_float_value(request.form.get('max_stock'), 0),
            reorder_point=parse_float_value(request.form.get('reorder_point') or request.form.get('safety_stock'), 0),
            expiry_date=expiry_date_parsed,
            alert_days=parse_int_value(request.form.get('alert_days'), 30, minimum=1, maximum=3650),
            price=initial_price,
            remark=((request.form.get('remark') or '').strip() or None),
            image=image_path
        )
        db.session.add(material)
        # BUG-2026-08-16-003：初始库存审计流水、库位账与物料在同一事务提交，
        # 消除原"先 commit 物料、再 commit 流水"双提交窗口（中间失败会留下
        # 库存已涨、流水缺失的不一致状态）。
        try:
            db.session.flush()
            # 统一多图：新增物料上传的图片写入 material_image 表（与手机端同目录/同表），
            # 并同步 Material.image 为该图作为 Web 列表主图。
            if image_path:
                from app import MaterialImage
                db.session.add(MaterialImage(material_id=material.id, image=image_path, sort_order=0))
                sync_material_primary_image(material)
            # BUG-2026-08-04-009: 新增物料带初始库存时补一条审计流水，保证库存台账/月报
            # 可追溯（与期初库存调整 opening_stock 语义一致）。仅在有初始库存时记录。
            if initial_stock and initial_stock > 0:
                _loc = None
                _default_wh = get_default_warehouse()
                if _default_wh:
                    _loc = _default_wh.name
                add_stock_transaction(
                    material,
                    initial_stock,
                    'opening',
                    reference_type='opening_stock',
                    reference_id=material.id,
                    location=_loc,
                    warehouse=_default_wh,  # B-2026-08-27：写入端统一落 warehouse_id
                    remark='新增物料初始库存',
                )
                # BUG-2026-08-16-003：开启库位管理时同步写库位账（初始库存归默认
                # 仓库，以仓库名作占位行），防止总账与库位账分叉。
                if _default_wh and location_management_enabled():
                    ok_inv, msg_inv = update_location_inventory(
                        material, _loc, initial_stock, warehouse=_default_wh)
                    if not ok_inv:
                        db.session.rollback()
                        return api_error(f'初始库存写入库位账失败：{msg_inv}')
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'物料创建失败: {e}')
            return jsonify({'status': 'error', 'msg': '物料创建失败，编码可能已存在'}), 500
        # BUG-2026-08-16-012：物料新增写结构化变更审计（new_data）
        log_audit(
            'add_material', 'material', material.id,
            target_name=f'{material.code} {material.name}',
            new_data={
                'code': material.code,
                'name': material.name,
                'spec': material.spec or '',
                'brand': material.brand or '',
                'category_id': material.category_id,
                'unit_id': material.unit_id,
                'supplier_id': material.supplier_id,
                'price': material.price or 0,
                'stock': material.stock or 0,
                'remark': material.remark or '',
            },
        )
        current_app.logger.info(f'物料创建成功：{material.code}')
        return jsonify({'status': 'success', 'msg': '物料新增成功'})

    @app.route('/material/<int:id>', methods=['GET'])
    @login_required
    def get_material(id):
        """Return a material record as JSON."""
        from app import Material, api_error
        material = db.session.get(Material, id)
        if not material:
            return api_error('物料不存在')

        return jsonify({
            'status': 'success',
            'material': {
                'id': material.id,
                'code': material.code,
                'name': material.name,
                'category_id': material.category_id,
                'unit_id': material.unit_id,
                'supplier_id': material.supplier_id,
                'brand': material.brand or '',
                'spec': material.spec or '',
                'purpose': material.purpose or '',
                'min_stock': material.min_stock or 0,
                'max_stock': material.max_stock or 0,
                'reorder_point': material.reorder_point or 0,
                'expiry_date': material.expiry_date.strftime('%Y-%m-%d') if material.expiry_date else '',
                'alert_days': material.alert_days or 30,
                'price': material.price or 0,
                'remark': material.remark or '',
                'image': material.image or ''
            }
        })

    @app.route('/material/<int:id>/image_candidates')
    @login_required
    def material_image_candidates(id):
        from app import Material, _search_material_images_online
        material = Material.query.options(joinedload(Material.category)).get_or_404(id)
        try:
            candidates, query = _search_material_images_online(material)
        except Exception as exc:
            current_app.logger.warning('Material image search failed: %s', exc)
            return jsonify({'status': 'error', 'msg': '网上找图失败，请稍后重试或手动上传图片'}), 502
        return jsonify({
            'status': 'success',
            'query': query,
            'candidates': candidates,
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/<int:id>/image_select', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def material_image_select(id):
        from app import Material, _save_material_image_from_url, log_operation
        material = Material.query.get_or_404(id)
        payload = request.get_json(silent=True) or {}
        image_url = (payload.get('image_url') or '').strip()
        if not image_url:
            return jsonify({'status': 'error', 'msg': '请选择图片'}), 400

        image_path, error = _save_material_image_from_url(material, image_url)
        if error:
            return jsonify({'status': 'error', 'msg': error}), 400

        material.image = image_path
        try:
            db.session.commit()
            log_operation('网上找图绑定物料图片', f'{material.code} {material.name}', 'material', material.id)
            return jsonify({
                'status': 'success',
                'msg': '图片已保存',
                'image': image_path,
                'image_url': url_for('static', filename=image_path),
            })
        except Exception as exc:
            db.session.rollback()
            current_app.logger.error('Material image bind failed: %s', exc)
            return jsonify({'status': 'error', 'msg': '保存物料图片失败'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/<int:id>/copy', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def copy_material(id):
        """Return a material draft for user review before creating a new record."""
        from app import Material, generate_material_copy_code, generate_material_copy_name
        try:
            source = db.session.get(Material, id)
            if not source:
                return jsonify({'status': 'error', 'msg': '物料不存在'}), 404

            # 生成新的物料编码
            try:
                suggested_code = generate_material_copy_code(source)
                suggested_name = generate_material_copy_name(source.name, source.spec or '')
            except ValueError as e:
                current_app.logger.error(f'生成复制物料草稿失败: {e}')
                return jsonify({'status': 'error', 'msg': str(e) or '生成复制草稿失败，请稍后重试'}), 400

            return jsonify({
                'status': 'success',
                'material': {
                    'source_id': source.id,
                    'code': '',
                    'suggested_code': suggested_code,
                    'name': suggested_name,
                    'source_name': source.name,
                    'category_id': source.category_id,
                    'unit_id': source.unit_id,
                    'supplier_id': source.supplier_id,
                    'brand': source.brand or '',
                    'spec': source.spec or '',
                    'purpose': source.purpose or '',
                    'min_stock': source.min_stock or 0,
                    'max_stock': source.max_stock or 0,
                    'reorder_point': source.reorder_point or 0,
                    'expiry_date': source.expiry_date.strftime('%Y-%m-%d') if source.expiry_date else '',
                    'alert_days': source.alert_days or 30,
                    'price': source.price or 0,
                    'remark': source.remark or '',
                }
            })
        except Exception as e:
            current_app.logger.error(f'复制物料失败: {e}')
            return jsonify({'status': 'error', 'msg': '复制物料失败，请稍后重试'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/edit/<int:id>', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def edit_material(id):
        """Update a material record."""
        from app import (
            BOM,
            BOMItem,
            InOrderItem,
            InventoryCheckItem,
            Material,
            MAX_REASONABLE_PRICE,
            OutOrderItem,
            PurchaseRequestItem,
            SubcontractItem,
            api_error,
            inventory_alert_enabled,
            log_audit,
            material_code_editable,
            material_name_spec_exists,
            parse_bounded_number,
            parse_date_value,
            parse_float_value,
            parse_int_value,
            save_upload_image,
        )
        from utils import sync_material_primary_image
        material = db.session.get(Material, id)
        if not material:
            return api_error('物料不存在')

        new_code = (request.form.get('code') or '').strip()
        new_name = (request.form.get('name') or '').strip()
        if not new_code:
            return api_error('请输入物料编码')
        if not new_name:
            return api_error('请输入物料名称')
        if len(new_code) > 50 or len(new_name) > 100:
            return jsonify({'status': 'error', 'msg': '物料编码不能超过 50 个字符，名称不能超过 100 个字符'}), 400

        existing = Material.query.filter_by(code=new_code).first()
        if existing and existing.id != id:
            return api_error('物料编码已存在')

        new_spec = (request.form.get('spec') or '').strip()
        new_brand = (request.form.get('brand') or '').strip()
        # BUG-F02-02 修复：物料编辑入口同样 6 字段长度校验（与新增保持一致）
        if len(new_code) > 50:
            return jsonify({'status': 'error', 'msg': f'物料编码不能超过 50 个字符（当前 {len(new_code)}）'}), 400
        if len(new_name) > 100:
            return jsonify({'status': 'error', 'msg': f'物料名称不能超过 100 个字符（当前 {len(new_name)}）'}), 400
        if len(new_brand) > 100:
            return jsonify({'status': 'error', 'msg': '品牌不能超过 100 个字符'}), 400
        if len(new_spec) > 100:
            return jsonify({'status': 'error', 'msg': f'物料规格不能超过 100 个字符（当前 {len(new_spec)}）'}), 400
        new_purpose = (request.form.get('purpose') or '').strip()
        if len(new_purpose) > 200:
            return jsonify({'status': 'error', 'msg': f'用途不能超过 200 个字符（当前 {len(new_purpose)}）'}), 400
        new_remark = (request.form.get('remark') or '').strip()
        if len(new_remark) > 500:
            return jsonify({'status': 'error', 'msg': f'备注不能超过 500 个字符（当前 {len(new_remark)}）'}), 400
        if material_name_spec_exists(new_name, new_spec, exclude_id=id):
            return api_error('物料名称和规格不能同时重复')

        image_file = request.files.get('image')
        image_path = material.image
        new_image_path = None
        if image_file and image_file.filename:
            new_image_path, image_error = save_upload_image(image_file, subfolder='material_images')
            if image_error:
                return jsonify({'status': 'error', 'msg': image_error}), 400
            image_path = new_image_path

        expiry_date = request.form.get('expiry_date')

        old_code = material.code
        old_name = material.name
        old_spec = material.spec  # BUG-2026-08-04-005: 赋值前捕获旧规格，否则 spec_changed 永为 False
        # BUG-2026-08-16-012：修改前捕获完整旧值，供结构化变更审计（含单价/角色等敏感字段）
        _audit_old = {
            'code': material.code, 'name': material.name, 'spec': material.spec or '',
            'brand': material.brand or '', 'category_id': material.category_id,
            'unit_id': material.unit_id, 'supplier_id': material.supplier_id,
            'price': material.price or 0, 'min_stock': material.min_stock or 0,
            'max_stock': material.max_stock or 0, 'remark': material.remark or '',
        }
        code_changed = (old_code != new_code)
        name_changed = (old_name != new_name)
        spec_changed = (old_spec != new_spec)  # 必须在 material.spec = new_spec 之前比较
        if code_changed and not material_code_editable():
            return api_error('系统参数已禁止修改物料编码')

        material.code = new_code
        material.name = new_name
        material.category_id = request.form.get('category_id') or None
        material.unit_id = request.form.get('unit_id') or None
        material.supplier_id = request.form.get('supplier_id') or None
        material.brand = new_brand or None
        material.spec = new_spec
        material.purpose = request.form.get('purpose')
        material.max_stock = parse_float_value(request.form.get('max_stock'), 0)
        if inventory_alert_enabled():
            material.min_stock = parse_float_value(request.form.get('min_stock'), 0)
            material.reorder_point = parse_float_value(request.form.get('reorder_point'), 0)
            material.alert_days = parse_int_value(request.form.get('alert_days'), 30, minimum=1, maximum=3650)
        material.expiry_date = parse_date_value(expiry_date)
        # BUG-2026-08-04-007: 编辑物料价格上限必须与新增一致（MAX_REASONABLE_PRICE），
        # 原来用 MAX_TRANSACTION_PRICE（1 万亿）可绕过新增时的 99999999.99 上限。
        material.price = parse_bounded_number(request.form.get('price'), 0, maximum=MAX_REASONABLE_PRICE)
        if material.price is None:
            db.session.rollback()
            return api_error(f'参考价格必须是 0 至 {MAX_REASONABLE_PRICE:,.2f} 的有限数字')
        material.remark = (request.form.get('remark') or '').strip() or None
        material.image = image_path

        # 级联更新所有关联单据
        cascade_info = []
        # BUG-2026-08-04-005: spec_changed 已在赋值前计算，此处不再重复比较

        if code_changed or name_changed or spec_changed:
            # 1. 更新采购申请明细
            pr_items = PurchaseRequestItem.query.filter_by(material_id=id).all()
            for item in pr_items:
                if code_changed and hasattr(item, 'material_code'):
                    item.material_code = new_code
                if name_changed and hasattr(item, 'material_name'):
                    item.material_name = new_name
                if spec_changed and hasattr(item, 'spec'):
                    item.spec = new_spec
            if pr_items:
                cascade_info.append(f'采购申请 {len(pr_items)} 条')

            # 2. 更新BOM
            bom_records = BOM.query.filter_by(product_code=old_code if code_changed else new_code).all()
            for bom in bom_records:
                if code_changed and hasattr(bom, 'product_code'):
                    bom.product_code = new_code
                if name_changed and hasattr(bom, 'product_name'):
                    bom.product_name = new_name
            if bom_records:
                cascade_info.append(f'BOM {len(bom_records)} 条')

            # 3~7 项均只统计关联条数用于提示（无冗余字段需同步）。
            # BUG-2026-09-05-001/R6：数据库与代码不同步（如库存盘点域缺
            # counted_by 列）时整段查询会抛 OperationalError，把"改个物料名称"
            # 打成 500。统计失败只记日志并跳过该项提示，不阻断保存。
            def _append_related_count(model, label):
                try:
                    n = model.query.filter_by(material_id=id).count()
                except Exception as _count_exc:
                    current_app.logger.warning(
                        '物料编辑级联统计跳过（%s，视为无关联）: %s', label, _count_exc)
                    return 0
                if n:
                    cascade_info.append(f'{label} {n} 条（仅关联，无冗余字段）')
                return n

            # 3. 入库单明细（InOrderItem 仅持有 material_id 外键，无需冗余同步）
            _append_related_count(InOrderItem, '入库单')
            # 4. 领料单明细（OutOrderItem 无冗余 code/name/spec 字段）
            _append_related_count(OutOrderItem, '领料单')
            # 5. 盘点单明细（InventoryCheckItem 无冗余字段）
            _append_related_count(InventoryCheckItem, '盘点单')
            # 6. 委外加工明细（SubcontractItem 无冗余字段）
            _append_related_count(SubcontractItem, '委外单')
            # 7. BOM子项明细（BOMItem 无冗余字段）
            _append_related_count(BOMItem, 'BOM子项')

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'物料更新失败: {e}')
            return jsonify({'status': 'error', 'msg': '物料更新失败'}), 500
        # BUG-2026-08-16-012：物料编辑写结构化变更审计（old_data + new_data，含单价变化）
        log_audit(
            'edit_material', 'material', id,
            target_name=f'{material.code} {material.name}',
            old_data=_audit_old,
            new_data={
                'code': material.code, 'name': material.name, 'spec': material.spec or '',
                'brand': material.brand or '', 'category_id': material.category_id,
                'unit_id': material.unit_id, 'supplier_id': material.supplier_id,
                'price': material.price or 0, 'min_stock': material.min_stock or 0,
                'max_stock': material.max_stock or 0, 'remark': material.remark or '',
            },
        )
        # 统一多图：编辑物料时上传的新图片追加到 material_image 表（与手机端同目录/同表），
        # 并同步 Material.image 为该图作为 Web 列表主图。
        if new_image_path:
            from app import MaterialImage
            current_count = MaterialImage.query.filter_by(material_id=id).count()
            db.session.add(MaterialImage(material_id=id, image=new_image_path, sort_order=current_count))
            sync_material_primary_image(material)
            db.session.commit()

        msg = '物料更新成功'
        if cascade_info:
            msg += f'，已同步更新：{"、".join(cascade_info)}'
        return jsonify({'status': 'success', 'msg': msg})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/delete_all', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_all_materials():
        """Disable destructive bulk material deletion in production."""
        return jsonify({'status': 'error', 'msg': '线上系统已禁用删除全部物料，请走停机维护流程'}), 403

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/fix_empty_fields', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def fix_empty_fields():
        """修复已导入物料的空分类/单位/供应商"""
        from app import Material, MaterialCategory, Unit
        default_unit = Unit.query.filter_by(name='个').first()
        if not default_unit:
            default_unit = Unit.query.first()
        default_cat = MaterialCategory.query.filter_by(name='默认分类').first()
        if not default_cat:
            default_cat = MaterialCategory.query.first()
        count = 0
        try:
            materials = Material.query.filter(
                db.or_(Material.unit_id.is_(None), Material.unit_id == 0,
                       Material.category_id.is_(None), Material.category_id == 0)
            ).all()
            for m in materials:
                if not m.unit_id or m.unit_id == 0:
                    m.unit_id = default_unit.id if default_unit else None
                if not m.category_id or m.category_id == 0:
                    m.category_id = default_cat.id if default_cat else None
                count += 1
            db.session.commit()
            return jsonify({'status': 'success', 'msg': f'已修复 {count} 条物料的空字段'})
        except Exception:
            db.session.rollback()
            return jsonify({'status': 'error', 'msg': '修复失败'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_material():
        from app import (
            AdjustmentOrderItem,
            AfterSaleOutOrderItem,
            AIDocumentItem,
            AIMaterialAlias,
            BOMItem,
            InOrderItem,
            InventoryCheckItem,
            InventoryCheckScanItem,
            LocationInventory,
            Material,
            OpeningStock,
            OutOrderItem,
            ProductionRequisitionItem,
            PurchaseOrderItem,
            PurchaseRequestItem,
            SalesOrderItem,
            StockTransaction,
            SubcontractIssueItem,
            SubcontractItem,
            SubcontractReceiveItem,
            TransferOrderItem,
            api_error,
            log_audit,
            log_operation,
        )
        import json
        # equest.json ?form data
        if request.is_json:
            ids = request.json.get('ids', [])
        else:
            #  form data
            ids_str = request.form.get('ids', '[]')
            try:
                ids = json.loads(ids_str)
            except (json.JSONDecodeError, TypeError):
                return api_error('删除参数格式错误')

        if not ids:
            return api_error('请选择要删除的物料')

        success_count = 0
        fail_count = 0
        # BUG-2026-08-16-012：记录被删物料快照，供结构化变更审计
        _deleted_audit = []

        for id in ids:
            material = db.session.get(Material, id)
            if material:
                # 引用完整性校验：被已完成单据或库存流水引用的物料不允许删除，
                # 避免历史单据变空壳、审计流水丢失。仅允许删除无业务引用的物料。
                # 逐表检查并分别 try/except：旧库可能缺少新增的 AI 关联表
                # （如 ai_material_alias / ai_document_item），若某张表查询抛错，
                # 只跳过该表（视为无引用），绝不能把整个删除接口打成 500 HTML。
                _ref_check_models = (
                    InOrderItem, OutOrderItem, StockTransaction, PurchaseOrderItem,
                    SalesOrderItem, ProductionRequisitionItem, SubcontractItem,
                    SubcontractIssueItem, SubcontractReceiveItem, BOMItem,
                    InventoryCheckItem, AfterSaleOutOrderItem, PurchaseRequestItem,
                    TransferOrderItem, AdjustmentOrderItem, InventoryCheckScanItem,
                    OpeningStock, AIMaterialAlias, AIDocumentItem,
                )
                _referenced = False
                for _model in _ref_check_models:
                    try:
                        if _model.query.filter_by(material_id=id).first():
                            _referenced = True
                            break
                    except Exception as _e:
                        current_app.logger.warning(
                            f"物料 {id} 引用检查跳过 {_model.__name__}（表或字段缺失，视为无引用）: {_e}")
                        continue
                if _referenced:
                    fail_count += 1
                    current_app.logger.info(f"跳过删除物料 {id}({material.code})：存在业务引用，建议改为停用")
                    continue
                try:
                    # 无业务引用时，先清理物料的库位辅助记录，再删除物料主数据。
                    LocationInventory.query.filter_by(material_id=id).delete()
                    # 主数据删除改用 raw SQL，避免 db.session.delete(material) 触发
                    # ORM 级联加载 ai_aliases / ai_document_items 等 backref——
                    # 旧库若缺这些新增的 AI 关联表，ORM 加载会抛 OperationalError 造成 500。
                    # 上面的引用检查已确认无业务引用，raw DELETE 不会触发外键冲突。
                    from sqlalchemy import text as _sa_text
                    db.session.expunge(material)
                    db.session.execute(
                        _sa_text("DELETE FROM material WHERE id = :m_id"),
                        {"m_id": id},
                    )
                    _deleted_audit.append({
                        'id': id, 'code': material.code, 'name': material.name,
                        'spec': material.spec or '', 'price': material.price or 0,
                    })
                    success_count += 1
                except Exception as e:
                    # 单条删除失败必须 rollback，否则 session 进入 PendingRollback 状态，
                    # 后续所有迭代都会持续失败，且最终 commit 也会失败
                    db.session.rollback()
                    current_app.logger.error(f"删除物料 {id} 失败: {str(e)}")
                    fail_count += 1
            else:
                fail_count += 1

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'数据库操作失败: {e}')
            return api_error(f'数据库操作失败: {str(e)}')

        # 记录操作日志
        log_operation('删除物料', f'批量删除物料，成功 {success_count} 条，失败 {fail_count} 条', 'material')
        # BUG-2026-08-16-012：物料删除写结构化变更审计（old_data = 被删物料快照）
        if _deleted_audit:
            log_audit(
                'delete_material', 'material', _deleted_audit[0]['id'],
                target_name=f'删除物料 {success_count} 条',
                old_data=_deleted_audit,
                reason=f'成功 {success_count} 条，失败 {fail_count} 条',
            )

        if success_count == 0:
            return api_error('删除失败，没有物料被删除')

        return jsonify({
            'status': 'success',
            'msg': f'物料删除完成，成功 {success_count} 条' + (f'，失败 {fail_count} 条' if fail_count > 0 else '')
        })

    @app.route('/material/download_template')
    @login_required
    def download_material_template():
        from app import inventory_alert_enabled
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '物料导入模板'
        headers = ['物料编码', '物料名称', '品牌', '规格', '单位', '分类', '供应商', '单价']
        example = ['MAT-001', '示例物料', 'ABB', '示例规格', '个', '原材料', '默认供应商', 0]
        if inventory_alert_enabled():
            headers.extend(['最低库存', '安全库存'])
            example.extend([0, 0])
        ws.append(headers)
        ws.append(example)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='material_import_template.xlsx', as_attachment=True)

    @app.route('/material/export')
    @login_required
    def export_material():
        from app import (
            Material,
            _material_low_stock_filter,
            export_max_rows,
            inventory_alert_enabled,
        )
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '物料数据'
        headers = ['物料编码', '物料名称', '品牌', '规格', '单位', '分类', '供应商', '当前库存', '单价']
        if inventory_alert_enabled():
            headers.extend(['最低库存', '安全库存'])
        ws.append(headers)
        search = request.args.get('search', '').strip()
        stock_filter = (request.args.get('stock_filter') or '').strip()
        if not inventory_alert_enabled() and stock_filter in {'low', 'normal'}:
            stock_filter = ''
        sort_by = request.args.get('sort', 'created_at')
        sort_order = request.args.get('order', 'desc')
        allowed_sorts = {'code', 'name', 'brand', 'spec', 'category_id', 'supplier_id', 'stock', 'price', 'created_at'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        if sort_order not in ('asc', 'desc'):
            sort_order = 'desc'
        query = Material.query.options(joinedload(Material.category), joinedload(Material.unit), joinedload(Material.supplier))
        if search:
            query = query.filter(
                db.or_(
                    Material.code.like(f'%{search}%'),
                    Material.name.like(f'%{search}%'),
                    Material.brand.like(f'%{search}%'),
                    Material.spec.like(f'%{search}%')
                )
            )
        if stock_filter == 'low':
            query = query.filter(_material_low_stock_filter())
        sort_column = getattr(Material, sort_by, Material.created_at)
        query = query.order_by(sort_column.asc() if sort_order == 'asc' else sort_column.desc())
        max_rows = export_max_rows()
        total_rows = query.count()
        if total_rows > max_rows:
            flash(f'当前筛选结果 {total_rows} 条，超过系统参数设置的导出上限 {max_rows} 条，请缩小筛选范围后再导出。', 'warning')
            return redirect(url_for('material_list', search=search, stock_filter=stock_filter, sort=sort_by, order=sort_order))
        for m in query.all():
            row = [m.code, m.name, m.brand or '', m.spec or '', m.unit.name if m.unit else '', m.category.name if m.category else '', m.supplier.name if m.supplier else '', m.stock or 0, m.price or 0]
            if inventory_alert_enabled():
                safety_stock = max(m.reorder_point or 0, m.min_stock or 0)
                row.extend([m.min_stock or 0, safety_stock])
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='materials.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/material/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_material():
        from app import (
            Material,
            MaterialCategory,
            MAX_REASONABLE_PRICE,
            Supplier,
            Unit,
            api_error,
            import_max_rows,
            inventory_alert_enabled,
            sanitize_text_input,
            validate_excel_extension,
            validate_excel_size,
        )
        file = request.files.get('file')
        if not file:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error('请选择要导入的物料文件')
            flash('请选择要导入的物料文件', 'danger')
            return redirect(url_for('material_list'))
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_ext_msg)
            flash(_ext_msg, 'danger')
            return redirect(url_for('material_list'))
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_size_msg)
            flash(_size_msg, 'danger')
            return redirect(url_for('material_list'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            header_row = [str(cell).strip() if cell else '' for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col_map = {}
            for idx, h in enumerate(header_row):
                if not h:
                    continue
                if '编码' in h or '代码' in h:
                    col_map['code'] = idx
                elif '名称' in h or '名字' in h:
                    col_map['name'] = idx
                elif '品牌' in h or h.lower() == 'brand':
                    col_map['brand'] = idx
                elif h == '规格' or '规格' in h:
                    col_map['spec'] = idx
                elif h == '单位' or '单位' in h:
                    col_map['unit'] = idx
                elif h == '分类' or '类别' in h or '分类' in h:
                    col_map['category'] = idx
                elif '供应商' in h:
                    col_map['supplier'] = idx
                elif '单价' in h or '价格' in h:
                    col_map['price'] = idx
                elif '安全库存' in h or '再订购点' in h or '再订货点' in h:
                    col_map['safety_stock'] = idx
                elif '最低库存' in h or '最小库存' in h:
                    col_map['min_stock'] = idx
            if 'code' not in col_map or 'name' not in col_map:
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(f'Excel表头缺少必要列（物料编码、物料名称）。检测到的表头：{", ".join(header_row)}')
                flash(f'Excel表头缺少必要列（物料编码、物料名称）。检测到的表头：{", ".join(header_row)}', 'danger')
                return redirect(url_for('material_list'))
            data_rows = max((ws.max_row or 1) - 1, 0)
            max_rows = import_max_rows()
            if data_rows > max_rows:
                msg = f'本次导入 {data_rows} 行，超过系统参数设置的导入上限 {max_rows} 行，请拆分文件后再导入'
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(msg)
                flash(msg, 'danger')
                return redirect(url_for('material_list'))
            count = 0
            skip = 0
            skip_details = []
            warnings = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # BUG-2026-08-04-008: 导入物料走 sanitize_text_input + 长度/价格校验，
                # 与 add_material 保持一致，防止 XSS/NUL、DB 列宽溢出、天价物料进入系统。
                # 长度校验在 sanitize_text_input 之前检查原始值，避免静默截断导致数据不符。
                raw_code = str(row[col_map['code']]).strip() if len(row) > col_map['code'] and row[col_map['code']] else ''
                raw_name = str(row[col_map['name']]).strip() if len(row) > col_map['name'] and row[col_map['name']] else ''
                if not raw_code or not raw_name:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：编码或名称为空')
                    continue
                if len(raw_code) > 50:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：编码不能超过50个字符')
                    continue
                if len(raw_name) > 100:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：名称不能超过100个字符')
                    continue
                code = sanitize_text_input(raw_code, max_len=50)
                name = sanitize_text_input(raw_name, max_len=100)
                if Material.query.filter_by(code=code).first():
                    skip += 1
                    skip_details.append(f'第{row_idx}行：编码{code}已存在')
                    continue
                raw_spec = str(row[col_map['spec']]).strip() if 'spec' in col_map and len(row) > col_map['spec'] and row[col_map['spec']] else ''
                raw_brand = str(row[col_map['brand']]).strip() if 'brand' in col_map and len(row) > col_map['brand'] and row[col_map['brand']] else ''
                if len(raw_spec) > 100:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：规格不能超过100个字符')
                    continue
                if len(raw_brand) > 100:
                    skip += 1
                    skip_details.append(f'第{row_idx}行：品牌不能超过100个字符')
                    continue
                spec = sanitize_text_input(raw_spec, max_len=100)
                brand = sanitize_text_input(raw_brand, max_len=100)
                unit_name = str(row[col_map['unit']]).strip() if 'unit' in col_map and len(row) > col_map['unit'] and row[col_map['unit']] else ''
                unit = None
                if unit_name:
                    unit = Unit.query.filter_by(name=unit_name).first()
                    if not unit:
                        unit = Unit.query.filter_by(code=unit_name).first()
                    if not unit:
                        unit = Unit.query.filter(db.func.lower(Unit.name) == db.func.lower(unit_name)).first()
                    if not unit:
                        unit = Unit.query.filter(db.func.lower(Unit.code) == db.func.lower(unit_name)).first()
                if unit_name and not unit:
                    unit = Unit(code=unit_name, name=unit_name)
                    db.session.add(unit)
                    db.session.flush()
                    warnings.append(f'自动创建单位：{unit_name}')
                cat_name = str(row[col_map['category']]).strip() if 'category' in col_map and len(row) > col_map['category'] and row[col_map['category']] else ''
                category = None
                if cat_name:
                    category = MaterialCategory.query.filter_by(name=cat_name).first()
                    if not category:
                        category = MaterialCategory.query.filter_by(code=cat_name).first()
                    if not category:
                        category = MaterialCategory.query.filter(db.func.lower(MaterialCategory.name) == db.func.lower(cat_name)).first()
                    if not category:
                        category = MaterialCategory.query.filter(db.func.lower(MaterialCategory.code) == db.func.lower(cat_name)).first()
                if cat_name and not category:
                    category = MaterialCategory(code=cat_name, name=cat_name)
                    db.session.add(category)
                    db.session.flush()
                    warnings.append(f'自动创建分类：{cat_name}')
                sup_name = str(row[col_map['supplier']]).strip() if 'supplier' in col_map and len(row) > col_map['supplier'] and row[col_map['supplier']] else ''
                supplier = None
                if sup_name:
                    supplier = Supplier.query.filter_by(name=sup_name).first()
                    if not supplier:
                        supplier = Supplier.query.filter_by(code=sup_name).first()
                    if not supplier:
                        supplier = Supplier.query.filter(db.func.lower(Supplier.name) == db.func.lower(sup_name)).first()
                    if not supplier:
                        supplier = Supplier.query.filter(db.func.lower(Supplier.code) == db.func.lower(sup_name)).first()
                if sup_name and not supplier:
                    supplier = Supplier(code=sup_name, name=sup_name)
                    db.session.add(supplier)
                    db.session.flush()
                    warnings.append(f'自动创建供应商：{sup_name}')
                # BUG-2026-08-04-008: 价格校验与 add_material 一致（0 至 MAX_REASONABLE_PRICE）。
                # 注意 parse_bounded_number 内部 parse_float_value 会把负数静默转 0，
                # 因此先显式检查原始值是否为负，负数行跳过并告知用户。
                raw_price = row[col_map['price']] if 'price' in col_map and len(row) > col_map['price'] and row[col_map['price']] else 0
                try:
                    _price_check = float(raw_price)
                except (ValueError, TypeError):
                    _price_check = 0
                if _price_check < 0 or _price_check > MAX_REASONABLE_PRICE or not math.isfinite(_price_check):
                    skip += 1
                    skip_details.append(f'第{row_idx}行：参考价格必须是 0 至 {MAX_REASONABLE_PRICE:,.2f} 的有限数字')
                    continue
                price_val = _price_check
                min_stock_val = 0
                safety_stock_val = 0
                if inventory_alert_enabled():
                    try:
                        min_stock_val = float(row[col_map['min_stock']]) if 'min_stock' in col_map and len(row) > col_map['min_stock'] and row[col_map['min_stock']] else 0
                    except (ValueError, TypeError):
                        min_stock_val = 0
                    try:
                        safety_stock_val = float(row[col_map['safety_stock']]) if 'safety_stock' in col_map and len(row) > col_map['safety_stock'] and row[col_map['safety_stock']] else min_stock_val
                    except (ValueError, TypeError):
                        safety_stock_val = min_stock_val
                    if safety_stock_val < min_stock_val:
                        safety_stock_val = min_stock_val
                material = Material(
                    code=code,
                    name=name,
                    brand=brand or None,
                    spec=spec,
                    stock=0,
                    price=price_val,
                    min_stock=min_stock_val,
                    reorder_point=safety_stock_val,
                    unit_id=unit.id if unit else None,
                    category_id=category.id if category else None,
                    supplier_id=supplier.id if supplier else None
                )
                db.session.add(material)
                count += 1
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error(f'导入失败：{str(e)}')
                flash(f'导入失败：{str(e)}', 'danger')
                return redirect(url_for('material_list'))
            msg = f'物料导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条'
            if skip_details:
                warnings.append(f'跳过详情：{"; ".join(skip_details[:20])}')
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                resp = {'status': 'success', 'msg': msg, 'count': count}
                if warnings:
                    resp['warnings'] = '；'.join(warnings)
                return jsonify(resp)
            flash(msg, 'success')
            for w in warnings:
                flash(w, 'warning')
        except Exception as e:
            db.session.rollback()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(f'物料导入失败：{str(e)}')
            flash('物料导入失败，请稍后重试', 'danger')
        return redirect(url_for('material_list'))

    @app.route('/material/print_label')
    @login_required
    def material_print_label_not_implemented():
        from app import api_error
        return api_error('物料标签打印功能未实现，请联系系统管理员', code=404)