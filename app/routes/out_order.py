#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 销售/领料/其他出库（out_order）域路由。
#
# 批量拆分模式：与售后出库（after_sale_out）域一致，采用「register_out_order_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 out_order_list、
# add_out_order、complete_out_order 等），
# 与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（OutOrder / OutOrderItem 模型、OutOrderPrintTemplate、
#   DocumentPushLine、各辅助函数 _get_order_list_filters / _apply_status_date_filters /
#   _apply_out_order_search / _apply_header_or_item_contract_filters / api_error /
#   _check_out_order_anomalies / _acquire_order_write_lock / _release_document_push_lines /
#   recalculate_order_total / is_future_date / parse_date_value /
#   validate_sales_warehouse / validate_sales_outbound_warehouse /
#   sync_sales_order_shipment / deduct_stock_atomic / deduct_location_inventory_atomic /
#   create_print_template / set_default_print_template / delete_print_template /
#   _render_html_print_content 等）
#   在各路由函数内延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志复用 register_out_order_routes(app) 传入的 app.logger（与 app.py 原实现一致）。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io
import json

from flask import abort, jsonify, render_template, request, send_file, url_for
from flask_login import login_required

from db import db
from utils import get_default_print_template, print_token_or_login_required, require_role


def _build_out_order_excel(order):
    """按用户指定样式生成领料单 Excel（无合计行）。

    列：物料编码|品牌|物料名称|规格|单位|数量|单价|金额|合同编号；
    表头：领料部门/日期；底部：领料。返回填充好数据的 .xlsx 字节流。
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '领料单'

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    title_font = Font(name='微软雅黑', size=16, bold=True)
    body_font = Font(name='微软雅黑', size=11)
    header_font = Font(name='微软雅黑', size=11, bold=True)

    for idx, width in enumerate([12, 10, 20, 14, 6, 8, 10, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.merge_cells('A1:I1')
    ws['A1'] = '领料单'
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    department = order.department.name if getattr(order, 'department', None) else (getattr(order, 'customer', None) or '')
    date_str = ''
    if getattr(order, 'date', None):
        try:
            date_str = order.date.strftime('%Y-%m-%d')
        except Exception:
            date_str = str(order.date)
    contract_no = getattr(order, 'contract_no', None) or ''

    ws.merge_cells('A2:E2')
    ws['A2'] = f'领料部门：{department or ""}'
    ws.merge_cells('F2:I2')
    ws['F2'] = f'日期：{date_str}'
    ws['A2'].font = body_font
    ws['A2'].alignment = left_align
    ws['F2'].font = body_font
    ws['F2'].alignment = left_align

    headers = ['物料编码', '品牌', '物料名称', '规格', '单位', '数量', '单价', '金额', '合同编号']
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=text)
        c.font = header_font
        c.alignment = center
        c.border = border

    row = 4
    for item in list(getattr(order, 'items', None) or []):
        material = getattr(item, 'material', None)
        vals = [
            (material.code if material is not None and getattr(material, 'code', None) else '') or '',
            (material.brand if material is not None and getattr(material, 'brand', None) else '') or '',
            (material.name if material is not None and getattr(material, 'name', None) else '') or '',
            (material.spec if material is not None and getattr(material, 'spec', None) else '') or '',
        ]
        unit = ''
        if material is not None and getattr(material, 'unit', None):
            unit = material.unit.name or ''
        for col, value in enumerate(vals, start=1):
            ws.cell(row=row, column=col, value=value)
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=float(getattr(item, 'quantity', None) or 0)).number_format = '0.##'
        ws.cell(row=row, column=7, value=float(getattr(item, 'price', None) or 0)).number_format = '0.00'
        ws.cell(row=row, column=8, value=float(getattr(item, 'amount', None) or 0)).number_format = '0.00'
        ws.cell(row=row, column=9, value=(getattr(item, 'contract_no', None) or contract_no or ''))
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.font = body_font
            c.border = border
            c.alignment = right_align if col in (6, 7, 8) else center
        row += 1

    # 预留空白明细行（至少补到第 11 行结束，便于打印/手写补录）
    while row < 12:
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.font = body_font
        row += 1

    # 「领料：」签名行：起点对齐「单价」列（G 列），向右延伸留出签字空间
    ws.merge_cells(f'G{row}:I{row}')
    ws.cell(row=row, column=7, value='领料：').alignment = Alignment(
        horizontal='left', vertical='center')
    for col in range(7, 10):
        ws.cell(row=row, column=col).font = body_font
    ws.row_dimensions[row].height = 22

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# no-test:reason=路由注册辅助函数，能力由 out_order_* 各路由测试覆盖
def register_out_order_routes(app):
    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/copy', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def copy_out_order(id):
        """复制领料/出库单为新草稿（不带销售来源关联，避免半关联脏数据）。"""
        from datetime import date
        from app import (OutOrder, OutOrderItem, api_error, generate_order_no,
                         log_operation, recalculate_order_total,
                         round_to_2_decimals)
        from flask_login import current_user
        from sqlalchemy.orm import joinedload
        source = OutOrder.query.options(
            joinedload(OutOrder.items).joinedload(OutOrderItem.material),
            joinedload(OutOrder.department),
        ).get_or_404(id)
        if not source.items:
            return api_error('原出库单没有明细，不能复制')

        business_type = source.business_type or '领料单'
        if business_type == '其他出库':
            prefix = 'OO'
        elif business_type == '销售出库':
            prefix = 'SO'
        else:
            prefix = 'OUT'

        try:
            remark_parts = [f'由出库单 {source.order_no} 复制生成']
            if source.remark:
                remark_parts.append(f'原备注：{source.remark}')
            if source.source_sales_order_id:
                remark_parts.append('已剥离销售订单来源，可按普通草稿继续编辑')
            new_order = OutOrder(
                order_no=generate_order_no(prefix),
                date=date.today(),
                department_id=source.department_id,
                customer=source.customer,
                business_type=business_type,
                warehouse=source.warehouse or '',
                location=getattr(source, 'location', '') or '',
                purpose=source.purpose,
                picker=source.picker,
                source_sales_order_id=None,
                remark='；'.join(remark_parts)[:200],
                contract_id=source.contract_id,
                contract_no=source.contract_no,
                project_name=source.project_name,
                status='pending',
                operator_id=current_user.id,
                total_amount=0,
            )
            db.session.add(new_order)
            db.session.flush()

            copied_count = 0
            for item in source.items:
                quantity = round_to_2_decimals(item.quantity or 0)
                if quantity <= 0:
                    continue
                price = round_to_2_decimals(item.price or 0)
                db.session.add(OutOrderItem(
                    out_order_id=new_order.id,
                    material_id=item.material_id,
                    source_sales_order_item_id=None,
                    quantity=quantity,
                    price=price,
                    amount=round_to_2_decimals(quantity * price),
                    remark=item.remark,
                    contract_id=item.contract_id,
                    contract_no=item.contract_no,
                    project_name=item.project_name,
                ))
                copied_count += 1

            if copied_count <= 0:
                db.session.rollback()
                return api_error('原出库单没有有效数量，不能复制')

            recalculate_order_total(new_order)
            db.session.commit()
            log_operation('复制出库单', f'{source.order_no} -> {new_order.order_no}', 'out_order', new_order.id)
            return jsonify({
                'status': 'success',
                'msg': '复制成功，已生成新的出库草稿',
                'id': new_order.id,
                'order_no': new_order.order_no,
                'redirect_url': url_for('out_order_detail', id=new_order.id),
            })
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'复制出库单失败: {e}')
            return api_error('复制失败，请稍后重试')

    @app.route('/out_order')
    @app.route('/other_out_order')
    @login_required
    def out_order_list():
        from app import (Material, OutOrder, OutOrderItem,
                         _apply_header_or_item_contract_filters,
                         _apply_out_order_search, _apply_status_date_filters,
                         _get_order_list_filters, get_active_warehouses,
                         get_default_warehouse, resolve_request_warehouse)
        from sqlalchemy.orm import joinedload
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        # per_page 必须有下限保护，传入 0 或负数会让 paginate 抛 ValueError 导致接口 500
        per_page = max(1, per_page)
        if per_page not in [10, 20, 50, 100, 200]:
            per_page = 20
        status_filter, search, date_start, date_end, sort_by, sort_order = _get_order_list_filters(('pending', 'completed'))
        allowed_sorts = {'order_no', 'date', 'department_id', 'customer', 'business_type', 'purpose', 'status', 'created_at', 'total_amount'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        sort_col = getattr(OutOrder, sort_by, OutOrder.created_at)
        # 按单据左连接明细展示，待完成但没有明细的单据也能查到。
        query = db.session.query(OutOrder, OutOrderItem).outerjoin(OutOrderItem, OutOrderItem.out_order_id == OutOrder.id).options(
            joinedload(OutOrder.department),
            joinedload(OutOrderItem.material).joinedload(Material.unit),
        )
        query = _apply_status_date_filters(query, OutOrder, status_filter, date_start, date_end)
        warehouse, warehouse_error = resolve_request_warehouse(request.args)
        if warehouse:
            query = query.filter(OutOrder.warehouse == warehouse.name)
        elif warehouse_error:
            query = query.filter(db.false())
        query = _apply_out_order_search(query, search)
        contract_no_filter = (request.args.get('contract_no') or '').strip()
        project_name_filter = (request.args.get('project_name') or '').strip()
        query = _apply_header_or_item_contract_filters(
            query, OutOrder, OutOrderItem, 'out_order_id',
            contract_no_filter=contract_no_filter,
            project_name_filter=project_name_filter,
        )
        # 领料明细默认排除"销售出库"（销售出库归销售管理，见 /sales/outflow_report），
        # 避免销售单据混入仓库领料明细。显式传 business_type=销售出库 时仍可查看。
        explicit_bt = '其他出库' if request.path == '/other_out_order' else (request.args.get('business_type') or '').strip()
        if not explicit_bt:
            query = query.filter(db.or_(OutOrder.business_type == '领料单', OutOrder.business_type.is_(None)))
        else:
            query = query.filter(OutOrder.business_type == explicit_bt)
        if sort_order == 'asc':
            query = query.order_by(sort_col.asc())
        else:
            query = query.order_by(sort_col.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        items = [
            type('OutOrderListRow', (), {
                'out_order': order,
                'material': item.material if item else None,
                'quantity': item.quantity if item else 0,
                'price': item.price if item else 0,
                'amount': item.amount if item else 0,
                'contract_no': item.contract_no if item else '',
                'project_name': item.project_name if item else '',
            })()
            for order, item in pagination.items
        ]
        filters = {
            'status': status_filter,
            'search': search,
            'date_start': date_start.strftime('%Y-%m-%d') if date_start else '',
            'date_end': date_end.strftime('%Y-%m-%d') if date_end else '',
            'business_type': explicit_bt,
            'contract_no': contract_no_filter,
            'project_name': project_name_filter,
            'warehouse_id': warehouse.id if warehouse else '',
        }
        page_title = '其他出库明细表' if explicit_bt == '其他出库' else '领料明细表'
        return render_template('out_order.html', items=items, pagination=pagination, sort_by=sort_by, sort_order=sort_order, per_page=per_page, filters=filters, page_title=page_title, warehouses=get_active_warehouses(), default_warehouse=get_default_warehouse())

    @app.route('/out_order/<int:id>')
    @login_required
    def out_order_detail(id):
        from datetime import date
        from app import (Department, DocumentPushLine, Material, OutOrder,
                         OutOrderItem, get_active_warehouses,
                         get_default_warehouse, get_recent_operation_logs,
                         location_management_enabled)
        from sqlalchemy.orm import joinedload
        order = OutOrder.query.options(joinedload(OutOrder.department), joinedload(OutOrder.items).joinedload(OutOrderItem.material).joinedload(Material.unit)).get_or_404(id)
        source_sales_orders = {}
        if order.source_sales_order:
            source_sales_orders[order.source_sales_order.id] = order.source_sales_order
        for item in order.items:
            if item.source_sales_order_item and item.source_sales_order_item.sales_order:
                source_order = item.source_sales_order_item.sales_order
                source_sales_orders[source_order.id] = source_order
        push_source = DocumentPushLine.query.filter_by(
            target_document_id=order.id, status='active'
        ).filter(DocumentPushLine.target_document_type.in_(('requisition', 'other_out'))).first()
        return render_template('out_order_detail.html', order=order, source_sales_orders=list(source_sales_orders.values()), push_source=push_source, operation_logs=get_recent_operation_logs('out_order', id), departments=Department.query.filter_by(status='active').order_by(Department.code.asc(), Department.id.asc()).all(), warehouses=get_active_warehouses(), default_warehouse=None if order.business_type == '销售出库' else get_default_warehouse(), location_management_enabled=location_management_enabled(), today=date.today())

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/update', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def update_out_order(id):
        """Update the header fields of a draft out/requisition order (items untouched)."""
        from app import (Department, OutOrder, api_error, assert_warehouse_active,
                         get_default_warehouse, is_future_date,
                         location_management_enabled, log_operation,
                         parse_date_value)
        order = OutOrder.query.get_or_404(id)
        if order.status != 'pending':
            return api_error('只有草稿状态的出库/领料单可以编辑')

        payload = request.get_json(silent=True)
        data = payload if isinstance(payload, dict) else request.form

        order_date = parse_date_value(data.get('date'), None)
        if not order_date:
            return api_error('日期格式不正确，请重新选择日期')
        if is_future_date(order_date):
            return jsonify({'status': 'error', 'msg': '出库日期不能晚于今天'}), 400

        # 销售出库单保持仓库与来源销售订单一致，不允许在此修改仓库/库位
        is_sale = order.business_type == '销售出库'
        if not is_sale:
            warehouse = (data.get('warehouse') or '').strip()
            if not warehouse:
                default_wh = get_default_warehouse()
                if default_wh:
                    warehouse = default_wh.name
            if not warehouse:
                return jsonify({'status': 'error', 'msg': '请选择仓库'}), 400
            ok, wh_msg = assert_warehouse_active(warehouse, allow_empty=False)
            if not ok:
                return jsonify({'status': 'error', 'msg': wh_msg}), 400
            order.warehouse = warehouse
            location = (data.get('location') or '').strip()
            if location_management_enabled() and not location:
                return jsonify({'status': 'error', 'msg': '请选择库位'}), 400
            order.location = location

        order.date = order_date
        department_id = (data.get('department_id') or '').strip()
        if department_id in ('', 'None', 'null'):
            order.department_id = None
        else:
            try:
                department_id = int(department_id)
            except (TypeError, ValueError):
                department_id = None
            department = db.session.get(Department, department_id) if department_id else None
            order.department_id = department.id if department else None
        order.customer = (data.get('customer') or '').strip() or None
        order.picker = (data.get('picker') or '').strip() or None
        order.purpose = (data.get('purpose') or '').strip() or None
        # 合同编号/工程名称：领料单（含下推生成的草稿）允许修改
        order.contract_no = (data.get('contract_no') or '').strip() or None
        order.project_name = (data.get('project_name') or '').strip() or None
        order.remark = (data.get('remark') or '').strip()

        try:
            db.session.commit()
            log_operation('编辑出库单', f'{order.business_type or "领料单"}：{order.order_no}', 'out_order', id)
            return jsonify({'status': 'success', 'msg': '保存成功'})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'编辑出库单失败: {e}')
            return api_error('保存失败，请稍后重试')

    @app.route('/out_order/add')
    @app.route('/other_out_order/add')
    @login_required
    def out_order_add_page():
        from datetime import date, datetime
        from app import (Customer, Department, Material, OutOrder, OutOrderItem, Unit,
                         generate_order_no, get_active_warehouses,
                         get_default_warehouse, location_management_enabled,
                         serialize_customer, serialize_material, serialize_unit)
        from sqlalchemy.orm import joinedload
        materials = Material.query.options(joinedload(Material.unit)).all()
        units = Unit.query.all()
        customers = Customer.query.order_by(Customer.code.asc(), Customer.id.asc()).all()
        order_id = request.args.get('order_id', type=int)
        order = None
        if order_id:
            order = OutOrder.query.options(
                joinedload(OutOrder.items).joinedload(OutOrderItem.material).joinedload(Material.unit)
            ).get_or_404(order_id)
            if order.status != 'pending':
                abort(409, '只有反提交后的草稿领料单可以编辑')
        order_type = 'other_out' if request.path == '/other_out_order/add' else (request.args.get('type') or '').strip().lower()
        is_sale_order = order.business_type == '销售出库' if order else order_type in ('sale', 'sales')
        is_other_out = order.business_type == '其他出库' if order else order_type in ('other', 'other_out')
        departments = Department.query.filter_by(status='active').all()
        warehouses = get_active_warehouses()
        order_no = order.order_no if order else generate_order_no('OO' if is_other_out else ('SO' if is_sale_order else 'OUT'))
        order_date = order.date.strftime('%Y-%m-%d') if order and order.date else datetime.now().strftime('%Y-%m-%d')
        return render_template('out_order_add.html',
                             materials=[serialize_material(material) for material in materials],
                             units=[serialize_unit(unit) for unit in units],
                             customers=[serialize_customer(customer) for customer in customers],
                             departments=departments,
                             warehouses=warehouses,
                             default_warehouse=None if is_sale_order else get_default_warehouse(),
                             location_management_enabled=location_management_enabled(),
                             is_sale_order=is_sale_order,
                             is_other_out=is_other_out,
                             default_business_type='其他出库' if is_other_out else ('销售出库' if is_sale_order else '领料单'),
                             # BUG-MENU-2026-07-29-A1: ?type=sale 是"销售出库"业务（单号前缀 SO），
                             # 原本 page_title="新增销售单"会让用户误以为是新建销售订单，改为"新增销售出库单"
                             page_title='新增其他出库单' if is_other_out else ('新增销售出库单' if is_sale_order else '新增领料单'),
                             party_label='客户/领用单位' if is_other_out else ('客户名称' if is_sale_order else '领料部门'),
                             party_required=not is_other_out,
                             return_list_url='/other_out_order' if is_other_out else '/out_order',
                             return_add_url='/other_out_order/add' if is_other_out else '/out_order/add',
                             prefill={
                                 'warehouse': order.warehouse if order else (request.args.get('warehouse') or '').strip(),
                                 'location': order.location if order else (request.args.get('location') or '').strip(),
                                 'purpose': order.purpose if order else (request.args.get('purpose') or '').strip(),
                                 'contract_id': str(order.contract_id or '') if order else (request.args.get('contract_id') or '').strip(),
                                 'contract_no': order.contract_no if order else (request.args.get('contract_no') or '').strip(),
                                 'project_name': order.project_name if order else (request.args.get('project_name') or '').strip(),
                                 'remark': order.remark if order else (request.args.get('remark') or '').strip(),
                                 'customer': order.customer if order else (request.args.get('customer') or '').strip(),
                                 'department_id': str(order.department_id or '') if order else (request.args.get('department_id') or '').strip(),
                                 'picker': order.picker if order else (request.args.get('picker') or '').strip(),
                                 'business_type': order.business_type if order else (request.args.get('business_type') or '').strip(),
                             },
                             order_id=order.id if order else None,
                             order_no=order_no,
                             order_date=order_date,
                             edit_items=[{
                                 'material_code': item.material.code,
                                 'quantity': item.quantity,
                                 'price': item.price,
                                 'contract_no': item.contract_no or '',
                                 'project_name': item.project_name or '',
                                 'remark': item.remark or '',
                             } for item in order.items] if order else [],
                             today=date.today())

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/add', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_out_order():
        from datetime import date
        from app import (DocumentPushLine, Material, OutOrder, OutOrderItem,
                         SalesOrder, api_error, generate_order_no,
                         get_default_warehouse, is_future_date,
                         location_management_enabled, log_operation,
                         parse_date_value, parse_float_value,
                         recalculate_order_total, round_to_2_decimals,
                         validate_inventory_warehouse,
                         validate_sales_warehouse)
        from flask_login import current_user
        try:
            payload = request.get_json(silent=True)
            data = payload if isinstance(payload, dict) else request.form
            order_id = data.get('order_id')
            if order_id in ('', 'None', 'null'):
                order_id = None
            elif order_id:
                try:
                    order_id = int(order_id)
                except (TypeError, ValueError):
                    order_id = None

            order_no = (data.get('order_no') or '').strip() or generate_order_no('OUT')
            order_date = parse_date_value(data.get('date'), date.today())
            if not order_date:
                return api_error('日期格式不正确，请重新选择日期')
            if is_future_date(order_date):
                return jsonify({'status': 'error', 'msg': '出库日期不能晚于今天'}), 400
            business_type = (data.get('business_type') or '').strip()
            if business_type == '生产出库':
                business_type = '领料单'
            if business_type not in ('领料单', '销售出库', '其他出库'):
                business_type = '领料单'
            department_id = data.get('department_id')
            if department_id is None or str(department_id).strip().lower() in ('', 'none', 'null'):
                department_id = None
            customer = (data.get('customer') or '').strip()
            picker = (data.get('picker') or '').strip()
            warehouse = (data.get('warehouse') or '').strip()
            location = (data.get('location') or '').strip()
            remark = (data.get('remark') or '').strip()
            contract_id = data.get('contract_id')
            contract_no = (data.get('contract_no') or '').strip()
            project_name = (data.get('project_name') or '').strip()

            sales_warehouse = None
            if business_type == '销售出库':
                sales_warehouse, warehouse_error = validate_sales_warehouse(
                    warehouse, data.get('warehouse_id')
                )
                if warehouse_error:
                    return jsonify({'status': 'error', 'msg': warehouse_error}), 400
                warehouse = sales_warehouse.name
            else:
                # BUG-2026-08-02-002 修复：领料单/其他出库仓库是必填字段，与库位管理无关。
                # 未填写时优先自动带入默认仓库，无默认仓库则拒绝保存。
                if not warehouse:
                    default_wh = get_default_warehouse()
                    if default_wh:
                        warehouse = default_wh.name
                if not warehouse:
                    return jsonify({'status': 'error', 'msg': '请选择仓库'}), 400
                # INV-AUDIT-005：领料单/其他出库仓库必须存在且 active
                wh_obj, wh_err = validate_inventory_warehouse(warehouse)
                if wh_err:
                    return jsonify({'status': 'error', 'msg': wh_err}), 400
                warehouse = wh_obj.name

            # AGENTS.md 规则二：开启库位管理时，库位为必填项
            if location_management_enabled() and not location:
                return jsonify({'status': 'error', 'msg': '请选择库位'}), 400

            # 转换department_id
            if department_id:
                try:
                    department_id = int(department_id) if department_id else None
                except (TypeError, ValueError):
                    department_id = None

            if order_id:
                order = db.session.get(OutOrder, order_id)
                if not order:
                    return api_error('领料单不存在，请刷新后重试')
                if order.status != 'pending':
                    return api_error('已完成的领料单不能修改')
                if DocumentPushLine.query.filter_by(
                    target_document_id=order.id, status='active'
                ).filter(DocumentPushLine.target_document_type.in_(('requisition', 'other_out'))).first():
                    return jsonify({'status': 'error', 'msg': '下推生成的目标草稿必须从来源单重新选择明细，不能通过普通编辑接口重建明细。'}), 409
                if business_type == '销售出库' and order.source_sales_order_id:
                    source_order = db.session.get(SalesOrder, order.source_sales_order_id)
                    if source_order and source_order.warehouse_id and source_order.warehouse_id != sales_warehouse.id:
                        return jsonify({'status': 'error', 'msg': '销售出库仓库必须与来源销售订单一致'}), 400
            else:
                order = OutOrder.query.filter_by(order_no=order_no).first()
                if order:
                    if order.status != 'pending':
                        return api_error('领料单号已存在，不能重复保存')
                else:
                    order = OutOrder(
                        order_no=order_no,
                        status='pending',
                        operator_id=current_user.id
                    )
            db.session.add(order)
            db.session.flush()

            order.order_no = order_no
            order.date = order_date
            order.business_type = business_type
            order.department_id = department_id
            order.customer = customer
            order.picker = picker or None
            order.warehouse = warehouse
            order.location = location
            order.remark = remark
            order.contract_id = int(contract_id) if contract_id else None
            order.contract_no = contract_no or None
            order.project_name = project_name or None

            submitted_items = []
            if isinstance(payload, dict):
                submitted_items = payload.get('items', []) or []
            elif request.form.get('items'):
                try:
                    submitted_items = json.loads(request.form.get('items', '[]'))
                except json.JSONDecodeError:
                    submitted_items = []

            if not isinstance(submitted_items, list) or not submitted_items:
                db.session.rollback()
                return jsonify({'status': 'error', 'msg': '出库单至少需要一条明细'}), 400

            if submitted_items:
                # SALES-AUDIT-005：编辑重建明细前，按 material_id 保留
                # source_sales_order_item_id 映射，否则重建后丢失来源关联，
                # sync_sales_order_shipment 无法按行级来源回写 shipped_quantity。
                source_item_by_material = {}
                for existing_item in list(order.items):
                    if getattr(existing_item, 'source_sales_order_item_id', None):
                        source_item_by_material[existing_item.material_id] = existing_item.source_sales_order_item_id
                for existing_item in list(order.items):
                    db.session.delete(existing_item)
                db.session.flush()
            else:
                source_item_by_material = {}

            for submitted_item in submitted_items:
                material_code = (submitted_item.get('code') or submitted_item.get('material_code') or '').strip()
                material = Material.query.filter_by(code=material_code).first()
                if not material:
                    db.session.rollback()
                    return api_error(f'物料 {material_code} 不存在')

                quantity = round_to_2_decimals(parse_float_value(submitted_item.get('quantity'), 0))
                if quantity <= 0:
                    db.session.rollback()
                    return api_error(f'物料 {material_code} 的数量必须大于0')

                price = round_to_2_decimals(parse_float_value(submitted_item.get('price'), material.price or 0))
                # SALES-AUDIT-005：优先用前端回传的来源，其次按 material_id 恢复
                preserved_source_id = (
                    int(submitted_item['source_sales_order_item_id'])
                    if submitted_item.get('source_sales_order_item_id')
                    else source_item_by_material.get(material.id)
                )
                db.session.add(OutOrderItem(
                    out_order_id=order.id,
                    material_id=material.id,
                    quantity=quantity,
                    price=price,
                    amount=round_to_2_decimals(quantity * price),
                    remark=(submitted_item.get('remark') or '').strip() or None,
                    contract_id=int(submitted_item.get('contract_id')) if submitted_item.get('contract_id') else None,
                    contract_no=(submitted_item.get('contract_no') or '').strip() or None,
                    project_name=(submitted_item.get('project_name') or '').strip() or None,
                    source_sales_order_item_id=preserved_source_id,
                ))

            recalculate_order_total(order)
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'数据库操作失败: {e}')
                return jsonify({'status': 'error', 'msg': '保存失败，请稍后重试'}), 500

            log_operation('保存领料单', f'领料单：{order_no}', 'out_order', order.id)
            app.logger.info(f'领料单创建成功：{order.order_no}')
            return jsonify({'status': 'success', 'msg': '保存成功', 'id': order.id, 'order_no': order.order_no})
        except Exception as e:
            db.session.rollback()
            app.logger.exception(f'保存出库单失败: {e}')
            return api_error('保存失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/item/add', methods=['POST'])
    @app.route('/out_order/<int:id>/add_item', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_out_order_item(id):
        from app import (Material, OutOrder, OutOrderItem, api_error,
                         parse_float_value, recalculate_order_total,
                         round_to_2_decimals)
        order = OutOrder.query.get_or_404(id)
        if order.status != 'pending':
            return api_error('只有待处理的领料单可以添加明细')

        material_code = (request.form.get('material_code') or request.form.get('code') or '').strip()
        if not material_code:
            return api_error('请选择物料后再添加')

        material = Material.query.filter_by(code=material_code).first()
        if not material:
            return api_error(f'物料 {material_code} 不存在')

        quantity = round_to_2_decimals(parse_float_value(request.form.get('quantity'), 0))
        if quantity <= 0:
            return api_error('数量必须大于0')

        price = round_to_2_decimals(parse_float_value(request.form.get('price'), material.price or 0))

        try:
            item = OutOrderItem(
                out_order_id=id,
                material_id=material.id,
                quantity=quantity,
                price=price,
                amount=round_to_2_decimals(quantity * price),
                remark=(request.form.get('remark') or '').strip() or None
            )
            db.session.add(item)
            recalculate_order_total(order)
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'数据库操作失败: {e}')
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            return jsonify({'status': 'success', 'msg': '明细添加成功', 'item_id': item.id})
        except Exception as e:
            db.session.rollback()
            return api_error('明细添加失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/delete_item/<int:item_id>', methods=['POST'])
    @app.route('/out_order/<int:id>/item/<int:item_id>/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_out_order_item(id, item_id):
        from app import (OutOrder, OutOrderItem, api_error,
                         recalculate_order_total)
        order = OutOrder.query.get_or_404(id)
        if order.status != 'pending':
            return api_error('只有待处理的领料单可以删除明细')

        item = OutOrderItem.query.get_or_404(item_id)
        if item.out_order_id != id:
            return api_error('明细不属于当前领料单')

        try:
            db.session.delete(item)
            recalculate_order_total(order)
            db.session.commit()
            return jsonify({'status': 'success', 'msg': '删除成功'})
        except Exception as e:
            db.session.rollback()
            return api_error('删除失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/item/update', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def update_out_order_item():
        from app import (Material, OutOrderItem, api_error, parse_float_value,
                         recalculate_order_total, round_to_2_decimals)
        item_id = request.form.get('id', type=int)
        if not item_id:
            return api_error('缺少明细ID')

        item = OutOrderItem.query.get_or_404(item_id)
        order = item.out_order
        if order.status != 'pending':
            return api_error('只有待处理的领料单可以修改明细')

        material_code = (request.form.get('code') or '').strip()
        if material_code:
            material = Material.query.filter_by(code=material_code).first()
            if not material:
                return api_error(f'物料 {material_code} 不存在')
            item.material_id = material.id

        quantity = round_to_2_decimals(parse_float_value(request.form.get('quantity'), item.quantity))
        if quantity <= 0:
            return api_error('数量必须大于0')

        price = round_to_2_decimals(parse_float_value(request.form.get('price'), item.price))

        try:
            item.quantity = quantity
            item.price = price
            item.amount = round_to_2_decimals(quantity * price)
            if 'remark' in request.form:
                item.remark = (request.form.get('remark') or '').strip() or None
            recalculate_order_total(order)
            db.session.commit()
            return jsonify({'status': 'success', 'msg': '保存成功'})
        except Exception as e:
            db.session.rollback()
            return api_error('保存失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/check_anomalies', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def check_out_order_anomalies(id):
        """检查出库单异常，返回异常列表供前端确认。"""
        from app import (OutOrder, _ai_call_llm_chat, _ai_llm_configured,
                         _check_out_order_anomalies, api_error)
        order = OutOrder.query.get_or_404(id)
        if order.status != 'pending':
            return api_error('该出库单已提交，不能检查异常')

        anomalies = _check_out_order_anomalies(order)

        # 如果有异常且启用了AI，添加AI原因分析
        if anomalies and _ai_llm_configured():
            for anomaly in anomalies:
                # 生成AI分析提示
                prompt = f"作为仓库管理专家，请分析以下出库异常情况并给出简短建议（50字内）：\n"
                prompt += f"异常类型：{anomaly['type']}\n"
                prompt += f"物料：{anomaly['material']}\n"
                prompt += f"当前值：{anomaly['current']}\n"
                prompt += f"历史均值：{anomaly['average']}\n"
                prompt += f"偏离度：{anomaly['deviation']}\n"
                prompt += f"可能原因和建议："

                try:
                    ai_suggestion = _ai_call_llm_chat(prompt)
                    if ai_suggestion:
                        anomaly['ai_suggestion'] = ai_suggestion.strip()
                except Exception as e:
                    app.logger.warning(f'AI异常分析失败: {e}')

        return jsonify({
            'status': 'success',
            'anomalies': anomalies,
            'has_anomalies': len(anomalies) > 0
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/complete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def complete_out_order(id):
        from app import (OutOrder, Warehouse, _acquire_order_write_lock,
                         _check_out_order_anomalies, allow_negative_stock,
                         api_error, assert_warehouse_active,
                         deduct_location_inventory_atomic, deduct_stock_atomic,
                         get_default_warehouse, get_warehouse_stock_quantities,
                         is_future_date, location_management_enabled,
                         log_operation, normalize_stock_quantity,
                         recalculate_order_total, resolve_inventory_warehouse_id,
                         sales_outbound_remaining_check,
                         sync_sales_order_shipment,
                         validate_sales_outbound_warehouse)
        from sqlalchemy.orm import selectinload
        order = OutOrder.query.get_or_404(id)
        if order.status != 'pending':
            return api_error('该领料单已提交，不能重复操作')
        if is_future_date(order.date):
            return jsonify({'status': 'error', 'msg': '出库日期不能晚于今天，请先修改单据日期'}), 400
        if not order.items:
            return api_error('请至少添加一条领料明细')
        if order.business_type == '销售出库':
            _, warehouse_error = validate_sales_outbound_warehouse(order)
            if warehouse_error:
                return jsonify({'status': 'error', 'msg': warehouse_error}), 400
        # BUG-2026-08-02-003 修复：仓库是出库单必填字段，与库位管理是否启用无关。
        # 存量未填仓库的 pending 单据完成时，先自动带入默认仓库，无默认仓库才拒绝。
        # 注意：此处只做读校验（fast-path），不修改 order.warehouse。
        # 因为下方 _acquire_order_write_lock 在 SQLite 分支会 db.session.rollback()，
        # 锁前的修改会被丢弃。实际赋值放到加锁后完成，保证 commit 时仓库已落库。
        if not order.warehouse and not get_default_warehouse():
            return api_error('请选择仓库')

        # 异常检测（force=true时跳过）
        force_submit = request.args.get('force', '').lower() in ('true', '1', 'yes')
        if not force_submit:
            anomalies = _check_out_order_anomalies(order)
            if anomalies:
                return jsonify({
                    'status': 'warning',
                    'msg': '检测到异常，请确认是否继续提交',
                    'anomalies': anomalies
                })

        try:
            # 加写锁并重新读取状态，避免多 worker 并发重复扣库存
            locked, ok = _acquire_order_write_lock(OutOrder, id, 'pending', selectinload(OutOrder.items))
            if not ok:
                return api_error('该领料单已提交，不能重复操作')
            order = locked
            if not order.items:
                db.session.rollback()
                return api_error('请至少添加一条领料明细')
            # 加锁后再做仓库赋值与必填校验，避免锁前修改被 rollback 丢弃。
            if not order.warehouse:
                default_wh = get_default_warehouse()
                if default_wh:
                    order.warehouse = default_wh.name
            if not order.warehouse:
                db.session.rollback()
                return api_error('请选择仓库')
            # BUG-2026-08-16-021：完成前复核仓库 active（销售出库走专属校验，其余走通用断言）
            if order.business_type != '销售出库':
                wh_ok, wh_err = assert_warehouse_active(order.warehouse, allow_empty=False)
                if not wh_ok:
                    db.session.rollback()
                    return api_error(wh_err or '仓库已停用，请先切换有效仓库')
            # SALES-AUDIT-008：草稿保存后仓库可能被停用，完成前必须复核
            # active 状态（对照 PUR-AUDIT-003 的 in_order.py:1389 修复模式）。
            if order.business_type == '销售出库':
                wh_ok, wh_err = validate_sales_outbound_warehouse(order)
                if not wh_ok:
                    db.session.rollback()
                    return api_error(wh_err or '仓库已停用，请先切换有效仓库')
            # P1-BUGFIX: 库位管理启用时 location 必填（AGENTS.md 规则二）
            if location_management_enabled() and not (order.location or '').strip():
                db.session.rollback()
                return api_error('库位管理已启用，请选择库位')
            use_location = bool(location_management_enabled() and (order.location or order.warehouse))
            # SALES-AUDIT-006：完成前校验每条有来源的明细数量不超过销售订单行
            # 未发货数量，防止"生成小数量草稿→编辑改大→完成"超量出库。
            if order.business_type == '销售出库':
                remaining_ok, remaining_err = sales_outbound_remaining_check(order)
                if not remaining_ok:
                    db.session.rollback()
                    return api_error(remaining_err or '出库数量超过销售订单未发货数量')
            # WMS-AUDIT-2026-08-28 (2): 库位管理关闭时按仓库维度校验库存。
            # 原实现只校验全局 Material.stock，多仓场景下 A 仓有货即可在 B 仓
            # 开单出库、把 B 仓库存扣成负数。与 transfer.py 调出校验对齐；
            # 开启库位管理时由 deduct_location_inventory_atomic 精确拦截。
            if not use_location and not allow_negative_stock():
                wh_obj = None
                if (order.warehouse or '').strip():
                    wh_key = order.warehouse.strip()
                    wh_obj = Warehouse.query.filter(
                        db.or_(Warehouse.name == wh_key, Warehouse.code == wh_key)
                    ).order_by(Warehouse.id.asc()).first()
                if wh_obj:
                    wh_stock = get_warehouse_stock_quantities(wh_obj)
                    for _chk_item in order.items:
                        if not _chk_item.material_id:
                            continue
                        _need = normalize_stock_quantity(_chk_item.quantity or 0)
                        if _need <= 0:
                            continue
                        _avail = wh_stock.get(_chk_item.material_id, 0)
                        if (_avail + 1e-9) < _need:
                            _code = (_chk_item.material.code if _chk_item.material
                                     else str(_chk_item.material_id))
                            db.session.rollback()
                            return api_error(
                                f'出库仓库 {order.warehouse} 库存不足：{_code}'
                                f'（需要 {_need:.2f}，可用 {_avail:.2f}）')
            for item in order.items:
                if not item.material_id:
                    continue
                material_code = item.material.code if item.material else str(item.material_id)
                # 原子扣总库存
                ok, err, _ = deduct_stock_atomic(
                    item.material_id, item.quantity or 0,
                    transaction_type='out',
                    reference_type='out_order',
                    reference_id=order.id,
                    warehouse=order.warehouse,
                )
                if not ok:
                    db.session.rollback()
                    return api_error(err or f'物料 {material_code} 库存不足')
                # 原子扣库位（优先 order.location，未启用库位管理时回退 order.warehouse）
                if use_location:
                    ok2, err2 = deduct_location_inventory_atomic(
                        item.material_id, order.location or order.warehouse, item.quantity or 0,
                        material_code_hint=material_code,
                        warehouse_id=resolve_inventory_warehouse_id(order.warehouse),
                    )
                    if not ok2:
                        db.session.rollback()
                        return api_error(err2 or '库位库存扣减失败')
            order.status = 'completed'
            sync_sales_order_shipment(order, quantity_sign=1)
            recalculate_order_total(order)
            db.session.commit()
            log_operation('完成领料单', f'领料单：{order.order_no}', 'out_order', id)
            app.logger.info(f'领料单完成：{order.order_no}')
            return jsonify({'status': 'success', 'msg': '提交成功'})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'领料单完成失败：{e}')
            return api_error('提交失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/revert', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def revert_out_order(id):
        from app import (OutOrder, _acquire_order_write_lock, add_stock,
                         api_error, location_management_enabled,
                         log_audit, log_operation, recalculate_order_total,
                         sync_sales_order_shipment, update_location_inventory)
        from sqlalchemy.orm import selectinload
        order = OutOrder.query.get_or_404(id)
        if order.status != 'completed':
            return api_error('只有已完成的领料单可以反提交')

        try:
            # 加写锁并重新读取状态，避免多 worker 并发反提交导致库存重复恢复
            locked, ok = _acquire_order_write_lock(OutOrder, id, 'completed', selectinload(OutOrder.items))
            if not ok:
                return api_error('该领料单已反提交，不能重复操作')
            order = locked
            for item in order.items:
                if not item.material or (item.quantity or 0) <= 0:
                    continue
                ok, err = add_stock(item.material, item.quantity or 0,
                                    transaction_type='revert_out',
                                    reference_type='out_order',
                                    reference_id=order.id,
                                    warehouse=order.warehouse)
                if not ok:
                    db.session.rollback()
                    return api_error(err or '库存恢复失败')
                # 同步还原库位库存（与 complete_out_order 对称），仅启用库位管理且有仓库时
                if location_management_enabled() and (order.location or order.warehouse):
                    loc_ok, loc_err = update_location_inventory(item.material, order.location or order.warehouse, item.quantity or 0, warehouse=order.warehouse)
                    if not loc_ok:
                        db.session.rollback()
                        return api_error(loc_err or '库位库存还原失败')
            order.status = 'pending'
            sync_sales_order_shipment(order, quantity_sign=-1)
            recalculate_order_total(order)
            db.session.commit()
            log_operation('反提交领料单', f'领料单：{order.order_no}', 'out_order', id)
            # BUG-2026-08-16-012：反提交领料单写结构化审计
            log_audit(
                'revert_out_order', 'out_order', id,
                target_name=order.order_no,
                old_data={'status': 'completed'},
                new_data={'status': 'pending'},
            )
            return jsonify({'status': 'success', 'msg': '操作完成'})
        except Exception as e:
            db.session.rollback()
            return api_error('操作失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/<int:id>/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_out_order(id):
        from app import (OutOrder, _acquire_order_write_lock,
                         _release_document_push_lines, api_error,
                         log_audit, log_operation)
        from sqlalchemy.orm import selectinload
        order = OutOrder.query.get_or_404(id)
        if order.status != 'pending':
            return api_error('只有待处理的领料单可以删除')

        try:
            # 重新锁定并校验草稿状态，防止并发完成后仍被物理删除。
            locked, ok = _acquire_order_write_lock(OutOrder, id, 'pending', [
                selectinload(OutOrder.items),
            ])
            if not ok:
                return jsonify({'status': 'error', 'msg': '该领料单状态已变更；已完成单请先反提交后再删除'}), 409
            order = locked

            _release_document_push_lines(
                'other_out' if order.business_type == '其他出库' else 'requisition',
                order.id, f'目标草稿 {order.order_no} 已删除'
            )
            for item in list(order.items):
                db.session.delete(item)
            db.session.delete(order)
            db.session.commit()
            log_operation('删除领料单', f'领料单：{order.order_no}', 'out_order', id)
            # BUG-2026-08-16-012：删除领料单写结构化审计
            log_audit(
                'delete_out_order', 'out_order', id,
                target_name=order.order_no,
                old_data={'order_no': order.order_no, 'warehouse': order.warehouse or ''},
                reason='草稿删除',
            )
            return jsonify({'status': 'success', 'msg': '删除成功'})
        except Exception as e:
            db.session.rollback()
            return api_error('删除失败，请稍后重试')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/batch_delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def batch_delete_out_order():
        from app import (OutOrder, _acquire_order_write_lock,
                         _release_document_push_lines, api_error,
                         log_operation)
        from sqlalchemy.orm import joinedload, selectinload
        payload = request.get_json(silent=True) or {}
        ids = payload.get('ids') or request.form.getlist('ids')
        ids = [int(item_id) for item_id in ids if str(item_id).isdigit()]
        if not ids:
            return api_error('请选择要删除的领料单')
        if len(ids) > 100:
            return jsonify({'status': 'error', 'msg': '单次批量操作不能超过 100 条，请分批处理'}), 400

        orders = OutOrder.query.options(joinedload(OutOrder.items)).filter(OutOrder.id.in_(ids)).all()
        blocked = [order.order_no for order in orders if order.status != 'pending']
        if blocked:
            return api_error('以下领料单已完成，不能删除：' + ', '.join(blocked))

        deleted_count = 0
        skipped = []
        # 逐条加写锁并独立提交，单点失败仅回滚自身，不影响其余单据。
        for order in list(orders):
            order_id = order.id
            order_no = order.order_no
            try:
                # 重新加锁并校验草稿状态，防止并发完成/反提交后状态已变更。
                locked, ok = _acquire_order_write_lock(
                    OutOrder, order_id, 'pending', selectinload(OutOrder.items)
                )
                if not ok or locked is None:
                    skipped.append(f'{order_no}(状态已变更)')
                    db.session.rollback()
                    continue
                order = locked
                _release_document_push_lines(
                    'other_out' if order.business_type == '其他出库' else 'requisition',
                    order.id, f'目标草稿 {order.order_no} 已批量删除'
                )
                for item in list(order.items):
                    db.session.delete(item)
                db.session.delete(order)
                db.session.commit()
                deleted_count += 1
                log_operation('批量删除领料单', f'领料单：{order_no}', 'out_order', order_id)
            except Exception:
                db.session.rollback()
                skipped.append(f'{order_no}(错误)')
                app.logger.exception('批量删除领料单失败: %s', order_no)

        msg = f'批量删除完成，共删除 {deleted_count} 张领料单'
        if skipped:
            msg += f'，跳过 {len(skipped)} 张：{", ".join(skipped[:10])}'
        return jsonify({
            'status': 'success',
            'msg': msg,
            'deleted': deleted_count,
            'skipped': skipped,
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order/batch_complete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def batch_complete_out_order():
        from app import (OutOrder, Warehouse, _acquire_order_write_lock,
                         _check_out_order_anomalies, allow_negative_stock,
                         api_error, deduct_stock_atomic, get_default_warehouse,
                         get_warehouse_stock_quantities, is_future_date,
                         is_stock_sufficient, location_management_enabled,
                         log_operation, normalize_stock_quantity,
                         recalculate_order_total,
                         sales_outbound_remaining_check,
                         sync_sales_order_shipment, update_location_inventory,
                         validate_sales_outbound_warehouse)
        from sqlalchemy.orm import joinedload, selectinload
        payload = request.get_json(silent=True) or {}
        ids = payload.get('ids') or request.form.getlist('ids')
        ids = [int(item_id) for item_id in ids if str(item_id).isdigit()]
        if not ids:
            return api_error('请选择要审核的领料单')
        if len(ids) > 100:
            return jsonify({'status': 'error', 'msg': '单次批量操作不能超过 100 条，请分批处理'}), 400
        orders = OutOrder.query.options(joinedload(OutOrder.items)).filter(OutOrder.id.in_(ids)).all()
        completed = 0
        skipped = []
        for order in list(orders):
            # 防止列表中重复 id 触发同一单据被处理两次
            order_id = order.id
            if order.status != 'pending':
                skipped.append(order.order_no)
                continue
            if not order.items:
                skipped.append(f'{order.order_no}(无明细)')
                continue
            # 重新加锁并校验状态，避免并发批量/单据完成请求重复审核同一张领料单
            # （重复审核会重复扣库存、重复推进销售单发货进度）
            locked, lock_ok = _acquire_order_write_lock(
                OutOrder, order_id, 'pending', selectinload(OutOrder.items)
            )
            if not lock_ok or locked is None:
                skipped.append(f'{order.order_no}(状态已变更)')
                db.session.rollback()
                continue
            order = locked
            # BUG-2026-08-16-004 修复：批量完成补齐与单据版 complete_out_order
            # 一致的业务校验，防止批量入口绕过单据完成门禁。
            # ① 未来日期拒绝（BUG-DATE-2026-07-27-001 同款规则）
            if is_future_date(order.date):
                skipped.append(f'{order.order_no}(出库日期晚于今天)')
                db.session.rollback()
                continue
            # BUG-2026-08-02-004 修复：批量完成时仓库必填校验，与单据版 complete_out_order 一致。
            # 未填仓库时自动带入默认仓库，无默认仓库则跳过本单（不阻断整批）。
            if not order.warehouse:
                default_wh = get_default_warehouse()
                if default_wh:
                    order.warehouse = default_wh.name
            if not order.warehouse:
                skipped.append(f'{order.order_no}(未填写仓库)')
                db.session.rollback()
                continue
            # ② 销售出库仓库有效性 + 与来源销售订单一致（SALES-AUDIT-008）
            if order.business_type == '销售出库':
                wh_ok, wh_err = validate_sales_outbound_warehouse(order)
                if not wh_ok:
                    skipped.append(f'{order.order_no}({wh_err or "仓库无效"})')
                    db.session.rollback()
                    continue
            # P1-BUGFIX: 库位管理启用时 location 必填（AGENTS.md 规则二）
            if location_management_enabled() and not (order.location or '').strip():
                skipped.append(f'{order.order_no}(未填写库位)')
                db.session.rollback()
                continue
            # ③ 销售出库超发拦截（SALES-AUDIT-006）：草稿改大后批量放行会超发
            if order.business_type == '销售出库':
                remaining_ok, remaining_err = sales_outbound_remaining_check(order)
                if not remaining_ok:
                    skipped.append(f'{order.order_no}({remaining_err or "出库数量超过销售订单未发货数量"})')
                    db.session.rollback()
                    continue
            # ④ 异常检测：批量无 force 交互通道，异常单一律跳过转人工单独审核
            anomalies = _check_out_order_anomalies(order)
            if anomalies:
                skipped.append(f'{order.order_no}(检测到异常，请单独审核)')
                db.session.rollback()
                continue
            stock_ok = True
            # WMS-AUDIT-2026-08-28 (2): 批量出库同样按仓库维度校验（与单张对齐）
            if not location_management_enabled() and not allow_negative_stock():
                wh_obj = None
                if (order.warehouse or '').strip():
                    wh_key = order.warehouse.strip()
                    wh_obj = Warehouse.query.filter(
                        db.or_(Warehouse.name == wh_key, Warehouse.code == wh_key)
                    ).order_by(Warehouse.id.asc()).first()
                if wh_obj:
                    wh_stock = get_warehouse_stock_quantities(wh_obj)
                    for _chk_item in order.items:
                        if not _chk_item.material_id:
                            continue
                        _need = normalize_stock_quantity(_chk_item.quantity or 0)
                        if _need <= 0:
                            continue
                        _avail = wh_stock.get(_chk_item.material_id, 0)
                        if (_avail + 1e-9) < _need:
                            _code = (_chk_item.material.code if _chk_item.material
                                     else str(_chk_item.material_id))
                            skipped.append(
                                f'{order.order_no}(出库仓库 {order.warehouse} '
                                f'库存不足：{_code} 需要 {_need:.2f}，可用 {_avail:.2f})')
                            stock_ok = False
                            break
            for item in order.items:
                stock = normalize_stock_quantity(item.material.stock or 0)
                quantity = normalize_stock_quantity(item.quantity or 0)
                if item.material and not allow_negative_stock() and not is_stock_sufficient(stock, quantity):
                    skipped.append(f'{order.order_no}(物料{item.material.code}库存不足)')
                    stock_ok = False
                    break
            if not stock_ok:
                db.session.rollback()
                continue
            try:
                for item in order.items:
                    # 使用原子扣减避免并发超卖，并检查返回值
                    ok, error_msg, _ = deduct_stock_atomic(item.material_id, item.quantity or 0,
                                 transaction_type='out',
                                 reference_type='out_order',
                                 reference_id=order.id,
                                 warehouse=order.warehouse)
                    if not ok:
                        raise ValueError(error_msg or f'物料 {item.material.code if item.material else ""} 库存不足')
                    # 同步库位库存（与单据版 complete_out_order 对称）
                    if location_management_enabled() and (order.location or order.warehouse):
                        loc_ok, loc_err = update_location_inventory(item.material, order.location or order.warehouse, -(item.quantity or 0), warehouse=order.warehouse)
                        if not loc_ok:
                            raise ValueError(loc_err or '库位库存扣减失败')
                order.status = 'completed'
                sync_sales_order_shipment(order, quantity_sign=1)
                recalculate_order_total(order)
                # 每张单据独立 commit，保证单点失败仅回滚自身，不影响后续单据
                db.session.commit()
                completed += 1
            except Exception as e:
                db.session.rollback()
                skipped.append(f'{order.order_no}(错误: {e})')
        msg = f'批量审核完成，共审核 {completed} 张领料单'
        if skipped:
            msg += f'，跳过 {len(skipped)} 张：{", ".join(skipped[:10])}'
        return jsonify({'status': 'success', 'msg': msg, 'completed': completed})

    @app.route('/out_order/<int:id>/preview_template')
    @login_required
    def preview_out_order_template(id):
        from app import (OutOrder, OutOrderPrintTemplate)
        order = OutOrder.query.get_or_404(id)
        template = get_default_print_template(OutOrderPrintTemplate)

        if template and template.template_type == 'excel' and template.excel_template_path:
            return jsonify({
                'status': 'success',
                'msg': '操作完成',
                'type': 'excel',
                'template_path': template.excel_template_path
            })

        # 打印模板仅允许 Excel：无 Excel 模板时预览内置打印页
        return jsonify({'status': 'success', 'msg': '操作完成', 'type': 'html',
                        'content': render_template('print_out.html', order=order)})

    def _render_out_order_print(id):
        # PRINT-ROUTING-F01-P3：抽出的未装饰实现，供 /print（ptoken 免登录）复用，
        # 避免直接调用带 @login_required 的视图函数导致 ptoken 通过外层仍被内层重定向。
        from app import (OutOrder, OutOrderPrintTemplate)
        order = OutOrder.query.get_or_404(id)
        template = get_default_print_template(OutOrderPrintTemplate)
        # 打印模板仅允许 Excel：历史 HTML 模板不再渲染，一律回退内置打印页
        if template and template.template_type == 'excel':
            return render_template('print_out_with_excel.html', order=order, template=template)
        return render_template('print_out.html', order=order)

    @app.route('/out_order/<int:id>/print_with_template')
    @login_required
    def print_out_order_with_template(id):
        return _render_out_order_print(id)

    @app.route('/out_order/<int:id>/print')
    @print_token_or_login_required(job_type='out_order')  # PRINT-ROUTING-F01-P3 + BUG-2026-08-24-002：ptoken 绑定目标出库单
    def print_out_order(id):
        return _render_out_order_print(id)

    @app.route('/out_order/<int:id>/print_excel')
    @login_required
    def print_out_order_excel(id):
        """按用户选定（或默认）Excel 打印模板填充并下载；无模板时回退内置版式。"""
        from app import OutOrder, OutOrderPrintTemplate
        from utils import resolve_print_template
        order = OutOrder.query.get_or_404(id)
        template = resolve_print_template(OutOrderPrintTemplate, request.args.get('template_id', type=int))
        if template and template.excel_template_path:
            import os
            from print_fill import build_filled_print_excel, template_file_abspath
            template_path = template_file_abspath(template.excel_template_path, app.static_folder)
            if template_path and os.path.exists(template_path):
                from datetime import datetime
                output = build_filled_print_excel(
                    template_path, order,
                    date_str=datetime.now().strftime('%Y-%m-%d'),
                )
                filename = f'{template.name or "领料单"}_{order.order_no or id}.xlsx'
                return send_file(
                    output,
                    download_name=filename,
                    as_attachment=True,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
        output = _build_out_order_excel(order)
        filename = f'领料单_{order.order_no or id}.xlsx'
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @app.route('/out_order_print_templates.json')
    @login_required
    def out_order_print_templates_json():
        """返回可选的 Excel 打印模板列表（用于打印时选择模板）。"""
        from app import OutOrderPrintTemplate
        templates = OutOrderPrintTemplate.query.filter_by(template_type='excel').order_by(
            OutOrderPrintTemplate.is_default.desc(), OutOrderPrintTemplate.updated_at.desc()
        ).all()
        return jsonify({
            'status': 'success',
            'templates': [{
                'id': t.id,
                'name': t.name,
                'is_default': bool(t.is_default),
                'has_file': bool(t.excel_template_path),
            } for t in templates],
        })

    @app.route('/out_order_print_template/<int:template_id>/download')
    @login_required
    def download_out_order_print_template(template_id):
        """下载指定打印模板的 Excel 文件。"""
        import os
        from app import OutOrderPrintTemplate, api_error
        from print_fill import template_file_abspath
        template = OutOrderPrintTemplate.query.get_or_404(template_id)
        if not template.excel_template_path:
            return api_error('打印模板缺少 Excel 文件'), 404
        template_path = template_file_abspath(template.excel_template_path, app.static_folder)
        if not template_path or not os.path.exists(template_path):
            return api_error('打印模板文件不存在'), 404
        from flask import send_from_directory
        return send_from_directory(
            os.path.dirname(template_path) or '.',
            os.path.basename(template_path),
            as_attachment=True,
            download_name=f'{template.name or "打印模板"}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @app.route('/out_order/print_template')
    @app.route('/out_order_print_template')
    @login_required
    def out_order_print_template_list():
        from app import OutOrderPrintTemplate, _print_template_query_from_args
        from utils import get_default_print_template
        query, filters, sort_by, sort_order = _print_template_query_from_args(OutOrderPrintTemplate)
        templates = query.all()
        default_template = get_default_print_template(OutOrderPrintTemplate)
        return render_template('out_order_print_template.html', templates=templates, filters=filters, sort_by=sort_by, sort_order=sort_order, default_template=default_template)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order_print_template/add', methods=['POST'])
    @require_role('admin')
    @login_required
    def add_out_order_print_template():
        from app import OutOrderPrintTemplate, create_print_template
        return create_print_template(OutOrderPrintTemplate, 'out_order_template')

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order_print_template/<int:template_id>/set_default', methods=['POST'])
    @require_role('admin')
    @login_required
    def set_default_out_order_print_template(template_id):
        from app import OutOrderPrintTemplate, set_default_print_template
        return set_default_print_template(OutOrderPrintTemplate, template_id)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/out_order_print_template/<int:template_id>/delete', methods=['POST'])
    @require_role('admin')
    @login_required
    def delete_out_order_print_template(template_id):
        from app import OutOrderPrintTemplate, delete_print_template
        return delete_print_template(OutOrderPrintTemplate, template_id)

    @app.route('/out_order/export')
    @login_required
    def export_out_order():
        from app import (Material, OutOrder, OutOrderItem,
                         _apply_out_order_search, _apply_status_date_filters,
                         _get_order_list_filters, resolve_request_warehouse)
        from openpyxl import Workbook
        from sqlalchemy.orm import joinedload, selectinload
        wb = Workbook()
        ws = wb.active
        ws.title = '领料单'
        ws.append(['单据编号', '日期', '领料部门', '领料人', '用途', '仓库', '物料编码', '物料名称', '规格', '合同单号', '工程名称', '单位', '数量', '单价', '金额', '状态', '备注'])
        status_filter, search, date_start, date_end, sort_by, sort_order = _get_order_list_filters(('pending', 'completed'))
        allowed_sorts = {'order_no', 'date', 'department_id', 'customer', 'business_type', 'purpose', 'status', 'created_at', 'total_amount'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        query = db.session.query(OutOrder).outerjoin(OutOrderItem, OutOrderItem.out_order_id == OutOrder.id).options(
            joinedload(OutOrder.department),
            selectinload(OutOrder.items).joinedload(OutOrderItem.material).joinedload(Material.unit),
        )
        query = _apply_status_date_filters(query, OutOrder, status_filter, date_start, date_end)
        query = _apply_out_order_search(query, search)
        warehouse, warehouse_error = resolve_request_warehouse(request.args)
        if warehouse_error:
            from app import api_error
            return api_error(warehouse_error, 400)
        query = query.filter(OutOrder.warehouse == warehouse.name)
        sort_col = getattr(OutOrder, sort_by, OutOrder.created_at)
        query = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc(), OutOrder.id.desc()).distinct()
        orders = query.all()
        for order in orders:
            if order.items:
                for item in order.items:
                    ws.append([
                        order.order_no,
                        order.date.strftime('%Y-%m-%d') if order.date else '',
                        order.customer or (order.department.name if order.department else '') or '',
                        order.picker or '',
                        order.business_type or order.purpose or '',
                        order.warehouse or '',
                        item.material.code if item.material else '',
                        item.material.name if item.material else '',
                        item.material.spec if item.material else '',
                        item.contract_no or '',
                        item.project_name or '',
                        item.material.unit.name if item.material and item.material.unit else '',
                        item.quantity or 0,
                        item.price or 0,
                        item.amount or 0,
                        '未审核/待完成' if order.status == 'pending' else ('已完成' if order.status == 'completed' else (order.status or '')),
                        order.remark or ''
                    ])
            else:
                ws.append([
                    order.order_no,
                    order.date.strftime('%Y-%m-%d') if order.date else '',
                    (order.department.name if order.department else '') or order.customer or '',
                    order.picker or '',
                    order.business_type or order.purpose or '',
                    order.warehouse or '',
                    '', '', '', '', '', 0, 0, 0,
                    '未审核/待完成' if order.status == 'pending' else ('已完成' if order.status == 'completed' else (order.status or '')),
                    order.remark or ''
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='out_orders.xlsx', as_attachment=True)

    @app.route('/out_order/<int:id>/export')
    @login_required
    def export_single_out_order(id):
        from app import OutOrder
        from openpyxl import Workbook
        order = OutOrder.query.get_or_404(id)
        wb = Workbook()
        ws = wb.active
        ws.title = '领料单'
        ws.append(['单据编号', '日期', '领料部门', '领料人', '用途', '仓库', '物料编码', '物料名称', '规格', '合同单号', '工程名称', '单位', '数量', '单价', '金额', '备注'])
        if order.items:
            for item in order.items:
                ws.append([
                    order.order_no,
                    order.date.strftime('%Y-%m-%d') if order.date else '',
                    (order.department.name if order.department else '') or order.customer or '',
                    order.picker or '',
                    order.business_type or order.purpose or '',
                    order.warehouse or '',
                    item.material.code if item.material else '',
                    item.material.name if item.material else '',
                    item.material.spec if item.material else '',
                    item.contract_no or '',
                    item.project_name or '',
                    item.material.unit.name if item.material and item.material.unit else '',
                    item.quantity or 0,
                    item.price or 0,
                    item.amount or 0,
                    order.remark or ''
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name=f'out_order_{order.order_no}.xlsx', as_attachment=True)
