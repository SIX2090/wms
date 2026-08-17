#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 报表（report）域路由。
#
# 批量拆分模式：与销售（sales）域一致，采用「register_report_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 report、purchase_report、
# report_dashboard、report_dashboard_ai_insights、report_view、report_api_query、
# report_inout_print、report_inout_export、report_stock_print、
# report_print_not_implemented 等），
# 与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / flask_login / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（REPORT_DEFINITIONS / REPORT_TYPE_ORDER /
#   build_report_dashboard_context / _get_report_definition / _build_report_filters /
#   _build_report_payload / _build_excel_response / _ai_llm_configured /
#   _ai_call_llm_chat / api_error / get_active_warehouses / get_default_warehouse /
#   inventory_alert_enabled / _material_alert_status_values / Supplier / Material /
#   InOrder / OutOrder 等）在各路由函数内延迟导入（请求期才执行），
#   避免 app.py 模块加载期触发循环导入。
# - 日志复用 register_report_routes(app) 传入的 app.logger（与 app.py 原实现一致）。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io
import json

from flask import abort, jsonify, render_template, request, send_file, url_for
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 report_* 各路由测试覆盖
def register_report_routes(app):
    @app.route('/report')
    @login_required
    def report():
        from app import REPORT_DEFINITIONS, REPORT_TYPE_ORDER
        report_cards = [
            {'report_type': report_type, **REPORT_DEFINITIONS[report_type]}
            for report_type in REPORT_TYPE_ORDER
        ]
        return render_template('report.html', report_cards=report_cards)

    @app.route('/purchase_report')
    @login_required
    def purchase_report():
        available_reports = [
            {
                'title': '采购订单执行统计表',
                'description': '按采购单和物料查看采购数量、已入库、未入库和逾期情况。',
                'url': url_for('report_view', report_type='purchase_order_execution'),
                'icon': 'bi-bag-check',
            },
            {
                'title': '供应商采购汇总表',
                'description': '按供应商汇总采购订单、金额、已入库和未入库。',
                'url': url_for('report_view', report_type='supplier_purchase_summary'),
                'icon': 'bi-building',
            },
            {
                'title': '物料采购汇总表',
                'description': '按物料编号、名称和规格汇总采购数量和金额。',
                'url': url_for('report_view', report_type='material_purchase_summary'),
                'icon': 'bi-box',
            },
            {
                'title': '采购价格分析表',
                'description': '对比物料、供应商、均价、最低价、最高价和最近采购价。',
                'url': url_for('report_view', report_type='purchase_price_analysis'),
                'icon': 'bi-currency-yen',
            },
            {
                'title': '采购入库明细报表',
                'description': '按入库单据查询采购入库流水，适合核对到货、数量和金额。',
                'url': url_for('report_view', report_type='in_detail'),
                'icon': 'bi-box-arrow-in-down',
            },
            {
                'title': '采购单执行跟踪',
                'description': '查看采购单下推、待入库、已入库和执行状态。',
                'url': url_for('purchase_order_list'),
                'icon': 'bi-bag-check',
            },
            {
                'title': '采购申请执行跟踪',
                'description': '查看采购申请转采购单后的执行进度。',
                'url': url_for('purchase_request_list'),
                'icon': 'bi-list-check',
            },
            {
                'title': '供应商档案',
                'description': '维护供应商资料，供采购单和采购统计使用。',
                'url': url_for('supplier_list'),
                'icon': 'bi-building',
            },
        ]
        missing_reports = [
            {'name': '采购入库统计表 / 明细表', 'gap': '当前已有通用入库明细，下一步可增加专门按采购业务聚合的入库统计。', 'priority': '高'},
            {'name': '采购价格波动趋势图', 'gap': '已补采购价格分析表，后续可增加按月趋势和同物料多供应商图表。', 'priority': '中'},
            {'name': '采购到货及时率 / 逾期未到货统计', 'gap': '缺少按交期跟踪供应商到货及时率和逾期未到货明细。', 'priority': '中'},
            {'name': '请购执行统计 / 请购转采购执行明细', 'gap': '已有采购申请下推采购单流程，但缺少转单率、未转单和执行进度统计。', 'priority': '中'},
            {'name': '采购退货统计', 'gap': '当前未形成采购退货业务闭环，因此暂缺退货数量、金额和供应商退货分析。', 'priority': '低'},
            {'name': '采购付款 / 应付类报表', 'gap': '当前系统没有财务应付模块，个人版可先不纳入本轮范围。', 'priority': '低'},
        ]
        return render_template(
            'purchase_report.html',
            available_reports=available_reports,
            missing_reports=missing_reports,
        )

    @app.route('/report/dashboard')
    @login_required
    def report_dashboard():
        from app import (build_report_dashboard_context, get_active_warehouses,
                         get_default_warehouse, resolve_request_warehouse)
        warehouses = get_active_warehouses()
        warehouse, warehouse_error = resolve_request_warehouse(request.args)
        if warehouse is None:
            return render_template(
                'report_dashboard.html',
                stats=None,
                chart_data=None,
                warehouses=warehouses,
                selected_warehouse_id=None,
            )
        stats, chart_data = build_report_dashboard_context(warehouse)
        return render_template(
            'report_dashboard.html',
            stats=stats,
            chart_data=chart_data,
            warehouses=warehouses,
            selected_warehouse_id=warehouse.id,
        )

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/report/dashboard/ai_insights', methods=['POST'])
    @login_required
    def report_dashboard_ai_insights():
        from app import (_ai_call_llm_chat, _ai_llm_configured, api_error,
                         build_report_dashboard_context, resolve_request_warehouse)
        warehouse, warehouse_error = resolve_request_warehouse(request.args)
        if warehouse is None:
            return api_error(warehouse_error or '请选择仓库', 400)
        if not _ai_llm_configured():
            return api_error('请先在系统设置中配置大模型API')
        stats, chart_data = build_report_dashboard_context(warehouse)
        
        # 计算环比增长率
        in_amount_mom = ((stats['month_in_amount'] - stats['last_month_in_amount']) / stats['last_month_in_amount'] * 100) if stats['last_month_in_amount'] > 0 else 0
        out_amount_mom = ((stats['month_out_amount'] - stats['last_month_out_amount']) / stats['last_month_out_amount'] * 100) if stats['last_month_out_amount'] > 0 else 0
        
        prompt = f"""你是仓库管理系统的数据分析师，请基于以下仪表盘数据生成简洁的业务洞察：

【核心指标】
- 本月入库金额：{stats.month_in_amount:.2f} 元，共 {stats.month_in_count} 笔单据
- 本月领料金额：{stats.month_out_amount:.2f} 元，共 {stats.month_out_count} 笔单据
- 库存总量：{stats.total_stock:.2f}，共 {stats.material_count} 种物料
- 库存金额：{stats.stock_value:.2f} 元

【环比分析】
- 入库金额环比：{in_amount_mom:+.1f}%（上月 {stats.last_month_in_amount:.2f} 元）
- 领料金额环比：{out_amount_mom:+.1f}%（上月 {stats.last_month_out_amount:.2f} 元）

【趋势数据】近10天出入库数量趋势：{json.dumps(chart_data.get('trend', {}), ensure_ascii=False)[:500]}

请生成4-6条有价值的业务洞察，包括：
1. 环比分析（本月vs上月，增长/下降原因分析）
2. 出入库趋势分析（近10天波动情况及可能原因）
3. 库存健康度评估（是否有积压或缺货风险）
4. 可执行的改进建议
要求语言简洁专业，每条不超过80字，用换行分隔。"""
        try:
            insights = _ai_call_llm_chat(prompt)
            if not insights:
                return api_error('AI生成失败，请稍后重试')
            return jsonify({'status': 'success', 'insights': insights})
        except Exception as e:
            app.logger.error(f'报表AI解读失败: {e}')
            return jsonify({'status': 'error', 'msg': '生成失败，请稍后重试'}), 500

    @app.route('/report/view/<report_type>')
    @login_required
    def report_view(report_type):
        from app import (Supplier, _get_report_definition, get_active_warehouses,
                         get_default_warehouse)
        definition = _get_report_definition(report_type)
        if definition is None:
            abort(404)
        suppliers = Supplier.query.all()
        # BUG-2026-08-02-014：报表筛选器加仓库下拉框，传 warehouses / default_warehouse 给模板
        return render_template(
            'report_view.html',
            report_type=report_type,
            title=definition['title'],
            suppliers=suppliers,
            warehouses=get_active_warehouses(),
            default_warehouse=get_default_warehouse(),
            summary_labels=definition['summary_labels'],
            summary_types=definition['summary_types'],
            filters_config=definition['filters'],
        )

    @app.route('/report/api/query')
    @app.route('/report/api/<report_type>')
    @login_required
    def report_api_query(report_type=None):
        from app import (_build_excel_response, _build_report_filters, _build_report_payload,
                         _get_report_definition, api_error)
        report_type = report_type or (request.args.get('report_type') or '').strip()
        definition = _get_report_definition(report_type)
        if definition is None:
            return jsonify({'status': 'error', 'msg': 'Unsupported report type'}), 400

        try:
            filters = _build_report_filters()
            # P1-BUGFIX：AGENTS.md 仓库必填规则——报表查询未指定仓库且无默认仓库时拒绝返回数据
            if not filters.get('warehouse_id'):
                return api_error('请选择仓库', 400)
            payload = _build_report_payload(report_type, filters)
        except ValueError as exc:
            # BUG-2026-08-16-019：业务异常详细记录，不把内部细节返回客户端
            app.logger.error(f'report_api_query ValueError({report_type}): {exc}', exc_info=True)
            return jsonify({'status': 'error', 'msg': '报表数据生成失败，请检查查询条件'}), 400

        if filters['export'] == 'excel':
            return _build_excel_response(report_type, payload['columns'], payload['all_rows'])

        return jsonify({
            'status': 'success',
            'title': payload['title'],
            'columns': payload['columns'],
            'data': payload['rows'],
            'summary': payload['summary'],
            'summary_labels': payload['summary_labels'],
            'summary_types': payload['summary_types'],
            'total': payload['total'],
            'page': filters['page'],
            'page_size': filters['page_size'],
        })

    @app.route('/report/print')
    @login_required
    def report_print_not_implemented():
        from app import api_error
        return api_error('报表打印功能未实现', code=404)

    @app.route('/report/inout/print')
    @login_required
    def report_inout_print():
        from openpyxl import Workbook
        from app import InOrder, OutOrder, api_error, resolve_request_warehouse
        # BUG-2026-08-12-005：旧版导出必须按仓库过滤（AGENTS.md 仓库必填）：
        # 显式 warehouse_id/code/name 校验，缺省时带入默认仓库，无默认仓库 400
        warehouse, wh_err = resolve_request_warehouse(request.args)
        if wh_err:
            return api_error(wh_err, 400)
        warehouse_name = warehouse.name or ''
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        wb = Workbook()

        # 入库单sheet
        ws_in = wb.active
        ws_in.title = '入库统计'
        ws_in.append(['单据编号', '日期', '供应商', '物料编码', '物料名称', '数量', '金额'])
        in_query = InOrder.query.filter(InOrder.warehouse == warehouse_name)
        if start_date:
            in_query = in_query.filter(InOrder.date >= start_date)
        if end_date:
            in_query = in_query.filter(InOrder.date <= end_date)
        for order in in_query.order_by(InOrder.date.desc()).all():
            for item in order.items:
                ws_in.append([
                    order.order_no,
                    order.date.strftime('%Y-%m-%d') if order.date else '',
                    order.supplier.name if order.supplier else '',
                    item.material.code if item.material else '',
                    item.material.name if item.material else '',
                    item.quantity or 0,
                    item.amount or 0
                ])
        
        # 领料单sheet
        ws_out = wb.create_sheet('领料统计')
        ws_out.append(['单据编号', '日期', '领料部门', '物料编码', '物料名称', '数量', '金额'])
        out_query = OutOrder.query.filter(OutOrder.warehouse == warehouse_name)
        if start_date:
            out_query = out_query.filter(OutOrder.date >= start_date)
        if end_date:
            out_query = out_query.filter(OutOrder.date <= end_date)
        for order in out_query.order_by(OutOrder.date.desc()).all():
            for item in order.items:
                ws_out.append([
                    order.order_no,
                    order.date.strftime('%Y-%m-%d') if order.date else '',
                    order.customer or '',
                    item.material.code if item.material else '',
                    item.material.name if item.material else '',
                    item.quantity or 0,
                    item.amount or 0
                ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='inout_report.xlsx', as_attachment=True)

    @app.route('/report/inout/export')
    @login_required
    def report_inout_export():
        return report_inout_print()

    @app.route('/report/stock/print')
    @login_required
    def report_stock_print():
        from openpyxl import Workbook
        from sqlalchemy.orm import joinedload
        from app import (Material, api_error, get_warehouse_stock_quantities,
                         inventory_alert_enabled, resolve_request_warehouse)
        # BUG-2026-08-12-005：库存导出必须输出仓库级库存，不得输出 Material.stock 跨仓总数；
        # 仓库解析规则与新版报表一致（显式参数校验 + 默认仓库回退，无默认 400）
        warehouse, wh_err = resolve_request_warehouse(request.args)
        if wh_err:
            return api_error(wh_err, 400)
        quantities = get_warehouse_stock_quantities(warehouse)
        wb = Workbook()
        ws = wb.active
        ws.title = '库存报表'
        headers = ['物料编码', '物料名称', '规格型号', '单位', '分类', '当前库存']
        if inventory_alert_enabled():
            headers.extend(['最低库存', '安全库存', '库存状态'])
        ws.append(headers)
        materials = Material.query.options(joinedload(Material.unit)).all()
        status_map = {
            'low': '低于最低库存',
            'danger': '低于安全库存',
            'normal': '正常',
            'disabled': '未启用预警',
        }
        for m in materials:
            stock = quantities.get(m.id, 0)
            row = [
                m.code or '',
                m.name or '',
                m.spec or '',
                m.unit.name if m.unit else '',
                m.category.name if m.category else '',
                stock,
            ]
            if inventory_alert_enabled():
                # 库存状态按仓库级数量判定，口径与 _material_alert_status_values 一致
                min_stock = m.min_stock or 0
                safety_stock = max(m.reorder_point or 0, min_stock)
                if min_stock <= 0 and safety_stock <= 0:
                    alert_status = 'disabled'
                elif stock <= min_stock:
                    alert_status = 'low'
                elif stock <= safety_stock:
                    alert_status = 'danger'
                else:
                    alert_status = 'normal'
                row.extend([min_stock, safety_stock, status_map.get(alert_status, '正常')])
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='stock_report.xlsx', as_attachment=True)
