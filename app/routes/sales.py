#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 销售（sales）域路由。
#
# 批量拆分模式：与售后出库（after_sale_out）域一致，采用「register_sales_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 sales_order_list、
# add_sales_order、sales_order_add、confirm_sales_order 等），
# 与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / flask_login / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（SalesOrder 模型、SalesOrderItem、Customer、Material、
#   OutOrder、OutOrderItem、AfterSaleOutOrder、Warehouse、Employee、StockTransaction、
#   各辅助函数 _workbook_response / validate_sales_warehouse / _read_import_sheet /
#   _get_excel_cell / _get_excel_number / _order_no_from_row / _parse_excel_date /
#   validate_excel_extension / validate_excel_size / round_to_2_decimals /
#   parse_float_value / recalculate_sales_order / generate_sales_order_no /
#   generate_order_no / recalculate_order_total / build_sales_outbound_draft /
#   resolve_active_sales_warehouse / _sales_report_orders / _sales_report_filters_context 等）
#   在各路由函数内延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志复用 register_sales_routes(app) 传入的 app.logger（与 app.py 原实现一致）。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import json

from flask import flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 sales_* 各路由测试覆盖
def register_sales_routes(app):
    @app.route('/sales/download_template')
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def download_sales_order_template():
        from datetime import date
        from app import _workbook_response
        return _workbook_response(
            'sales_order_import_template.xlsx',
            '销售订单导入',
            ['销售订单号', '订单日期', '客户名称', '交货日期', '发货仓库', '业务员', '项目号', '币别', '结算方式', '物料编码', '物料名称', '规格', '单位', '数量', '含税单价', '税率', '批次号', '序列号', '备注'],
            [['SO240001', date.today().isoformat(), '示例客户', date.today().isoformat(), '', '张三', 'PRJ-001', 'CNY', '月结30天', 'A001', '示例物料', '', '', 1, 100, 0.13, '', '', '导入后为草稿']],
        )

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/import', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def import_sales_orders():
        from datetime import date
        from flask_login import current_user
        from app import (Employee, Material, SalesOrder, SalesOrderItem,
                         _find_or_create_customer, _find_or_create_material,
                         _get_excel_cell, _get_excel_number, _import_result,
                         _order_no_from_row, _parse_excel_date, _read_import_sheet,
                         recalculate_sales_order, round_to_2_decimals,
                         validate_excel_extension, validate_excel_size, validate_sales_warehouse)
        file = request.files.get('file')
        if not file:
            return jsonify({'status': 'error', 'msg': '请选择销售订单 Excel 文件'}), 400
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return jsonify({'status': 'error', 'msg': _ext_msg}), 400
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return jsonify({'status': 'error', 'msg': _size_msg}), 400
        aliases = {
            'order_no': ['销售订单号', '订单号', '单据编号'],
            'date': ['订单日期', '日期'],
            'customer': ['客户名称', '客户'],
            'delivery_date': ['交货日期', '交期'],
            'warehouse': ['发货仓库', '仓库'],
            'salesperson': ['业务员', '业务员名称'],
            'project_no': ['项目号', '项目编号'],
            'currency': ['币别', '币种'],
            'settlement_method': ['结算方式', '结算'],
            'material_code': ['物料编码', '物料代码'],
            'material_name': ['物料名称', '物料'],
            'spec': ['规格', '规格型号'],
            'unit': ['单位'],
            'quantity': ['数量', '订单数量'],
            'price': ['含税单价', '单价', '价格'],
            'tax_rate': ['税率'],
            'batch_no': ['批次号', '批次'],
            'serial_no': ['序列号', '序列号/批号'],
            'remark': ['备注'],
        }
        try:
            ws, col_map, header_row = _read_import_sheet(file, aliases)
            required = {'customer', 'material_code', 'quantity'}
            if not required.issubset(col_map):
                return jsonify({'status': 'error', 'msg': f'Excel 缺少必填列：客户名称、物料编码、数量；当前表头：{", ".join(header_row)}'}), 400
            orders_by_no = {}
            order_count = 0
            item_count = 0
            skip = 0
            skip_details = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                order_no = _order_no_from_row(row, col_map, 'order_no', 'SO')
                customer_name = _get_excel_cell(row, col_map, 'customer')
                material_code = _get_excel_cell(row, col_map, 'material_code')
                quantity = _get_excel_number(row, col_map, 'quantity')
                if not customer_name or not material_code or quantity <= 0:
                    skip += 1
                    skip_details.append(f'第 {row_idx} 行：客户、物料编码或数量无效')
                    continue
                order = orders_by_no.get(order_no)
                if not order:
                    existing = SalesOrder.query.filter_by(order_no=order_no).first()
                    if existing:
                        skip += 1
                        skip_details.append(f'第 {row_idx} 行：销售订单号 {order_no} 已存在')
                        continue
                    customer = _find_or_create_customer(customer_name)
                    if not customer:
                        skip += 1
                        skip_details.append(f'第 {row_idx} 行：无法创建客户 {customer_name}')
                        continue
                    salesperson_name = _get_excel_cell(row, col_map, 'salesperson')
                    salesperson = Employee.query.filter_by(name=salesperson_name).first() if salesperson_name else None
                    currency = _get_excel_cell(row, col_map, 'currency') or 'CNY'
                    warehouse_value = _get_excel_cell(row, col_map, 'warehouse')
                    warehouse, warehouse_error = validate_sales_warehouse(warehouse_value)
                    if warehouse_error:
                        skip += 1
                        skip_details.append(f'第 {row_idx} 行：{warehouse_error}')
                        continue
                    order = SalesOrder(order_no=order_no, date=_parse_excel_date(_get_excel_cell(row, col_map, 'date'), date.today()), delivery_date=_parse_excel_date(_get_excel_cell(row, col_map, 'delivery_date'), None) if _get_excel_cell(row, col_map, 'delivery_date') else None, customer_id=customer.id, warehouse=warehouse.name, warehouse_id=warehouse.id, salesperson_id=salesperson.id if salesperson else None, project_no=_get_excel_cell(row, col_map, 'project_no'), currency=currency, settlement_method=_get_excel_cell(row, col_map, 'settlement_method'), status='draft', shipment_status='pending', remark=_get_excel_cell(row, col_map, 'remark'), operator_id=current_user.id)
                    db.session.add(order)
                    db.session.flush()
                    orders_by_no[order_no] = order
                    order_count += 1
                material = Material.query.filter_by(code=material_code).first() or _find_or_create_material(material_code, _get_excel_cell(row, col_map, 'material_name'), _get_excel_cell(row, col_map, 'spec'), _get_excel_cell(row, col_map, 'unit'))
                if not material:
                    skip += 1
                    skip_details.append(f'第 {row_idx} 行：无法创建物料 {material_code}')
                    continue
                price = _get_excel_number(row, col_map, 'price')
                tax_rate = _get_excel_number(row, col_map, 'tax_rate') if 'tax_rate' in col_map else 0.13
                if tax_rate <= 0:
                    tax_rate = 0.13
                included_amount = round_to_2_decimals(quantity * price)
                untaxed_amount = round_to_2_decimals(included_amount / (1 + tax_rate)) if (1 + tax_rate) > 0 else included_amount
                tax_amount = round_to_2_decimals(included_amount - untaxed_amount)
                untaxed_price = round(price / (1 + tax_rate), 4) if quantity > 0 and (1 + tax_rate) > 0 else 0
                db.session.add(SalesOrderItem(sales_order_id=order.id, material_id=material.id, quantity=quantity, price=price, tax_rate=tax_rate, untaxed_price=untaxed_price, untaxed_amount=untaxed_amount, tax_amount=tax_amount, tax_included_amount=included_amount, amount=included_amount, batch_no=_get_excel_cell(row, col_map, 'batch_no'), serial_no=_get_excel_cell(row, col_map, 'serial_no'), remark=_get_excel_cell(row, col_map, 'remark')))
                item_count += 1
            for order in orders_by_no.values():
                recalculate_sales_order(order)
                order.status = 'draft'
                order.shipment_status = 'pending'
            db.session.commit()
            return _import_result('销售订单', order_count, item_count, skip, skip_details, {'draft_only': True})
        except Exception as exc:
            db.session.rollback()
            app.logger.exception('导入销售订单失败')
            return jsonify({'status': 'error', 'msg': f'导入失败：{exc}'}), 500

    @app.route('/sales')
    @login_required
    def sales_order_list():
        from datetime import date, datetime
        from sqlalchemy.orm import joinedload
        from app import (Customer, Employee, Material, SalesOrder, SalesOrderItem,
                         _apply_header_or_item_contract_filters, sales_shipment_status_label,
                         sales_status_label)
        search = (request.args.get('search') or '').strip()
        status = (request.args.get('status') or '').strip()
        customer_id = request.args.get('customer_id', type=int)
        salesperson_id = request.args.get('salesperson_id', type=int)
        date_start = request.args.get('date_start') or ''
        date_end = request.args.get('date_end') or ''
        sort_by = request.args.get('sort', 'date')
        sort_order = request.args.get('order', 'desc')

        # 排序字段白名单
        allowed_sorts = {'order_no', 'date', 'delivery_date', 'status', 'total_amount', 'created_at'}
        if sort_by not in allowed_sorts:
            sort_by = 'date'
        if sort_order not in ['asc', 'desc']:
            sort_order = 'desc'

        query = SalesOrder.query.join(Customer).outerjoin(SalesOrderItem, SalesOrderItem.sales_order_id == SalesOrder.id).outerjoin(Material, SalesOrderItem.material_id == Material.id).distinct()
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(SalesOrder.order_no.like(like), Customer.name.like(like), Customer.code.like(like), Material.code.like(like), Material.name.like(like), SalesOrder.project_no.like(like)))
        contract_no_filter = (request.args.get('contract_no') or '').strip()
        project_name_filter = (request.args.get('project_name') or '').strip()
        query = _apply_header_or_item_contract_filters(
            query, SalesOrder, SalesOrderItem, 'sales_order_id',
            contract_no_filter=contract_no_filter,
            project_name_filter=project_name_filter,
        )
        if status:
            query = query.filter(SalesOrder.status == status)
        if customer_id:
            query = query.filter(SalesOrder.customer_id == customer_id)
        if salesperson_id:
            query = query.filter(SalesOrder.salesperson_id == salesperson_id)
        if date_start:
            try:
                query = query.filter(SalesOrder.date >= datetime.strptime(date_start, '%Y-%m-%d').date())
            except ValueError:
                date_start = ''
        if date_end:
            try:
                query = query.filter(SalesOrder.date <= datetime.strptime(date_end, '%Y-%m-%d').date())
            except ValueError:
                date_end = ''

        # 排序
        sort_col = getattr(SalesOrder, sort_by, SalesOrder.date)
        query = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc())

        # SM-P6-03-2: 每页条数选择器（白名单 20/50/100/200，默认 20，避免极端值导致性能问题）
        requested_per_page = request.args.get('per_page', 20, type=int)
        per_page = requested_per_page if requested_per_page in (20, 50, 100, 200) else 20
        pagination = query.options(joinedload(SalesOrder.customer), joinedload(SalesOrder.salesperson)).paginate(page=request.args.get('page', 1, type=int), per_page=per_page, error_out=False)

        # 计算汇总数据
        summary = {
            'pending_count': 0,
            'partial_count': 0,
            'pending_amount': 0,
            'partial_amount': 0,
            'overdue_count': 0,
            'month_amount': 0
        }

        # 统计待发货单据和金额
        pending_orders = SalesOrder.query.filter(SalesOrder.status.in_(['draft', 'confirmed']), SalesOrder.shipment_status == 'pending').all()
        summary['pending_count'] = len(pending_orders)
        summary['pending_amount'] = sum(float(order.remaining_amount or 0) for order in pending_orders)

        # 统计部分发货
        partial_orders = SalesOrder.query.filter(SalesOrder.status.in_(['draft', 'confirmed']), SalesOrder.shipment_status == 'partial').all()
        summary['partial_count'] = len(partial_orders)
        summary['partial_amount'] = sum(float(order.remaining_amount or 0) for order in partial_orders)

        # 统计逾期未发货
        today = date.today()
        overdue_orders = SalesOrder.query.filter(
            SalesOrder.status.in_(['draft', 'confirmed']),
            SalesOrder.shipment_status != 'shipped',
            SalesOrder.delivery_date < today,
            SalesOrder.delivery_date.isnot(None)
        ).all()
        summary['overdue_count'] = len(overdue_orders)

        # 统计本月销售金额
        month_start = today.replace(day=1)
        month_orders = SalesOrder.query.filter(
            SalesOrder.date >= month_start,
            SalesOrder.status != 'cancelled'
        ).all()
        summary['month_amount'] = sum(float(order.total_amount or 0) for order in month_orders)

        return render_template('sales_order.html',
            pagination=pagination,
            customers=Customer.query.order_by(Customer.code.asc()).all(),
            employees=Employee.query.order_by(Employee.id.asc()).all(),
            filters={
                'search': search,
                'status': status,
                'customer_id': customer_id,
                'salesperson_id': salesperson_id,
                'date_start': date_start,
                'date_end': date_end,
                'contract_no': contract_no_filter,
                'project_name': project_name_filter,
            },
            status_label=sales_status_label,
            shipment_status_label=sales_shipment_status_label,
            summary=summary,
            sort_by=sort_by,
            sort_order=sort_order,
            today=today
        )

    @app.route('/sales/outbound_selection')
    @login_required
    def sales_outbound_selection_page():
        from app import Customer, Employee, get_active_warehouses
        return render_template(
            'sales_outbound_selection.html',
            customers=Customer.query.order_by(Customer.code.asc(), Customer.id.asc()).all(),
            employees=Employee.query.order_by(Employee.id.asc()).all(),
            warehouses=get_active_warehouses(),
        )

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/create_outbound_from_selection', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def create_sales_outbound_from_selection():
        from datetime import date
        from flask_login import current_user
        from sqlalchemy.orm import joinedload
        from app import (OutOrder, OutOrderItem, SalesOrder, SalesOrderItem,
                         STOCK_COMPARE_EPSILON, generate_order_no, log_operation,
                         parse_float_value, recalculate_order_total, round_to_2_decimals,
                         validate_sales_warehouse)
        payload = request.get_json(silent=True) or {}
        selected_items = payload.get('items') or []
        if not isinstance(selected_items, list) or not selected_items:
            return jsonify({'status': 'error', 'msg': '请选择要下推的销售订单明细'}), 400

        selected_qty_by_item_id = {}
        for row in selected_items:
            if not isinstance(row, dict):
                continue
            item_id = row.get('sales_order_item_id') or row.get('item_id') or row.get('id')
            try:
                item_id = int(item_id)
            except (TypeError, ValueError):
                continue
            quantity = round_to_2_decimals(parse_float_value(row.get('quantity'), 0))
            if quantity > 0:
                selected_qty_by_item_id[item_id] = round_to_2_decimals(selected_qty_by_item_id.get(item_id, 0) + quantity)
        if not selected_qty_by_item_id:
            return jsonify({'status': 'error', 'msg': '请选择大于 0 的下推数量'}), 400

        source_items = SalesOrderItem.query.options(
            joinedload(SalesOrderItem.sales_order).joinedload(SalesOrder.customer),
            joinedload(SalesOrderItem.material),
        ).filter(SalesOrderItem.id.in_(selected_qty_by_item_id.keys())).all()
        if len(source_items) != len(selected_qty_by_item_id):
            return jsonify({'status': 'error', 'msg': '部分销售订单明细不存在，请刷新后重试'}), 400

        # SQLite does not honor SELECT ... FOR UPDATE. Acquire a short database
        # write lock before the final pending-draft check, then reload source rows
        # so two worker processes cannot both reserve the same sales lines.
        if db.engine.dialect.name == 'sqlite':
            db.session.rollback()
            db.session.connection().exec_driver_sql('BEGIN IMMEDIATE')
            source_items = SalesOrderItem.query.options(
                joinedload(SalesOrderItem.sales_order).joinedload(SalesOrder.customer),
                joinedload(SalesOrderItem.material),
            ).filter(SalesOrderItem.id.in_(selected_qty_by_item_id.keys())).all()
        else:
            source_items = SalesOrderItem.query.with_for_update().options(
                joinedload(SalesOrderItem.sales_order).joinedload(SalesOrder.customer),
                joinedload(SalesOrderItem.material),
            ).filter(SalesOrderItem.id.in_(selected_qty_by_item_id.keys())).all()

        order_ids = set()
        customer_ids = set()
        warehouses = set()
        warehouse_ids = set()
        conversions = []
        for item in source_items:
            order = item.sales_order
            if not order or order.status != 'confirmed' or order.shipment_status == 'shipped':
                return jsonify({'status': 'error', 'msg': '只能选择已确认且仍有待发货数量的销售订单明细'}), 400
            warehouse, warehouse_error = validate_sales_warehouse(order.warehouse, order.warehouse_id)
            if warehouse_error:
                return jsonify({'status': 'error', 'msg': f'销售订单 {order.order_no} {warehouse_error}'}), 400
            order.warehouse = warehouse.name
            order.warehouse_id = warehouse.id
            quantity = selected_qty_by_item_id[item.id]
            remaining = round_to_2_decimals((item.quantity or 0) - (item.shipped_quantity or 0))
            if quantity - remaining > STOCK_COMPARE_EPSILON:
                return jsonify({'status': 'error', 'msg': f'销售订单 {order.order_no} 的下推数量不能超过未发货数量'}), 400
            pending = OutOrder.query.filter(
                db.or_(OutOrder.source_sales_order_id == order.id, OutOrder.purpose == f'来源销售订单 {order.order_no}'),
                OutOrder.status == 'pending',
            ).first()
            if pending:
                return jsonify({'status': 'error', 'msg': f'销售订单 {order.order_no} 已存在待处理销售出库草稿，请先处理'}), 400
            order_ids.add(order.id)
            customer_ids.add(order.customer_id)
            warehouses.add(warehouse.name)
            warehouse_ids.add(warehouse.id)
            conversions.append((item, order, quantity))

        if len(customer_ids) != 1:
            return jsonify({'status': 'error', 'msg': '一张销售出库单只能选择同一客户的销售订单明细'}), 400
        if len(warehouses) != 1 or len(warehouse_ids) != 1:
            return jsonify({'status': 'error', 'msg': '一张销售出库单只能选择同一发货仓库的销售订单明细'}), 400

        try:
            order_nos = sorted({order.order_no for _, order, _ in conversions})
            customer = source_items[0].sales_order.customer
            outbound = OutOrder(
                order_no=generate_order_no('OU'),
                date=date.today(),
                customer=customer.name if customer else '',
                business_type='销售出库',
                warehouse=next(iter(warehouses)),
                purpose='选单生成销售出库',
                source_sales_order_id=next(iter(order_ids)) if len(order_ids) == 1 else None,
                remark=(payload.get('remark') or '').strip() or ('由销售订单选单生成：' + '、'.join(order_nos)),
                status='pending',
                operator_id=current_user.id,
            )
            db.session.add(outbound)
            db.session.flush()
            for item, order, quantity in conversions:
                price = round_to_2_decimals(item.price or 0)
                db.session.add(OutOrderItem(
                    out_order_id=outbound.id,
                    material_id=item.material_id,
                    source_sales_order_item_id=item.id,
                    quantity=quantity,
                    price=price,
                    amount=round_to_2_decimals(quantity * price),
                    remark=f'来源销售订单 {order.order_no}',
                ))
                order.shipment_order_no = outbound.order_no
            recalculate_order_total(outbound)
            db.session.commit()
            log_operation('销售订单选单生成销售出库单', f'{", ".join(order_nos)} -> {outbound.order_no}', 'out_order', outbound.id)
            return jsonify({'status': 'success', 'id': outbound.id, 'order_no': outbound.order_no, 'redirect_url': url_for('out_order_detail', id=outbound.id)})
        except Exception:
            db.session.rollback()
            app.logger.exception('销售订单选单生成销售出库单失败')
            return jsonify({'status': 'error', 'msg': '生成销售出库草稿失败，请稍后重试'}), 500

    @app.route('/sales/add')
    @login_required
    def sales_order_add_page():
        from datetime import date
        from sqlalchemy.orm import joinedload
        from app import (Customer, Employee, Material, generate_sales_order_no,
                         get_active_warehouses, serialize_material)
        return render_template('sales_order_add.html', order_no=generate_sales_order_no(), order_date=date.today().isoformat(), customers=Customer.query.order_by(Customer.code.asc()).all(), warehouses=get_active_warehouses(), materials=[serialize_material(material) for material in Material.query.options(joinedload(Material.unit)).order_by(Material.code.asc()).all()], employees=Employee.query.order_by(Employee.id.asc()).all(), default_tax_rate=0.13)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/add', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def sales_order_add():
        from datetime import date, datetime
        from flask_login import current_user
        from app import (Customer, Employee, Material, SalesOrder, SalesOrderItem,
                         generate_sales_order_no, log_operation, parse_float_value,
                         recalculate_sales_order, round_to_2_decimals, validate_sales_warehouse)
        payload = request.get_json(silent=True) or request.form
        try:
            customer_id = int(payload.get('customer_id') or 0)
            customer = db.session.get(Customer, customer_id)
            if not customer:
                return jsonify({'status': 'error', 'msg': '请选择客户'}), 400
            items = payload.get('items', [])
            if isinstance(items, str):
                items = json.loads(items or '[]')
            if not items:
                return jsonify({'status': 'error', 'msg': '请至少添加一条销售明细'}), 400
            salesperson_id = int(payload.get('salesperson_id') or 0) if payload.get('salesperson_id') else None
            if salesperson_id and not db.session.get(Employee, salesperson_id):
                salesperson_id = None
            warehouse, warehouse_error = validate_sales_warehouse(
                payload.get('warehouse'), payload.get('warehouse_id')
            )
            if warehouse_error:
                return jsonify({'status': 'error', 'msg': warehouse_error}), 400
            _contract_id_raw = (payload.get('contract_id') or '').strip()
            order = SalesOrder(
                order_no=(payload.get('order_no') or '').strip() or generate_sales_order_no(),
                customer_id=customer.id,
                operator_id=current_user.id,
                date=datetime.strptime(payload.get('date') or date.today().isoformat(), '%Y-%m-%d').date(),
                delivery_date=datetime.strptime(payload.get('delivery_date'), '%Y-%m-%d').date() if payload.get('delivery_date') else None,
                warehouse=warehouse.name,
                warehouse_id=warehouse.id,
                remark=(payload.get('remark') or '').strip(),
                salesperson_id=salesperson_id,
                project_no=(payload.get('project_no') or '').strip() or None,
                currency=(payload.get('currency') or '').strip() or 'CNY',
                settlement_method=(payload.get('settlement_method') or '').strip() or None,
                contract_id=int(_contract_id_raw) if _contract_id_raw else None,
                contract_no=(payload.get('contract_no') or '').strip() or None,
                project_name=(payload.get('project_name') or '').strip() or None,
                status='draft',
            )
            db.session.add(order)
            db.session.flush()
            for data in items:
                material = Material.query.filter_by(code=(data.get('code') or data.get('material_code') or '').strip()).first()
                quantity = round_to_2_decimals(parse_float_value(data.get('quantity'), 0))
                if not material or quantity <= 0:
                    db.session.rollback()
                    return jsonify({'status': 'error', 'msg': '物料编码不存在或数量无效'}), 400
                price = round_to_2_decimals(parse_float_value(data.get('price'), material.price or 0))
                tax_rate = parse_float_value(data.get('tax_rate'), 0.13)
                if tax_rate < 0:
                    tax_rate = 0
                db.session.add(SalesOrderItem(
                    sales_order_id=order.id,
                    material_id=material.id,
                    quantity=quantity,
                    price=price,
                    amount=round_to_2_decimals(quantity * price),
                    tax_rate=tax_rate,
                    batch_no=(data.get('batch_no') or '').strip() or None,
                    serial_no=(data.get('serial_no') or '').strip() or None,
                    remark=(data.get('remark') or '').strip() or None,
                    contract_id=int(data.get('contract_id')) if data.get('contract_id') else None,
                    contract_no=(data.get('contract_no') or '').strip() or None,
                    project_name=(data.get('project_name') or '').strip() or None,
                ))
            recalculate_sales_order(order)
            db.session.commit()
            log_operation('保存销售订单', f'销售订单：{order.order_no}', 'sales_order', order.id)
            return jsonify({'status': 'success', 'id': order.id, 'order_no': order.order_no})
        except (ValueError, TypeError, json.JSONDecodeError):
            db.session.rollback()
            return jsonify({'status': 'error', 'msg': '日期或明细格式不正确'}), 400
        except Exception:
            db.session.rollback()
            app.logger.exception('保存销售订单失败')
            return jsonify({'status': 'error', 'msg': '保存失败，请稍后重试'}), 500

    @app.route('/sales/<int:id>')
    @login_required
    def sales_order_detail(id):
        from datetime import date
        from sqlalchemy.orm import joinedload
        from app import (AfterSaleOutOrder, Material, OutOrder, OutOrderItem, SalesOrder,
                         SalesOrderItem, sales_shipment_status_label, sales_status_label)
        order = SalesOrder.query.options(joinedload(SalesOrder.customer), joinedload(SalesOrder.salesperson), joinedload(SalesOrder.operator), joinedload(SalesOrder.items).joinedload(SalesOrderItem.material).joinedload(Material.unit)).get_or_404(id)
        related_sales_out_orders = OutOrder.query.options(
            joinedload(OutOrder.items).joinedload(OutOrderItem.material).joinedload(Material.unit),
        ).filter(
            db.or_(
                OutOrder.source_sales_order_id == order.id,
                OutOrder.purpose == f'来源销售订单 {order.order_no}',
            ),
            OutOrder.business_type == '销售出库',
        ).order_by(OutOrder.date.desc(), OutOrder.id.desc()).all()
        related_after_sale_orders = AfterSaleOutOrder.query.filter(
            AfterSaleOutOrder.source_sales_order_id == order.id,
        ).order_by(AfterSaleOutOrder.date.desc(), AfterSaleOutOrder.id.desc()).all()
        return render_template(
            'sales_order_detail.html',
            order=order,
            related_sales_out_orders=related_sales_out_orders,
            related_after_sale_orders=related_after_sale_orders,
            today=date.today(),
            status_label=sales_status_label,
            shipment_status_label=sales_shipment_status_label,
        )

    @app.route('/sales/<int:id>/edit', methods=['GET'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def sales_order_edit_page(id):
        from sqlalchemy.orm import joinedload
        from app import (Customer, Employee, Material, SalesOrder, SalesOrderItem, Warehouse)
        order = SalesOrder.query.options(joinedload(SalesOrder.items).joinedload(SalesOrderItem.material).joinedload(Material.unit)).get_or_404(id)
        if order.status != 'draft':
            flash('仅草稿状态的销售订单可修改，请先反确认或新建订单', 'warning')
            return redirect(url_for('sales_order_detail', id=id))
        customers = Customer.query.order_by(Customer.code.asc()).all()
        employees = Employee.query.order_by(Employee.id.asc()).all()
        warehouses = Warehouse.query.order_by(Warehouse.id.asc()).all()
        materials = Material.query.order_by(Material.code.asc()).all()
        material_list = [{'code': m.code, 'name': m.name, 'spec': m.spec or '', 'unit': m.unit.name if m.unit else '', 'price': float(m.price or 0)} for m in materials]
        # 已有明细项转 JSON 供前端回填
        existing_items = [{'code': item.material.code if item.material else '', 'quantity': item.quantity, 'price': item.price, 'tax_rate': item.tax_rate, 'batch_no': item.batch_no or '', 'serial_no': item.serial_no or '', 'remark': item.remark or '', 'contract_id': item.contract_id, 'contract_no': item.contract_no or order.contract_no or '', 'project_name': item.project_name or order.project_name or ''} for item in order.items]
        return render_template('sales_order_edit.html', order=order, customers=customers, employees=employees, warehouses=warehouses, materials=materials, material_list=material_list, existing_items=existing_items, default_tax_rate=0.13)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/<int:id>/edit', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def sales_order_edit(id):
        from datetime import date, datetime
        from app import (Customer, Employee, Material, SalesOrder, SalesOrderItem,
                         log_operation, parse_float_value, recalculate_sales_order,
                         round_to_2_decimals, validate_sales_warehouse)
        order = SalesOrder.query.get_or_404(id)
        if order.status != 'draft':
            return jsonify({'status': 'error', 'msg': '仅草稿状态的销售订单可修改'}), 400
        payload = request.get_json(silent=True) or request.form
        try:
            customer_id = int(payload.get('customer_id') or 0)
            customer = db.session.get(Customer, customer_id)
            if not customer:
                return jsonify({'status': 'error', 'msg': '请选择客户'}), 400
            items = payload.get('items', [])
            if isinstance(items, str):
                items = json.loads(items or '[]')
            if not items:
                return jsonify({'status': 'error', 'msg': '请至少保留一条销售明细'}), 400
            salesperson_id = int(payload.get('salesperson_id') or 0) if payload.get('salesperson_id') else None
            if salesperson_id and not db.session.get(Employee, salesperson_id):
                salesperson_id = None
            warehouse, warehouse_error = validate_sales_warehouse(
                payload.get('warehouse'), payload.get('warehouse_id')
            )
            if warehouse_error:
                return jsonify({'status': 'error', 'msg': warehouse_error}), 400
            # 更新订单头
            order.customer_id = customer.id
            order.date = datetime.strptime(payload.get('date') or date.today().isoformat(), '%Y-%m-%d').date()
            order.delivery_date = datetime.strptime(payload.get('delivery_date'), '%Y-%m-%d').date() if payload.get('delivery_date') else None
            order.warehouse = warehouse.name
            order.warehouse_id = warehouse.id
            order.remark = (payload.get('remark') or '').strip()
            order.salesperson_id = salesperson_id
            order.project_no = (payload.get('project_no') or '').strip() or None
            order.currency = (payload.get('currency') or '').strip() or 'CNY'
            order.settlement_method = (payload.get('settlement_method') or '').strip() or None
            _edit_contract_id = (payload.get('contract_id') or '').strip()
            order.contract_id = int(_edit_contract_id) if _edit_contract_id else None
            order.contract_no = (payload.get('contract_no') or '').strip() or None
            order.project_name = (payload.get('project_name') or '').strip() or None
            # 删除旧明细，重建新明细（草稿状态无出库记录，可安全重建）
            SalesOrderItem.query.filter_by(sales_order_id=order.id).delete()
            db.session.flush()
            for data in items:
                material = Material.query.filter_by(code=(data.get('code') or data.get('material_code') or '').strip()).first()
                quantity = round_to_2_decimals(parse_float_value(data.get('quantity'), 0))
                if not material or quantity <= 0:
                    db.session.rollback()
                    return jsonify({'status': 'error', 'msg': '物料编码不存在或数量无效'}), 400
                price = round_to_2_decimals(parse_float_value(data.get('price'), material.price or 0))
                tax_rate = parse_float_value(data.get('tax_rate'), 0.13)
                if tax_rate < 0:
                    tax_rate = 0
                db.session.add(SalesOrderItem(
                    sales_order_id=order.id,
                    material_id=material.id,
                    quantity=quantity,
                    price=price,
                    amount=round_to_2_decimals(quantity * price),
                    tax_rate=tax_rate,
                    batch_no=(data.get('batch_no') or '').strip() or None,
                    serial_no=(data.get('serial_no') or '').strip() or None,
                    remark=(data.get('remark') or '').strip() or None,
                    contract_id=int(data.get('contract_id')) if data.get('contract_id') else None,
                    contract_no=(data.get('contract_no') or '').strip() or None,
                    project_name=(data.get('project_name') or '').strip() or None,
                ))
            recalculate_sales_order(order)
            db.session.commit()
            log_operation('修改销售订单', f'销售订单：{order.order_no}', 'sales_order', order.id)
            return jsonify({'status': 'success', 'id': order.id, 'order_no': order.order_no})
        except (ValueError, TypeError, json.JSONDecodeError):
            db.session.rollback()
            return jsonify({'status': 'error', 'msg': '日期或明细格式不正确'}), 400
        except Exception:
            db.session.rollback()
            app.logger.exception('修改销售订单失败')
            return jsonify({'status': 'error', 'msg': '保存失败，请稍后重试'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/<int:id>/confirm', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def confirm_sales_order(id):
        from app import SalesOrder, log_operation, validate_sales_warehouse
        order = SalesOrder.query.get_or_404(id)
        if order.status != 'draft' or not order.items:
            return jsonify({'status': 'error', 'msg': '只有有明细的草稿订单可以确认'}), 400
        warehouse, warehouse_error = validate_sales_warehouse(order.warehouse, order.warehouse_id)
        if warehouse_error:
            return jsonify({'status': 'error', 'msg': warehouse_error}), 400
        order.warehouse = warehouse.name
        order.warehouse_id = warehouse.id
        order.status = 'confirmed'
        db.session.commit()
        log_operation('确认销售订单', f'销售订单：{order.order_no}', 'sales_order', order.id)
        return jsonify({'status': 'success', 'msg': '销售订单已确认'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/<int:id>/create_outbound', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def create_sales_outbound_draft(id):
        from sqlalchemy.orm import joinedload
        from app import (_acquire_order_write_lock, SalesOrder, SalesOrderItem,
                         STOCK_COMPARE_EPSILON, build_sales_outbound_draft,
                         log_operation, parse_float_value, round_to_2_decimals,
                         validate_sales_warehouse)
        # SALES-AUDIT-003：加写锁并重读状态，防止并发重复下推导致超扣。
        # 对照 create_sales_outbound_from_selection 的 BEGIN IMMEDIATE / FOR UPDATE。
        order, lock_ok = _acquire_order_write_lock(
            SalesOrder, id, ('confirmed', 'closed'),
            joinedload(SalesOrder.items).joinedload(SalesOrderItem.material),
        )
        if not lock_ok:
            return jsonify({'status': 'error', 'msg': '请先确认销售订单或订单已被并发修改'}), 400
        warehouse, warehouse_error = validate_sales_warehouse(order.warehouse, order.warehouse_id)
        if warehouse_error:
            db.session.rollback()
            return jsonify({'status': 'error', 'msg': warehouse_error}), 400
        order.warehouse = warehouse.name
        order.warehouse_id = warehouse.id
        try:
            payload = request.get_json(silent=True) or {}
            selected_qty_by_item_id = None
            if isinstance(payload.get('items'), list):
                selected_qty_by_item_id = {}
                for row in payload['items']:
                    if not isinstance(row, dict):
                        continue
                    try:
                        item_id = int(row.get('sales_order_item_id') or row.get('item_id'))
                    except (TypeError, ValueError):
                        continue
                    quantity = round_to_2_decimals(parse_float_value(row.get('quantity'), 0))
                    if quantity > STOCK_COMPARE_EPSILON:
                        selected_qty_by_item_id[item_id] = round_to_2_decimals(selected_qty_by_item_id.get(item_id, 0) + quantity)
            outbound, result = build_sales_outbound_draft(order, selected_qty_by_item_id)
            if result == 'invalid_selection':
                db.session.rollback()
                return jsonify({'status': 'error', 'msg': '请至少选择一条大于 0 的未发货明细'}), 400
            if result == 'over_quantity':
                db.session.rollback()
                return jsonify({'status': 'error', 'msg': '出库数量不能超过销售订单未发货数量'}), 400
            if result == 'completed':
                db.session.commit()
                return jsonify({'status': 'error', 'msg': '销售订单已全部发货'}), 400
            db.session.commit()
            log_operation('生成销售出库草稿', f'{order.order_no} -> {outbound.order_no}', 'out_order', outbound.id)
            return jsonify({'status': 'success', 'id': outbound.id, 'order_no': outbound.order_no, 'url': url_for('out_order_detail', id=outbound.id), 'existing': result == 'existing'})
        except Exception:
            db.session.rollback()
            app.logger.exception('生成销售出库草稿失败')
            return jsonify({'status': 'error', 'msg': '生成销售出库草稿失败，请稍后重试'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/batch_confirm', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def batch_confirm_sales_orders():
        from app import SalesOrder, log_operation, validate_sales_warehouse
        payload = request.get_json(silent=True) or {}
        raw_ids = payload.get('ids') or request.form.getlist('ids')
        ids = [int(value) for value in raw_ids if str(value).isdigit()]
        if not ids:
            return jsonify({'status': 'error', 'msg': '请先选择销售订单'}), 400
        orders = SalesOrder.query.filter(SalesOrder.id.in_(ids)).all()
        confirmed = 0
        skipped = []
        for order in orders:
            if order.status != 'draft' or not order.items:
                skipped.append(order.order_no)
                continue
            warehouse, warehouse_error = validate_sales_warehouse(order.warehouse, order.warehouse_id)
            if warehouse_error:
                skipped.append(f'{order.order_no}(仓库无效)')
                continue
            order.warehouse = warehouse.name
            order.warehouse_id = warehouse.id
            order.status = 'confirmed'
            order.shipment_status = 'pending'
            confirmed += 1
        db.session.commit()
        for order in orders:
            if order.status == 'confirmed' and order.order_no not in skipped:
                log_operation('批量确认销售订单', f'销售订单：{order.order_no}', 'sales_order', order.id)
        return jsonify({'status': 'success', 'msg': f'已确认 {confirmed} 张销售订单，跳过 {len(skipped)} 张', 'confirmed': confirmed, 'skipped': skipped})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/batch_create_outbound', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def batch_create_sales_outbound():
        from app import (_acquire_order_write_lock, SalesOrder, build_sales_outbound_draft,
                         log_operation, validate_sales_warehouse)
        payload = request.get_json(silent=True) or {}
        raw_ids = payload.get('ids') or request.form.getlist('ids')
        ids = [int(value) for value in raw_ids if str(value).isdigit()]
        if not ids:
            return jsonify({'status': 'error', 'msg': '请先选择销售订单'}), 400
        # SALES-AUDIT-003：逐张加写锁重读状态后下推，每张订单独立提交，
        # 防止并发重复生成草稿导致超扣。（_acquire_order_write_lock 内部会
        # 先 rollback 再 BEGIN IMMEDIATE，因此每张订单必须在同一锁事务内提交。）
        pre_orders = SalesOrder.query.filter(SalesOrder.id.in_(ids)).all()
        order_no_by_id = {o.id: o.order_no for o in pre_orders}
        created = []
        skipped = []
        for oid in ids:
            try:
                order, lock_ok = _acquire_order_write_lock(
                    SalesOrder, oid, ('confirmed', 'closed'),
                )
                if not lock_ok:
                    skipped.append(f'{order_no_by_id.get(oid, oid)}(状态不允许或已被并发修改)')
                    continue
                if order.shipment_status == 'shipped':
                    skipped.append(f'{order.order_no}(已全部发货)')
                    db.session.rollback()
                    continue
                warehouse, warehouse_error = validate_sales_warehouse(order.warehouse, order.warehouse_id)
                if warehouse_error:
                    skipped.append(f'{order.order_no}(仓库无效)')
                    db.session.rollback()
                    continue
                order.warehouse = warehouse.name
                order.warehouse_id = warehouse.id
                outbound, result = build_sales_outbound_draft(order)
                if result == 'completed':
                    skipped.append(f'{order.order_no}(已全部发货)')
                    db.session.rollback()
                    continue
                db.session.commit()
                created.append({'sales_order_no': order.order_no, 'outbound_id': outbound.id, 'outbound_no': outbound.order_no, 'existing': result == 'existing'})
            except Exception:
                db.session.rollback()
                app.logger.exception('批量生成销售出库草稿失败 order_id=%s', oid)
                skipped.append(f'{order_no_by_id.get(oid, oid)}(生成失败)')
        for item in created:
            log_operation('批量生成销售出库草稿', f"{item['sales_order_no']} -> {item['outbound_no']}", 'out_order', item['outbound_id'])
        return jsonify({'status': 'success', 'msg': f'生成 {len(created)} 张销售出库草稿，跳过 {len(skipped)} 张', 'created': created, 'skipped': skipped})

    @app.route('/sales/<int:id>/print')
    @login_required
    def print_sales_order(id):
        from sqlalchemy.orm import joinedload, selectinload
        from app import (Material, SalesOrder, SalesOrderItem, _fmt_date, _material_row_common,
                         _operator_name, _render_generic_document_print,
                         sales_shipment_status_label)
        order = SalesOrder.query.options(
            joinedload(SalesOrder.customer),
            joinedload(SalesOrder.salesperson),
            joinedload(SalesOrder.operator),
            selectinload(SalesOrder.items).joinedload(SalesOrderItem.material).joinedload(Material.unit),
        ).get_or_404(id)
        rows = [_material_row_common(
            item,
            extra={
                'tax_rate': '{:.0f}%'.format((item.tax_rate or 0) * 100),
                'untaxed_amount': item.untaxed_amount or 0,
                'tax_amount': item.tax_amount or 0,
                'tax_included_amount': item.tax_included_amount or item.amount or 0,
                'batch_no': item.batch_no or '',
                'serial_no': item.serial_no or '',
            },
        ) for item in order.items]
        return _render_generic_document_print({
            'title': '销售订单',
            'subtitle': 'SALES ORDER',
            'number_label': '销售订单号',
            'number': order.order_no,
            'date_label': '订单日期',
            'date': _fmt_date(order.date),
            'status': order.status,
            'info': [
                ('客户名称', order.customer.name if order.customer else ''),
                ('交货日期', _fmt_date(order.delivery_date)),
                ('发货仓库', order.warehouse or ''),
                ('业务员', order.salesperson.name if order.salesperson else ''),
                ('项目号', order.project_no or ''),
                # SALES-AUDIT-009：补齐合同编号与工程名称
                ('合同编号', order.contract_no or ''),
                ('工程名称', order.project_name or ''),
                ('币别', order.currency or 'CNY'),
                ('结算方式', order.settlement_method or ''),
                ('发货状态', sales_shipment_status_label(order.shipment_status)),
                ('制单人', _operator_name(order)),
                ('未税金额', f'{order.untaxed_amount or 0:.2f}'),
                ('税额', f'{order.tax_amount or 0:.2f}'),
                ('含税金额', f'{order.total_amount or 0:.2f}'),
            ],
            'remark': order.remark or '',
            'columns': [
                ('code', '物料编码', ''),
                ('name', '物料名称', ''),
                ('spec', '规格', ''),
                ('unit', '单位', 'center'),
                ('quantity', '订单数量', 'right'),
                ('price', '含税单价', 'right money'),
                ('tax_rate', '税率', 'center'),
                ('untaxed_amount', '未税金额', 'right money'),
                ('tax_amount', '税额', 'right money'),
                ('tax_included_amount', '含税金额', 'right money'),
                ('batch_no', '批次号', ''),
                ('serial_no', '序列号', ''),
                ('remark', '备注', ''),
            ],
            'rows': rows,
            'total_amount': order.total_amount or sum(row.get('tax_included_amount', 0) or 0 for row in rows),
            'signatures': ['制单', '销售确认', '仓库', '客户确认'],
        })

    @app.route('/sales/outbound')
    @login_required
    def sales_outbound_list():
        from sqlalchemy.orm import joinedload, selectinload
        from app import OutOrder, OutOrderItem
        page = max(1, request.args.get('page', 1, type=int))
        requested_per_page = request.args.get('per_page', 30, type=int)
        per_page = requested_per_page if requested_per_page in (20, 50, 100, 200) else 30
        status = (request.args.get('status') or '').strip()
        search = (request.args.get('search') or '').strip()
        sort_by = request.args.get('sort', 'date')
        sort_order = request.args.get('order', 'desc')
        query = OutOrder.query.filter(OutOrder.business_type == '销售出库')
        if status in ('pending', 'completed'):
            query = query.filter(OutOrder.status == status)
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(OutOrder.order_no.like(like), OutOrder.customer.like(like), OutOrder.purpose.like(like)))
        if sort_by == 'order_no':
            query = query.order_by(OutOrder.order_no.desc() if sort_order == 'desc' else OutOrder.order_no.asc())
        elif sort_by == 'status':
            query = query.order_by(OutOrder.status.desc() if sort_order == 'desc' else OutOrder.status.asc())
        else:
            query = query.order_by(OutOrder.date.desc() if sort_order == 'desc' else OutOrder.date.asc())
        query = query.order_by(OutOrder.id.desc())
        pagination = query.options(
            joinedload(OutOrder.source_sales_order),
            selectinload(OutOrder.items).joinedload(OutOrderItem.material),
        ).paginate(page=page, per_page=per_page, error_out=False)
        return render_template('sales_outbound_list.html', pagination=pagination, status=status, search=search, sort_by=sort_by, sort_order=sort_order)

    @app.route('/sales/outbound/export')
    @login_required
    def export_sales_outbound():
        from datetime import datetime
        import io
        from openpyxl import Workbook
        from sqlalchemy.orm import joinedload, selectinload
        from app import Material, OutOrder, OutOrderItem, _require_report_warehouse
        # SALES-AUDIT-007：仓库必填门禁，未提供且无默认仓库时返回 400
        selected_warehouse, _wh_err = _require_report_warehouse()
        if not selected_warehouse:
            return jsonify({'status': 'error', 'msg': _wh_err or '请选择仓库'}), 400
        warehouse = selected_warehouse.name
        wb = Workbook()
        ws = wb.active
        ws.title = '销售出库'
        # SALES-AUDIT-009：补齐合同单号与工程名称（来自关联销售订单）
        ws.append(['出库单号', '日期', '客户', '合同单号', '工程名称', '仓库', '来源销售订单', '物料编码', '物料名称', '规格', '单位', '数量', '单价', '金额', '状态'])
        status = (request.args.get('status') or '').strip()
        search = (request.args.get('search') or '').strip()
        query = OutOrder.query.filter(OutOrder.business_type == '销售出库', OutOrder.warehouse == warehouse)
        if status in ('pending', 'completed'):
            query = query.filter(OutOrder.status == status)
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(OutOrder.order_no.like(like), OutOrder.customer.like(like), OutOrder.purpose.like(like)))
        orders = query.options(
            joinedload(OutOrder.source_sales_order),
            selectinload(OutOrder.items).joinedload(OutOrderItem.material).joinedload(Material.unit),
        ).order_by(OutOrder.date.desc(), OutOrder.id.desc()).all()
        for order in orders:
            if order.items:
                for item in order.items:
                    ws.append([
                        order.order_no,
                        order.date.strftime('%Y-%m-%d') if order.date else '',
                        order.customer or '',
                        (order.source_sales_order.contract_no or '') if order.source_sales_order else '',
                        (order.source_sales_order.project_name or '') if order.source_sales_order else '',
                        order.warehouse or '',
                        order.source_sales_order.order_no if order.source_sales_order else '',
                        item.material.code if item.material else '',
                        item.material.name if item.material else '',
                        item.material.spec if item.material else '',
                        item.material.unit.name if item.material and item.material.unit else '',
                        item.quantity or 0,
                        item.price or 0,
                        item.amount or 0,
                        '已完成' if order.status == 'completed' else '草稿'
                    ])
            else:
                ws.append([order.order_no, order.date.strftime('%Y-%m-%d') if order.date else '', order.customer or '', (order.source_sales_order.contract_no or '') if order.source_sales_order else '', (order.source_sales_order.project_name or '') if order.source_sales_order else '', order.warehouse or '', order.source_sales_order.order_no if order.source_sales_order else '', '', '', '', '', '', '', '', '已完成' if order.status == 'completed' else '草稿'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='销售出库_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx')

    @app.route('/sales/dashboard')
    @login_required
    def sales_dashboard():
        from datetime import date
        from app import (OutOrder, SalesOrder, STOCK_COMPARE_EPSILON, round_to_2_decimals,
                         sales_shipment_status_label, sales_status_label)
        today = date.today()
        month_start = today.replace(day=1)
        active_orders = SalesOrder.query.filter(SalesOrder.status != 'cancelled').all()
        month_orders = [order for order in active_orders if order.date and order.date >= month_start]
        pending_orders = [order for order in active_orders if order.shipment_status != 'shipped']
        draft_orders = [order for order in active_orders if order.status == 'draft']
        ready_orders = [order for order in active_orders if order.status == 'confirmed' and order.shipment_status in ('pending', 'partial')]
        overdue_orders = [order for order in pending_orders if order.delivery_date and order.delivery_date < today]
        pending_outbounds = OutOrder.query.filter_by(business_type='销售出库', status='pending').order_by(OutOrder.date.asc(), OutOrder.id.asc()).limit(12).all()
        shortage_items = []
        for order in ready_orders:
            for item in order.items:
                remaining = round_to_2_decimals((item.quantity or 0) - (item.shipped_quantity or 0))
                stock = float(item.material.stock or 0) if item.material else 0
                if remaining > STOCK_COMPARE_EPSILON and stock + STOCK_COMPARE_EPSILON < remaining:
                    shortage_items.append({'order': order, 'item': item, 'remaining': remaining, 'stock': round_to_2_decimals(stock)})
        shortage_items.sort(key=lambda row: (row['stock'] - row['remaining'], row['order'].delivery_date or date.max))
        customer_summary = {}
        for order in month_orders:
            key = order.customer.name if order.customer else '未设置客户'
            row = customer_summary.setdefault(key, {'name': key, 'orders': 0, 'amount': 0, 'pending_amount': 0})
            row['orders'] += 1
            row['amount'] += order.total_amount or 0
            if order.shipment_status != 'shipped':
                row['pending_amount'] += order.total_amount or 0
        return render_template(
            'sales_dashboard.html',
            month_orders=month_orders,
            today=today,
            pending_orders=sorted(pending_orders, key=lambda order: (order.delivery_date or date.max, order.id))[:12],
            draft_orders=sorted(draft_orders, key=lambda order: order.id, reverse=True)[:12],
            ready_orders=sorted(ready_orders, key=lambda order: (order.delivery_date or date.max, order.id))[:12],
            overdue_orders=overdue_orders,
            pending_outbounds=pending_outbounds,
            shortage_items=shortage_items[:12],
            month_amount=round_to_2_decimals(sum(order.total_amount or 0 for order in month_orders)),
            pending_amount=round_to_2_decimals(sum(order.total_amount or 0 for order in pending_orders)),
            pending_count=len(pending_orders),
            overdue_count=len(overdue_orders),
            customer_summary=sorted(customer_summary.values(), key=lambda row: row['amount'], reverse=True)[:10],
            status_label=sales_status_label,
            shipment_status_label=sales_shipment_status_label,
        )

    @app.route('/sales/exceptions')
    @login_required
    def sales_exceptions():
        """Read-only sales exception workbench; every row drills into a source document."""
        from datetime import date
        from sqlalchemy.orm import joinedload, selectinload
        from app import (OutOrder, OutOrderItem, SalesOrder, SalesOrderItem, STOCK_COMPARE_EPSILON)
        kind = (request.args.get('kind') or '').strip()
        today = date.today()
        exceptions = []
        orders = SalesOrder.query.filter(SalesOrder.status != 'cancelled').options(
            joinedload(SalesOrder.customer),
            selectinload(SalesOrder.items).joinedload(SalesOrderItem.material),
        ).order_by(SalesOrder.delivery_date.asc(), SalesOrder.id.asc()).all()
        for order in orders:
            if order.delivery_date and order.delivery_date < today and order.shipment_status != 'shipped':
                exceptions.append({
                    'kind': 'overdue', 'label': '逾期未发货', 'severity': '高', 'order': order,
                    'item': None, 'outbound': None, 'detail': f'交货日期 {order.delivery_date}',
                })
            for item in order.items:
                quantity = float(item.quantity or 0)
                shipped = float(item.shipped_quantity or 0)
                remaining = max(quantity - shipped, 0)
                if shipped > quantity + STOCK_COMPARE_EPSILON:
                    exceptions.append({
                        'kind': 'over_shipped', 'label': '超发数量', 'severity': '高', 'order': order,
                        'item': item, 'outbound': None,
                        'detail': f'订单 {quantity:g}，已发货 {shipped:g}',
                    })
                if remaining > STOCK_COMPARE_EPSILON and item.material and float(item.material.stock or 0) + STOCK_COMPARE_EPSILON < remaining:
                    exceptions.append({
                        'kind': 'shortage', 'label': '库存不足', 'severity': '高', 'order': order,
                        'item': item, 'outbound': None,
                        'detail': f'待发 {remaining:g}，现存 {float(item.material.stock or 0):g}',
                    })
                if float(item.price or 0) <= 0:
                    exceptions.append({
                        'kind': 'price', 'label': '销售价异常', 'severity': '中', 'order': order,
                        'item': item, 'outbound': None, 'detail': '含税单价为 0 或负数',
                    })

        outbound_items = OutOrderItem.query.join(OutOrder).filter(
            OutOrder.business_type == '销售出库',
            OutOrder.status.in_(('pending', 'completed')),
        ).options(
            joinedload(OutOrderItem.out_order).joinedload(OutOrder.source_sales_order),
            joinedload(OutOrderItem.material),
            joinedload(OutOrderItem.source_sales_order_item).joinedload(SalesOrderItem.sales_order),
        ).all()
        for item in outbound_items:
            if not item.source_sales_order_item_id and not item.out_order.source_sales_order_id:
                exceptions.append({
                    'kind': 'missing_source', 'label': '缺少销售来源', 'severity': '中', 'order': None,
                    'item': item, 'outbound': item.out_order, 'detail': '出库明细未关联销售订单或销售订单行',
                })

        if kind:
            exceptions = [row for row in exceptions if row['kind'] == kind]
        severity_order = {'高': 0, '中': 1, '低': 2}
        exceptions.sort(key=lambda row: (severity_order.get(row['severity'], 9), row['order'].delivery_date if row['order'] else date.max, row['detail']))
        return render_template(
            'sales_exceptions.html', exceptions=exceptions, kind=kind,
            counts={value: sum(1 for row in exceptions if row['kind'] == value) for value in ('overdue', 'shortage', 'over_shipped', 'missing_source', 'price')},
        )

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/<int:id>/cancel', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def cancel_sales_order(id):
        from app import OutOrder, SalesOrder, STOCK_COMPARE_EPSILON, log_operation
        order = SalesOrder.query.get_or_404(id)
        if order.status not in ('draft', 'confirmed') or any((item.shipped_quantity or 0) > STOCK_COMPARE_EPSILON for item in order.items):
            return jsonify({'status': 'error', 'msg': '已发货或当前状态不允许取消销售订单'}), 400
        # SALES-AUDIT-002：存在 pending 销售出库草稿时禁止取消，否则会留下
        # 指向 cancelled 订单的孤儿草稿，完成后触发 SALES-AUDIT-001 复活链。
        pending_outbound = OutOrder.query.filter_by(
            source_sales_order_id=order.id, status='pending'
        ).first()
        if pending_outbound:
            return jsonify({
                'status': 'error',
                'msg': f'存在未完成的销售出库草稿 {pending_outbound.order_no}，请先处理后再取消'
            }), 400
        order.status = 'cancelled'
        order.shipment_status = 'pending'
        db.session.commit()
        log_operation('取消销售订单', f'销售订单：{order.order_no}', 'sales_order', order.id)
        return jsonify({'status': 'success', 'msg': '销售订单已取消'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/<int:id>/delete', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def delete_sales_order(id):
        from app import SalesOrder, _acquire_order_write_lock, api_error, log_operation
        from sqlalchemy.orm import selectinload
        order = SalesOrder.query.get_or_404(id)
        if order.status != 'draft':
            return jsonify({'status': 'error', 'msg': '只有草稿销售订单可以删除'}), 400
        try:
            # P1-BUGFIX: 重新锁定并校验 draft 状态，防止并发确认后仍被物理删除
            locked, ok = _acquire_order_write_lock(
                SalesOrder, id, 'draft',
                selectinload(SalesOrder.items),
            )
            if not ok:
                return jsonify({'status': 'error', 'msg': '该销售订单状态已变更，请刷新后重试'}), 409
            order = locked
            for item in list(order.items):
                db.session.delete(item)
            db.session.delete(order)
            db.session.commit()
            log_operation('删除销售订单', f'销售订单：{order.order_no}', 'sales_order', id)
            return jsonify({'status': 'success', 'msg': '销售订单已删除'})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'删除销售订单失败: {e}')
            return jsonify({'status': 'error', 'msg': '操作失败'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/<int:id>/copy', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def copy_sales_order(id):
        from datetime import date
        from flask_login import current_user
        from app import (SalesOrder, SalesOrderItem, generate_sales_order_no,
                         log_operation, recalculate_sales_order)
        order = SalesOrder.query.get_or_404(id)
        try:
            new_order = SalesOrder(
                order_no=generate_sales_order_no(),
                customer_id=order.customer_id,
                operator_id=current_user.id,
                date=date.today(),
                delivery_date=order.delivery_date,
                warehouse=order.warehouse,
                warehouse_id=order.warehouse_id,
                remark=order.remark,
                salesperson_id=order.salesperson_id,
                project_no=order.project_no,
                currency=order.currency,
                settlement_method=order.settlement_method,
                status='draft',
                shipment_status='pending',
            )
            db.session.add(new_order)
            db.session.flush()
            for item in order.items:
                new_item = SalesOrderItem(
                    sales_order_id=new_order.id,
                    material_id=item.material_id,
                    quantity=item.quantity,
                    price=item.price,
                    tax_rate=item.tax_rate,
                    batch_no=item.batch_no,
                    serial_no=item.serial_no,
                    remark=item.remark,
                )
                db.session.add(new_item)
            db.session.commit()
            recalculate_sales_order(new_order)
            db.session.commit()
            log_operation('复制销售订单', f'从 {order.order_no} 复制到 {new_order.order_no}', 'sales_order', new_order.id)
            return jsonify({'status': 'success', 'msg': f'已复制为新销售订单 {new_order.order_no}', 'id': new_order.id})
        except Exception as exc:
            db.session.rollback()
            app.logger.exception('复制销售订单失败')
            return jsonify({'status': 'error', 'msg': f'复制失败：{exc}'}), 500

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/sales/batch_delete', methods=['POST'])
    @require_role('warehouse', 'purchase', 'sales')
    @login_required
    def batch_delete_sales_orders():
        from app import SalesOrder, _acquire_order_write_lock, api_error, log_operation
        from sqlalchemy.orm import selectinload
        payload = request.get_json(silent=True) or {}
        ids = payload.get('ids', [])
        if not ids:
            return jsonify({'status': 'error', 'msg': '请选择要删除的销售订单'}), 400
        # 预筛：状态校验
        for order_id in ids:
            order = db.session.get(SalesOrder, order_id)
            if not order:
                continue
            if order.status != 'draft':
                return jsonify({'status': 'error', 'msg': f'{order.order_no} 不是草稿状态，无法删除'}), 400
        deleted = []
        errors = []
        try:
            # P1-BUGFIX: 逐张加写锁再校验状态，防止并发确认后仍被物理删除。
            # _acquire_order_write_lock 内部 rollback 会撤销前一张未提交的 delete，
            # 因此每张单必须在循环内独立 commit。
            for order_id in ids:
                locked, ok = _acquire_order_write_lock(
                    SalesOrder, order_id, 'draft',
                    selectinload(SalesOrder.items),
                )
                if not ok:
                    db.session.rollback()
                    return jsonify({'status': 'error', 'msg': f'销售订单 {order_id} 状态已变更，请刷新后重试'}), 409
                order = locked
                if order.status != 'draft':
                    db.session.rollback()
                    errors.append(f'{order.order_no} 不是草稿状态，无法删除')
                    continue
                order_no = order.order_no
                for item in list(order.items):
                    db.session.delete(item)
                db.session.delete(order)
                db.session.commit()
                deleted.append(order_no)
                log_operation('批量删除销售订单', f'删除草稿销售订单：{order_no}', 'sales_order', order_id)
            msg = f'成功删除 {len(deleted)} 张销售订单'
            if errors:
                msg += '；' + '；'.join(errors)
            return jsonify({'status': 'success', 'msg': msg})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'批量删除销售订单失败: {e}')
            return jsonify({'status': 'error', 'msg': '操作失败'}), 500

    @app.route('/sales/export')
    @login_required
    def export_sales_orders():
        from datetime import datetime
        from sqlalchemy.orm import joinedload
        from app import (Customer, Material, SalesOrder, SalesOrderItem,
                         _require_report_warehouse, sales_shipment_status_label, sales_status_label)
        # SALES-AUDIT-007：仓库必填门禁，未提供且无默认仓库时返回 400
        selected_warehouse, _wh_err = _require_report_warehouse()
        if not selected_warehouse:
            return jsonify({'status': 'error', 'msg': _wh_err or '请选择仓库'}), 400
        search = (request.args.get('search') or '').strip()
        status = (request.args.get('status') or '').strip()
        customer_id = request.args.get('customer_id', type=int)
        salesperson_id = request.args.get('salesperson_id', type=int)
        date_start = request.args.get('date_start') or ''
        date_end = request.args.get('date_end') or ''
        query = SalesOrder.query.join(Customer).outerjoin(SalesOrderItem, SalesOrderItem.sales_order_id == SalesOrder.id).outerjoin(Material, SalesOrderItem.material_id == Material.id).distinct()
        query = query.filter(db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)))
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(SalesOrder.order_no.like(like), Customer.name.like(like), Customer.code.like(like), Material.code.like(like), Material.name.like(like), SalesOrder.project_no.like(like)))
        contract_no_filter = (request.args.get('contract_no') or '').strip()
        if contract_no_filter:
            query = query.filter(SalesOrder.contract_no.like(f'%{contract_no_filter}%'))
        if status:
            query = query.filter(SalesOrder.status == status)
        if customer_id:
            query = query.filter(SalesOrder.customer_id == customer_id)
        if salesperson_id:
            query = query.filter(SalesOrder.salesperson_id == salesperson_id)
        if date_start:
            try:
                query = query.filter(SalesOrder.date >= datetime.strptime(date_start, '%Y-%m-%d').date())
            except ValueError:
                pass
        if date_end:
            try:
                query = query.filter(SalesOrder.date <= datetime.strptime(date_end, '%Y-%m-%d').date())
            except ValueError:
                pass
        orders = query.options(joinedload(SalesOrder.customer), joinedload(SalesOrder.salesperson), joinedload(SalesOrder.items).joinedload(SalesOrderItem.material)).order_by(SalesOrder.date.desc(), SalesOrder.id.desc()).all()
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '销售订单'
        # SALES-AUDIT-009：补齐合同单号与工程名称（与 PUR-AUDIT-004 对齐）
        sheet.append(['订单号', '订单日期', '客户', '合同单号', '工程名称', '业务员', '项目号', '仓库', '交货日期', '未税金额', '税额', '含税金额', '已发货金额', '待发货金额', '订单状态', '发货状态', '币别', '结算方式', '备注'])
        for order in orders:
            sheet.append([
                order.order_no,
                order.date.isoformat() if order.date else '',
                order.customer.name if order.customer else '',
                order.contract_no or '',
                order.project_name or '',
                order.salesperson.name if order.salesperson else '',
                order.project_no or '',
                order.warehouse or '',
                order.delivery_date.isoformat() if order.delivery_date else '',
                float(order.untaxed_amount or 0),
                float(order.tax_amount or 0),
                float(order.total_amount or 0),
                float(order.shipped_amount or 0),
                float(order.remaining_amount or 0),
                sales_status_label(order.status),
                sales_shipment_status_label(order.shipment_status),
                order.currency or 'CNY',
                order.settlement_method or '',
                order.remark or '',
            ])
        from flask import Response
        import io
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return Response(output.read(), headers={'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'Content-Disposition': 'attachment; filename=sales_orders.xlsx'}), 200

    @app.route('/sales/report/export')
    @login_required
    def export_sales_report():
        from datetime import date, datetime
        import io
        from app import (SalesOrder, _require_report_warehouse, round_to_2_decimals,
                         sales_shipment_status_label, sales_status_label)
        # SALES-AUDIT-007：仓库必填门禁，未提供且无默认仓库时返回 400
        selected_warehouse, _wh_err = _require_report_warehouse()
        if not selected_warehouse:
            return jsonify({'status': 'error', 'msg': _wh_err or '请选择仓库'}), 400
        date_start = request.args.get('date_start') or date.today().replace(day=1).isoformat()
        date_end = request.args.get('date_end') or date.today().isoformat()
        try:
            start = datetime.strptime(date_start, '%Y-%m-%d').date()
            end = datetime.strptime(date_end, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'status': 'error', 'msg': '日期格式不正确'}), 400
        orders = SalesOrder.query.filter(
            SalesOrder.date >= start, SalesOrder.date <= end, SalesOrder.status != 'cancelled',
            db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)),
        ).order_by(SalesOrder.date.asc(), SalesOrder.id.asc()).all()
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '销售订单执行'
        # SALES-AUDIT-009：补齐合同单号与工程名称
        sheet.append(['销售订单号', '订单日期', '客户', '合同单号', '工程名称', '仓库', '订单状态', '发货状态', '物料编码', '物料名称', '规格', '单位', '订单数量', '已发货数量', '未发货数量', '含税单价', '含税金额', '备注'])
        for order in orders:
            for item in order.items:
                shipped = item.shipped_quantity or 0
                sheet.append([
                    order.order_no,
                    order.date.isoformat() if order.date else '',
                    order.customer.name if order.customer else '',
                    order.contract_no or '',
                    order.project_name or '',
                    order.warehouse or '',
                    sales_status_label(order.status),
                    sales_shipment_status_label(order.shipment_status),
                    item.material.code if item.material else '',
                    item.material.name if item.material else '',
                    item.material.spec if item.material else '',
                    item.material.unit.name if item.material and item.material.unit else '',
                    item.quantity or 0,
                    shipped,
                    round_to_2_decimals((item.quantity or 0) - shipped),
                    item.price or 0,
                    item.amount or 0,
                    item.remark or order.remark or '',
                ])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name=f'sales_execution_{start:%Y%m%d}_{end:%Y%m%d}.xlsx', as_attachment=True)

    @app.route('/sales/reconciliation')
    @login_required
    def sales_reconciliation_report():
        from sqlalchemy import func
        from sqlalchemy.orm import selectinload
        from app import (SalesOrder, SalesOrderItem, StockTransaction, Warehouse,
                         _require_report_warehouse, STOCK_COMPARE_EPSILON, round_to_2_decimals)
        warehouse_id = request.args.get('warehouse_id', type=int)
        selected_warehouse, _wh_err = _require_report_warehouse()
        if selected_warehouse:
            warehouse_id = selected_warehouse.id
        # SALES-AUDIT-007：仓库必填，未提供且无默认仓库时返回空结果
        if not selected_warehouse:
            orders = []
        else:
            query = SalesOrder.query.filter(SalesOrder.status != 'cancelled')
            query = query.filter(db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)))
            orders = query.options(selectinload(SalesOrder.items), selectinload(SalesOrder.outbound_orders)).order_by(SalesOrder.date.desc(), SalesOrder.id.desc()).all()
        rows = []
        for order in orders:
            completed_orders = [outbound for outbound in order.outbound_orders if outbound.status == 'completed']
            completed_qty = 0
            completed_amount = 0
            inventory_qty = 0
            for outbound in completed_orders:
                for item in outbound.items:
                    if item.source_sales_order_item_id:
                        source = db.session.get(SalesOrderItem, item.source_sales_order_item_id)
                        if not source or source.sales_order_id != order.id:
                            continue
                    completed_qty += float(item.quantity or 0)
                    completed_amount += float(item.amount or 0)
                    inventory_qty += abs(float(db.session.query(func.coalesce(func.sum(StockTransaction.quantity), 0)).filter(
                        StockTransaction.reference_type == 'out_order', StockTransaction.reference_id == outbound.id,
                        StockTransaction.material_id == item.material_id, StockTransaction.transaction_type == 'out'
                    ).scalar() or 0))
            shipped_qty = sum(float(item.shipped_quantity or 0) for item in order.items)
            shipped_amount = float(order.shipped_amount or 0)
            rows.append({'order': order, 'shipped_qty': round_to_2_decimals(shipped_qty), 'completed_qty': round_to_2_decimals(completed_qty), 'inventory_qty': round_to_2_decimals(inventory_qty), 'quantity_diff': round_to_2_decimals(shipped_qty - completed_qty), 'amount_diff': round_to_2_decimals(shipped_amount - completed_amount), 'ok': abs(shipped_qty - completed_qty) <= STOCK_COMPARE_EPSILON and abs(shipped_amount - completed_amount) <= 0.01})
        total_orders = len(rows)
        passed_orders = sum(1 for r in rows if r['ok'])
        failed_orders = total_orders - passed_orders
        total_shipped = sum(r['shipped_qty'] for r in rows)
        total_completed = sum(r['completed_qty'] for r in rows)
        total_diff = sum(abs(r['quantity_diff']) for r in rows)
        chart_data = {
            'summary': {
                'total': total_orders,
                'passed': passed_orders,
                'failed': failed_orders,
                'passed_pct': round(passed_orders / total_orders * 100, 1) if total_orders > 0 else 0,
                'total_shipped': round(total_shipped, 2),
                'total_completed': round(total_completed, 2),
                'total_diff': round(total_diff, 2),
            },
            'warehouse_stats': {},
            'monthly_trend': []
        }
        return render_template('sales_reconciliation_report.html', rows=rows, warehouses=Warehouse.query.filter_by(status='active').order_by(Warehouse.code.asc()).all(), warehouse_id=warehouse_id or '', chart_data=chart_data)

    @app.route('/sales/reconciliation/export')
    @login_required
    def export_sales_reconciliation_report():
        import io
        from sqlalchemy.orm import selectinload
        from app import SalesOrder, _require_report_warehouse, STOCK_COMPARE_EPSILON
        from openpyxl import Workbook
        # SALES-AUDIT-007：仓库必填门禁，未提供且无默认仓库时返回 400
        selected_warehouse, _wh_err = _require_report_warehouse()
        if not selected_warehouse:
            return jsonify({'status': 'error', 'msg': _wh_err or '请选择仓库'}), 400
        query = SalesOrder.query.filter(SalesOrder.status != 'cancelled')
        query = query.filter(db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)))
        workbook = Workbook(); sheet = workbook.active; sheet.title = '销售对账'; sheet.append(['销售订单号', '仓库', '订单已发货数量', '出库完成数量', '订单已发货金额', '对账状态'])
        for order in query.options(selectinload(SalesOrder.items), selectinload(SalesOrder.outbound_orders)).all():
            completed_qty = sum(float(item.quantity or 0) for outbound in order.outbound_orders if outbound.status == 'completed' for item in outbound.items if not item.source_sales_order_item_id or (item.source_sales_order_item and item.source_sales_order_item.sales_order_id == order.id))
            completed_amount = sum(float(item.amount or 0) for outbound in order.outbound_orders if outbound.status == 'completed' for item in outbound.items if not item.source_sales_order_item_id or (item.source_sales_order_item and item.source_sales_order_item.sales_order_id == order.id))
            ok = abs(float(order.shipped_amount or 0) - completed_amount) <= 0.01 and abs(sum(float(item.shipped_quantity or 0) for item in order.items) - completed_qty) <= STOCK_COMPARE_EPSILON
            sheet.append([order.order_no, order.warehouse or '', sum(float(item.shipped_quantity or 0) for item in order.items), completed_qty, float(order.shipped_amount or 0), '通过' if ok else '不一致'])
        output = io.BytesIO(); workbook.save(output); output.seek(0)
        return send_file(output, download_name='sales_reconciliation.xlsx', as_attachment=True)

    @app.route('/sales/report')
    @login_required
    def sales_report():
        from datetime import date, datetime
        from app import (Customer, Employee, Material, SalesOrder, SalesOrderItem, Warehouse,
                         _require_report_warehouse, round_to_2_decimals, sales_status_label)
        date_start = request.args.get('date_start') or (date.today().replace(day=1).isoformat())
        date_end = request.args.get('date_end') or date.today().isoformat()
        drill_customer_id = request.args.get('customer_id', type=int)
        drill_material_code = (request.args.get('material_code') or '').strip()
        salesperson_id = request.args.get('salesperson_id', type=int)
        status = (request.args.get('status') or '').strip()
        shipment_status = (request.args.get('shipment_status') or '').strip()
        project_no = (request.args.get('project_no') or '').strip()
        warehouse_id = request.args.get('warehouse_id', type=int)
        warehouse = (request.args.get('warehouse') or '').strip()
        selected_warehouse, _wh_err = _require_report_warehouse()
        if selected_warehouse:
            warehouse_id = selected_warehouse.id
            warehouse = selected_warehouse.name
        start = datetime.strptime(date_start, '%Y-%m-%d').date()
        end = datetime.strptime(date_end, '%Y-%m-%d').date()
        if not selected_warehouse:
            orders = []
        else:
            query = SalesOrder.query.filter(SalesOrder.date >= start, SalesOrder.date <= end, SalesOrder.status != 'cancelled')
            if drill_customer_id:
                query = query.filter(SalesOrder.customer_id == drill_customer_id)
            if salesperson_id:
                query = query.filter(SalesOrder.salesperson_id == salesperson_id)
            if status:
                query = query.filter(SalesOrder.status == status)
            if shipment_status:
                query = query.filter(SalesOrder.shipment_status == shipment_status)
            if project_no:
                query = query.filter(SalesOrder.project_no.like(f'%{project_no}%'))
            query = query.filter(db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)))
            orders = query.order_by(SalesOrder.date.asc(), SalesOrder.id.asc()).all()
        # 明细钻取：按物料筛选时只显示匹配明细行
        drill_material_id = None
        if drill_material_code:
            material = Material.query.filter_by(code=drill_material_code).first()
            if material:
                drill_material_id = material.id
        by_customer = {}
        by_material = {}
        by_salesperson = {}
        total_untaxed = 0
        total_tax = 0
        for order in orders:
            customer_key = order.customer.name if order.customer else '未设置客户'
            customer_row = by_customer.setdefault(customer_key, {'name': customer_key, 'customer_id': order.customer_id, 'orders': 0, 'quantity': 0, 'amount': 0, 'untaxed_amount': 0, 'tax_amount': 0})
            customer_row['orders'] += 1
            customer_row['amount'] += order.total_amount or 0
            customer_row['untaxed_amount'] += order.untaxed_amount or 0
            customer_row['tax_amount'] += order.tax_amount or 0
            salesperson_key = order.salesperson.name if order.salesperson else '未设置业务员'
            salesperson_row = by_salesperson.setdefault(salesperson_key, {'name': salesperson_key, 'orders': 0, 'amount': 0, 'untaxed_amount': 0, 'tax_amount': 0})
            salesperson_row['orders'] += 1
            salesperson_row['amount'] += order.total_amount or 0
            salesperson_row['untaxed_amount'] += order.untaxed_amount or 0
            salesperson_row['tax_amount'] += order.tax_amount or 0
            for item in order.items:
                customer_row['quantity'] += item.quantity or 0
                material_key = item.material.code if item.material else '-'
                material_row = by_material.setdefault(material_key, {'code': material_key, 'name': item.material.name if item.material else '-', 'quantity': 0, 'amount': 0, 'untaxed_amount': 0, 'tax_amount': 0})
                material_row['quantity'] += item.quantity or 0
                material_row['amount'] += item.tax_included_amount or item.amount or 0
                material_row['untaxed_amount'] += item.untaxed_amount or 0
                material_row['tax_amount'] += item.tax_amount or 0
        shipped_quantity = round_to_2_decimals(sum(item.shipped_quantity or 0 for order in orders for item in order.items))
        shipped_amount = round_to_2_decimals(sum((item.shipped_quantity or 0) * (item.price or 0) for order in orders for item in order.items))
        total_amount = round_to_2_decimals(sum(order.total_amount or 0 for order in orders))
        total_untaxed = round_to_2_decimals(sum(order.untaxed_amount or 0 for order in orders))
        total_tax = round_to_2_decimals(sum(order.tax_amount or 0 for order in orders))
        # 钻取明细行
        drill_items = []
        if drill_material_id:
            for order in orders:
                for item in order.items:
                    if item.material_id == drill_material_id:
                        drill_items.append({
                            'order_no': order.order_no,
                            'order_id': order.id,
                            'date': order.date,
                            'customer': order.customer.name if order.customer else '-',
                            'salesperson': order.salesperson.name if order.salesperson else '-',
                            'quantity': item.quantity or 0,
                            'price': item.price or 0,
                            'tax_rate': item.tax_rate or 0,
                            'untaxed_amount': item.untaxed_amount or 0,
                            'tax_amount': item.tax_amount or 0,
                            'tax_included_amount': item.tax_included_amount or item.amount or 0,
                            'shipped_quantity': item.shipped_quantity or 0,
                        })
        return render_template('sales_report.html', date_start=date_start, date_end=date_end, drill_customer_id=drill_customer_id, drill_material_code=drill_material_code, salesperson_id=salesperson_id, status=status, shipment_status=shipment_status, project_no=project_no, warehouse=warehouse, warehouse_id=warehouse_id or '', drill_material_name=(Material.query.get(drill_material_id).name if drill_material_id else ''), drill_items=drill_items, customers=Customer.query.order_by(Customer.code.asc()).all(), employees=Employee.query.order_by(Employee.id.asc()).all(), warehouses=Warehouse.query.filter_by(status='active').order_by(Warehouse.code.asc()).all(), orders=orders, by_customer=sorted(by_customer.values(), key=lambda row: row['amount'], reverse=True), by_material=sorted(by_material.values(), key=lambda row: row['amount'], reverse=True), by_salesperson=sorted(by_salesperson.values(), key=lambda row: row['amount'], reverse=True), total_amount=total_amount, total_untaxed=total_untaxed, total_tax=total_tax, shipped_amount=shipped_amount, pending_amount=round_to_2_decimals(total_amount - shipped_amount), shipped_quantity=shipped_quantity, total_orders=len(orders), status_label=sales_status_label)

    @app.route('/sales/outflow_report')
    @login_required
    def sales_outflow_report():
        """销售出库明细表：基于 OutOrder(business_type='销售出库') 的实际出库记录。"""
        from datetime import date, datetime
        import re
        from app import (OutOrder, SalesOrder, Warehouse, _require_report_warehouse,
                         round_to_2_decimals, sales_status_label)
        date_start = request.args.get('date_start') or (date.today().replace(day=1).isoformat())
        date_end = request.args.get('date_end') or date.today().isoformat()
        search = (request.args.get('search') or '').strip()
        warehouse_id = request.args.get('warehouse_id', type=int)
        warehouse = (request.args.get('warehouse') or '').strip()
        selected_warehouse, _wh_err = _require_report_warehouse()
        if selected_warehouse:
            warehouse_id = selected_warehouse.id
            warehouse = selected_warehouse.name
        customer_name = (request.args.get('customer') or '').strip()
        try:
            start = datetime.strptime(date_start, '%Y-%m-%d').date()
            end = datetime.strptime(date_end, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'status': 'error', 'msg': '日期格式不正确'}), 400
        if not selected_warehouse:
            out_orders = []
        else:
            query = OutOrder.query.filter(
                OutOrder.business_type == '销售出库',
                OutOrder.date >= start,
                OutOrder.date <= end,
            )
            if search:
                like = f'%{search}%'
                query = query.filter(db.or_(OutOrder.order_no.like(like), OutOrder.customer.like(like), OutOrder.purpose.like(like), OutOrder.remark.like(like)))
            query = query.filter(OutOrder.warehouse == warehouse)
            if customer_name:
                query = query.filter(OutOrder.customer.like(f'%{customer_name}%'))
            out_orders = query.order_by(OutOrder.date.desc(), OutOrder.id.desc()).all()
        rows = []
        total_quantity = 0
        total_amount = 0
        total_untaxed = 0
        total_tax = 0
        for oo in out_orders:
            # 通过外键或 purpose 关联销售订单，获取税率信息
            sales_order = None
            if oo.source_sales_order_id:
                sales_order = SalesOrder.query.get(oo.source_sales_order_id)
            if not sales_order:
                source_match = re.search(r'来源销售订单\s+([^\s]+)', oo.purpose or '')
                if source_match:
                    sales_order = SalesOrder.query.filter_by(order_no=source_match.group(1)).first()
            # 建立 material_id -> sales_item 映射，用于获取税率
            sales_item_map = {}
            if sales_order:
                sales_item_map = {item.material_id: item for item in sales_order.items}
            for item in (oo.items or []):
                material = item.material
                line_amount = item.amount or round_to_2_decimals((item.quantity or 0) * (item.price or 0))
                # 从销售订单明细获取税率，计算未税/税额
                sales_item = sales_item_map.get(item.material_id) if sales_item_map else None
                tax_rate = sales_item.tax_rate if sales_item and sales_item.tax_rate else 0.13
                untaxed_amount = round_to_2_decimals(line_amount / (1 + tax_rate)) if (1 + tax_rate) > 0 else line_amount
                tax_amount = round_to_2_decimals(line_amount - untaxed_amount)
                rows.append({
                    'out_date': oo.date,
                    'out_order_no': oo.order_no,
                    'out_order_id': oo.id,
                    'out_status': oo.status,
                    'customer': oo.customer or '',
                    'warehouse': oo.warehouse or '',
                    'material_code': material.code if material else '',
                    'material_name': material.name if material else '',
                    'spec': material.spec if material else '',
                    'unit': material.unit.name if material and material.unit else '',
                    'quantity': item.quantity or 0,
                    'price': item.price or 0,
                    'tax_rate': tax_rate,
                    'untaxed_amount': untaxed_amount,
                    'tax_amount': tax_amount,
                    'amount': line_amount,
                    'purpose': oo.purpose or '',
                    'remark': item.remark or oo.remark or '',
                })
                if oo.status == 'completed':
                    total_quantity += item.quantity or 0
                    total_amount += line_amount
                    total_untaxed += untaxed_amount
                    total_tax += tax_amount
        return render_template('sales_outflow_report.html', date_start=date_start, date_end=date_end, search=search, warehouse=warehouse, warehouse_id=warehouse_id or '', customer=customer_name, warehouses=Warehouse.query.filter_by(status='active').order_by(Warehouse.code.asc()).all(), rows=rows, total_quantity=round_to_2_decimals(total_quantity), total_amount=round_to_2_decimals(total_amount), total_untaxed=round_to_2_decimals(total_untaxed), total_tax=round_to_2_decimals(total_tax), total_rows=len(rows), status_label=sales_status_label)

    @app.route('/sales/outflow_report/export')
    @login_required
    def export_sales_outflow_report():
        from datetime import date, datetime
        import io
        import re
        from openpyxl import Workbook
        from app import (OutOrder, SalesOrder, _require_report_warehouse, round_to_2_decimals)
        # SALES-AUDIT-007：仓库必填门禁，未提供且无默认仓库时返回 400
        selected_warehouse, _wh_err = _require_report_warehouse()
        if not selected_warehouse:
            return jsonify({'status': 'error', 'msg': _wh_err or '请选择仓库'}), 400
        warehouse = selected_warehouse.name
        date_start = request.args.get('date_start') or (date.today().replace(day=1).isoformat())
        date_end = request.args.get('date_end') or date.today().isoformat()
        search = (request.args.get('search') or '').strip()
        customer_name = (request.args.get('customer') or '').strip()
        try:
            start = datetime.strptime(date_start, '%Y-%m-%d').date()
            end = datetime.strptime(date_end, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'status': 'error', 'msg': '日期格式不正确'}), 400
        query = OutOrder.query.filter(OutOrder.business_type == '销售出库', OutOrder.date >= start, OutOrder.date <= end)
        if search:
            like = f'%{search}%'
            query = query.filter(db.or_(OutOrder.order_no.like(like), OutOrder.customer.like(like), OutOrder.purpose.like(like), OutOrder.remark.like(like)))
        query = query.filter(OutOrder.warehouse == warehouse)
        if customer_name:
            query = query.filter(OutOrder.customer.like(f'%{customer_name}%'))
        out_orders = query.order_by(OutOrder.date.asc(), OutOrder.id.asc()).all()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '销售出库明细'
        # SALES-AUDIT-009：补齐合同单号与工程名称（来自关联销售订单）
        sheet.append(['出库日期', '出库单号', '客户', '合同单号', '工程名称', '仓库', '物料编码', '物料名称', '规格', '单位', '数量', '含税单价', '税率', '未税金额', '税额', '含税金额', '状态', '来源', '备注'])
        for oo in out_orders:
            # 关联销售订单获取税率
            sales_order = None
            if oo.source_sales_order_id:
                sales_order = SalesOrder.query.get(oo.source_sales_order_id)
            if not sales_order:
                source_match = re.search(r'来源销售订单\s+([^\s]+)', oo.purpose or '')
                if source_match:
                    sales_order = SalesOrder.query.filter_by(order_no=source_match.group(1)).first()
            sales_item_map = {item.material_id: item for item in sales_order.items} if sales_order else {}
            for item in (oo.items or []):
                material = item.material
                line_amount = item.amount or round_to_2_decimals((item.quantity or 0) * (item.price or 0))
                sales_item = sales_item_map.get(item.material_id) if sales_item_map else None
                tax_rate = sales_item.tax_rate if sales_item and sales_item.tax_rate else 0.13
                untaxed_amount = round_to_2_decimals(line_amount / (1 + tax_rate)) if (1 + tax_rate) > 0 else line_amount
                tax_amount = round_to_2_decimals(line_amount - untaxed_amount)
                sheet.append([
                    oo.date.isoformat() if oo.date else '',
                    oo.order_no,
                    oo.customer or '',
                    (sales_order.contract_no or '') if sales_order else '',
                    (sales_order.project_name or '') if sales_order else '',
                    oo.warehouse or '',
                    material.code if material else '',
                    material.name if material else '',
                    material.spec if material else '',
                    material.unit.name if material and material.unit else '',
                    item.quantity or 0,
                    item.price or 0,
                    tax_rate,
                    untaxed_amount,
                    tax_amount,
                    line_amount,
                    '已完成' if oo.status == 'completed' else '待完成',
                    oo.purpose or '',
                    item.remark or oo.remark or '',
                ])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name=f'sales_outflow_{start:%Y%m%d}_{end:%Y%m%d}.xlsx', as_attachment=True)

    @app.route('/sales/trend_report')
    @login_required
    def sales_trend_report():
        """销售趋势分析表：按月聚合销售订单金额、数量和发货进度。"""
        from datetime import date, timedelta
        from app import (SalesOrder, Warehouse, _require_report_warehouse, round_to_2_decimals)
        months_back = request.args.get('months', 12, type=int)
        months_back = max(1, min(months_back, 60))
        warehouse_id = request.args.get('warehouse_id', type=int)
        selected_warehouse, _wh_err = _require_report_warehouse()
        today = date.today()
        start_month = (today.replace(day=1) - timedelta(days=months_back * 31)).replace(day=1)
        # SALES-AUDIT-007：仓库必填，未提供且无默认仓库时返回空结果
        if not selected_warehouse:
            orders = []
        else:
            orders = SalesOrder.query.filter(
                SalesOrder.date >= start_month, SalesOrder.status != 'cancelled',
                db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)),
            ).order_by(SalesOrder.date.asc()).all()
        by_month = {}
        for order in orders:
            if not order.date:
                continue
            key = order.date.strftime('%Y-%m')
            row = by_month.setdefault(key, {'month': key, 'orders': 0, 'untaxed_amount': 0, 'tax_amount': 0, 'amount': 0, 'quantity': 0, 'shipped_amount': 0, 'customers': set()})
            row['orders'] += 1
            row['untaxed_amount'] += order.untaxed_amount or 0
            row['tax_amount'] += order.tax_amount or 0
            row['amount'] += order.total_amount or 0
            row['shipped_amount'] += order.shipped_amount or 0
            if order.customer_id:
                row['customers'].add(order.customer_id)
            for item in order.items:
                row['quantity'] += item.quantity or 0
        months = sorted(by_month.keys())
        rows = []
        prev_amount = None
        for m in months:
            row = by_month[m]
            row['customers'] = len(row['customers'])
            row['untaxed_amount'] = round_to_2_decimals(row['untaxed_amount'])
            row['tax_amount'] = round_to_2_decimals(row['tax_amount'])
            row['amount'] = round_to_2_decimals(row['amount'])
            row['shipped_amount'] = round_to_2_decimals(row['shipped_amount'])
            if prev_amount and prev_amount > 0:
                row['growth'] = round_to_2_decimals((row['amount'] - prev_amount) / prev_amount * 100)
            else:
                row['growth'] = None
            rows.append(row)
            prev_amount = row['amount']
        return render_template('sales_trend_report.html', months_back=months_back, warehouse_id=selected_warehouse.id if selected_warehouse else '', warehouse=selected_warehouse.name if selected_warehouse else '', warehouses=Warehouse.query.filter_by(status='active').order_by(Warehouse.code.asc()).all(), rows=rows, total_orders=sum(r['orders'] for r in rows), total_amount=round_to_2_decimals(sum(r['amount'] for r in rows)), total_untaxed=round_to_2_decimals(sum(r['untaxed_amount'] for r in rows)), total_tax=round_to_2_decimals(sum(r['tax_amount'] for r in rows)), total_shipped=round_to_2_decimals(sum(r['shipped_amount'] for r in rows)))

    @app.route('/sales/trend_report/export')
    @login_required
    def export_sales_trend_report():
        from datetime import date, timedelta
        import io
        from openpyxl import Workbook
        from app import (SalesOrder, _require_report_warehouse, round_to_2_decimals)
        # SALES-AUDIT-007：仓库必填门禁，未提供且无默认仓库时返回 400
        selected_warehouse, _wh_err = _require_report_warehouse()
        if not selected_warehouse:
            return jsonify({'status': 'error', 'msg': _wh_err or '请选择仓库'}), 400
        months_back = request.args.get('months', 12, type=int)
        months_back = max(1, min(months_back, 60))
        today = date.today()
        start_month = (today.replace(day=1) - timedelta(days=months_back * 31)).replace(day=1)
        orders = SalesOrder.query.filter(
            SalesOrder.date >= start_month, SalesOrder.status != 'cancelled',
            db.or_(SalesOrder.warehouse_id == selected_warehouse.id, db.and_(SalesOrder.warehouse_id.is_(None), SalesOrder.warehouse == selected_warehouse.name)),
        ).all()
        by_month = {}
        for order in orders:
            if not order.date:
                continue
            key = order.date.strftime('%Y-%m')
            row = by_month.setdefault(key, {'month': key, 'orders': 0, 'untaxed_amount': 0, 'tax_amount': 0, 'amount': 0, 'quantity': 0, 'shipped_amount': 0, 'customers': set()})
            row['orders'] += 1
            row['untaxed_amount'] += order.untaxed_amount or 0
            row['tax_amount'] += order.tax_amount or 0
            row['amount'] += order.total_amount or 0
            row['shipped_amount'] += order.shipped_amount or 0
            if order.customer_id:
                row['customers'].add(order.customer_id)
            for item in order.items:
                row['quantity'] += item.quantity or 0
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '销售趋势分析'
        sheet.append(['月份', '订单数', '客户数', '数量', '未税金额', '税额', '含税金额', '已发货金额', '环比增长%'])
        prev_amount = None
        for m in sorted(by_month.keys()):
            row = by_month[m]
            growth = ''
            if prev_amount and prev_amount > 0:
                growth = round_to_2_decimals((row['amount'] - prev_amount) / prev_amount * 100)
            sheet.append([m, row['orders'], len(row['customers']), round_to_2_decimals(row['quantity']), round_to_2_decimals(row['untaxed_amount']), round_to_2_decimals(row['tax_amount']), round_to_2_decimals(row['amount']), round_to_2_decimals(row['shipped_amount']), growth])
            prev_amount = row['amount']
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name=f'sales_trend_{months_back}months.xlsx', as_attachment=True)

    @app.route('/sales/execution_report')
    @login_required
    def sales_execution_report():
        from app import (Customer, SalesOrder, Warehouse, _sales_report_filters_context,
                         _sales_report_orders, round_to_2_decimals, sales_status_label)
        orders = _sales_report_orders()
        rows = []
        for order in orders:
            total_qty = sum((item.quantity or 0) for item in order.items)
            shipped_qty = sum((item.shipped_quantity or 0) for item in order.items)
            remaining_qty = max(total_qty - shipped_qty, 0)
            amount = float(order.total_amount or 0)
            shipped_amount = float(order.shipped_amount or 0)
            rows.append({
                'order': order,
                'quantity': round_to_2_decimals(total_qty),
                'shipped_quantity': round_to_2_decimals(shipped_qty),
                'remaining_quantity': round_to_2_decimals(remaining_qty),
                'amount': round_to_2_decimals(amount),
                'shipped_amount': round_to_2_decimals(shipped_amount),
                'remaining_amount': round_to_2_decimals(max(amount - shipped_amount, 0)),
                'execution_rate': round_to_2_decimals(shipped_amount / amount * 100) if amount else 0,
            })
        return render_template(
            'sales_execution_report.html',
            rows=rows,
            filters=_sales_report_filters_context(),
            customers=Customer.query.order_by(Customer.code.asc(), Customer.id.asc()).all(),
            warehouses=Warehouse.query.filter_by(status='active').order_by(Warehouse.code.asc()).all(),
            status_label=sales_status_label,
            total_orders=len(rows),
            total_amount=round_to_2_decimals(sum(row['amount'] for row in rows)),
            shipped_amount=round_to_2_decimals(sum(row['shipped_amount'] for row in rows)),
            remaining_amount=round_to_2_decimals(sum(row['remaining_amount'] for row in rows)),
        )

    @app.route('/sales/execution_report/export')
    @login_required
    def export_sales_execution_report():
        import io
        from openpyxl import Workbook
        from app import _sales_report_orders, round_to_2_decimals
        rows = []
        for order in _sales_report_orders():
            total_qty = sum((item.quantity or 0) for item in order.items)
            shipped_qty = sum((item.shipped_quantity or 0) for item in order.items)
            amount = float(order.total_amount or 0)
            shipped_amount = float(order.shipped_amount or 0)
            # SALES-AUDIT-009：补齐合同单号与工程名称
            rows.append([
                order.order_no, order.date.strftime('%Y-%m-%d') if order.date else '',
                order.customer.name if order.customer else '',
                order.contract_no or '', order.project_name or '',
                order.status,
                round_to_2_decimals(total_qty), round_to_2_decimals(shipped_qty),
                round_to_2_decimals(max(total_qty - shipped_qty, 0)),
                round_to_2_decimals(amount), round_to_2_decimals(shipped_amount),
                round_to_2_decimals(max(amount - shipped_amount, 0)),
                round_to_2_decimals(shipped_amount / amount * 100) if amount else 0,
            ])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '销售订单执行'
        sheet.append(['销售订单号', '订单日期', '客户', '合同单号', '工程名称', '状态', '订单数量', '已发货数量', '待发货数量', '订单金额', '已发货金额', '待发货金额', '金额执行率%'])
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name='sales_execution_report.xlsx', as_attachment=True)

    @app.route('/sales/price_analysis')
    @login_required
    def sales_price_analysis():
        from app import (Customer, SalesOrder, Warehouse, _sales_report_filters_context,
                         _sales_report_orders, round_to_2_decimals)
        aggregates = {}
        for order in _sales_report_orders():
            customer_id = order.customer_id
            for item in order.items:
                material = item.material
                if not material:
                    continue
                key = material.id
                row = aggregates.setdefault(key, {
                    'code': material.code, 'name': material.name, 'spec': material.spec or '',
                    'quantity': 0, 'amount': 0, 'prices': [], 'customers': set(), 'lines': 0,
                })
                quantity = float(item.quantity or 0)
                price = float(item.price or 0)
                row['quantity'] += quantity
                row['amount'] += float(item.tax_included_amount or item.amount or 0)
                row['prices'].append(price)
                row['customers'].add(customer_id)
                row['lines'] += 1
        rows = []
        for row in aggregates.values():
            prices = row.pop('prices')
            row['avg_price'] = round_to_2_decimals(sum(prices) / len(prices)) if prices else 0
            row['min_price'] = round_to_2_decimals(min(prices)) if prices else 0
            row['max_price'] = round_to_2_decimals(max(prices)) if prices else 0
            row['quantity'] = round_to_2_decimals(row['quantity'])
            row['amount'] = round_to_2_decimals(row['amount'])
            row['customers'] = len({value for value in row['customers'] if value})
            rows.append(row)
        rows.sort(key=lambda row: row['amount'], reverse=True)
        return render_template(
            'sales_price_analysis.html', rows=rows, filters=_sales_report_filters_context(),
            customers=Customer.query.order_by(Customer.code.asc(), Customer.id.asc()).all(),
            warehouses=Warehouse.query.filter_by(status='active').order_by(Warehouse.code.asc()).all(),
            total_quantity=round_to_2_decimals(sum(row['quantity'] for row in rows)),
            total_amount=round_to_2_decimals(sum(row['amount'] for row in rows)),
        )

    @app.route('/sales/price_analysis/export')
    @login_required
    def export_sales_price_analysis():
        import io
        from openpyxl import Workbook
        from app import _sales_report_orders, round_to_2_decimals
        rows = []
        for order in _sales_report_orders():
            for item in order.items:
                material = item.material
                if material:
                    rows.append((material.code, material.name, material.spec or '', float(item.quantity or 0), float(item.price or 0), float(item.tax_included_amount or item.amount or 0)))
        grouped = {}
        for code, name, spec, quantity, price, amount in rows:
            row = grouped.setdefault(code, {'code': code, 'name': name, 'spec': spec, 'quantity': 0, 'amount': 0, 'prices': []})
            row['quantity'] += quantity
            row['amount'] += amount
            row['prices'].append(price)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '销售价格分析'
        sheet.append(['物料编码', '物料名称', '规格', '销售数量', '平均成交价', '最低成交价', '最高成交价', '含税金额'])
        for row in sorted(grouped.values(), key=lambda value: value['amount'], reverse=True):
            prices = row['prices']
            sheet.append([row['code'], row['name'], row['spec'], round_to_2_decimals(row['quantity']), round_to_2_decimals(sum(prices) / len(prices)), round_to_2_decimals(min(prices)), round_to_2_decimals(max(prices)), round_to_2_decimals(row['amount'])])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name='sales_price_analysis.xlsx', as_attachment=True)