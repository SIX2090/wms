#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 单位/供应商的模板下载、导出、导入（unit_supplier_import）域路由。
#
# 批量拆分模式：为避免 endpoint 前缀化导致大量 url_for 引用改动，
# 采用「register_unit_supplier_import_routes(app)」直接在 app 上注册路由，
# endpoint 名保持不变，与 app.py 内原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（Unit、Supplier、api_error、_get_master_list_filters、
#   _apply_simple_search、_apply_master_order 等）在各路由函数内延迟导入
#   （请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import flash, jsonify, redirect, request, send_file, url_for
from flask_login import login_required

from db import db
from utils import require_role, validate_excel_extension, validate_excel_size


# no-test:reason=路由注册辅助函数，能力由 unit_supplier_import_* 各路由测试覆盖
def register_unit_supplier_import_routes(app):
    @app.route('/unit/download_template')
    @login_required
    def download_unit_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '计量单位导入模板'
        ws.append(['单位编号', '单位名称'])
        ws.append(['U-001', '个'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='unit_template.xlsx', as_attachment=True)

    @app.route('/unit/export')
    @login_required
    def export_unit():
        from openpyxl import Workbook
        from app import (
            Unit,
            _apply_master_order,
            _apply_simple_search,
            _get_master_list_filters,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '计量单位数据'
        ws.append(['单位编号', '单位名称'])
        search, status_filter, sort_by, sort_order = _get_master_list_filters('code')
        query = _apply_simple_search(Unit.query, Unit, search, ['code', 'name'])
        query, _ = _apply_master_order(query, Unit, sort_by, sort_order, {'id', 'code', 'name', 'created_at'}, 'code')
        for u in query.all():
            ws.append([u.code, u.name])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='units.xlsx', as_attachment=True)

    @app.route('/unit/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_unit():
        from app import Unit, api_error
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的单位文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not code or not name:
                    skip += 1
                    continue
                if Unit.query.filter_by(code=code).first() or Unit.query.filter_by(name=name).first():
                    skip += 1
                    continue
                unit = Unit(code=code, name=name)
                db.session.add(unit)
                count += 1
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
                return jsonify({'status': 'error', 'msg': '操作失败'}), 500
            msg = f'单位导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条（重复或格式错误）'
            return jsonify({'status': 'success', 'msg': msg})
        except Exception:
            db.session.rollback()
            return api_error('导入失败，请稍后重试')

    @app.route('/supplier/download_template')
    @login_required
    def download_supplier_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '供应商导入模板'
        ws.append(['供应商编号', '供应商名称', '联系人', '电话', '地址'])
        ws.append(['SUP-001', '示例供应商', '张三', '13800138000', '广州市天河区'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='supplier_template.xlsx', as_attachment=True)

    @app.route('/supplier/export')
    @login_required
    def export_supplier():
        from openpyxl import Workbook
        from app import (
            Supplier,
            _apply_master_advanced_filters,
            _apply_master_order,
            _get_master_list_filters,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = '供应商数据'
        ws.append(['供应商编号', '供应商名称', '联系人', '电话', '地址'])
        search, status_filter, sort_by, sort_order = _get_master_list_filters('code')
        # AI-WMS-FILTER-003：与列表页共用同一筛选入口，避免「页面筛了、导出没筛」
        _biz_parts = []
        for _rel in ('in_orders', 'purchase_orders', 'subcontract_orders'):
            _r = getattr(Supplier, _rel, None)
            if _r is not None and hasattr(_r, 'any'):
                _biz_parts.append(_r.any())
        query, _adv = _apply_master_advanced_filters(
            Supplier.query, Supplier, ['code', 'name', 'contact', 'phone', 'address'],
            business_expr=db.or_(*_biz_parts) if _biz_parts else None)
        query, _ = _apply_master_order(query, Supplier, sort_by, sort_order, {'id', 'code', 'name', 'contact', 'phone', 'created_at'}, 'code')
        for s in query.all():
            ws.append([s.code, s.name, s.contact or '', s.phone or '', s.address or ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='suppliers.xlsx', as_attachment=True)

    @app.route('/supplier/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_supplier():
        from app import Supplier, api_error
        file = request.files.get('file')
        if not file:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error('请选择要导入的供应商文件')
            flash('请选择要导入的供应商文件', 'danger')
            return redirect(url_for('supplier_list'))
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_ext_msg)
            flash(_ext_msg, 'danger')
            return redirect(url_for('supplier_list'))
        # m-03：限制 Excel 上传 ≤ 5MB
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error(_size_msg)
            flash(_size_msg, 'danger')
            return redirect(url_for('supplier_list'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not code or not name:
                    skip += 1
                    continue
                if Supplier.query.filter_by(code=code).first() or Supplier.query.filter_by(name=name).first():
                    skip += 1
                    continue
                sup = Supplier(
                    code=code,
                    name=name,
                    contact=str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    phone=str(row[3]).strip() if len(row) > 3 and row[3] else '',
                    address=str(row[4]).strip() if len(row) > 4 and row[4] else ''
                )
                db.session.add(sup)
                count += 1
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return api_error('导入失败')
                flash('导入失败，请稍后重试', 'danger')
                return redirect(url_for('supplier_list'))
            msg = f'供应商导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条（重复或格式错误）'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'success', 'msg': msg, 'count': count})
            flash(msg, 'success')
        except Exception:
            db.session.rollback()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return api_error('供应商导入失败')
            flash('供应商导入失败，请稍后重试', 'danger')
        return redirect(url_for('supplier_list'))