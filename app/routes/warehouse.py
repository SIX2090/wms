#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 仓库（warehouse）域路由。
#
# 批量拆分模式：与员工/供应商/物料分类域一致，采用「register_<domain>_routes(app)」
# 直接在 app 上注册路由，endpoint 名保持不变（如 warehouse_list），与 app.py 内
# 原有 url_for 引用完全兼容。
#
# - 模块级只导入稳定依赖（flask / db / utils），不导入 app，避免循环导入。
# - app.py 内部定义（Warehouse 模型、SalesOrder 等模型、辅助函数等）在各路由函数内
#   延迟导入（请求期才执行），避免 app.py 模块加载期触发循环导入。
# - 日志统一使用 current_app.logger 替代 app.logger。
# 注意：本文件顶部不用多行 """docstring""" 作为模块说明，会触发 lint 脚本
# strip_py_comments 把多行字符串折叠成一行、导致行号偏移、豁免注释检测失效。
from __future__ import annotations

import io

from flask import current_app, jsonify, render_template, request, send_file
from flask_login import login_required

from db import db
from utils import require_role


# no-test:reason=路由注册辅助函数，能力由 warehouse_* 各路由测试覆盖
def register_warehouse_routes(app):
    @app.route('/warehouse')
    @login_required
    def warehouse_list():
        from app import Warehouse, _warehouse_query_from_args
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        # per_page 必须有下限保护，传入 0 或负数会让 paginate 抛 ValueError 导致接口 500
        per_page = max(1, per_page)
        if per_page not in [10, 20, 50, 100, 200]:
            per_page = 20
        query, filters, sort_by, sort_order = _warehouse_query_from_args()
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        warehouses = pagination.items
        return render_template('warehouse.html', warehouses=warehouses, pagination=pagination, filters=filters, sort_by=sort_by, sort_order=sort_order, per_page=per_page)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/warehouse/add', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def add_warehouse():
        from app import Warehouse, api_error
        code = request.form.get('code', '').strip()
        if not code:
            return api_error('请输入仓库编码')

        name = request.form.get('name', '').strip()
        if not name:
            return api_error('请输入仓库名称')

        if Warehouse.query.filter_by(code=code).first():
            return api_error('仓库编码已存在')

        if Warehouse.query.filter_by(name=name).first():
            return api_error('仓库名称已存在')

        warehouse = Warehouse(
            code=code,
            name=name,
            type=request.form.get('type', '').strip() or None,
            location=request.form.get('location', '').strip() or None,
            status=request.form.get('status', 'active').strip(),
            remark=request.form.get('remark', '').strip() or None
        )
        # W4：后端校验状态值域，防止绕过前端写入任意状态
        if warehouse.status not in ('active', 'inactive'):
            return api_error('状态仅支持 active/inactive')
        try:
            db.session.add(warehouse)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'新增仓库失败: {e}')
            return jsonify({'status': 'error', 'msg': '新增失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '新增成功'})

    @app.route('/warehouse/<int:id>')
    @login_required
    def get_warehouse(id):
        from app import Warehouse
        warehouse = db.session.get(Warehouse, id)
        if not warehouse:
            return jsonify({'status': 'error', 'msg': '仓库不存在'}), 404

        return jsonify({
            'status': 'success',
            'warehouse': {
                'id': warehouse.id,
                'code': warehouse.code,
                'name': warehouse.name,
                'type': warehouse.type,
                'location': warehouse.location,
                'status': warehouse.status,
                'remark': warehouse.remark
            }
        })

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/warehouse/<int:id>/edit', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def edit_warehouse(id):
        from app import (AfterSaleOutOrder, InOrder, OutOrder, ProductionRequisition,
                         SalesOrder, SubcontractOrder, TransferOrder,
                         Warehouse, api_error)
        warehouse = db.session.get(Warehouse, id)
        if not warehouse:
            return jsonify({'status': 'error', 'msg': '仓库不存在'}), 404

        code = request.form.get('code', '').strip()
        if not code:
            return api_error('请输入仓库编码')

        name = request.form.get('name', '').strip()
        if not name:
            return api_error('请输入仓库名称')

        # 检查编码是否重复（排除自己）
        existing = Warehouse.query.filter(Warehouse.code == code, Warehouse.id != id).first()
        if existing:
            return api_error('仓库编码已存在')

        # 检查名称是否重复（排除自己）
        existing = Warehouse.query.filter(Warehouse.name == name, Warehouse.id != id).first()
        if existing:
            return api_error('仓库名称已存在')

        status = request.form.get('status', 'active').strip()
        # W4：后端校验状态值域，防止绕过前端写入任意状态（仅 active/inactive 合法）
        if status not in ('active', 'inactive'):
            return api_error('状态仅支持 active/inactive')

        old_name = warehouse.name or ''
        warehouse.code = code
        warehouse.name = name
        warehouse.type = request.form.get('type', '').strip() or None
        warehouse.location = request.form.get('location', '').strip() or None
        warehouse.status = status
        warehouse.remark = request.form.get('remark', '').strip() or None

        try:
            # W2：仓库改名时同步所有单据表冗余的仓库文本字段（仅命中旧名称的行，
            # 存编号/库位的行不受影响），避免改名后按新名称查不到历史单据。
            # 与库存报表/明细的名称+编号任一匹配口径互补。
            if old_name and old_name != name:
                # 采购订单（PurchaseOrder）本身不记录仓库，不参与同步
                for model, column in [
                    (InOrder, InOrder.warehouse),
                    (OutOrder, OutOrder.warehouse),
                    (SubcontractOrder, SubcontractOrder.warehouse),
                    (ProductionRequisition, ProductionRequisition.warehouse),
                    (AfterSaleOutOrder, AfterSaleOutOrder.warehouse),
                    (SalesOrder, SalesOrder.warehouse),
                ]:
                    model.query.filter(column == old_name).update(
                        {column: name}, synchronize_session=False
                    )
                # 调拨单的 from/to 仓库字段同样同步
                TransferOrder.query.filter(TransferOrder.from_warehouse == old_name).update(
                    {TransferOrder.from_warehouse: name}, synchronize_session=False
                )
                TransferOrder.query.filter(TransferOrder.to_warehouse == old_name).update(
                    {TransferOrder.to_warehouse: name}, synchronize_session=False
                )
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'编辑仓库失败: {e}')
            return jsonify({'status': 'error', 'msg': '编辑失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '编辑成功'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/warehouse/<int:id>/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def delete_warehouse(id):
        from app import Warehouse, _format_delete_blockers, _warehouse_delete_blockers, api_error
        warehouse = db.session.get(Warehouse, id)
        if not warehouse:
            return jsonify({'status': 'error', 'msg': '仓库不存在'}), 404

        blockers = _warehouse_delete_blockers(warehouse)
        if blockers:
            return api_error('该仓库已有业务数据，不能删除：' + _format_delete_blockers(blockers))

        try:
            db.session.delete(warehouse)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'删除仓库失败: {e}')
            return jsonify({'status': 'error', 'msg': '删除失败，请稍后重试'}), 500
        return jsonify({'status': 'success', 'msg': '删除成功'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/warehouse/delete', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def batch_delete_warehouse_master():
        from app import Warehouse, _format_delete_blockers, _warehouse_delete_blockers
        ids = (request.get_json(silent=True) or {}).get('ids', [])
        deleted = 0
        blocked = []
        for item_id in ids:
            warehouse = db.session.get(Warehouse, int(item_id)) if str(item_id).isdigit() else None
            if warehouse:
                blockers = _warehouse_delete_blockers(warehouse)
                if blockers:
                    blocked.append(f'{warehouse.name}（{_format_delete_blockers(blockers)}）')
                    continue
                db.session.delete(warehouse)
                deleted += 1
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'批量删除仓库失败: {e}')
            return jsonify({'status': 'error', 'msg': '删除失败，请稍后重试'}), 500
        if blocked:
            return jsonify({
                'status': 'error' if deleted == 0 else 'success',
                'msg': f'已删除 {deleted} 个仓库，以下仓库有关联数据未删除：' + '；'.join(blocked)
            })
        return jsonify({'status': 'success', 'msg': f'已删除 {deleted} 个仓库'})

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/warehouse/<int:warehouse_id>/set_default', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def warehouse_set_default(warehouse_id):
        """M-03：设为默认仓。同一事务内先把所有 is_default=True 置 False，再把当前行置 True。"""
        from app import Warehouse
        warehouse = db.session.get(Warehouse, warehouse_id)
        if not warehouse:
            return jsonify({'status': 'error', 'msg': '仓库不存在'}), 404
        if warehouse.status != 'active':
            return jsonify({'status': 'error', 'msg': '停用状态的仓库不能设为默认'}), 400
        try:
            Warehouse.query.filter(Warehouse.is_default.is_(True)).update({Warehouse.is_default: False}, synchronize_session=False)
            warehouse.is_default = True
            db.session.commit()
            return jsonify({'status': 'success', 'msg': f'已将“{warehouse.name}”设为默认仓'})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'设置默认仓失败: {e}')
            return jsonify({'status': 'error', 'msg': '设置失败，请稍后重试'}), 500

    @app.route('/warehouse/api/list')
    @login_required
    def warehouse_api_list():
        """仓库列表API - 只返回启用的仓库"""
        from app import Warehouse
        warehouses = Warehouse.query.filter_by(status='active').order_by(Warehouse.code).all()
        return jsonify({
            'warehouses': [{
                'id': w.id,
                'code': w.code,
                'name': w.name,
                'type': w.type
            } for w in warehouses]
        })

    @app.route('/warehouse/download_template')
    @login_required
    def download_warehouse_template():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '仓库导入模板'
        ws.append(['仓库编码', '仓库名称', '仓库类型', '仓库位置', '状态', '备注'])
        ws.append(['WH001', '材料仓', '原料仓', '一楼', 'active', ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='warehouse_template.xlsx', as_attachment=True)

    @app.route('/warehouse/export')
    @login_required
    def export_warehouse():
        from openpyxl import Workbook
        from app import Warehouse, _warehouse_query_from_args
        wb = Workbook()
        ws = wb.active
        ws.title = '仓库数据'
        ws.append(['仓库编码', '仓库名称', '仓库类型', '仓库位置', '状态', '备注'])
        query, _, _, _ = _warehouse_query_from_args()
        for warehouse in query.all():
            status_label = '启用' if warehouse.status == 'active' else ('停用' if warehouse.status == 'inactive' else (warehouse.status or ''))
            ws.append([warehouse.code, warehouse.name, warehouse.type or '', warehouse.location or '', status_label, warehouse.remark or ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='warehouses.xlsx', as_attachment=True)

    # pydantic:reason=存量路由从 app.py 原样迁移，保持行为不变，pydantic 迁移另行任务
    @app.route('/warehouse/import', methods=['POST'])
    @require_role('warehouse')
    @login_required
    def import_warehouse():
        from app import Warehouse, api_error, validate_excel_extension, validate_excel_size
        file = request.files.get('file')
        if not file:
            return api_error('请选择要导入的仓库文件')
        _ext_ok, _ext_msg = validate_excel_extension(file.filename)
        if not _ext_ok:
            return api_error(_ext_msg)
        # m-03：限制 Excel 上传 ≤ 5MB，避免大文件读入内存导致 OOM/超时
        _size_ok, _size_msg = validate_excel_size(file)
        if not _size_ok:
            return api_error(_size_msg)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row and row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not code or not name:
                    skip += 1
                    continue
                if Warehouse.query.filter_by(code=code).first() or Warehouse.query.filter_by(name=name).first():
                    skip += 1
                    continue
                warehouse = Warehouse(
                    code=code,
                    name=name,
                    type=str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    location=str(row[3]).strip() if len(row) > 3 and row[3] else '',
                    status=str(row[4]).strip() if len(row) > 4 and row[4] else 'active',
                    remark=str(row[5]).strip() if len(row) > 5 and row[5] else ''
                )
                db.session.add(warehouse)
                count += 1
            db.session.commit()
            msg = f'仓库导入成功，共导入 {count} 条'
            if skip:
                msg += f'，跳过 {skip} 条（重复或格式错误）'
            return jsonify({'status': 'success', 'msg': msg, 'count': count})
        except Exception:
            db.session.rollback()
            return api_error('仓库导入失败')