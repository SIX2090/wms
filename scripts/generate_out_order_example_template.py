# -*- coding: utf-8 -*-
"""生成「领料单打印模板示例.xlsx」（标准领料单版式，含占位符）。

版式与 print_out_with_excel.html / _build_out_order_excel 保持一致：
- 标题：领料单（A1:I1 合并）
- 表头：领料部门（A2:E2）/ 日期（F2:I2）
- 列：物料编码|品牌|物料名称|规格|单位|数量|单价|金额|合同编号
- 第 4 行为明细模板行（{item.*} 占位符），下方 7 行空白示例行
- 末尾签名行：「领料：」起点对齐「单价」列（G 列），G:I 合并留签字空间

用法：python scripts/generate_out_order_example_template.py
产出：app/static/templates/领料单打印模板示例.xlsx
"""
from __future__ import annotations

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(ROOT, 'app', 'static', 'templates', '领料单打印模板示例.xlsx')


def build(path=OUT_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = '领料单'

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    title_font = Font(name='微软雅黑', size=16, bold=True)
    body_font = Font(name='微软雅黑', size=11)
    header_font = Font(name='微软雅黑', size=11, bold=True)

    for idx, width in enumerate([12, 10, 20, 14, 6, 8, 10, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.merge_cells('A1:I1')
    ws['A1'] = '领料单'
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:E2')
    ws['A2'] = '领料部门：{order.customer}'
    ws.merge_cells('F2:I2')
    ws['F2'] = '日期：{order.date}'
    for cell in ('A2', 'F2'):
        ws[cell].font = body_font
        ws[cell].alignment = left_align

    headers = ['物料编码', '品牌', '物料名称', '规格', '单位', '数量', '单价', '金额', '合同编号']
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=text)
        c.font = header_font
        c.alignment = center
        c.border = border

    item_placeholders = [
        '{item.material.code}', '{item.material.brand}', '{item.material.name}',
        '{item.material.spec}', '{item.material.unit.name}', '{item.quantity}',
        '{item.price}', '{item.amount}', '{item.contract_no}',
    ]
    for col, text in enumerate(item_placeholders, start=1):
        c = ws.cell(row=4, column=col, value=text)
        c.font = body_font
        c.alignment = center
        c.border = border

    # 空白示例行 7 行（打印时按实际明细自动增删行）
    for r in range(5, 12):
        for col in range(1, 10):
            c = ws.cell(row=r, column=col)
            c.font = body_font
            c.border = border

    # 签名行：领料：起点对齐「单价」列（G 列），G:I 合并留签字空间
    ws.merge_cells('G12:I12')
    ws['G12'] = '领料：'
    ws['G12'].font = body_font
    ws['G12'].alignment = left_align
    ws.row_dimensions[12].height = 22

    wb.save(path)
    return path


if __name__ == '__main__':
    path = build()
    print('generated:', path)
    sys.exit(0)
