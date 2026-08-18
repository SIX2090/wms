# -*- coding: utf-8 -*-
"""BUG-2026-08-12-005 回归：旧版报表导出（/report/inout/print、/report/inout/export、
/report/stock/print）必须按仓库隔离，不得导出跨仓数据。

规则依据（AGENTS.md 仓库必填规则 + WMS_FULL_AUDIT_REPAIR_PROMPT FIX-2）：
- 旧导出 URL 必须接受与新版报表一致的 warehouse_id 参数；
- 未传仓库时带入 active 默认仓库；无默认仓库统一返回 400「请选择仓库」；
- 入库/出库导出必须按 InOrder.warehouse / OutOrder.warehouse 精确过滤；
- 库存导出必须输出仓库级库存（LocationInventory/流水净额），不得输出 Material.stock 跨仓总数；
- 库存查询页不得再生成无仓库参数的旧导出链接。
"""
from __future__ import annotations

import io
import os
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")

import app as app_module  # noqa: E402
from app import db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False

TODAY = date.today()


def _reset_db():
    db.drop_all()
    db.create_all()


def _make_client():
    client = app_module.app.test_client()
    client.post(
        "/login",
        data={"username": "admin", "password": "admin"},
        content_type="application/x-www-form-urlencoded",
    )
    return client


def _seed_admin():
    from werkzeug.security import generate_password_hash
    from app import User
    db.session.add(User(username="admin", password_hash=generate_password_hash("admin"),
                        role="admin", must_change_password=False))
    db.session.commit()


def _seed_warehouse(code, name, is_default=False):
    from app import Warehouse
    w = Warehouse(code=code, name=name, status="active", is_default=is_default)
    db.session.add(w)
    db.session.commit()
    return w


def _seed_material(code, name, stock=0, min_stock=0):
    from app import Material
    m = Material(code=code, name=name, stock=stock, min_stock=min_stock)
    db.session.add(m)
    db.session.commit()
    return m


def _seed_location_stock(material, warehouse_name, quantity):
    from app import LocationInventory
    db.session.add(LocationInventory(material_id=material.id, location=warehouse_name, quantity=quantity))
    db.session.commit()


def _enable_location_management():
    from app import set_system_setting
    set_system_setting("location_management_enabled", "1")
    db.session.commit()


def _seed_in_order(order_no, warehouse_name, material, quantity):
    from app import InOrder, InOrderItem
    order = InOrder(order_no=order_no, warehouse=warehouse_name, status="completed", date=TODAY)
    db.session.add(order)
    db.session.flush()
    db.session.add(InOrderItem(in_order_id=order.id, material_id=material.id,
                               quantity=quantity, price=2, amount=quantity * 2))
    db.session.commit()
    return order


def _seed_out_order(order_no, warehouse_name, material, quantity):
    from app import OutOrder, OutOrderItem
    order = OutOrder(order_no=order_no, warehouse=warehouse_name, status="completed", date=TODAY)
    db.session.add(order)
    db.session.flush()
    db.session.add(OutOrderItem(out_order_id=order.id, material_id=material.id,
                                quantity=quantity, price=2, amount=quantity * 2))
    db.session.commit()
    return order


def _read_xlsx(response):
    from openpyxl import load_workbook
    assert response.status_code == 200, response.status_code
    return load_workbook(io.BytesIO(response.data))


def _sheet_rows(workbook, sheet_name):
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


class _BaseExport:
    """公共夹具：仓库 A（默认）/ B；各自入库、出库、仓库级库存。"""

    def _setup(self, with_default=True):
        with app_module.app.app_context():
            _reset_db()
            _seed_admin()
            wh_a = _seed_warehouse("WHA", "仓库A", is_default=with_default)
            wh_b = _seed_warehouse("WHB", "仓库B")
            m1 = _seed_material("M001", "6204轴承", stock=100, min_stock=0)
            _enable_location_management()
            # 仓库级库存：A=5 / B=95（全局 stock=100 不得出现在导出中）
            _seed_location_stock(m1, "仓库A", 5)
            _seed_location_stock(m1, "仓库B", 95)
            # 单据：A/B 各一笔入库、一笔出库
            _seed_in_order("IN-A-01", "仓库A", m1, 10)
            _seed_in_order("IN-B-01", "仓库B", m1, 20)
            _seed_out_order("OUT-A-01", "仓库A", m1, 3)
            _seed_out_order("OUT-B-01", "仓库B", m1, 7)
            return wh_a.id, wh_b.id, m1.id


class TestInoutExportWarehouseScope(_BaseExport):
    """/report/inout/print 与 /report/inout/export 必须按仓库过滤。"""

    def test_inout_print_explicit_warehouse_a_excludes_b(self):
        wh_a_id, _wh_b_id, _m1_id = self._setup()
        client = _make_client()
        resp = client.get(f"/report/inout/print?warehouse_id={wh_a_id}")
        wb = _read_xlsx(resp)
        _, in_rows = _sheet_rows(wb, "入库统计")
        _, out_rows = _sheet_rows(wb, "领料统计")
        in_nos = [row[0] for row in in_rows]
        out_nos = [row[0] for row in out_rows]
        assert in_nos == ["IN-A-01"], in_rows
        assert out_nos == ["OUT-A-01"], out_rows

    def test_inout_print_explicit_warehouse_b_excludes_a(self):
        _wh_a_id, wh_b_id, _m1_id = self._setup()
        client = _make_client()
        resp = client.get(f"/report/inout/print?warehouse_id={wh_b_id}")
        wb = _read_xlsx(resp)
        _, in_rows = _sheet_rows(wb, "入库统计")
        _, out_rows = _sheet_rows(wb, "领料统计")
        assert [row[0] for row in in_rows] == ["IN-B-01"], in_rows
        assert [row[0] for row in out_rows] == ["OUT-B-01"], out_rows
        # 仓库 B 的数量必须出现，仓库 A 的数量不得混入
        # 列结构：单据编号(0), 日期(1), 供应商/领料部门(2), 合同编号(3), 工程名称(4), 物料编码(5), 物料名称(6), 数量(7), 金额(8)
        assert [row[7] for row in in_rows] == [20], in_rows
        assert [row[7] for row in out_rows] == [7], out_rows

    def test_inout_print_default_warehouse_fallback(self):
        self._setup(with_default=True)
        client = _make_client()
        resp = client.get("/report/inout/print")
        wb = _read_xlsx(resp)
        _, in_rows = _sheet_rows(wb, "入库统计")
        _, out_rows = _sheet_rows(wb, "领料统计")
        assert [row[0] for row in in_rows] == ["IN-A-01"], in_rows
        assert [row[0] for row in out_rows] == ["OUT-A-01"], out_rows

    def test_inout_print_no_default_returns_400(self):
        self._setup(with_default=False)
        client = _make_client()
        resp = client.get("/report/inout/print")
        assert resp.status_code == 400, resp.status_code
        payload = resp.get_json()
        assert payload and "请选择仓库" in (payload.get("msg") or "")

    def test_inout_export_matches_print_behavior(self):
        wh_a_id, _wh_b_id, _m1_id = self._setup()
        client = _make_client()
        resp = client.get(f"/report/inout/export?warehouse_id={wh_a_id}")
        wb = _read_xlsx(resp)
        _, in_rows = _sheet_rows(wb, "入库统计")
        _, out_rows = _sheet_rows(wb, "领料统计")
        assert [row[0] for row in in_rows] == ["IN-A-01"], in_rows
        assert [row[0] for row in out_rows] == ["OUT-A-01"], out_rows


class TestStockPrintWarehouseScope(_BaseExport):
    """/report/stock/print 必须输出仓库级库存而非全局 Material.stock。"""

    def test_stock_print_uses_warehouse_quantity_not_global(self):
        wh_a_id, _wh_b_id, _m1_id = self._setup()
        client = _make_client()
        resp = client.get(f"/report/stock/print?warehouse_id={wh_a_id}")
        wb = _read_xlsx(resp)
        header, rows = _sheet_rows(wb, "库存报表")
        stock_idx = header.index("当前库存")
        code_idx = header.index("物料编码")
        by_code = {row[code_idx]: row[stock_idx] for row in rows}
        # 仓库 A 库位库存 = 5；全局 100 / 仓库 B 95 均不得出现
        assert by_code.get("M001") == 5, rows

    def test_stock_print_warehouse_b_quantity(self):
        _wh_a_id, wh_b_id, _m1_id = self._setup()
        client = _make_client()
        resp = client.get(f"/report/stock/print?warehouse_id={wh_b_id}")
        wb = _read_xlsx(resp)
        header, rows = _sheet_rows(wb, "库存报表")
        stock_idx = header.index("当前库存")
        code_idx = header.index("物料编码")
        by_code = {row[code_idx]: row[stock_idx] for row in rows}
        assert by_code.get("M001") == 95, rows

    def test_stock_print_no_default_returns_400(self):
        self._setup(with_default=False)
        client = _make_client()
        resp = client.get("/report/stock/print")
        assert resp.status_code == 400, resp.status_code
        payload = resp.get_json()
        assert payload and "请选择仓库" in (payload.get("msg") or "")


class TestStockQueryTemplateLink:
    """库存查询页不得再生成无仓库参数的旧导出链接。"""

    def test_stock_query_print_link_carries_warehouse_id(self):
        with app_module.app.app_context():
            _reset_db()
            _seed_admin()
            wh_a = _seed_warehouse("WHA", "仓库A", is_default=True)
            wh_a_id = wh_a.id
        client = _make_client()
        resp = client.get(f"/stock_query?warehouse_id={wh_a_id}")
        assert resp.status_code == 200
        html = resp.get_data(as_text=True)
        assert 'href="/report/stock/print"' not in html, "旧导出链接未携带 warehouse_id"
        assert f"/report/stock/print?warehouse_id={wh_a_id}" in html, html[:2000]
