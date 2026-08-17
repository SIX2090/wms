# -*- coding: utf-8 -*-
"""BUG-2026-08-17-005 回归：库存状态与旧列表导出必须按仓库隔离。"""
from __future__ import annotations

import io
import os
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")
os.environ.setdefault("WMS_SKIP_AUTO_UPDATE", "1")

import app as app_module  # noqa: E402
from app import (  # noqa: E402
    AdjustmentOrder,
    AdjustmentOrderItem,
    InOrder,
    InOrderItem,
    InventoryCheck,
    InventoryCheckItem,
    Material,
    OutOrder,
    OutOrderItem,
    User,
    Warehouse,
    add_stock,
    db,
    set_system_setting,
)

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


def _reset_db():
    db.drop_all()
    db.create_all()


def _login(client):
    client.post("/login", data={"username": "admin", "password": "admin"})


def _read_rows(response):
    from openpyxl import load_workbook

    assert response.status_code == 200, response.status_code
    worksheet = load_workbook(io.BytesIO(response.data)).active
    rows = list(worksheet.iter_rows(values_only=True))
    return rows[0], rows[1:]


def _seed_warehouses_and_material():
    from werkzeug.security import generate_password_hash

    db.session.add_all([
        User(username="admin", password_hash=generate_password_hash("admin"),
             role="admin", must_change_password=False),
        Warehouse(code="WHA", name="仓库A", status="active", is_default=True),
        Warehouse(code="WHB", name="仓库B", status="active"),
    ])
    db.session.commit()
    material = Material(code="M001", name="测试轴承", stock=100, min_stock=5, price=10)
    db.session.add(material)
    db.session.commit()
    return (
        Warehouse.query.filter_by(code="WHA").first(),
        Warehouse.query.filter_by(code="WHB").first(),
        material,
    )


class TestWarehouseStockStatusAndLegacyExports:
    def test_stock_query_and_legacy_exports_use_selected_warehouse(self):
        with app_module.app.app_context():
            _reset_db()
            warehouse_a, warehouse_b, material = _seed_warehouses_and_material()
            set_system_setting("inventory_alert_enabled", "1")
            set_system_setting("location_management_enabled", "0")
            with app_module.app.test_request_context():
                add_stock(material, 10, "in", "in_order", 1, warehouse=warehouse_b)
            db.session.add_all([
                InOrder(order_no="IN-A", warehouse=warehouse_a.name, date=date.today(), status="completed"),
                InOrder(order_no="IN-B", warehouse=warehouse_b.name, date=date.today(), status="completed"),
                OutOrder(order_no="OUT-A", warehouse=warehouse_a.name, date=date.today(), status="completed"),
                OutOrder(order_no="OUT-B", warehouse=warehouse_b.name, date=date.today(), status="completed"),
                InventoryCheck(check_no="CHECK-A", warehouse=warehouse_a.name, date=date.today(), status="completed"),
                InventoryCheck(check_no="CHECK-B", warehouse=warehouse_b.name, date=date.today(), status="completed"),
                AdjustmentOrder(adjustment_no="ADJ-A", warehouse=warehouse_a.name, date=date.today(), adjustment_type="surplus", status="completed"),
                AdjustmentOrder(adjustment_no="ADJ-B", warehouse=warehouse_b.name, date=date.today(), adjustment_type="loss", status="completed"),
            ])
            db.session.flush()
            for order in InOrder.query.all():
                db.session.add(InOrderItem(in_order_id=order.id, material_id=material.id, quantity=1, price=10, amount=10))
            for order in OutOrder.query.all():
                db.session.add(OutOrderItem(out_order_id=order.id, material_id=material.id, quantity=1, price=10, amount=10))
            for check in InventoryCheck.query.all():
                db.session.add(InventoryCheckItem(inventory_check_id=check.id, material_id=material.id,
                                                  system_stock=1, actual_stock=1, difference=0))
            for order in AdjustmentOrder.query.all():
                db.session.add(AdjustmentOrderItem(adjustment_order_id=order.id, material_id=material.id,
                                                   quantity=1, reason="测试"))
            db.session.commit()

            client = app_module.app.test_client()
            _login(client)

            low_a = client.get(f"/stock_query?warehouse_id={warehouse_a.id}&stock_filter=low").get_data(as_text=True)
            normal_a = client.get(f"/stock_query?warehouse_id={warehouse_a.id}&stock_filter=normal").get_data(as_text=True)
            low_b = client.get(f"/stock_query?warehouse_id={warehouse_b.id}&stock_filter=low").get_data(as_text=True)
            normal_b = client.get(f"/stock_query?warehouse_id={warehouse_b.id}&stock_filter=normal").get_data(as_text=True)
            assert "M001" in low_a
            assert "M001" not in normal_a
            assert "M001" not in low_b
            assert "M001" in normal_b

            for url, included, excluded in (
                (f"/in_order/export?warehouse_id={warehouse_a.id}", "IN-A", "IN-B"),
                (f"/out_order/export?warehouse_id={warehouse_a.id}", "OUT-A", "OUT-B"),
                (f"/check/export?warehouse_id={warehouse_a.id}", "CHECK-A", "CHECK-B"),
                (f"/adjustment/export?warehouse_id={warehouse_a.id}", "ADJ-A", "ADJ-B"),
            ):
                header, rows = _read_rows(client.get(url))
                values = [value for row in rows for value in row]
                assert included in values, (url, rows)
                assert excluded not in values, (url, rows)
                if "check/export" in url or "adjustment/export" in url:
                    warehouse_index = header.index("仓库")
                    assert {row[warehouse_index] for row in rows} == {warehouse_a.name}
