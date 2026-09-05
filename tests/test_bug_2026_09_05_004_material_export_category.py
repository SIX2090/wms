# -*- coding: utf-8 -*-
"""BUG-2026-09-05-004 回归：物料导出忽略 category_id 筛选。

问题：物料列表页左侧分类树筛选后点「导出」（app.js
buildCurrentFilteredUrl 会把 category_id 拼进导出 URL），但
export_material 此前只读 search/stock_filter/sort/order，完全不读
category_id——用户选「电线」分类导出，得到的却是全量物料，
页面与导出结果不一致。与 BUG-2026-08-29-003（供应商文本筛选后导出
与页面不一致）同根因：页面筛选条件未被导出接口接收（R6 同类消费点）。

修复：export_material 读取 category_id，逐字镜像 material_list 的
分类子树展开（build_category_tree_rows + descendants），按
Material.category_id.in_(子树 ids) 过滤，页导一致；超限闪回重定向
同样携带 category_id 不丢筛选。

覆盖：
T1. 带 category_id 导出：仅含该分类（含子分类）物料，其他分类不出现
T2. 不带 category_id：全量导出（原行为不变）
T3. category_id 与 search 叠加：两个条件同时生效
T4. 子分类展开：选父分类导出包含子分类物料（与列表页口径一致）
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")

import app as app_module  # noqa: E402
from app import db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False

_ctx = app_module.app.app_context()

import pytest as _pytest  # noqa: E402


@_pytest.fixture(autouse=True, scope="module")
def _release_app_ctx_after_module():
    _ctx.push()
    yield
    try:
        _ctx.pop()
    except Exception:
        pass


def _reset_db():
    db.drop_all()
    db.create_all()


def _seed_scene():
    from werkzeug.security import generate_password_hash
    from app import Material, MaterialCategory, User
    _reset_db()
    u = User(username="admin", password_hash=generate_password_hash("admin"),
             role="admin", must_change_password=False)
    db.session.add(u)
    cat_wire = MaterialCategory(code="DX", name="电线")
    db.session.add(cat_wire)
    db.session.flush()
    cat_cable = MaterialCategory(code="DL", name="电缆", parent_id=cat_wire.id)
    cat_other = MaterialCategory(code="QT", name="其他")
    db.session.add_all([cat_cable, cat_other])
    db.session.flush()
    mats = [
        Material(code="W001", name="BV电线", stock=10, category_id=cat_wire.id),
        Material(code="W002", name="护套电线", stock=20, category_id=cat_wire.id),
        Material(code="C001", name="电力电缆", stock=5, category_id=cat_cable.id),
        Material(code="Q001", name="螺丝", stock=100, category_id=cat_other.id),
        Material(code="N001", name="无分类物料", stock=1, category_id=None),
    ]
    db.session.add_all(mats)
    db.session.commit()
    return cat_wire, cat_cable, cat_other


def _client():
    client = app_module.app.test_client()
    r = client.post("/login", data={"username": "admin", "password": "admin"},
                    follow_redirects=False)
    assert r.status_code in (302, 303), r.get_data(as_text=True)
    return client


def _export_codes(client, query=""):
    from openpyxl import load_workbook
    r = client.get(f"/material/export{query}")
    assert r.status_code == 200, r.get_data(as_text=True)
    wb = load_workbook(io.BytesIO(r.data))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return {row[0] for row in rows}


# T1 ───────────────────────────────────────────────────────────────
def test_t1_export_with_category_only_that_category():
    """带 category_id：仅该分类（含子分类）物料入导出。"""
    cat_wire, _c, _o = _seed_scene()
    codes = _export_codes(_client(), f"?category_id={cat_wire.id}")
    assert codes == {"W001", "W002", "C001"}


# T2 ───────────────────────────────────────────────────────────────
def test_t2_export_without_category_full_set():
    """不带 category_id：全量导出（原行为不变）。"""
    _seed_scene()
    codes = _export_codes(_client())
    assert codes == {"W001", "W002", "C001", "Q001", "N001"}


# T3 ───────────────────────────────────────────────────────────────
def test_t3_category_and_search_stack():
    """category_id 与 search 叠加：分类子树 ∩ 关键词。"""
    cat_wire, _c, _o = _seed_scene()
    codes = _export_codes(_client(), f"?category_id={cat_wire.id}&search=护套")
    assert codes == {"W002"}


# T4 ───────────────────────────────────────────────────────────────
def test_t4_leaf_category_excludes_sibling():
    """选叶子分类「电缆」：只导出电缆，不含父分类「电线」物料。"""
    _w, cat_cable, _o = _seed_scene()
    codes = _export_codes(_client(), f"?category_id={cat_cable.id}")
    assert codes == {"C001"}
