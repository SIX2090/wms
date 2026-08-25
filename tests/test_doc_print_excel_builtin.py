# -*- coding: utf-8 -*-
"""PRINT-TEMPLATE-F03（A1）回归测试：全模块内置 Excel 打印模板。

需求（2026-08-25）：所有单据、列表、报表的打印模块都要支持在线 Excel
格式编辑（参考简道云 Excel 打印模板）。前提是每个模块都有规范的内置
Excel 模板，启动时幂等同步进 excel_print_template 表，模板中心即可对
其在线编辑。

测试用例：
  T1. 全部注册目标的内置模板可通过 validate_template_file 安全校验，
      且含明细占位符行（引擎可识别）
  T2. ensure_builtin_excel_doc_templates 幂等：首次建行+生成文件，
      二次运行无任何变化；内置副本文件丢失后再次运行自动补文件并回指
  T3. render_doc_excel_print 用内置模板正确填充单据（表头/明细扩展/合计）
  T4. render_table_excel_print 多工作表报表按工作表名分别填充
  T5. ensure 对缺失的 excel_print_template 表 / 缺失的 DB 文件静默跳过
"""
from __future__ import annotations

import io
import os
import sqlite3
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent.parent / "app"
sys.path.insert(0, str(APP_DIR))

import doc_print_excel as dpe  # noqa: E402
from print_fill import validate_template_file  # noqa: E402


# -------------------- T1：内置模板生成合法性 --------------------

ALL_TARGET_CODES = sorted(
    list(dpe.DOC_EXCEL_PRINT_TYPES) + list(dpe.TABLE_EXCEL_PRINT_TYPES)
    + list(dpe.LABEL_EXCEL_PRINT_TYPES))


@pytest.mark.parametrize("target_code", ALL_TARGET_CODES)
def test_generate_builtin_template(target_code):
    content = dpe.generate_builtin_template(target_code)
    assert content is not None, f'{target_code} 未生成内置模板'
    raw = content.read()
    assert validate_template_file(raw) == '', f'{target_code} 内置模板未通过安全校验'
    wb = load_workbook(io.BytesIO(raw))
    # 每个工作表都必须有引擎可识别的 {item.} 明细占位符行
    for ws in wb.worksheets:
        texts = [c.value for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str)]
        assert any('{item.' in t for t in texts), \
            f'{target_code}/{ws.title} 缺少明细占位符行'
        # 明细行下方禁止合并单元格（openpyxl insert_rows 不移动合并区）
        item_row = next(
            r for r in range(1, (ws.max_row or 1) + 1)
            if any(isinstance(ws.cell(r, c).value, str)
                   and '{item.' in ws.cell(r, c).value
                   for c in range(1, (ws.max_column or 1) + 1)))
        for rng in ws.merged_cells.ranges:
            assert rng.min_row <= item_row, \
                f'{target_code}/{ws.title} 明细行下方存在合并单元格 {rng}'


def test_unregistered_target_code_returns_none():
    assert dpe.generate_builtin_template('no_such_target') is None


# -------------------- T2：ensure 幂等同步 --------------------

def _create_db(db_path):
    conn = sqlite3.connect(str(db_path))
    conn.execute(
        "CREATE TABLE excel_print_template ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " name VARCHAR(100) NOT NULL,"
        " target_type VARCHAR(30) NOT NULL,"
        " target_code VARCHAR(80) NOT NULL,"
        " template_type VARCHAR(20) NOT NULL DEFAULT 'excel',"
        " excel_template_path VARCHAR(500) NOT NULL,"
        " is_default BOOLEAN,"
        " created_at DATETIME, updated_at DATETIME)")
    conn.commit()
    conn.close()


def _snapshot(db_path):
    conn = sqlite3.connect(str(db_path))
    rows = conn.execute(
        "SELECT name, target_type, target_code, template_type,"
        " excel_template_path, is_default FROM excel_print_template"
        " ORDER BY target_type, target_code, id").fetchall()
    conn.close()
    return rows


@pytest.fixture()
def sandbox(tmp_path):
    db_path = tmp_path / 'test.db'
    _create_db(db_path)
    static_folder = tmp_path / 'static'
    static_folder.mkdir()
    return str(db_path), str(static_folder)


def test_ensure_builtin_excel_doc_templates(sandbox):
    db_path, static_folder = sandbox
    dpe.ensure_builtin_excel_doc_templates(db_path, static_folder)
    rows = _snapshot(db_path)
    assert len(rows) == len(ALL_TARGET_CODES)
    codes = {r[2] for r in rows}
    assert codes == set(ALL_TARGET_CODES)
    for name, target_type, code, tpl_type, path, is_default in rows:
        assert name.startswith('系统默认') and name.endswith('模板')
        assert tpl_type == 'excel'
        assert is_default == 1
        assert path.startswith('/static/uploads/print_templates/builtin_')
        rel = path[len('/static/'):]
        assert os.path.exists(os.path.join(static_folder, rel))


def test_ensure_is_idempotent(sandbox):
    db_path, static_folder = sandbox
    dpe.ensure_builtin_excel_doc_templates(db_path, static_folder)
    first = _snapshot(db_path)
    dpe.ensure_builtin_excel_doc_templates(db_path, static_folder)
    assert _snapshot(db_path) == first


def test_ensure_repairs_lost_builtin_file(sandbox):
    db_path, static_folder = sandbox
    dpe.ensure_builtin_excel_doc_templates(db_path, static_folder)
    rows = _snapshot(db_path)
    victim = next(r for r in rows if r[2] == 'transfer')
    rel = victim[4][len('/static/'):]
    os.remove(os.path.join(static_folder, rel))
    # 用户自建模板（非 builtin_ 前缀）即使路径不同也不被动
    conn = sqlite3.connect(db_path)
    conn.execute(
        "INSERT INTO excel_print_template (name, target_type, target_code,"
        " template_type, excel_template_path, is_default, created_at, updated_at)"
        " VALUES ('我的调拨模板', 'document', 'transfer', 'excel',"
        " '/static/uploads/print_templates/user_custom.xlsx', 0,"
        " datetime('now'), datetime('now'))")
    conn.commit()
    conn.close()
    dpe.ensure_builtin_excel_doc_templates(db_path, static_folder)
    after = _snapshot(db_path)
    assert len(after) == len(rows) + 1
    repaired = next(r for r in after if r[2] == 'transfer'
                    and r[4].startswith('/static/uploads/print_templates/builtin_'))
    assert os.path.exists(os.path.join(static_folder, repaired[4][len('/static/'):]))
    # 用户自建模板路径未被改写
    custom = next(r for r in after if r[0] == '我的调拨模板')
    assert custom[4] == '/static/uploads/print_templates/user_custom.xlsx'
    assert custom[5] == 0  # 已有默认模板，用户模板不会被抬为默认


# -------------------- T3：单据填充 --------------------

def _fake_order(**overrides):
    unit = SimpleNamespace(name='件')
    material = SimpleNamespace(code='M-001', brand='品牌A', name='螺丝',
                               spec='M6*20', unit=unit)
    items = [
        SimpleNamespace(material=material, unit=unit, quantity=2, price=1.5,
                        amount=3.0, remark='第一批', contract_no='HT-1'),
        SimpleNamespace(material=material, unit=unit, quantity=3, price=2.0,
                        amount=6.0, remark='', contract_no='HT-1'),
        SimpleNamespace(material=material, unit=unit, quantity=5, price=1.0,
                        amount=5.0, remark='尾批', contract_no='HT-1'),
    ]
    order = SimpleNamespace(
        id=9, order_no='PO-2026-0009', transfer_no='DB-2026-0009',
        date='2026-08-25', status='completed', remark='测试单据',
        from_warehouse='原料仓', to_warehouse='成品仓', warehouse='原料仓',
        operator=SimpleNamespace(username='张三'),
        supplier=SimpleNamespace(name='供应商X', contact='李四', phone='138'),
        total_amount=14.0, items=items,
    )
    for key, value in overrides.items():
        setattr(order, key, value)
    return order


def test_render_doc_excel_print(tmp_path, monkeypatch):
    monkeypatch.setattr(dpe, 'resolve_excel_template', lambda *a, **k: None)
    order = _fake_order()
    result = dpe.render_doc_excel_print(
        'transfer', order, static_folder=str(tmp_path), date_str='2026-08-25')
    assert result is not None
    output, filename = result
    assert filename.endswith('.xlsx') and '调拨单' in filename
    wb = load_workbook(output)
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    # 表头字段已填充
    assert 'DB-2026-0009' in texts
    assert '原料仓' in texts and '成品仓' in texts
    # 3 条明细正确扩展
    data_rows = [r for r in range(1, (ws.max_row or 1) + 1)
                 if ws.cell(r, 1).value == 'M-001']
    assert len(data_rows) == 3
    # 合计行：数量合计 10（引擎按 item.quantity 求和）
    total_row = next(r for r in range(1, (ws.max_row or 1) + 1)
                     if ws.cell(r, 1).value == '合计')
    assert float(ws.cell(total_row, 6).value) == pytest.approx(10.0)
    # 签名行
    assert any(isinstance(t, str) and t.startswith('制单：张三') for t in texts)


def test_render_doc_excel_print_unknown_code_returns_none(tmp_path):
    assert dpe.render_doc_excel_print(
        'nope', _fake_order(), static_folder=str(tmp_path)) is None
    assert dpe.render_doc_excel_print(
        'transfer', None, static_folder=str(tmp_path)) is None


# -------------------- T4：列表/报表填充 --------------------

def test_render_table_excel_print(tmp_path, monkeypatch):
    monkeypatch.setattr(dpe, 'resolve_excel_template', lambda *a, **k: None)
    rows_in = [SimpleNamespace(order_no='RK-1', date='2026-08-01', supplier='供应商X',
                               contract_no='HT-1', project_name='工程A',
                               material_code='M-001', material_name='螺丝',
                               quantity=2, amount=3.0)]
    rows_out = [SimpleNamespace(order_no='LL-1', date='2026-08-02', supplier='生产部',
                                contract_no='', project_name='工程A',
                                material_code='M-002', material_name='螺母',
                                quantity=4, amount=2.0)]
    result = dpe.render_table_excel_print(
        'report_inout', {'入库统计': rows_in, '领料统计': rows_out},
        static_folder=str(tmp_path), date_str='2026-08-25')
    assert result is not None
    output, filename = result
    assert '出入库统计报表' in filename
    wb = load_workbook(output)
    ws_in = wb['入库统计']
    assert any(ws_in.cell(r, 1).value == 'RK-1'
               for r in range(1, (ws_in.max_row or 1) + 1))
    ws_out = wb['领料统计']
    assert any(ws_out.cell(r, 1).value == 'LL-1'
               for r in range(1, (ws_out.max_row or 1) + 1))
    # 入库表不被领料数据污染
    assert not any(ws_in.cell(r, 1).value == 'LL-1'
                   for r in range(1, (ws_in.max_row or 1) + 1))


# -------------------- T4.5：模板记录解析 --------------------

class _FakeQuery:
    """最小化的 SQLAlchemy query 替身（链式 filter_by/order_by + first）。"""

    def __init__(self, result=None, exc=None):
        self._result = result
        self._exc = exc

    def filter_by(self, **kwargs):
        return self

    def order_by(self, *args):
        return self

    def first(self):
        if self._exc is not None:
            raise self._exc
        return self._result


def test_resolve_excel_template(monkeypatch):
    sentinel = SimpleNamespace(id=1, name='模板')
    fake_module = SimpleNamespace(
        ExcelPrintTemplate=SimpleNamespace(
            query=_FakeQuery(result=sentinel), updated_at=None))
    monkeypatch.setitem(sys.modules, 'app', fake_module)
    assert dpe.resolve_excel_template('document', 'transfer') is sentinel
    # ORM 查询异常（无应用上下文/表缺失）→ 回退 None 而非抛错
    broken_module = SimpleNamespace(
        ExcelPrintTemplate=SimpleNamespace(
            query=_FakeQuery(exc=RuntimeError('no app context')),
            updated_at=None))
    monkeypatch.setitem(sys.modules, 'app', broken_module)
    assert dpe.resolve_excel_template('document', 'transfer') is None


# -------------------- T5：ensure 异常环境静默跳过 --------------------

def test_ensure_skips_missing_table_and_db(tmp_path):
    # DB 文件不存在
    dpe.ensure_builtin_excel_doc_templates(
        str(tmp_path / 'missing.db'), str(tmp_path))
    # 表不存在
    empty_db = tmp_path / 'empty.db'
    sqlite3.connect(str(empty_db)).close()
    dpe.ensure_builtin_excel_doc_templates(str(empty_db), str(tmp_path))
    conn = sqlite3.connect(str(empty_db))
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()
    conn.close()
    assert tables == []
