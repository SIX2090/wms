# -*- coding: utf-8 -*-
"""PRINT-TEMPLATE-F03（A2）回归测试：全模块单据 Excel 模板打印路由。

需求（2026-08-25）：所有单据的打印模块都要支持在线 Excel 格式编辑。
A1 已同步内置模板；A2 把 10 种单据的打印出口接到模板体系：
- GET /{prefix}/<id>/print_excel 按所选（或默认）模板填充下载 .xlsx
- GET /doc_print_templates/<target_code>.json 可选模板列表（含在线编辑入口）

测试用例：
  T1. 10 条 print_excel 路由全部注册（GET）
  T2. 模板列表 JSON：返回默认模板与在线编辑链接；未知 target_code 404
  T3. 调拨单 print_excel 端到端：按内置模板填充，表头/明细/合计正确
  T4. 未登录访问 print_excel 被重定向到登录页
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

import app as m  # noqa: E402

EXPECTED_RULES = {
    '/check/<int:id>/print_excel': 'doc_print_excel_check',
    '/transfer/<int:id>/print_excel': 'doc_print_excel_transfer',
    '/requisition/<int:id>/print_excel': 'doc_print_excel_requisition',
    '/purchase_order/<int:id>/print_excel': 'doc_print_excel_purchase_order',
    '/sales/<int:id>/print_excel': 'doc_print_excel_sales_order',
    '/adjustment/<int:id>/print_excel': 'doc_print_excel_adjustment',
    '/subcontract/<int:id>/print_excel': 'doc_print_excel_subcontract',
    '/subcontract_issue/<int:id>/print_excel': 'doc_print_excel_subcontract_issue',
    '/subcontract_receive/<int:id>/print_excel': 'doc_print_excel_subcontract_receive',
    '/after_sale_out/<int:id>/print_excel': 'doc_print_excel_after_sale_out',
}


def _client():
    m.app.config['TESTING'] = True
    m.app.config['WTF_CSRF_ENABLED'] = False
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
        from werkzeug.security import generate_password_hash
        m.db.session.add(m.User(
            username='admin', password_hash=generate_password_hash('admin'),
            role='admin', must_change_password=False))
        m.db.session.commit()
    client = m.app.test_client()
    client.post('/login', data={'username': 'admin', 'password': 'admin'})
    return client


def test_register_doc_print_excel_routes():
    rules = {r.rule: r.endpoint for r in m.app.url_map.iter_rules()
             if r.endpoint.startswith('doc_print_excel_')}
    for rule, endpoint in EXPECTED_RULES.items():
        assert rules.get(rule) == endpoint, f'缺少路由 {rule}'
        methods = next(r.methods for r in m.app.url_map.iter_rules(endpoint)
                       if r.rule == rule)
        assert 'GET' in methods
    assert any(r.rule == '/doc_print_templates/<target_code>.json'
               for r in m.app.url_map.iter_rules())


def test_doc_print_templates_json():
    client = _client()
    with m.app.app_context():
        import doc_print_excel as dpe
        dpe.ensure_builtin_excel_doc_templates(
            m._resolve_sqlite_db_path() or ':memory:', m.app.static_folder)
        # 内存库下 raw sqlite 同步不可达，直接 ORM 建行
        if not m.ExcelPrintTemplate.query.filter_by(target_code='transfer').first():
            path = dpe._builtin_template_abspath(m.app.static_folder, 'transfer')
            m.db.session.add(m.ExcelPrintTemplate(
                name='系统默认调拨单模板', target_type='document',
                target_code='transfer', template_type='excel',
                excel_template_path='/static/uploads/print_templates/'
                                    + os.path.basename(path),
                is_default=True))
            m.db.session.commit()
    resp = client.get('/doc_print_templates/transfer.json')
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['status'] == 'success'
    assert data['target_type'] == 'document'
    assert len(data['templates']) >= 1
    tpl = data['templates'][0]
    assert tpl['is_default'] is True
    assert tpl['has_file'] is True
    assert tpl['edit_url'].endswith(f"/{tpl['id']}/edit")
    assert tpl['edit_url'].startswith('/global_print_template/')
    # 未知 target_code → 404
    assert client.get('/doc_print_templates/no_such.json').status_code == 404


def test_doc_print_excel_transfer_e2e():
    client = _client()
    with m.app.app_context():
        unit = m.Unit(code='JIAN', name='件')
        m.db.session.add(unit)
        m.db.session.flush()
        material = m.Material(code='M-001', name='螺丝', spec='M6*20',
                              brand='品牌A', unit_id=unit.id)
        m.db.session.add(material)
        m.db.session.flush()
        order = m.TransferOrder(
            transfer_no='DB-2026-0801', from_warehouse='原料仓',
            to_warehouse='成品仓', from_location='原料仓', to_location='成品仓',
            status='completed', remark='A2测试', operator_id=1)
        m.db.session.add(order)
        m.db.session.flush()
        m.db.session.add(m.TransferOrderItem(
            transfer_order_id=order.id, material_id=material.id,
            quantity=4, unit_id=unit.id, price=2.0, amount=8.0, remark='批1'))
        m.db.session.add(m.TransferOrderItem(
            transfer_order_id=order.id, material_id=material.id,
            quantity=6, unit_id=unit.id, price=1.0, amount=6.0, remark=''))
        m.db.session.commit()
        order_id = order.id
    resp = client.get(f'/transfer/{order_id}/print_excel')
    assert resp.status_code == 200
    assert resp.headers['Content-Type'].startswith(
        'application/vnd.openxmlformats-officedocument')
    assert 'DB-2026-0801' in resp.headers.get('Content-Disposition', '')
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert 'DB-2026-0801' in texts
    assert '原料仓' in texts and '成品仓' in texts
    data_rows = [r for r in range(1, (ws.max_row or 1) + 1)
                 if ws.cell(r, 1).value == 'M-001']
    assert len(data_rows) == 2
    total_row = next(r for r in range(1, (ws.max_row or 1) + 1)
                     if ws.cell(r, 1).value == '合计')
    assert float(ws.cell(total_row, 6).value) == 10.0
    assert any(isinstance(t, str) and t.startswith('制单：admin') for t in texts)


def test_doc_print_excel_requires_login():
    m.app.config['TESTING'] = True
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
    anon = m.app.test_client()
    resp = anon.get('/transfer/1/print_excel')
    assert resp.status_code in (301, 302, 401)
    assert anon.get('/doc_print_templates/transfer.json').status_code in (301, 302, 401)
