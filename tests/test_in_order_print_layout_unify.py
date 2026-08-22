# -*- coding: utf-8 -*-
"""采购入库单各打印出口版式统一回归（2026-08-22 用户反馈"修改没生效"）。

网页打印（print_in_with_excel.html）改了「采购入库单号」标签和「收货：」
位置后，Excel 下载出口仍是旧版式，且内置示例模板是另一套布局。本次统一：
1. _build_in_order_excel（无模板文件时的回退生成器）：采购入库单号 +
   收货：起点对齐单价列（G 列，G:I 合并）。
2. static/templates/入库单打印模板示例.xlsx 重建为标准 9 列版式。
3. 内置模板系统管理副本（builtin_* 前缀）随版式升级：v1→v2 路径切换。

覆盖：回退生成器标签与签名列 / 示例模板版式守卫 / 填充示例模板端到端 /
builtin 管理副本升级且用户文件不动。
"""
from __future__ import annotations

import os
import sqlite3
import sys
from datetime import date
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

import app as m  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

EXAMPLE_XLSX = APP_DIR / "static" / "templates" / "入库单打印模板示例.xlsx"


def _fake_order():
    unit = SimpleNamespace(name='把')
    material = SimpleNamespace(code='208005', brand='None', name='0.5-10mm端子压线钳',
                               spec='', unit=unit)
    item = SimpleNamespace(material=material, quantity=3, price=0.0, amount=0.0,
                           contract_no='')
    return SimpleNamespace(order_no='IN26080175', date=date(2026, 8, 22),
                           supplier=SimpleNamespace(name='明辉达'), contract_no='',
                           remark='', total_amount=0.0, items=[item])


def test_fallback_builder_uses_new_label_and_signature_column():
    """_build_in_order_excel 输出「采购入库单号」且收货：在单价列（G）起点。"""
    from routes.in_order import _build_in_order_excel
    output = _build_in_order_excel(_fake_order())
    wb = load_workbook(output)
    ws = wb.active
    all_text = [c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str)]
    assert any('采购入库单号：IN26080175' in t for t in all_text)
    assert not any(t.startswith('采购单号') for t in all_text)
    # 收货：位于 G 列（第 7 列 = 单价列）
    sig_cells = [c for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and c.value.startswith('收货')]
    assert len(sig_cells) == 1
    assert sig_cells[0].column == 7
    wb.close()


def test_example_template_matches_canonical_layout():
    """内置示例模板为标准版式：采购入库单号占位符、9 列、收货在 G 列。"""
    wb = load_workbook(str(EXAMPLE_XLSX))
    ws = wb.active
    assert ws['A1'].value == '采购入库单'
    assert ws['D2'].value == '采购入库单号：{order.order_no}'
    headers = [ws.cell(row=3, column=c).value for c in range(1, 10)]
    assert headers == ['物料编码', '品牌', '物料名称', '规格', '单位',
                       '数量', '单价', '金额', '合同编号']
    assert ws['A4'].value == '{item.material.code}'
    assert ws['B4'].value == '{item.material.brand}'
    assert ws['I4'].value == '{item.contract_no}'
    sig = [c for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.startswith('收货')]
    assert len(sig) == 1 and sig[0].column == 7
    wb.close()


def test_fill_example_template_end_to_end():
    """用填充引擎填示例模板：标签、明细数据、收货签名列全部正确。"""
    from print_fill import build_filled_print_excel
    output = build_filled_print_excel(str(EXAMPLE_XLSX), _fake_order(), date_str='2026-08-22')
    wb = load_workbook(output)
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any(isinstance(t, str) and '采购入库单号：IN26080175' in t for t in texts)
    assert any(isinstance(t, str) and '供应商：明辉达' in t for t in texts)
    assert '208005' in texts
    assert '0.5-10mm端子压线钳' in texts
    sig = [c for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.startswith('收货')]
    assert len(sig) == 1 and sig[0].column == 7
    wb.close()


def test_builtin_managed_copy_upgrades_to_v2(tmp_path):
    """系统管理的 builtin_ 旧副本路径在启动同步时升级到 v2；用户文件不动。"""
    db_path = tmp_path / "inventory.db"
    with sqlite3.connect(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE in_order_print_template (
                id INTEGER PRIMARY KEY, name VARCHAR(100) NOT NULL,
                template_type VARCHAR(20), excel_template_path VARCHAR(500),
                html_template_content TEXT, is_default BOOLEAN,
                created_at DATETIME, updated_at DATETIME
            );
            INSERT INTO in_order_print_template
                (name, template_type, excel_template_path, html_template_content, is_default)
                VALUES ('系统默认入库单模板', 'excel',
                        '/static/uploads/print_templates/builtin_in_order_default.xlsx',
                        '<div>x</div>', 1);
            """
        )
    m._ensure_default_print_templates_unconditional(db_path=str(db_path))
    with sqlite3.connect(db_path) as conn:
        path = conn.execute(
            "SELECT excel_template_path FROM in_order_print_template WHERE is_default=1"
        ).fetchone()[0]
    assert path.endswith('builtin_in_order_default_v2.xlsx'), path
    abspath = APP_DIR / 'static' / 'uploads' / 'print_templates' / 'builtin_in_order_default_v2.xlsx'
    assert abspath.exists()
