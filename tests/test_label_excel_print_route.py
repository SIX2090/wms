# -*- coding: utf-8 -*-
"""PRINT-TEMPLATE-F04（A3）回归测试：标签 Excel 模板打印出口。

需求（2026-08-25）：标签打印也支持 Excel 在线编辑模板（参考简道云），
批量标签打印页可按所选（或默认）Excel 模板下载 .xlsx 标签（含条码图）。

测试用例：
  T1. /label/batch_print_excel 路由已注册且需登录（未登录跳转/401）
  T2. 端到端：按内置物料标签模板填充，每物料一行且嵌入条码图片
  T3. /doc_print_templates/material_label.json 返回 label 类型模板列表
  T4. 指定 template_id 时使用用户模板；未知物料 id 返回空明细但不 500
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

import app as m  # noqa: E402


def _client():
    m.app.config['TESTING'] = True
    m.app.config['WTF_CSRF_ENABLED'] = False
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
        from werkzeug.security import generate_password_hash
        m.db.session.add(m.User(
            username='admin', password_hash=generate_password_hash('admin'),
            role='admin', must_change_password=False))
        unit = m.Unit(code='TAO', name='套')
        m.db.session.add(unit)
        m.db.session.flush()
        m.db.session.add(m.Material(code='M-001', name='轴承', spec='6204',
                                    unit_id=unit.id, stock=100, price=12.5))
        m.db.session.add(m.Material(code='M-002', name='螺母', spec='M8',
                                    unit_id=unit.id, stock=500, price=0.5))
        m.db.session.commit()
    client = m.app.test_client()
    client.post('/login', data={'username': 'admin', 'password': 'admin'})
    return client


def test_batch_print_excel_route_registered():
    rules = {r.rule for r in m.app.url_map.iter_rules()}
    assert '/label/batch_print_excel' in rules


def test_batch_print_excel_requires_login():
    m.app.config['TESTING'] = True
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
    anon = m.app.test_client()
    resp = anon.get('/label/batch_print_excel?ids=1')
    assert resp.status_code in (301, 302, 401)


def test_batch_print_excel_e2e():
    client = _client()
    resp = client.get('/label/batch_print_excel?ids=1,2')
    assert resp.status_code == 200
    assert resp.headers['Content-Type'].startswith(
        'application/vnd.openxmlformats-officedocument')
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row
             if c.value is not None]
    assert 'M-001' in texts and '轴承' in texts
    assert 'M-002' in texts and '螺母' in texts
    # 每个物料一张条码图（内置模板 {img_barcode:item.barcode} 列）
    assert len(ws._images) == 2
    assert not any(isinstance(t, str) and '{item.' in t for t in texts)


def test_batch_print_excel_unknown_ids_no_500():
    client = _client()
    resp = client.get('/label/batch_print_excel?ids=999')
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    assert len(wb.active._images) == 0


def test_doc_print_templates_material_label_json():
    client = _client()
    with m.app.app_context():
        import doc_print_excel as dpe
        path = dpe._builtin_template_abspath(m.app.static_folder,
                                             'material_label')
        if not m.ExcelPrintTemplate.query.filter_by(
                target_code='material_label').first():
            m.db.session.add(m.ExcelPrintTemplate(
                name='系统默认物料标签模板', target_type='label',
                target_code='material_label', template_type='excel',
                excel_template_path='/static/uploads/print_templates/'
                                    + os.path.basename(path),
                is_default=True))
            m.db.session.commit()
    resp = client.get('/doc_print_templates/material_label.json')
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['status'] == 'success'
    assert data['target_type'] == 'label'
    assert len(data['templates']) >= 1
    assert data['templates'][0]['edit_url'].startswith(
        '/global_print_template/')
