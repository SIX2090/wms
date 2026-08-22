# -*- coding: utf-8 -*-
"""领料单打印签名位回归（2026-08-22 用户反馈）。

与采购入库单一致：「领料：」签名位移到「单价」列下方（G 列起点）。
覆盖三个打印出口：
1. 网页打印 print_out_with_excel.html：tfoot 无边框签名行（colspan 6+3）。
2. Excel 回退生成器 _build_out_order_excel：领料：在 G 列、G:I 合并。
3. 内置示例模板 领料单打印模板示例.xlsx：标准 9 列版式、领料：在 G12。
"""
from __future__ import annotations

import os
import sys
from datetime import date
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

import app as m  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

EXAMPLE_XLSX = APP_DIR / "static" / "templates" / "领料单打印模板示例.xlsx"


def _fake_order():
    unit = SimpleNamespace(name='斤')
    material = SimpleNamespace(code='207008', brand='None', name='不锈钢螺丝',
                               spec='10*65', unit=unit)
    item = SimpleNamespace(material=material, quantity=4, price=0.0, amount=0.0,
                           contract_no='HD260707')
    return SimpleNamespace(order_no='OU26080001', date=date(2026, 8, 22),
                           department=None, customer='生产部', contract_no='',
                           remark='', total_amount=0.0, items=[item])


def _webpage_html():
    """渲染领料单网页打印页（excel 类型默认模板 → print_out_with_excel.html）。"""
    m.app.config['TESTING'] = True
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
        from werkzeug.security import generate_password_hash
        user = m.User(username='admin', password_hash=generate_password_hash('admin'),
                      role='admin', must_change_password=False)
        m.db.session.add(user)
        unit = m.Unit(code='JIN', name='斤')
        cat = m.MaterialCategory(code='DEFAULT', name='默认分类')
        m.db.session.add_all([unit, cat])
        m.db.session.flush()
        material = m.Material(code='207008', name='不锈钢螺丝', brand='None',
                              spec='10*65', unit_id=unit.id, category_id=cat.id, price=0)
        m.db.session.add(material)
        m.db.session.flush()
        order = m.OutOrder(order_no='OU26080001', customer='生产部',
                           warehouse='主仓库', date=date(2026, 8, 22),
                           status='completed', operator_id=user.id)
        m.db.session.add(order)
        m.db.session.flush()
        m.db.session.add(m.OutOrderItem(out_order_id=order.id, material_id=material.id,
                                        quantity=4, price=0, amount=0, contract_no='HD260707'))
        m.db.session.add(m.OutOrderPrintTemplate(
            name='系统默认领料单模板', template_type='excel', is_default=True,
            html_template_content=m.DEFAULT_OUT_ORDER_HTML_TEMPLATE))
        m.db.session.commit()
        order_id = order.id
    m.app.config['WTF_CSRF_ENABLED'] = False
    client = m.app.test_client()
    client.post('/login', data={'username': 'admin', 'password': 'admin'})
    resp = client.get(f'/out_order/{order_id}/print')
    assert resp.status_code == 200
    return resp.data.decode('utf-8')


def test_webpage_signature_row_at_price_column():
    """网页打印：领料：在 tfoot 无边框行，起点对齐单价列（colspan 6+3）。"""
    html = _webpage_html()
    assert '<td colspan="6" class="sig-cell"></td>' in html
    assert '<td colspan="3" class="sig-cell">领料：</td>' in html
    assert 'signature-row' not in html


def test_fallback_builder_signature_at_price_column():
    """Excel 回退生成器：领料：在 G 列（第 7 列 = 单价列）。"""
    from routes.out_order import _build_out_order_excel
    output = _build_out_order_excel(_fake_order())
    wb = load_workbook(output)
    ws = wb.active
    sig = [c for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.startswith('领料：')]
    assert len(sig) == 1
    assert sig[0].column == 7
    wb.close()


def test_example_template_canonical_layout():
    """内置示例模板：领料部门/日期表头、9 列、领料：在 G 列。"""
    wb = load_workbook(str(EXAMPLE_XLSX))
    ws = wb.active
    assert ws['A1'].value == '领料单'
    assert ws['A2'].value == '领料部门：{order.customer}'
    headers = [ws.cell(row=3, column=c).value for c in range(1, 10)]
    assert headers == ['物料编码', '品牌', '物料名称', '规格', '单位',
                       '数量', '单价', '金额', '合同编号']
    assert ws['A4'].value == '{item.material.code}'
    sig = [c for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.startswith('领料：')]
    assert len(sig) == 1 and sig[0].column == 7
    wb.close()


def test_fill_example_template_end_to_end():
    """填充示例模板：部门、明细、合同编号、签名列全部正确。"""
    from print_fill import build_filled_print_excel
    output = build_filled_print_excel(str(EXAMPLE_XLSX), _fake_order(), date_str='2026-08-22')
    wb = load_workbook(output)
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any(isinstance(t, str) and '领料部门：生产部' in t for t in texts)
    assert '207008' in texts and 'HD260707' in texts
    sig = [c for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.startswith('领料：')]
    assert len(sig) == 1 and sig[0].column == 7
    wb.close()
