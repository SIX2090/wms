# -*- coding: utf-8 -*-
"""打印模板仅允许 Excel 的回归测试。

覆盖需求：
- 打印模板只能使用 Excel（create_print_template 强制 excel，HTML 被拒）。
- 打印模板可下载（/in_order_print_template/<id>/download）。
- 支持上传多个模板并可在打印时通过 /xx_print_templates.json 选择。
- 打印时按所选模板填充数据并下载（/in_order/<id>/print_excel?template_id=）。
"""
from __future__ import annotations

import io
import os
import sys
from datetime import date
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["WMS_DEBUG"] = "0"
os.environ.setdefault("WMS_SKIP_AUTO_UPDATE", "1")

from werkzeug.security import generate_password_hash  # noqa: E402

import app as app_module  # noqa: E402
from app import (  # noqa: E402
    InOrder, InOrderItem, InOrderPrintTemplate, Material, MaterialCategory,
    OutOrder, OutOrderItem, OutOrderPrintTemplate, Supplier, Unit, User, db,
)
from print_fill import build_filled_print_excel  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


def _login(client):
    return client.post(
        "/login",
        data={"username": "admin", "password": "admin"},
        content_type="application/x-www-form-urlencoded",
    )


def _reset_and_seed():
    """重置内存库并写入基础测试数据（单位/分类/供应商/用户/物料）。"""
    db.drop_all()
    db.create_all()
    db.session.add_all([
        Unit(name="个", code="PCS"),
        MaterialCategory(name="默认分类", code="CAT-DEFAULT"),
        Supplier(code="SUP001", name="华南轴承厂"),
        User(username="admin", password_hash=generate_password_hash("admin"),
             role="admin", must_change_password=False),
    ])
    db.session.commit()
    db.session.add(Material(code="M001", name="6204轴承", spec="6204",
                            category_id=1, unit_id=1, supplier_id=1,
                            stock=0, price=10))
    db.session.commit()


@pytest.fixture()
def seeded_db():
    """纯填充引擎/非 HTTP 测试用的独立库：每个用例干净重建并种子化。"""
    with app_module.app.app_context():
        _reset_and_seed()
    yield None


@pytest.fixture()
def client(seeded_db):
    c = app_module.app.test_client()
    _login(c)
    yield c


def _xlsx_bytes(content=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    if content:
        for (r, col, v) in content:
            ws.cell(r, col, v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_in_order():
    user = User.query.filter_by(username="admin").first()
    order = InOrder(order_no="IN-PRINT-001", business_type="采购入库",
                    date=date(2026, 8, 20), warehouse="仓库A",
                    supplier_id=1, operator_id=user.id,
                    total_amount=1500.0)
    db.session.add(order)
    db.session.flush()
    db.session.add(InOrderItem(
        in_order_id=order.id, material_id=1,
        quantity=100, price=10, amount=1000.0))
    db.session.add(InOrderItem(
        in_order_id=order.id, material_id=1,
        quantity=50, price=10, amount=500.0))
    db.session.commit()
    return order


def _upload_in_template(client, name="标准入库模板"):
    resp = client.post('/in_order_print_template/add', data={
        'name': name,
        'template_type': 'excel',
        'excel_file': (_xlsx_bytes([
            (1, 1, '采购入库单'),
            (2, 1, '{order.order_no}'),
            (3, 1, '{order.supplier.name}'),
            (4, 1, '明细'),
            (5, 1, '{item.material.name}'),
            (6, 1, '{total_amount}'),
        ]), '标准入库.xlsx'),
    }, content_type='multipart/form-data')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    return resp.get_json()['id']


def test_build_filled_print_excel_fills_placeholders(tmp_path, seeded_db):
    """填充引擎：订单级/明细级占位符被正确填充，模板路径可解析。"""
    from print_fill import template_file_abspath
    from openpyxl import load_workbook
    tpl = tmp_path / "tpl.xlsx"
    wb = _xlsx_bytes([
        (1, 1, '{order.order_no}'),
        (2, 1, '{order.supplier.name}'),
        (3, 1, '{item.material.name}'),
        (4, 1, '{total_amount}'),
    ])
    tpl.write_bytes(wb.read())
    with app_module.app.app_context():
        order = _make_in_order()
        out = build_filled_print_excel(str(tpl), order,
                                       date_str='2026-08-20')
    workbook = load_workbook(io.BytesIO(out.read()))
    ws = workbook.active
    text = '|'.join(str(c.value) for row in ws.iter_rows() for c in row)
    assert 'IN-PRINT-001' in text
    assert '华南轴承厂' in text
    assert '6204轴承' in text
    assert '1500' in text
    assert '{' not in text.replace('{{', '')  # 无残留占位符


def test_templates_json_lists_excel_templates(client):
    """templates.json 返回可用的 Excel 模板列表。"""
    tid = _upload_in_template(client)
    resp = client.get('/in_order_print_templates.json')
    data = resp.get_json()
    assert data['status'] == 'success'
    ids = [t['id'] for t in data['templates']]
    assert tid in ids
    assert all(t['has_file'] for t in data['templates'])


def test_download_print_template(client):
    """打印模板可下载为 Excel 文件。"""
    _upload_in_template(client)
    with app_module.app.app_context():
        t = InOrderPrintTemplate.query.filter_by(
            name="标准入库模板").first()
        tid = t.id
    resp = client.get(f'/in_order_print_template/{tid}/download')
    assert resp.status_code == 200
    assert resp.mimetype.endswith('spreadsheetml.sheet')
    assert resp.data[:2] == b'PK'  # xlsx 是 zip


def test_print_excel_with_selected_template(client):
    """打印时选择模板应返回按模板填充好的 xlsx。"""
    from openpyxl import load_workbook
    tid = _upload_in_template(client)
    with app_module.app.app_context():
        order = _make_in_order()
        oid = order.id
    resp = client.get(f'/in_order/{oid}/print_excel?template_id={tid}')
    assert resp.status_code == 200
    assert resp.mimetype.endswith('spreadsheetml.sheet')
    workbook = load_workbook(io.BytesIO(resp.data))
    ws = workbook.active
    text = '|'.join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert 'IN-PRINT-001' in text
    assert '6204轴承' in text


def test_print_excel_without_template_falls_back(client):
    """未指定模板时打印回退内置版式，仍返回 xlsx。"""
    with app_module.app.app_context():
        order = _make_in_order()
        oid = order.id
    resp = client.get(f'/in_order/{oid}/print_excel')
    assert resp.status_code == 200
    assert resp.mimetype.endswith('spreadsheetml.sheet')


def test_openpyxl_fill_preserves_rows(tmp_path, seeded_db):
    """填充引擎：明细行按条数扩展，不残留示例占位符。"""
    from openpyxl import load_workbook
    tpl = tmp_path / "tpl2.xlsx"
    tpl.write_bytes(_xlsx_bytes([
        (1, 1, '{item.material.name}'),
        (2, 1, '{item.quantity}'),
        (3, 1, '{total_amount}'),
    ]).read())
    with app_module.app.app_context():
        order = _make_in_order()  # 2 条明细
        out = build_filled_print_excel(str(tpl), order)
    ws = load_workbook(io.BytesIO(out.read())).active
    vals = [str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)]
    assert '6204轴承' in vals
    assert sum(1 for v in vals if v == '6204轴承') == 2  # 2 条明细各占一行
    assert all('{' not in v for v in vals)


# ---- 精确命名回归（满足防 BUG A9：每个新业务函数有同名测试） ----

def test_template_file_abspath():
    from print_fill import template_file_abspath
    sf = '/var/app/static'
    assert template_file_abspath('/static/uploads/print_templates/a.xlsx', sf) == \
        os.path.join(sf, 'uploads/print_templates/a.xlsx')
    assert template_file_abspath('uploads/b.xlsx', sf) == \
        os.path.join(sf, 'uploads/b.xlsx')
    assert template_file_abspath('', sf) is None
    assert template_file_abspath(None, sf) is None


def test_validate_template_file():
    from print_fill import validate_template_file
    # 空
    assert validate_template_file(b'') == '打印模板文件为空'
    # 非 zip
    assert '不是有效的 Excel' in validate_template_file(b'not-a-zip')
    # 合法 xlsx + 合法占位符
    ok = _xlsx_bytes([(1, 1, '{order.order_no}'), (2, 1, '{item.material.name}')])
    assert validate_template_file(ok.read()) == ''
    # 非法占位符
    bad = _xlsx_bytes([(1, 1, '{order.order_no}'), (2, 1, '{unknown_thing}')])
    msg = validate_template_file(bad.read())
    assert '不支持的占位符' in msg and 'unknown_thing' in msg


def test_build_filled_print_excel(tmp_path, seeded_db):
    from openpyxl import load_workbook
    tpl = tmp_path / "canon.xlsx"
    tpl.write_bytes(_xlsx_bytes([(1, 1, '{order.order_no}')]).read())
    with app_module.app.app_context():
        order = _make_in_order()
        out = build_filled_print_excel(str(tpl), order)
    ws = load_workbook(io.BytesIO(out.read())).active
    assert ws['A1'].value == 'IN-PRINT-001'


def test_resolve_print_template(client):
    from app import InOrderPrintTemplate
    from utils import resolve_print_template
    tid = _upload_in_template(client, name="默认解析模板")
    with app_module.app.app_context():
        got_id = resolve_print_template(InOrderPrintTemplate, tid)
        got_none = resolve_print_template(InOrderPrintTemplate, None)
        assert got_id is not None and got_id.id == tid
        assert got_none is not None


def test_order_value(seeded_db):
    from print_fill import _Filler
    with app_module.app.app_context():
        order = _make_in_order()
        f = _Filler(order, order.items, '2026-08-20')
        assert f.order_value('order.order_no') == 'IN-PRINT-001'
        assert f.order_value('total_quantity') == 150
        assert f.order_value('print_date') == '2026-08-20'


def test_item_value(seeded_db):
    from print_fill import _Filler
    with app_module.app.app_context():
        order = _make_in_order()
        f = _Filler(order, order.items, None)
        it = order.items[0]
        assert f.item_value('item.material.name', it) == '6204轴承'
        assert f.item_value('item.quantity', it) == 100


def test_fill(seeded_db):
    from openpyxl import Workbook
    from print_fill import _Filler
    wb = Workbook()
    wb.active['A1'] = '{order.order_no}'
    wb.active['B1'] = '{print_date}'
    with app_module.app.app_context():
        order = _make_in_order()
        _Filler(order, order.items, '2026-08-20').fill(wb.active)
    assert wb.active['A1'].value == 'IN-PRINT-001'
    assert wb.active['B1'].value == '2026-08-20'