# -*- coding: utf-8 -*-
"""打印模板在线编辑（PRINT-TEMPLATE-F02）回归测试。

覆盖：
- GET /{prefix}_print_template/<id>/grid 读取模板网格 JSON
- POST /{prefix}_print_template/<id>/grid 回写单元格值并落盘
- 非法占位符回写被 400 拒绝且文件不变（原子性）
- 不存在的工作表名 400
- GET /{prefix}_print_template/<id>/edit 编辑器页面渲染 200
"""
from __future__ import annotations

import io
import json
import os
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["WMS_DEBUG"] = "0"

from werkzeug.security import generate_password_hash  # noqa: E402

import app as app_module  # noqa: E402
from app import InOrderPrintTemplate, Unit, User, db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


@pytest.fixture()
def client():
    app_module.app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
    with app_module.app.app_context():
        db.drop_all()
        db.create_all()
        db.session.add_all([
            Unit(name="个", code="PCS"),
            User(username="admin",
                 password_hash=generate_password_hash("admin"),
                 role="admin", must_change_password=False),
        ])
        db.session.commit()
    c = app_module.app.test_client()
    c.post("/login", data={"username": "admin", "password": "admin"},
           content_type="application/x-www-form-urlencoded")
    return c


def _xlsx_bytes(cells=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.cell(2, 1, "单号：{order.order_no}")
    ws.cell(3, 1, "{item.material.name}")
    for (r, c, v) in (cells or []):
        ws.cell(r, c, v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _upload_template(client):
    resp = client.post('/in_order_print_template/add', data={
        'name': '在线编辑测试模板',
        'template_type': 'excel',
        'excel_file': (_xlsx_bytes(), '在线编辑模板.xlsx'),
    }, content_type='multipart/form-data')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    payload = json.loads(resp.get_data(as_text=True))
    assert payload['status'] == 'success'
    return payload['id']


def _grid_path(template_id):
    return f'/in_order_print_template/{template_id}/grid'


def _grid_read(client, template_id):
    return client.get(_grid_path(template_id))


def _grid_write(client, template_id, sheets):
    return client.post(_grid_path(template_id),
                       data=json.dumps({'sheets': sheets}),
                       content_type='application/json')


def test_grid_read_returns_sheet_cells(client):
    template_id = _upload_template(client)
    resp = _grid_read(client, template_id)
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload['status'] == 'success'
    assert payload['template']['name'] == '在线编辑测试模板'
    sheets = payload['sheets']
    assert len(sheets) == 1
    text = ' '.join(str(cell['value']) for cell in sheets[0]['cells'])
    assert '采购入库单' in text
    assert '{order.order_no}' in text
    assert '{item.material.name}' in text


def test_edit_page_renders(client):
    template_id = _upload_template(client)
    resp = client.get(f'/in_order_print_template/{template_id}/edit')
    assert resp.status_code == 200
    text = resp.get_data(as_text=True)
    assert '在线编辑打印模板' in text


def test_grid_write_updates_file(client):
    template_id = _upload_template(client)
    resp = _grid_write(client, template_id, [
        {'name': '模板', 'upserts': [[1, 1, '采购入库单（已改）']], 'del_rows': []},
    ])
    assert resp.status_code == 200, resp.get_data(as_text=True)
    assert resp.get_json()['status'] == 'success'

    # 重新读取应显示新值
    payload = _grid_read(client, template_id).get_json()
    values = [cell['value'] for cell in payload['sheets'][0]['cells']]
    assert '采购入库单（已改）' in values
    assert '采购入库单' not in values


def test_grid_write_illegal_placeholder_rejected_file_unchanged(client):
    template_id = _upload_template(client)
    resp = _grid_write(client, template_id, [
        {'name': '模板', 'upserts': [[1, 2, '{bad.token}']], 'del_rows': []},
    ])
    assert resp.status_code == 400
    assert '不支持的占位符' in resp.get_json()['msg']

    payload = _grid_read(client, template_id).get_json()
    values = [cell['value'] for cell in payload['sheets'][0]['cells']]
    assert '{bad.token}' not in values
    assert '采购入库单' in values


def test_grid_write_unknown_sheet_rejected(client):
    template_id = _upload_template(client)
    resp = _grid_write(client, template_id, [
        {'name': '不存在的表', 'upserts': [[1, 1, 'x']], 'del_rows': []},
    ])
    assert resp.status_code == 400
    assert '不存在' in resp.get_json()['msg']


def test_grid_write_delete_row(client):
    template_id = _upload_template(client)
    # 追加第 4 行数据占位，然后删除该行
    resp = _grid_write(client, template_id, [
        {'name': '模板', 'upserts': [[4, 1, '{print_date}']], 'del_rows': [4]},
    ])
    assert resp.status_code == 200, resp.get_data(as_text=True)
    payload = _grid_read(client, template_id).get_json()
    rows = {cell['row'] for cell in payload['sheets'][0]['cells']}
    assert 4 not in rows


# ---------- 引擎级单元测试（满足 A9：serialize/apply 直接覆盖） ----------

def _write_workbook(path, cells):
    from openpyxl import load_workbook
    with open(path, 'wb') as f:
        f.write(_xlsx_bytes(cells).read())
    wb = load_workbook(path)
    ws = wb.active
    # 制造一个合并区域：C1:E1
    ws.merge_cells('C1:E1')
    wb.save(path)
    wb.close()


def test_serialize_print_template_grid(tmp_path):
    from print_fill import serialize_print_template_grid
    p = str(tmp_path / 't.xlsx')
    _write_workbook(p, [])
    grid = serialize_print_template_grid(p)
    sheets = grid['sheets']
    assert len(sheets) == 1
    assert sheets[0]['name'] == '模板'
    text = ' '.join(str(c['value']) for c in sheets[0]['cells'])
    assert '采购入库单' in text
    # 合并区域 C1:E1 应被序列化为锚点
    anchors = [m for m in sheets[0]['merges']
               if m['row'] == 1 and m['col'] == 3]
    assert anchors and anchors[0]['colspan'] == 3


def test_apply_print_template_grid(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / 't.xlsx')
    _write_workbook(p, [])
    out = apply_print_template_grid(p, [
        {'name': '模板',
         'upserts': [[1, 1, '已编辑标题'], [2, 1, '{order.supplier_name}']],
         'del_rows': [3]},
    ])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(1, 1).value == '已编辑标题'
    assert ws.cell(2, 1).value == '{order.supplier_name}'
    wb.close()


def test_register_print_template_editor_routes(tmp_path):
    # A9：注册函数本身需要对应测试——验证入库/出库两组路由均被挂载。
    from routes.print_template_editor import register_print_template_editor_routes
    from app import app as _app
    urls = {str(r.rule) for r in _app.url_map.iter_rules()}
    for prefix in ('in_order', 'out_order'):
        assert f'/{prefix}_print_template/<int:template_id>/edit' in urls
        assert f'/{prefix}_print_template/<int:template_id>/grid' in urls