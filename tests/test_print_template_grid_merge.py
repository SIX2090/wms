# -*- coding: utf-8 -*-
"""打印模板网格合并区回写（PRINT-TEMPLATE-F05-A3）测试。

覆盖：
- merges/unmerges roundtrip
- 新增合并与现存/互相互交 → 400（不落盘）
- del_rows 与合并区相交 → 400（防 openpyxl delete_rows 不平移合并区的坏文件）
- 删除合并区上方行后，新增合并按平移落到正确行号
- 越界/单格合并 → 400
- 路由级：合并保存后重新读取 anchors 正确；非法合并 400 且文件不变
"""
from __future__ import annotations

import io
import json
import os
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)


def _base_workbook(path):
    """6 行 4 列：A1:C1 标题合并 + 若干占位行。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.merge_cells("A1:C1")
    ws.cell(2, 1, "单号：{order.order_no}")
    ws.cell(3, 1, "{item.material.name}")
    ws.cell(4, 1, "备用行1")
    ws.cell(5, 1, "备用行2")
    ws.cell(6, 1, "打印日期：{print_date}")
    wb.save(path)
    wb.close()


def _merged_ranges(ws):
    return sorted((r.min_row, r.min_col, r.max_row, r.max_col)
                  for r in ws.merged_cells.ranges)


def test_apply_merge_roundtrip(tmp_path):
    from openpyxl import load_workbook
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    out = apply_print_template_grid(p, [{
        "name": "模板",
        "merges": [{"row": 4, "col": 1, "rowspan": 2, "colspan": 2}],
    }])
    wb = load_workbook(out)
    ranges = _merged_ranges(wb.active)
    assert (1, 1, 1, 3) in ranges          # 原合并保留
    assert (4, 1, 5, 2) in ranges          # 新合并生效
    wb.close()


def test_apply_unmerge_roundtrip(tmp_path):
    from openpyxl import load_workbook
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    out = apply_print_template_grid(p, [{
        "name": "模板",
        "unmerges": [[1, 1]],
    }])
    wb = load_workbook(out)
    assert _merged_ranges(wb.active) == []
    wb.close()


def test_apply_unmerge_invalid_anchor_rejected(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    with pytest.raises(ValueError, match="左上角"):
        apply_print_template_grid(p, [{"name": "模板", "unmerges": [[2, 2]]}])


def test_apply_merge_overlap_existing_rejected(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    with pytest.raises(ValueError, match="重叠"):
        apply_print_template_grid(p, [{
            "name": "模板",
            "merges": [{"row": 1, "col": 2, "rowspan": 2, "colspan": 2}],
        }])


def test_apply_merge_overlap_each_other_rejected(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    with pytest.raises(ValueError, match="互相重叠"):
        apply_print_template_grid(p, [{
            "name": "模板",
            "merges": [
                {"row": 3, "col": 1, "rowspan": 2, "colspan": 2},
                {"row": 4, "col": 2, "rowspan": 2, "colspan": 2},
            ],
        }])


def test_apply_merge_out_of_range_rejected(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    with pytest.raises(ValueError, match="超出上限"):
        apply_print_template_grid(p, [{
            "name": "模板",
            "merges": [{"row": 1, "col": 99, "rowspan": 1, "colspan": 5}],
        }])
    with pytest.raises(ValueError, match="至少跨 2 格"):
        apply_print_template_grid(p, [{
            "name": "模板",
            "merges": [{"row": 3, "col": 3, "rowspan": 1, "colspan": 1}],
        }])


def test_del_rows_intersect_merge_rejected(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    with pytest.raises(ValueError, match="合并单元格"):
        apply_print_template_grid(p, [{"name": "模板", "del_rows": [1]}])
    with pytest.raises(ValueError, match="新增合并区域重叠"):
        apply_print_template_grid(p, [{
            "name": "模板",
            "del_rows": [4],
            "merges": [{"row": 4, "col": 1, "rowspan": 2, "colspan": 2}],
        }])


def test_merge_shifted_after_row_delete(tmp_path):
    """删除合并区上方第 2 行后，原第 4-5 行的新合并应平移到第 3-4 行。"""
    from openpyxl import load_workbook
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    out = apply_print_template_grid(p, [{
        "name": "模板",
        "del_rows": [2],
        "merges": [{"row": 4, "col": 1, "rowspan": 2, "colspan": 2}],
    }])
    wb = load_workbook(out)
    ranges = _merged_ranges(wb.active)
    assert (1, 1, 1, 3) in ranges
    assert (3, 1, 4, 2) in ranges
    wb.close()


def test_merge_atomic_on_failure_file_unchanged(tmp_path):
    p = str(tmp_path / "t.xlsx")
    _base_workbook(p)
    before = os.path.getsize(p)
    from print_fill import apply_print_template_grid
    with pytest.raises(ValueError):
        apply_print_template_grid(p, [{
            "name": "模板",
            "unmerges": [[1, 1]],
            "merges": [{"row": 1, "col": 1, "rowspan": 1, "colspan": 2},
                       {"row": 1, "col": 2, "rowspan": 2, "colspan": 2}],
        }])
    assert os.path.getsize(p) == before


# ---------- 路由级 ----------

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["WMS_DEBUG"] = "0"

from werkzeug.security import generate_password_hash  # noqa: E402

import app as app_module  # noqa: E402
from app import Unit, User, db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


@pytest.fixture()
def client():
    app_module.app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
    with app_module.app.app_context():
        db.drop_all()
        db.create_all()
        db.session.add_all([
            Unit(name="个", code="PCS"),
            User(username="admin",
                 password_hash=generate_password_hash("admin"),
                 role="admin", must_change_password=False),
        ])
        db.session.commit()
    c = app_module.app.test_client()
    c.post("/login", data={"username": "admin", "password": "admin"},
           content_type="application/x-www-form-urlencoded")
    return c


def _upload_template(client):
    buf = io.BytesIO()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.merge_cells("A1:C1")
    ws.cell(2, 1, "单号：{order.order_no}")
    ws.cell(3, 1, "{item.material.name}")
    ws.cell(4, 1, "备用")
    wb.save(buf)
    buf.seek(0)
    resp = client.post('/in_order_print_template/add', data={
        'name': '合并回写测试模板',
        'template_type': 'excel',
        'excel_file': (buf, '合并模板.xlsx'),
    }, content_type='multipart/form-data')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    return resp.get_json()['id']


def test_route_grid_write_merge_roundtrip(client):
    template_id = _upload_template(client)
    resp = client.post(f'/in_order_print_template/{template_id}/grid',
                       data=json.dumps({'sheets': [{
                           'name': '模板',
                           'merges': [{'row': 3, 'col': 2,
                                       'rowspan': 2, 'colspan': 2}],
                       }]}), content_type='application/json')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    payload = client.get(
        f'/in_order_print_template/{template_id}/grid').get_json()
    merges = payload['sheets'][0]['merges']
    assert {'row': 3, 'col': 2, 'rowspan': 2, 'colspan': 2} in merges
    assert {'row': 1, 'col': 1, 'rowspan': 1, 'colspan': 3} in merges


def test_route_grid_write_invalid_merge_400(client):
    template_id = _upload_template(client)
    resp = client.post(f'/in_order_print_template/{template_id}/grid',
                       data=json.dumps({'sheets': [{
                           'name': '模板',
                           'merges': [{'row': 1, 'col': 1,
                                       'rowspan': 2, 'colspan': 2}],
                       }]}), content_type='application/json')
    assert resp.status_code == 400
    # 未落盘：合并区仍只有 A1:C1
    payload = client.get(
        f'/in_order_print_template/{template_id}/grid').get_json()
    merges = payload['sheets'][0]['merges']
    assert merges == [{'row': 1, 'col': 1, 'rowspan': 1, 'colspan': 3}]
