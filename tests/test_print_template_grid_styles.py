# -*- coding: utf-8 -*-
"""打印模板网格样式序列化（PRINT-TEMPLATE-F05-A1）测试。

覆盖 serialize_print_template_grid 的样式/尺寸输出：
- 字体（加粗/斜体/下划线/字体名/字号/颜色）、背景色、对齐、自动换行、边框
- 空值但带边框的单元格（合计行边框格）也会输出
- col_widths / row_heights 仅输出显式设置的尺寸
- 合并锚点单元格同样携带 style
- theme/indexed 颜色不输出
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)


def _styled_workbook(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.cell(1, 1).font = Font(name="微软雅黑", size=16, bold=True,
                              color="FFFF0000")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="FFFFF2CC")
    thin = Side(style="thin")
    # 空值但带边框的合计行单元格
    ws.cell(3, 2).border = Border(top=thin, right=thin, bottom=thin, left=thin)
    ws.cell(4, 1, "{item.material.name}")
    ws.cell(4, 1).font = Font(italic=True, underline="single")
    ws.merge_cells("A1:C1")
    ws.column_dimensions["B"].width = 20.5
    ws.row_dimensions[1].height = 30.0
    wb.save(path)
    wb.close()


def _cell_map(sheet):
    return {(c["row"], c["col"]): c for c in sheet["cells"]}


def test_serialize_outputs_font_fill_alignment(tmp_path):
    from print_fill import serialize_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _styled_workbook(p)
    grid = serialize_print_template_grid(p)
    sheet = grid["sheets"][0]
    cells = _cell_map(sheet)
    title = cells[(1, 1)]
    assert title["merged"] is True
    style = title["style"]
    assert style["bold"] is True
    assert style["font_name"] == "微软雅黑"
    assert style["font_size"] == 16.0
    assert style["font_color"] == "#FF0000"
    assert style["bg_color"] == "#FFF2CC"
    assert style["h_align"] == "center"
    assert style["v_align"] == "center"
    assert style["wrap"] is True


def test_serialize_outputs_empty_cell_with_border(tmp_path):
    from print_fill import serialize_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _styled_workbook(p)
    sheet = serialize_print_template_grid(p)["sheets"][0]
    cells = _cell_map(sheet)
    # 空值但带边框的单元格必须出现在网格里（否则边框样式丢失）
    assert (3, 2) in cells
    border = cells[(3, 2)]["style"]["border"]
    assert border == {"top": "thin", "right": "thin",
                      "bottom": "thin", "left": "thin"}
    item = cells[(4, 1)]
    assert item["style"]["italic"] is True
    assert item["style"]["underline"] is True


def test_serialize_outputs_dimensions(tmp_path):
    from print_fill import serialize_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _styled_workbook(p)
    sheet = serialize_print_template_grid(p)["sheets"][0]
    assert sheet["col_widths"] == {"2": 20.5}
    assert sheet["row_heights"] == {"1": 30.0}


def test_serialize_skips_theme_color_and_default_style(tmp_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles.colors import Color
    from print_fill import serialize_print_template_grid
    p = str(tmp_path / "t2.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "纯文本无样式")
    ws.cell(2, 1, "主题色")
    ws.cell(2, 1).font = Font(color=Color(theme=1))
    wb.save(p)
    wb.close()
    sheet = serialize_print_template_grid(p)["sheets"][0]
    cells = _cell_map(sheet)
    # 默认样式单元格不输出 style 键
    assert "style" not in cells[(1, 1)]
    # theme 颜色不输出 font_color
    assert "font_color" not in cells[(2, 1)].get("style", {})
    assert sheet["col_widths"] == {}
    assert sheet["row_heights"] == {}


# ==================== F05-A2：样式/尺寸回写（apply + 路由） ====================

def _plain_workbook(path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.cell(2, 1, "单号：{order.order_no}")
    ws.cell(3, 1, "{item.material.name}")
    wb.save(path)
    wb.close()


def test_apply_styles_roundtrip(tmp_path):
    from openpyxl import load_workbook
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _plain_workbook(p)
    out = apply_print_template_grid(p, [{
        "name": "模板",
        "styles": [
            [1, 1, {"bold": True, "font_size": 18, "font_color": "#FF0000",
                    "bg_color": "#FFF2CC", "h_align": "center",
                    "v_align": "center", "wrap": True,
                    "font_name": "微软雅黑"}],
            [3, 1, {"italic": True, "underline": True,
                    "border": {"top": "thin", "bottom": "double"}}],
        ],
        "col_widths": {"2": 22.5},
        "row_heights": {"1": 36.0},
    }])
    wb = load_workbook(out)
    ws = wb.active
    c = ws.cell(1, 1)
    assert c.font.bold is True
    assert c.font.size == 18
    assert c.font.color.rgb == "FFFF0000"
    assert c.font.name == "微软雅黑"
    assert c.fill.patternType == "solid"
    assert c.fill.fgColor.rgb == "FFFFF2CC"
    assert c.alignment.horizontal == "center"
    assert c.alignment.vertical == "center"
    assert c.alignment.wrap_text is True
    c3 = ws.cell(3, 1)
    assert c3.font.italic is True
    assert c3.font.underline == "single"
    assert c3.border.top.style == "thin"
    assert c3.border.bottom.style == "double"
    assert c3.border.left.style is None
    assert ws.column_dimensions["B"].width == 22.5
    assert ws.row_dimensions[1].height == 36.0
    wb.close()


def test_apply_styles_clear_semantics(tmp_path):
    from openpyxl import load_workbook
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _styled_workbook(p)  # A1: bold/font_size=16/font_color=#FF0000 于 (1,1)
    out = apply_print_template_grid(p, [{
        "name": "模板",
        "styles": [[1, 1, {"bold": None, "font_color": None, "italic": True}]],
    }])
    wb = load_workbook(out)
    c = wb.active.cell(1, 1)
    assert not c.font.bold            # None=清除
    assert c.font.color is None or c.font.color.rgb in (None, "00000000")
    assert c.font.italic is True      # 有值=设置
    assert c.font.size == 16          # 缺失=保持
    wb.close()


def test_apply_styles_invalid_rejected_file_unchanged(tmp_path):
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _plain_workbook(p)
    import os as _os
    before = _os.path.getsize(p)
    bad_cases = [
        [1, 1, {"font_size": 100}],                    # 字号超上限
        [1, 1, {"font_color": "red"}],                 # 非法颜色
        [1, 1, {"h_align": "middle"}],                 # 非法对齐
        [1, 1, {"border": {"top": "ultra"}}],          # 非法边样式
        [1, 1, {"unknown_key": 1}],                    # 非法样式键
        [9999, 1, {"bold": True}],                     # 坐标超限
    ]
    import pytest
    for bad in bad_cases:
        with pytest.raises(ValueError):
            apply_print_template_grid(p, [{"name": "模板", "styles": [bad]}])
    # 尺寸校验
    with pytest.raises(ValueError):
        apply_print_template_grid(p, [{"name": "模板", "col_widths": {"1": 500}}])
    with pytest.raises(ValueError):
        apply_print_template_grid(p, [{"name": "模板", "row_heights": {"1": -3}}])
    # 文件未被改动
    assert _os.path.getsize(p) == before


def test_apply_legacy_payload_compatible(tmp_path):
    from openpyxl import load_workbook
    from print_fill import apply_print_template_grid
    p = str(tmp_path / "t.xlsx")
    _plain_workbook(p)
    out = apply_print_template_grid(p, [
        {"name": "模板", "upserts": [[1, 1, "改"]], "del_rows": []},
    ])
    wb = load_workbook(out)
    assert wb.active.cell(1, 1).value == "改"
    wb.close()


# ---------- 路由级：pydantic 400 + 落盘 roundtrip ----------

import io  # noqa: E402
import json  # noqa: E402

import pytest  # noqa: E402

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["WMS_DEBUG"] = "0"

from werkzeug.security import generate_password_hash  # noqa: E402

import app as app_module  # noqa: E402
from app import Unit, User, db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


@pytest.fixture()
def client():
    app_module.app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
    with app_module.app.app_context():
        db.drop_all()
        db.create_all()
        db.session.add_all([
            Unit(name="个", code="PCS"),
            User(username="admin",
                 password_hash=generate_password_hash("admin"),
                 role="admin", must_change_password=False),
        ])
        db.session.commit()
    c = app_module.app.test_client()
    c.post("/login", data={"username": "admin", "password": "admin"},
           content_type="application/x-www-form-urlencoded")
    return c


def _upload_template(client):
    buf = io.BytesIO()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.cell(2, 1, "单号：{order.order_no}")
    wb.save(buf)
    buf.seek(0)
    resp = client.post('/in_order_print_template/add', data={
        'name': '样式回写测试模板',
        'template_type': 'excel',
        'excel_file': (buf, '样式模板.xlsx'),
    }, content_type='multipart/form-data')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    return resp.get_json()['id']


def test_route_grid_write_styles_roundtrip(client):
    template_id = _upload_template(client)
    resp = client.post(f'/in_order_print_template/{template_id}/grid',
                       data=json.dumps({'sheets': [{
                           'name': '模板',
                           'styles': [[1, 1, {'bold': True, 'font_size': 20,
                                              'h_align': 'center'}]],
                           'col_widths': {'1': 30},
                           'row_heights': {'1': 40},
                       }]}), content_type='application/json')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    payload = client.get(
        f'/in_order_print_template/{template_id}/grid').get_json()
    sheet = payload['sheets'][0]
    cells = {(c['row'], c['col']): c for c in sheet['cells']}
    style = cells[(1, 1)]['style']
    assert style['bold'] is True
    assert style['font_size'] == 20.0
    assert style['h_align'] == 'center'
    assert sheet['col_widths'] == {'1': 30.0}
    assert sheet['row_heights'] == {'1': 40.0}


def test_route_grid_write_invalid_style_400(client):
    template_id = _upload_template(client)
    resp = client.post(f'/in_order_print_template/{template_id}/grid',
                       data=json.dumps({'sheets': [{
                           'name': '模板',
                           'styles': [[1, 1, {'font_size': 999}]],
                       }]}), content_type='application/json')
    assert resp.status_code == 400
    # 非法样式不落盘：重新读取应无 bold/size 样式
    payload = client.get(
        f'/in_order_print_template/{template_id}/grid').get_json()
    cells = {(c['row'], c['col']): c
             for c in payload['sheets'][0]['cells']}
    assert 'style' not in cells.get((1, 1), {})


def test_to_engine_payload():
    """A9：to_engine_payload 必须保留「缺失=保持 / 显式 null=清除」语义。"""
    from routes.print_template_editor import PrintTemplateGridRequest
    req = PrintTemplateGridRequest.model_validate({'sheets': [{
        'name': '模板',
        'styles': [
            [1, 1, {'bold': True}],          # 只给 bold → 其余键不出现
            [2, 1, {'font_color': None}],    # 显式 null → 保留 None（清除）
        ],
        'col_widths': {'3': 18},
    }]})
    payload = req.sheets[0].to_engine_payload()
    assert payload['name'] == '模板'
    assert payload['styles'][0][2] == {'bold': True}
    assert payload['styles'][1][2] == {'font_color': None}
    assert payload['col_widths'] == {'3': 18.0}
    assert payload['row_heights'] == {}
    assert payload['upserts'] == [] and payload['del_rows'] == []
