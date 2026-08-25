# -*- coding: utf-8 -*-
"""PRINT-TEMPLATE-F04（A2）回归测试：图片占位符引擎 + 标签 Excel 模板。

需求（2026-08-25）：所有打印模板（含标签）支持 Excel 在线编辑，参考简道云
打印模板。标签模板需要条码/二维码图片占位符：
  {img_barcode:item.barcode} / {img_qrcode:item.code} / {img_barcode:order.order_no}

测试用例：
  T1. 明细级条码占位符：填充后嵌入图片、占位符文本清除、文本字段正常填充
  T2. 明细级二维码占位符 + 订单级条码占位符（纯订单模板行）
  T3. 条码数据为空/非法时回退数据文本，不输出占位符原文、不抛异常
  T4. 上传/在线编辑白名单：图片占位符合法，畸形 {img_barcode:foo} 拒绝
  T5. 内置物料标签模板通过安全校验且含图片占位符与明细占位符行
  T6. render_label_excel_print 填充 dict 行并嵌入条码图片
  T7. 零明细时图片占位符被清空、不残留占位符
"""
from __future__ import annotations

import io
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

APP_DIR = Path(__file__).resolve().parent.parent / "app"
sys.path.insert(0, str(APP_DIR))

import doc_print_excel as dpe  # noqa: E402
from print_fill import (  # noqa: E402
    _Filler,
    apply_print_template_grid,
    build_filled_print_excel,
    validate_template_file,
)


def _make_template(cells):
    """按 {(row, col): value} 构造模板文件，返回路径。"""
    wb = Workbook()
    ws = wb.active
    ws.title = '标签'
    for (r, c), value in cells.items():
        ws.cell(r, c, value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_tmp(tmp_path, raw):
    path = tmp_path / 'tmpl.xlsx'
    path.write_bytes(raw)
    return str(path)


# -------------------- T1：明细级条码占位符 --------------------

def test_item_barcode_image_embedded(tmp_path):
    raw = _make_template({
        (1, 1): '物料编码', (1, 2): '条码',
        (2, 1): '{item.code}', (2, 2): '{img_barcode:item.barcode}',
    })
    path = _write_tmp(tmp_path, raw)
    order = SimpleNamespace(total_amount=0)
    items = [SimpleNamespace(code='M-001', barcode='6901234567890'),
             SimpleNamespace(code='M-002', barcode='6901234567891')]
    output = build_filled_print_excel(path, order, items=items,
                                      date_str='2026-08-25')
    wb = load_workbook(output)
    ws = wb.active
    assert len(ws._images) == 2  # 每个物料一张条码图
    texts = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert not any('{img_barcode' in t for t in texts)
    assert not any('{item.' in t for t in texts)
    assert 'M-001' in texts and 'M-002' in texts


def test_item_qrcode_and_order_barcode(tmp_path):
    raw = _make_template({
        (1, 1): '单号条码：', (1, 2): '{img_barcode:order.order_no}',
        (2, 1): '{item.code}', (2, 2): '{img_qrcode:item.code}',
    })
    path = _write_tmp(tmp_path, raw)
    order = SimpleNamespace(order_no='RK-2026-0001', total_amount=0)
    items = [SimpleNamespace(code='M-001')]
    output = build_filled_print_excel(path, order, items=items,
                                      date_str='2026-08-25')
    wb = load_workbook(output)
    ws = wb.active
    assert len(ws._images) == 2  # 订单条码 1 张 + 明细二维码 1 张
    texts = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert not any('{img_' in t for t in texts)


# -------------------- T3：空数据回退 --------------------

def test_empty_barcode_falls_back_to_text(tmp_path):
    raw = _make_template({
        (1, 1): '{item.code}', (1, 2): '{img_barcode:item.barcode}',
    })
    path = _write_tmp(tmp_path, raw)
    order = SimpleNamespace(total_amount=0)
    items = [SimpleNamespace(code='M-003', barcode='')]
    output = build_filled_print_excel(path, order, items=items,
                                      date_str='2026-08-25')
    wb = load_workbook(output)
    ws = wb.active
    assert len(ws._images) == 0
    texts = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert not any('{img_barcode' in t for t in texts)


def test_zero_items_clears_image_placeholder(tmp_path):
    raw = _make_template({
        (1, 1): '{item.code}', (1, 2): '{img_barcode:item.barcode}',
    })
    path = _write_tmp(tmp_path, raw)
    order = SimpleNamespace(total_amount=0)
    output = build_filled_print_excel(path, order, items=[],
                                      date_str='2026-08-25')
    wb = load_workbook(output)
    ws = wb.active
    assert len(ws._images) == 0
    texts = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert not any('{img_' in t or '{item.' in t for t in texts)


# -------------------- T4：白名单 --------------------

def test_validate_template_file_accepts_image_placeholders(tmp_path):
    raw = _make_template({
        (1, 1): '{item.code}', (1, 2): '{img_barcode:item.barcode}',
        (1, 3): '{img_qrcode:order.order_no}',
    })
    assert validate_template_file(raw) == ''


def test_validate_template_file_rejects_bad_image_placeholder(tmp_path):
    raw = _make_template({(1, 1): '{img_barcode:foo}'})
    assert validate_template_file(raw) != ''
    raw2 = _make_template({(1, 1): '{img_svg:item.code}'})
    assert validate_template_file(raw2) != ''


def test_grid_editor_accepts_image_placeholder(tmp_path):
    raw = _make_template({(1, 1): '占位'})
    path = _write_tmp(tmp_path, raw)
    output = apply_print_template_grid(path, [{
        'name': '标签',
        'upserts': [[1, 1, '{img_barcode:item.barcode}']],
        'del_rows': [],
    }])
    wb = load_workbook(output)
    assert wb.active.cell(1, 1).value == '{img_barcode:item.barcode}'
    with pytest.raises(ValueError):
        apply_print_template_grid(path, [{
            'name': '标签',
            'upserts': [[1, 1, '{img_barcode:evil}']],
            'del_rows': [],
        }])


# -------------------- T5：内置物料标签模板 --------------------

def test_builtin_material_label_template():
    content = dpe.generate_builtin_template('material_label')
    assert content is not None
    raw = content.read()
    assert validate_template_file(raw) == ''
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert any('{img_barcode:item.barcode}' in t for t in texts)
    assert any('{item.code}' in t for t in texts)


# -------------------- T6：render_label_excel_print --------------------

def test_render_label_excel_print(tmp_path, monkeypatch):
    monkeypatch.setattr(dpe, 'resolve_excel_template', lambda *a, **k: None)
    rows = [
        {'code': 'M-001', 'name': '轴承', 'spec': '6204', 'unit_name': '套',
         'category_name': '轴承类', 'supplier_name': '鑫达', 'price': 12.5,
         'stock': 100, 'barcode': '6901234567890'},
        SimpleNamespace(code='M-002', name='螺母', spec='M8', unit_name='个',
                        category_name='紧固件', supplier_name='鑫达',
                        price=0.5, stock=500, barcode='6901234567891'),
    ]
    result = dpe.render_label_excel_print(
        'material_label', rows, static_folder=str(tmp_path),
        date_str='2026-08-25')
    assert result is not None
    output, filename = result
    assert '物料标签' in filename and filename.endswith('.xlsx')
    wb = load_workbook(output)
    ws = wb.active
    assert len(ws._images) == 2
    texts = [c.value for row in ws.iter_rows() for c in row
             if c.value is not None]
    assert 'M-001' in texts and '轴承' in texts and '鑫达' in texts
    assert 'M-002' in texts and '螺母' in texts
    assert not any(isinstance(t, str) and '{item.' in t for t in texts)


def test_render_label_excel_print_unknown_code(tmp_path):
    assert dpe.render_label_excel_print(
        'nope', [], static_folder=str(tmp_path)) is None
