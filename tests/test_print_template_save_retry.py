# -*- coding: utf-8 -*-
"""打印模板在线保存的「目标被占用」重试回归测试（AI-WMS-REPLACE-RETRY）。

背景：Windows 上 os.replace 会因目标 .xlsx 仍被占用而抛 [WinError 5] 拒绝访问。
占用者通常是还没被 GC 回收的 openpyxl workbook —— 部分模板读取路径没有显式
close()，句柄只能等回收才释放。首次 replace 失败并不代表文件真的有问题，
但旧实现会直接回滚，管理员看到的是「保存失败，请稍后重试」。

_grid_write 现在的做法是：先 gc.collect() 释放句柄，再按指数退避重试，
只有持续被占用才当作真实错误抛出。本测试用 monkeypatch 模拟占用，锁住两点：
  1. 首次 WinError 5 不会让保存失败（重试后成功）
  2. 非占用类错误（winerror 不在 5/13/32 白名单内）不被吞掉，仍如实报 500
"""
from __future__ import annotations

import io
import json
import os
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["WMS_DEBUG"] = "0"

from werkzeug.security import generate_password_hash  # noqa: E402

import app as app_module  # noqa: E402
from app import Unit, User, db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


@pytest.fixture()
def client():
    app_module.app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
    with app_module.app.app_context():
        db.drop_all()
        db.create_all()
        db.session.add_all([
            Unit(name="个", code="PCS"),
            User(username="admin",
                 password_hash=generate_password_hash("admin"),
                 role="admin", must_change_password=False),
        ])
        db.session.commit()
    c = app_module.app.test_client()
    c.post("/login", data={"username": "admin", "password": "admin"},
           content_type="application/x-www-form-urlencoded")
    return c


def _xlsx_bytes():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.cell(1, 1, "采购入库单")
    ws.cell(2, 1, "单号：{order.order_no}")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _upload_template(client):
    resp = client.post('/in_order_print_template/add', data={
        'name': '占用重试测试模板',
        'template_type': 'excel',
        'excel_file': (_xlsx_bytes(), '占用重试模板.xlsx'),
    }, content_type='multipart/form-data')
    assert resp.status_code == 200, resp.get_data(as_text=True)
    payload = json.loads(resp.get_data(as_text=True))
    assert payload['status'] == 'success'
    return payload['id']


def _grid_write(client, template_id):
    return client.post(
        f'/in_order_print_template/{template_id}/grid',
        data=json.dumps({'sheets': [
            {'name': '模板', 'upserts': [[1, 1, '采购入库单（已改）']], 'del_rows': []},
        ]}),
        content_type='application/json')


def test_grid_write_retries_when_target_locked_then_succeeds(client, monkeypatch):
    """首次 WinError 5（目标被占用）后应重试成功，而不是让用户看到保存失败。"""
    real_replace = os.replace
    calls = {'n': 0}

    def fake_replace(src, dst, *args, **kwargs):
        if 'print_templates' not in str(dst):
            return real_replace(src, dst, *args, **kwargs)
        calls['n'] += 1
        if calls['n'] == 1:
            err = PermissionError(5, '拒绝访问。')
            err.winerror = 5
            raise err
        return real_replace(src, dst, *args, **kwargs)

    monkeypatch.setattr(os, 'replace', fake_replace)
    template_id = _upload_template(client)
    resp = _grid_write(client, template_id)
    assert resp.status_code == 200, resp.get_data(as_text=True)
    assert resp.get_json()['status'] == 'success'
    # 至少尝试了两次：第一次被占用，重试后成功
    assert calls['n'] >= 2


def test_grid_write_does_not_swallow_non_lock_error(client, monkeypatch):
    """非占用类 OSError 不能被重试逻辑吞掉，必须如实报 500。"""
    real_replace = os.replace

    def fake_replace(src, dst, *args, **kwargs):
        if 'print_templates' not in str(dst):
            return real_replace(src, dst, *args, **kwargs)
        err = OSError(2, 'No such file or directory')
        err.winerror = 2
        raise err

    monkeypatch.setattr(os, 'replace', fake_replace)
    template_id = _upload_template(client)
    resp = _grid_write(client, template_id)
    assert resp.status_code == 500
    assert '保存失败' in resp.get_json()['msg']
