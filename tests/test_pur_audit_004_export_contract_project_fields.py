# -*- coding: utf-8 -*-
"""
PUR-AUDIT-004 回归测试：采购入库明细表与领料单明细表导出必须含合同单号、工程名称。

需求：用户要求「采购入库明细表，领料单明细表，导出采购入库明细表，
导出领料单明细表必须有合同单号，工程名称字段」。

涉及 4 个导出函数：
  - /in_order/batch_export        batch_export_in_order
  - /in_order/<id>/export        export_single_in_order
  - /out_order/export            export_out_order
  - /out_order/<id>/export       export_single_out_order

测试用例：
  T1. 采购入库单张导出含「合同单号」「工程名称」列，且行级值与 item 一致
  T2. 采购入库批量导出含「合同单号」「工程名称」列，且行级值与 item 一致
  T3. 领料单单张导出含「合同单号」「工程名称」列，且行级值与 item 一致
  T4. 领料单批量导出含「合同单号」「工程名称」列，且行级值与 item 一致
"""
from __future__ import annotations

import io
import os
import re
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")
os.environ.setdefault("WMS_SKIP_AUTO_UPDATE", "1")

import app as app_module  # noqa: E402
from app import (  # noqa: E402
    db, Warehouse, User, Material, MaterialCategory, Unit, Supplier,
    InOrder, InOrderItem, OutOrder, OutOrderItem,
)

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False

TODAY = date.today()


def _reset_db():
    db.drop_all()
    db.create_all()


def _seed_base():
    from werkzeug.security import generate_password_hash
    unit = Unit(name="个", code="PCS")
    cat = MaterialCategory(name="默认分类", code="CAT-DEFAULT")
    sup = Supplier(code="SUP001", name="供应商甲")
    wh = Warehouse(code="WHA", name="仓库A", is_default=True, status="active")
    user = User(
        username="admin",
        password_hash=generate_password_hash("admin"),
        role="admin",
        must_change_password=False,
    )
    mat = Material(
        code="M001", name="测试物料", spec="S1",
        category=cat, unit=unit, supplier=sup,
        stock=0, price=10, min_stock=0, max_stock=9999, reorder_point=0,
    )
    db.session.add_all([unit, cat, sup, wh, user, mat])
    db.session.commit()
    return {"mat": mat, "wh": wh, "sup": sup, "user": user}


def _make_client():
    client = app_module.app.test_client()
    login_page = client.get("/login").get_data(as_text=True)
    m = re.search(r'name="csrf_token".*?value="([^"]+)"', login_page)
    token = m.group(1) if m else ""
    client.post(
        "/login",
        data={"username": "admin", "password": "admin", "csrf_token": token},
    )
    return client


def _seed_in_order(order_no, material, contract_no, project_name):
    """创建一张带合同/工程字段的采购入库单（草稿）。"""
    order = InOrder(
        order_no=order_no,
        warehouse="仓库A",
        supplier_id=material.supplier_id,
        status="pending",
        date=TODAY,
    )
    db.session.add(order)
    db.session.flush()
    item = InOrderItem(
        in_order_id=order.id,
        material_id=material.id,
        quantity=10,
        price=10,
        amount=100,
        contract_no=contract_no,
        project_name=project_name,
    )
    db.session.add(item)
    db.session.commit()
    return order, item


def _seed_out_order(order_no, material, contract_no, project_name):
    """创建一张带合同/工程字段的领料单（草稿）。"""
    order = OutOrder(
        order_no=order_no,
        warehouse="仓库A",
        status="pending",
        date=TODAY,
    )
    db.session.add(order)
    db.session.flush()
    item = OutOrderItem(
        out_order_id=order.id,
        material_id=material.id,
        quantity=5,
        price=10,
        amount=50,
        contract_no=contract_no,
        project_name=project_name,
    )
    db.session.add(item)
    db.session.commit()
    return order, item


def _read_xlsx(response):
    from openpyxl import load_workbook
    assert response.status_code == 200, response.status_code
    return load_workbook(io.BytesIO(response.data))


def _find_col(header_row, label):
    for idx, val in enumerate(header_row):
        if val == label:
            return idx
    return -1


CONTRACT_HEADER_VARIANTS = ("合同单号", "合同编号")
PROJECT_HEADER_VARIANTS = ("工程名称",)


def _find_contract_col(header_row):
    for label in CONTRACT_HEADER_VARIANTS:
        idx = _find_col(header_row, label)
        if idx >= 0:
            return idx
    return -1


def _find_project_col(header_row):
    for label in PROJECT_HEADER_VARIANTS:
        idx = _find_col(header_row, label)
        if idx >= 0:
            return idx
    return -1


class TestInOrderSingleExport:
    """T1：采购入库单张导出含合同单号、工程名称。"""

    def test_single_in_order_export_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, item = _seed_in_order(
                "IN-EXPORT-001", seed["mat"], "HD260814001", "一号厂房改造"
            )
            client = _make_client()

            resp = client.get(f"/in_order/{order.id}/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"导出表头缺少合同单号/合同编号列：{header}"
            assert project_idx >= 0, f"导出表头缺少工程名称列：{header}"

            # 第一条数据行
            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814001", data_row
            assert data_row[project_idx] == "一号厂房改造", data_row


class TestInOrderBatchExport:
    """T2：采购入库批量导出含合同单号、工程名称。"""

    def test_batch_in_order_export_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order1, item1 = _seed_in_order(
                "IN-EXPORT-002", seed["mat"], "HD260814002", "二号厂房改造"
            )
            order2, item2 = _seed_in_order(
                "IN-EXPORT-003", seed["mat"], "HD260814003", "三号厂房改造"
            )
            client = _make_client()

            resp = client.post(
                "/in_order/batch_export",
                json={"ids": [order1.id, order2.id]},
            )
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"批量导出表头缺少合同单号/合同编号列：{header}"
            assert project_idx >= 0, f"批量导出表头缺少工程名称列：{header}"

            data_rows = rows[1:]
            assert len(data_rows) == 2, f"应导出 2 条明细：{data_rows}"
            values = {row[contract_idx]: row[project_idx] for row in data_rows}
            assert values.get("HD260814002") == "二号厂房改造", values
            assert values.get("HD260814003") == "三号厂房改造", values


class TestOutOrderSingleExport:
    """T3：领料单单张导出含合同单号、工程名称。"""

    def test_single_out_order_export_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, item = _seed_out_order(
                "OUT-EXPORT-001", seed["mat"], "HD260814001", "一号厂房改造"
            )
            client = _make_client()

            resp = client.get(f"/out_order/{order.id}/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"导出表头缺少合同单号/合同编号列：{header}"
            assert project_idx >= 0, f"导出表头缺少工程名称列：{header}"

            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814001", data_row
            assert data_row[project_idx] == "一号厂房改造", data_row


class TestOutOrderBatchExport:
    """T4：领料单批量导出含合同单号、工程名称。"""

    def test_batch_out_order_export_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order1, _ = _seed_out_order(
                "OUT-EXPORT-002", seed["mat"], "HD260814002", "二号厂房改造"
            )
            order2, _ = _seed_out_order(
                "OUT-EXPORT-003", seed["mat"], "HD260814003", "三号厂房改造"
            )
            client = _make_client()

            # 领料单 /out_order/export 是 GET，无 ids 入参，按列表过滤导出全部。
            resp = client.get("/out_order/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"批量导出表头缺少合同单号/合同编号列：{header}"
            assert project_idx >= 0, f"批量导出表头缺少工程名称列：{header}"

            data_rows = rows[1:]
            assert len(data_rows) >= 2, f"应导出至少 2 条明细：{data_rows}"
            values = {row[contract_idx]: row[project_idx] for row in data_rows}
            assert values.get("HD260814002") == "二号厂房改造", values
            assert values.get("HD260814003") == "三号厂房改造", values
