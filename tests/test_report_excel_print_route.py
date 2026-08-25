# -*- coding: utf-8 -*-
"""PRINT-TEMPLATE-F04（A4）回归测试：报表统一 Excel 模板打印出口。

需求（2026-08-25）：所有报表支持 Excel 在线编辑模板（参考简道云打印模板），
报表页可按所选（或默认）模板下载 .xlsx；无用户模板回退动态内置模板。

测试用例：
  T1. /report/<type>/print_excel 路由注册且需登录；未知报表类型 400
  T2. 仓库必填：未指定仓库且无默认仓库 → 400「请选择仓库」（AGENTS.md 规则）
  T3. 端到端：库存报表按动态内置模板填充（标题/列头/数据行），
      并自动把内置模板登记进 excel_print_template（幂等，模板中心可见可编辑）
  T4. /doc_print_templates/report_inventory.json 返回报表模板列表
  T5. 动态内置模板生成：空 columns 返回 None，正常 columns 通过安全校验
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

import app as m  # noqa: E402


def _client(with_warehouse=True):
    m.app.config['TESTING'] = True
    m.app.config['WTF_CSRF_ENABLED'] = False
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
        from werkzeug.security import generate_password_hash
        m.db.session.add(m.User(
            username='admin', password_hash=generate_password_hash('admin'),
            role='admin', must_change_password=False))
        if with_warehouse:
            m.db.session.add(m.Warehouse(
                code='WH001', name='材料仓', status='active', is_default=True))
            unit = m.Unit(code='JIAN', name='件')
            m.db.session.add(unit)
            m.db.session.flush()
            m.db.session.add(m.Material(code='M-001', name='螺丝',
                                        spec='M6*20', unit_id=unit.id,
                                        stock=10, price=1.5))
        m.db.session.commit()
    client = m.app.test_client()
    client.post('/login', data={'username': 'admin', 'password': 'admin'})
    return client


def test_route_registered_and_login_required():
    rules = {r.rule for r in m.app.url_map.iter_rules()}
    assert '/report/<report_type>/print_excel' in rules
    m.app.config['TESTING'] = True
    with m.app.app_context():
        m.db.drop_all()
        m.db.create_all()
    anon = m.app.test_client()
    resp = anon.get('/report/inventory/print_excel')
    assert resp.status_code in (301, 302, 401)


def test_unknown_report_type_400():
    client = _client()
    resp = client.get('/report/no_such_type/print_excel')
    assert resp.status_code == 400


def test_warehouse_required():
    client = _client(with_warehouse=False)
    resp = client.get('/report/inventory/print_excel')
    assert resp.status_code == 400
    assert resp.get_json()['msg'] == '请选择仓库'


def test_inventory_report_print_excel_e2e():
    client = _client()
    resp = client.get('/report/inventory/print_excel')
    assert resp.status_code == 200
    assert resp.headers['Content-Type'].startswith(
        'application/vnd.openxmlformats-officedocument')
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    texts = [c.value for row in ws.iter_rows() for c in row
             if c.value is not None]
    assert '库存报表' in texts
    assert 'M-001' in texts and '螺丝' in texts
    assert not any(isinstance(t, str) and '{item.' in t for t in texts)
    # 自动登记内置模板到模板中心（幂等）
    with m.app.app_context():
        rows = m.ExcelPrintTemplate.query.filter_by(
            target_type='report', target_code='report_inventory').all()
        assert len(rows) == 1
        assert rows[0].is_default is True
        assert rows[0].excel_template_path.startswith(
            '/static/uploads/print_templates/builtin_report_inventory')
    # 第二次打印不重复登记
    resp2 = client.get('/report/inventory/print_excel')
    assert resp2.status_code == 200
    with m.app.app_context():
        assert m.ExcelPrintTemplate.query.filter_by(
            target_type='report', target_code='report_inventory').count() == 1


def test_doc_print_templates_report_json():
    client = _client()
    client.get('/report/inventory/print_excel')  # 触发自动登记
    resp = client.get('/doc_print_templates/report_inventory.json')
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['status'] == 'success'
    assert data['target_type'] == 'report'
    assert len(data['templates']) == 1
    assert data['templates'][0]['edit_url'].startswith(
        '/global_print_template/')


def test_generate_report_builtin_template():
    import doc_print_excel as dpe
    from print_fill import validate_template_file
    assert dpe.generate_report_builtin_template('空', []) is None
    content = dpe.generate_report_builtin_template('库存报表', [
        {'field': 'code', 'title': '物料编码'},
        {'field': 'name', 'title': '物料名称'},
    ])
    assert content is not None
    raw = content.read()
    assert validate_template_file(raw) == ''
    wb = load_workbook(io.BytesIO(raw))
    texts = [c.value for row in wb.active.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert '{item.code}' in texts and '{item.name}' in texts


def test_report_target_code():
    import doc_print_excel as dpe
    assert dpe.report_target_code('inventory') == 'report_inventory'
    assert dpe.report_target_code('ledger') == 'report_ledger'


def test_render_report_excel_print(tmp_path, monkeypatch):
    import doc_print_excel as dpe
    monkeypatch.setattr(dpe, 'resolve_excel_template', lambda *a, **k: None)
    result = dpe.render_report_excel_print(
        'inventory', '库存报表',
        [{'field': 'code', 'title': '物料编码'},
         {'field': 'name', 'title': '物料名称'},
         {'field': 'stock', 'title': '库存数量', 'type': 'number'}],
        [{'code': 'M-001', 'name': '螺丝', 'stock': 10},
         {'code': 'M-002', 'name': '螺母', 'stock': 20}],
        static_folder=str(tmp_path), date_str='2026-08-25')
    assert result is not None
    output, filename, template_path = result
    assert '库存报表' in filename
    assert template_path.endswith('builtin_report_inventory_default.xlsx')
    wb = load_workbook(output)
    texts = [c.value for row in wb.active.iter_rows() for c in row
             if c.value is not None]
    assert '库存报表' in texts
    assert 'M-001' in texts and 'M-002' in texts
    assert 10 in texts and 20 in texts  # dict 行数值类型保留
    assert not any(isinstance(t, str) and '{item.' in t for t in texts)
    # 再次调用复用已生成的内置模板文件（幂等）
    result2 = dpe.render_report_excel_print(
        'inventory', '库存报表',
        [{'field': 'code', 'title': '物料编码'}], [],
        static_folder=str(tmp_path), date_str='2026-08-25')
    assert result2 is not None
