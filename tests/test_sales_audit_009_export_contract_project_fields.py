# -*- coding: utf-8 -*-
"""SALES-AUDIT-009 回归测试：销售导出/打印补齐合同单号与工程名称。

需求（与 PUR-AUDIT-004 一致）：销售订单导出必须有合同单号、工程名称字段。
SalesOrder 模型已有 contract_no/project_name/project_no 字段，列表页与
详情页已展示，但导出/打印环节遗漏。

涉及 6 个导出/打印函数（4 个 P1 必修 + 2 个 P2 可选补充）：
  P1：
    - /sales/export              export_sales_orders
    - /sales_report/export       export_sales_report
    - /sales/execution_report/export  export_sales_execution_report
    - /sales/<id>/print          print_sales_order（info 含合同编号/工程名称）
  P2：
    - /sales/outflow_report/export   export_sales_outflow_report
    - /sales/outbound/export         export_sales_outbound

测试用例：
  T1. export_sales_orders 含合同单号、工程名称列，且行级值与订单一致
  T2. export_sales_report 含合同单号、工程名称列，且行级值与订单一致
  T3. export_sales_execution_report 含合同单号、工程名称列，且行级值与订单一致
  T4. print_sales_order 渲染的 HTML 含「合同编号」「工程名称」及订单值
  T5. export_sales_outflow_report 含合同单号、工程名称列，来自关联销售订单
  T6. export_sales_outbound 含合同单号、工程名称列，来自关联销售订单
"""
from __future__ import annotations

import io
import os
import re
import sys
from datetime import date
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")
os.environ.setdefault("WMS_SKIP_AUTO_UPDATE", "1")

from werkzeug.security import generate_password_hash  # noqa: E402

import app as app_module  # noqa: E402
from app import (  # noqa: E402
    db, Warehouse, User, Material, MaterialCategory, Unit, Supplier,
    Customer, SalesOrder, SalesOrderItem, OutOrder, OutOrderItem,
)

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False

TODAY = date.today()


def _reset_db():
    db.drop_all()
    db.create_all()


def _seed_base():
    unit = Unit(name="个", code="PCS")
    cat = MaterialCategory(name="默认分类", code="CAT-DEFAULT")
    sup = Supplier(code="SUP001", name="供应商甲")
    wh = Warehouse(code="WHA", name="仓库A", is_default=True, status="active")
    user = User(
        username="admin",
        password_hash=generate_password_hash("admin"),
        role="admin",
        must_change_password=False,
    )
    cust = Customer(code="C001", name="测试客户")
    mat = Material(
        code="M001", name="测试物料", spec="S1",
        category=cat, unit=unit, supplier=sup,
        stock=100, price=10, min_stock=0, max_stock=9999, reorder_point=0,
    )
    db.session.add_all([unit, cat, sup, wh, user, cust, mat])
    db.session.commit()
    return {"mat": mat, "wh": wh, "cust": cust, "user": user}


def _make_client():
    client = app_module.app.test_client()
    login_page = client.get("/login").get_data(as_text=True)
    m = re.search(r'name="csrf_token".*?value="([^"]+)"', login_page)
    token = m.group(1) if m else ""
    client.post(
        "/login",
        data={"username": "admin", "password": "admin", "csrf_token": token},
    )
    return client


def _seed_sales_order(order_no, customer, material, contract_no, project_name,
                     project_no="PROJ-001", status="confirmed"):
    """创建一张带合同/工程字段的销售订单。"""
    order = SalesOrder(
        order_no=order_no,
        customer_id=customer.id,
        warehouse="仓库A",
        date=TODAY,
        status=status,
        shipment_status="pending",
        contract_no=contract_no,
        project_name=project_name,
        project_no=project_no,
        untaxed_amount=100,
        tax_amount=13,
        total_amount=113,
    )
    db.session.add(order)
    db.session.flush()
    item = SalesOrderItem(
        sales_order_id=order.id,
        material_id=material.id,
        quantity=10,
        shipped_quantity=0,
        price=10,
        amount=100,
        tax_rate=0.13,
    )
    db.session.add(item)
    db.session.commit()
    return order, item


def _seed_sales_outbound(order_no, sales_order, material, qty=5):
    """创建一张销售出库单（草稿），关联到销售订单。"""
    outbound = OutOrder(
        order_no=order_no,
        warehouse="仓库A",
        business_type="销售出库",
        source_sales_order_id=sales_order.id,
        customer=sales_order.customer.name if sales_order.customer else "",
        status="pending",
        date=TODAY,
    )
    db.session.add(outbound)
    db.session.flush()
    item = OutOrderItem(
        out_order_id=outbound.id,
        material_id=material.id,
        quantity=qty,
        price=10,
        amount=qty * 10,
    )
    db.session.add(item)
    db.session.commit()
    return outbound


def _read_xlsx(response):
    from openpyxl import load_workbook
    assert response.status_code == 200, (
        f"导出失败 status={response.status_code} body={response.get_data(as_text=True)[:200]}"
    )
    return load_workbook(io.BytesIO(response.data))


def _find_col(header_row, label):
    for idx, val in enumerate(header_row):
        if val == label:
            return idx
    return -1


def _find_contract_col(header_row):
    """支持「合同单号」「合同编号」两种命名。"""
    for label in ("合同单号", "合同编号"):
        idx = _find_col(header_row, label)
        if idx >= 0:
            return idx
    return -1


def _find_project_col(header_row):
    return _find_col(header_row, "工程名称")


class TestExportSalesOrders:
    """T1：export_sales_orders 含合同单号、工程名称列。"""

    def test_export_sales_orders_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, _ = _seed_sales_order(
                "SO-EXP-001", seed["cust"], seed["mat"],
                "HD260814001", "一号厂房改造",
            )
            client = _make_client()
            resp = client.get("/sales/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"表头缺少合同单号列：{header}"
            assert project_idx >= 0, f"表头缺少工程名称列：{header}"
            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814001", data_row
            assert data_row[project_idx] == "一号厂房改造", data_row


class TestExportSalesReport:
    """T2：export_sales_report 含合同单号、工程名称列。"""

    def test_export_sales_report_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, _ = _seed_sales_order(
                "SO-EXP-002", seed["cust"], seed["mat"],
                "HD260814002", "二号厂房改造",
            )
            client = _make_client()
            resp = client.get("/sales/report/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"表头缺少合同单号列：{header}"
            assert project_idx >= 0, f"表头缺少工程名称列：{header}"
            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814002", data_row
            assert data_row[project_idx] == "二号厂房改造", data_row


class TestExportSalesExecutionReport:
    """T3：export_sales_execution_report 含合同单号、工程名称列。"""

    def test_export_sales_execution_report_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, _ = _seed_sales_order(
                "SO-EXP-003", seed["cust"], seed["mat"],
                "HD260814003", "三号厂房改造",
            )
            client = _make_client()
            resp = client.get("/sales/execution_report/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"表头缺少合同单号列：{header}"
            assert project_idx >= 0, f"表头缺少工程名称列：{header}"
            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814003", data_row
            assert data_row[project_idx] == "三号厂房改造", data_row


class TestPrintSalesOrder:
    """T4：print_sales_order 渲染的 HTML 含合同编号、工程名称。"""

    def test_print_sales_order_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, _ = _seed_sales_order(
                "SO-PRINT-001", seed["cust"], seed["mat"],
                "HD260814004", "四号厂房改造",
            )
            client = _make_client()
            resp = client.get(f"/sales/{order.id}/print")
            assert resp.status_code == 200, resp.status_code
            html = resp.get_data(as_text=True)
            # info 列表渲染为「合同编号」「工程名称」
            assert "合同编号" in html, "打印模板缺少「合同编号」标签"
            assert "工程名称" in html, "打印模板缺少「工程名称」标签"
            assert "HD260814004" in html, "打印模板缺少合同编号值"
            assert "四号厂房改造" in html, "打印模板缺少工程名称值"


class TestExportSalesOutflowReport:
    """T5：export_sales_outflow_report 含合同单号、工程名称列（来自关联销售订单）。"""

    def test_export_sales_outflow_report_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, _ = _seed_sales_order(
                "SO-OUTFLOW-001", seed["cust"], seed["mat"],
                "HD260814005", "五号厂房改造",
            )
            outbound = _seed_sales_outbound("OUT-OUTFLOW-001", order, seed["mat"])
            client = _make_client()
            resp = client.get("/sales/outflow_report/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"表头缺少合同单号列：{header}"
            assert project_idx >= 0, f"表头缺少工程名称列：{header}"
            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814005", data_row
            assert data_row[project_idx] == "五号厂房改造", data_row


class TestExportSalesOutbound:
    """T6：export_sales_outbound 含合同单号、工程名称列（来自关联销售订单）。"""

    def test_export_sales_outbound_has_contract_and_project(self):
        with app_module.app.app_context():
            _reset_db()
            seed = _seed_base()
            order, _ = _seed_sales_order(
                "SO-OUTBOUND-001", seed["cust"], seed["mat"],
                "HD260814006", "六号厂房改造",
            )
            outbound = _seed_sales_outbound("OUT-EXP-001", order, seed["mat"])
            client = _make_client()
            resp = client.get("/sales/outbound/export")
            wb = _read_xlsx(resp)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows, "导出内容不应为空"
            header = rows[0]
            contract_idx = _find_contract_col(header)
            project_idx = _find_project_col(header)
            assert contract_idx >= 0, f"表头缺少合同单号列：{header}"
            assert project_idx >= 0, f"表头缺少工程名称列：{header}"
            data_row = rows[1]
            assert data_row[contract_idx] == "HD260814006", data_row
            assert data_row[project_idx] == "六号厂房改造", data_row
