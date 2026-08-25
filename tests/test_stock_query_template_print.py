# -*- coding: utf-8 -*-
"""PRINT-TEMPLATE-F04（A6）回归测试：库存查询列表接入 Excel 模板打印。

需求（2026-08-25）：列表打印也支持 Excel 在线编辑模板（参考简道云），
/report/stock/print 在显式 template_id 或存在默认模板记录时按模板渲染；
无模板记录时保持 BUG-2026-08-12-005 的硬编码导出行为不变（含仓库隔离）。

测试用例：
  T1. 无模板记录：回退硬编码导出（表头含 当前库存，文件名 stock_report.xlsx）
  T2. 显式 template_id：按模板渲染（含模板标题与 库存金额 列，仓库级库存）
  T3. 存在默认模板记录（无 template_id）：自动按模板渲染
  T4. 模板渲染同样遵守仓库必填（无仓库 400）
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")

import app as app_module  # noqa: E402
from app import db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False


def _client():
    with app_module.app.app_context():
        db.drop_all()
        db.create_all()
        from werkzeug.security import generate_password_hash
        from app import Material, Unit, Warehouse
        db.session.add(app_module.User(
            username="admin",
            password_hash=generate_password_hash("admin"),
            role="admin", must_change_password=False))
        db.session.add(Warehouse(code="WHA", name="甲仓", status="active",
                                 is_default=True))
        unit = Unit(code="JIAN", name="件")
        db.session.add(unit)
        db.session.flush()
        db.session.add(Material(code="M-001", name="螺丝", spec="M6*20",
                                unit_id=unit.id, stock=99, price=2.0))
        db.session.commit()
    client = app_module.app.test_client()
    client.post("/login", data={"username": "admin", "password": "admin"})
    return client


def _add_stock_template(name="系统默认库存查询列表模板", is_default=True):
    """登记一条 stock_query 模板记录（指向内置模板文件），返回模板 id。"""
    import doc_print_excel as dpe
    path = dpe._builtin_template_abspath(app_module.app.static_folder,
                                         "stock_query")
    tpl = app_module.ExcelPrintTemplate(
        name=name, target_type="list", target_code="stock_query",
        template_type="excel",
        excel_template_path="/static/uploads/print_templates/"
                            + os.path.basename(path),
        is_default=is_default)
    db.session.add(tpl)
    db.session.commit()
    return tpl.id


def test_legacy_export_without_template():
    client = _client()
    resp = client.get("/report/stock/print?warehouse_id=1")
    assert resp.status_code == 200
    assert "stock_report.xlsx" in resp.headers.get("Content-Disposition", "")
    wb = load_workbook(io.BytesIO(resp.data))
    first_row = [c.value for c in wb.active[1]]
    assert "当前库存" in first_row  # 硬编码表头


def test_template_print_with_template_id():
    client = _client()
    with app_module.app.app_context():
        template_id = _add_stock_template()
    resp = client.get(
        f"/report/stock/print?warehouse_id=1&template_id={template_id}")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    texts = [c.value for row in wb.active.iter_rows() for c in row
             if c.value is not None]
    assert "库存查询列表" in texts  # 模板标题（硬编码导出无此行）
    assert "库存金额" in texts
    assert "M-001" in texts and "螺丝" in texts
    assert "甲仓" in texts
    assert not any(isinstance(t, str) and "{item." in t for t in texts)


def test_template_print_with_default_template_record():
    client = _client()
    with app_module.app.app_context():
        _add_stock_template()
    resp = client.get("/report/stock/print?warehouse_id=1")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    texts = [c.value for row in wb.active.iter_rows() for c in row
             if c.value is not None]
    assert "库存查询列表" in texts


def test_template_print_warehouse_required():
    m = app_module
    m.app.config["TESTING"] = True
    with m.app.app_context():
        db.drop_all()
        db.create_all()
        from werkzeug.security import generate_password_hash
        db.session.add(m.User(
            username="admin",
            password_hash=generate_password_hash("admin"),
            role="admin", must_change_password=False))
        db.session.commit()
        _add_stock_template()
    client = m.app.test_client()
    client.post("/login", data={"username": "admin", "password": "admin"})
    resp = client.get("/report/stock/print")
    assert resp.status_code == 400
