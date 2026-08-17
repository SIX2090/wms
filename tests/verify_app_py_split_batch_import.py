# -*- coding: utf-8 -*-
"""
app.py 拆分回归测试：批量导入（batch_import）域路由迁移到 routes/batch_import.py。

register-on-app 模式（register_batch_import_routes(app)），endpoint 名与 URL 不变。

验收点：
P1. 核心 endpoint 已注册，且无 batch_import.xxx 前缀重复。
P2. GET /batch_import 返回 200。
P3. /import/out_order 未上传文件时返回 api_error（JSON）。
P4. 基础资料 import/export stub 路由存在并跳转（3xx）到 /batch_import。
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")

import app as app_module  # noqa: E402
from app import db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False

ENDPOINTS = [
    "import_out_order",
    "import_in_order",
    "batch_import_page",
    "user_import_stub",
    "user_export_stub",
    "label_template_import_stub",
    "label_template_export_stub",
    "opening_stock_import_stub",
    "opening_stock_export_stub",
]


def _reset_db():
    db.drop_all()
    db.create_all()


def _seed_admin():
    from werkzeug.security import generate_password_hash
    from app import User
    u = User(username="admin", password_hash=generate_password_hash("admin"),
             role="admin", must_change_password=False)
    db.session.add(u)
    db.session.commit()


def _make_client():
    return app_module.app.test_client()


def _login(client):
    return client.post(
        "/login",
        data={"username": "admin", "password": "admin"},
        content_type="application/x-www-form-urlencoded",
    )


def _setup():
    with app_module.app.app_context():
        _reset_db()
        _seed_admin()
    client = _make_client()
    _login(client)
    return client


def test_endpoints_registered():
    rules = {r.endpoint for r in app_module.app.url_map.iter_rules()}
    for ep in ENDPOINTS:
        assert ep in rules, f"endpoint {ep} 未注册"
    # register-on-app 模式不应产生 batch_import.xxx 前缀的 endpoint
    assert not any(ep.startswith("batch_import.") for ep in rules)


def test_batch_import_page_returns_200():
    client = _setup()
    resp = client.get("/batch_import")
    assert resp.status_code == 200, resp.status_code


def test_import_out_order_no_file_returns_error():
    client = _setup()
    resp = client.post(
        "/import/out_order",
        data={},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert resp.status_code == 400, resp.status_code
    payload = resp.get_json()
    assert payload is not None
    assert payload.get("status") == "error"


def test_import_in_order_no_file_returns_error():
    client = _setup()
    resp = client.post(
        "/import/in_order",
        data={},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert resp.status_code == 400, resp.status_code
    payload = resp.get_json()
    assert payload is not None
    assert payload.get("status") == "error"


def test_import_out_order_valid_excel_succeeds():
    client = _setup()
    # 构造合法领料单 Excel（含基础数据：部门、物料、单位、出库单）
    from app import Department, Material, OutOrder, OutOrderItem, Unit
    with app_module.app.app_context():
        dept = Department(code="D-001", name="生产车间")
        db.session.add(dept)
        db.session.commit()

        mat = Material(code="MAT001", name="示例物料", spec="规格A")
        unit = Unit(code="个", name="个")
        mat.unit = unit
        db.session.add(mat)
        db.session.add(unit)
        db.session.commit()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "领料单"
    ws.append(["单据编号", "日期", "用途", "部门", "物料编码", "物料名称", "规格", "单位", "数量", "单价", "金额", "备注"])
    ws.append(["CK20240101001", "2024-01-01", "领料单", "生产车间", "MAT001", "示例物料", "规格A", "个", "10", "5.00", "50.00", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    data = {"file": (buf, "out_order.xlsx")}
    resp = client.post(
        "/import/out_order",
        data=data,
        content_type="multipart/form-data",
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload is not None
    assert payload.get("status") == "success", payload
    assert payload.get("count") == 1, payload


def test_import_in_order_valid_excel_succeeds():
    client = _setup()
    # 构造合法入库单 Excel（含基础数据：供应商、物料、单位、入库单）
    from app import InOrder, InOrderItem, Material, Supplier, Unit, Customer, Warehouse
    with app_module.app.app_context():
        sup = Supplier(code="SUP-001", name="示例供应商")
        customer = Customer(code="CUS-001", name="示例客户")
        warehouse = Warehouse(code="WH-001", name="一号仓库", is_default=True, status="active")
        db.session.add_all([sup, customer, warehouse])
        db.session.commit()

        mat = Material(code="MAT002", name="入库物料", spec="规格B")
        unit = Unit(code="个", name="个")
        mat.unit = unit
        db.session.add(mat)
        db.session.add(unit)
        db.session.commit()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "入库单"
    ws.append(["单据编号", "日期", "业务类型", "用途", "仓库", "供应商", "客户", "物料编码", "物料名称", "规格", "单位", "数量", "单价", "金额", "客供", "合同编号", "工程名称", "备注"])
    ws.append(["RK20240101001", "2024-01-01", "采购入库", "采购到货", "一号仓库", "示例供应商", "", "MAT002", "入库物料", "规格B", "个", "100", "10.00", "1000.00", "否", "HT-001", "项目A", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    data = {"file": (buf, "in_order.xlsx")}
    resp = client.post(
        "/import/in_order",
        data=data,
        content_type="multipart/form-data",
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload is not None
    assert payload.get("status") == "success", payload
    assert payload.get("count") == 1, payload
    with app_module.app.app_context():
        order = InOrder.query.filter_by(order_no="RK20240101001").one()
        item = InOrderItem.query.filter_by(in_order_id=order.id).one()
        assert order.business_type == "采购入库"
        assert order.warehouse == "一号仓库"
        assert order.supplier.name == "示例供应商"
        assert order.contract_no == "HT-001"
        assert order.project_name == "项目A"
        assert item.is_customer_supplied is False


def test_import_in_order_other_type_preserves_customer_and_customer_supplied():
    client = _setup()
    from app import Customer, InOrder, InOrderItem, Material, Unit, Warehouse
    with app_module.app.app_context():
        customer = Customer(code="CUS-002", name="客供客户")
        warehouse = Warehouse(code="WH-002", name="二号仓库", is_default=True, status="active")
        unit = Unit(code="个", name="个")
        material = Material(code="MAT003", name="客供物料", spec="规格C", unit=unit)
        db.session.add_all([customer, warehouse, unit, material])
        db.session.commit()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["单据编号", "日期", "业务类型", "用途", "仓库", "供应商", "客户", "物料编码", "数量", "单价", "客供"])
    ws.append(["OI20240101001", "2024-01-01", "其他入库", "客供料", "二号仓库", "", "客供客户", "MAT003", "5", "2", "是"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = client.post("/import/in_order", data={"file": (buf, "other_in_order.xlsx")}, content_type="multipart/form-data", headers={"X-Requested-With": "XMLHttpRequest"})
    assert resp.status_code == 200, resp.get_data(as_text=True)
    with app_module.app.app_context():
        order = InOrder.query.filter_by(order_no="OI20240101001").one()
        item = InOrderItem.query.filter_by(in_order_id=order.id).one()
        assert order.business_type == "其他入库"
        assert order.customer.name == "客供客户"
        assert order.supplier_id is None
        assert item.is_customer_supplied is True


def test_stub_routes_redirect():
    client = _setup()
    for url in ["/user/export", "/label_template/export", "/opening_stock/export"]:
        resp = client.get(url)
        assert resp.status_code in (301, 302, 303, 307, 308), f"{url} -> {resp.status_code}"
        assert "/batch_import" in resp.headers.get("Location", ""), f"{url} -> {resp.headers.get('Location')}"
    for url in ["/user/import", "/label_template/import", "/opening_stock/import"]:
        resp = client.post(url)
        assert resp.status_code in (301, 302, 303, 307, 308), f"{url} -> {resp.status_code}"
        assert "/batch_import" in resp.headers.get("Location", ""), f"{url} -> {resp.headers.get('Location')}"