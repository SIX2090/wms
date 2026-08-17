# -*- coding: utf-8 -*-
"""
app.py 拆分回归测试：导出（export）域路由迁移到 routes/export.py。

register-on-app 模式（register_export_routes(app)），endpoint 名与 URL 不变。

验收点：
P1. 核心 endpoint 已注册，且无 export.xxx 前缀重复。
P2. URL 路径保持不变（/export/... 与 /xxx/export 均在）。
P3. 模板导出（/export/template/*）返回可下载的 xlsx（200）。
P4. 数据导出（/export/in_order 等）在登录后返回 xlsx（200）。
"""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))
os.chdir(APP_DIR)

os.environ.setdefault("WMS_BOOTSTRAP_PASSWORD", "admin")
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["WMS_DATABASE_URI"] = "sqlite:///:memory:"
os.environ.setdefault("WMS_DEBUG", "0")

import app as app_module  # noqa: E402
from app import db  # noqa: E402

app_module.app.config["TESTING"] = True
app_module.app.config["WTF_CSRF_ENABLED"] = False

ENDPOINTS = [
    "export_bom_template",
    "export_requisition_template",
    "export_subcontract_template",
    "export_subcontract_issue_template",
    "export_subcontract_receive_template",
    "export_adjustment_template",
    "export_check_template",
    "export_purchase_request_template",
    "export_purchase_order_template",
    "export_in_order",
    "export_purchase_request",
    "export_after_sale_out",
    "export_material_template",
    "export_in_order_template",
    "export_out_order_template",
]

TEMPLATE_URLS = [
    "/export/template/bom",
    "/export/template/requisition",
    "/export/template/subcontract",
    "/export/template/subcontract_issue",
    "/export/template/subcontract_receive",
    "/export/template/adjustment",
    "/export/template/check",
    "/export/template/purchase_request",
    "/export/template/purchase_order",
    "/export/template/material",
    "/export/template/in_order",
    "/export/template/out_order",
]

# 数据导出同时保留备用 URL（/xxx/export 兼容路由）
ALIAS_URLS = [
    "/export/in_order",
    "/in_order/export",
    "/export/purchase_request",
    "/purchase_request/export",
    "/export/after_sale_out",
    "/after_sale_out/export",
]


def _reset_db():
    db.drop_all()
    db.create_all()


def _seed_admin():
    from werkzeug.security import generate_password_hash
    from app import User, Warehouse
    db.session.add(Warehouse(code="WH-001", name="一号仓库", is_default=True, status="active"))
    u = User(username="admin", password_hash=generate_password_hash("admin"),
             role="admin", must_change_password=False)
    db.session.add(u)
    db.session.commit()


def _make_client():
    return app_module.app.test_client()


def _login(client):
    return client.post(
        "/login",
        data={"username": "admin", "password": "admin"},
        content_type="application/x-www-form-urlencoded",
    )


def _setup():
    with app_module.app.app_context():
        _reset_db()
        _seed_admin()
    client = _make_client()
    _login(client)
    return client


def test_endpoints_registered():
    rules = {r.endpoint for r in app_module.app.url_map.iter_rules()}
    for ep in ENDPOINTS:
        assert ep in rules, f"endpoint {ep} 未注册"
    # register-on-app 模式不应产生 export.xxx 前缀的 endpoint
    assert not any(ep.startswith("export.") for ep in rules)


def test_template_urls_return_xlsx():
    client = _setup()
    for url in TEMPLATE_URLS:
        resp = client.get(url)
        assert resp.status_code == 200, f"{url} -> {resp.status_code}"
        assert "application" in resp.content_type or "octet-stream" in resp.content_type, f"{url} content_type={resp.content_type}"


def test_alias_urls_resolve():
    client = _setup()
    for url in ALIAS_URLS:
        resp = client.get(url)
        assert resp.status_code == 200, f"{url} -> {resp.status_code}"


def test_export_in_order_returns_xlsx():
    client = _setup()
    resp = client.get("/export/in_order")
    assert resp.status_code == 200
    assert "octet-stream" in resp.content_type or "xlsx" in resp.content_type or "spreadsheet" in resp.content_type


def test_other_in_export_filters_business_type_and_preserves_customer_fields():
    from datetime import date
    from openpyxl import load_workbook
    from app import Customer, InOrder, InOrderItem, Material, Supplier

    client = _setup()
    with app_module.app.app_context():
        supplier = Supplier(code="SUP-001", name="供应商甲")
        customer = Customer(code="CUS-001", name="客户甲")
        material = Material(code="MAT-001", name="物料甲", stock=0, price=1)
        db.session.add_all([supplier, customer, material])
        db.session.flush()
        purchase = InOrder(order_no="IN-PURCHASE", date=date.today(), warehouse="一号仓库", business_type="采购入库", supplier_id=supplier.id, status="pending")
        other = InOrder(order_no="IN-OTHER", date=date.today(), warehouse="一号仓库", business_type="其他入库", customer_id=customer.id, location="A-01", status="pending")
        db.session.add_all([purchase, other])
        db.session.flush()
        db.session.add_all([
            InOrderItem(in_order_id=purchase.id, material_id=material.id, quantity=1, price=1, amount=1),
            InOrderItem(in_order_id=other.id, material_id=material.id, quantity=2, price=1, amount=2, is_customer_supplied=True),
        ])
        db.session.commit()

    response = client.get("/in_order/export?warehouse_id=1&type=other_in")
    assert response.status_code == 200, response.get_data(as_text=True)
    rows = list(load_workbook(io.BytesIO(response.data)).active.iter_rows(values_only=True))
    header, exported = rows[0], rows[1:]
    order_index = header.index("单据编号")
    customer_index = header.index("客户")
    supplied_index = header.index("客供")
    assert [row[order_index] for row in exported] == ["IN-OTHER"]
    assert exported[0][customer_index] == "客户甲"
    assert exported[0][supplied_index] == "是"